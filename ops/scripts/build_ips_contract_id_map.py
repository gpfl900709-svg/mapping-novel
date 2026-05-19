from __future__ import annotations

import argparse
import csv
import json
import re
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

from ips.core.auth import load_site_credentials, perform_api_login, resolve_env_path
from ips.sites import get_site
from ips_title_rules import resolve_default_ips_export_path


PROJECT_ROOT = Path(__file__).resolve().parents[1] / "SIAAN Project"
DEFAULT_SHEET_NAME = "콘텐츠 목록"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output" / "ips_contract_id_map"


@dataclass(frozen=True)
class ContentTarget:
    cid: str
    content_name: str
    manager_name: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read the latest IPS 콘텐츠 목록 workbook and map CID -> 계약ID "
            "through the same KIPM API used by [7021] 콘텐츠상세 > 정산정보 > 계약목록."
        ),
    )
    parser.add_argument(
        "--ips-export",
        default=str(resolve_default_ips_export_path()),
        help="Path to the IPS 콘텐츠 목록 workbook.",
    )
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_SHEET_NAME,
        help="Workbook sheet name to read.",
    )
    parser.add_argument(
        "--manager",
        default="",
        help="Optional 담당자명 filter from the workbook.",
    )
    parser.add_argument(
        "--cid",
        action="append",
        default=[],
        help="Optional CID filter. Repeatable.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional max CID count after filters are applied.",
    )
    parser.add_argument(
        "--output",
        default="",
        help="Optional flat CSV output path.",
    )
    parser.add_argument(
        "--json-output",
        default="",
        help="Optional nested JSON output path.",
    )
    parser.add_argument(
        "--env-file",
        default="",
        help="Optional env file path for IPS credentials.",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=100,
        help="Page size to request from the 계약목록 API.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Resolve mappings and print the summary without writing files.",
    )
    return parser.parse_args()


def safe_suffix(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z가-힣]+", "_", str(value or "").strip()).strip("_")
    return cleaned or "all"


def build_output_paths(args: argparse.Namespace) -> tuple[Path | None, Path | None]:
    csv_output = Path(args.output) if args.output else None
    json_output = Path(args.json_output) if args.json_output else None

    if csv_output is None and json_output is None:
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = safe_suffix(args.manager) if args.manager else "all"
        csv_output = DEFAULT_OUTPUT_DIR / f"{stamp}__ips_contract_id_map_{suffix}.csv"
        json_output = DEFAULT_OUTPUT_DIR / f"{stamp}__ips_contract_id_map_{suffix}.json"
    elif csv_output is not None and json_output is None and csv_output.suffix.lower() == ".csv":
        json_output = csv_output.with_suffix(".json")

    return csv_output, json_output


def read_workbook_targets(path: Path, *, sheet_name: str) -> list[ContentTarget]:
    warnings.filterwarnings("ignore", message="Workbook contains no default style")
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    header_index = {header: index for index, header in enumerate(headers)}

    required_columns = ("콘텐츠ID", "콘텐츠명", "담당자명")
    missing = [column for column in required_columns if column not in header_index]
    if missing:
        raise KeyError(f"필수 컬럼이 없습니다: {', '.join(missing)}")

    seen: set[str] = set()
    targets: list[ContentTarget] = []
    for values in rows:
        cid = str(values[header_index["콘텐츠ID"]] or "").strip()
        if not cid or cid in seen:
            continue
        seen.add(cid)
        targets.append(
            ContentTarget(
                cid=cid,
                content_name=str(values[header_index["콘텐츠명"]] or "").strip(),
                manager_name=str(values[header_index["담당자명"]] or "").strip(),
            ),
        )
    return targets


def filter_targets(
    targets: list[ContentTarget],
    *,
    manager_name: str,
    selected_cids: list[str],
    limit: int,
) -> tuple[list[ContentTarget], list[str]]:
    manager_name = str(manager_name or "").strip()
    selected = {str(cid or "").strip() for cid in selected_cids if str(cid or "").strip()}

    filtered: list[ContentTarget] = []
    seen_cids: set[str] = set()
    for target in targets:
        if manager_name and target.manager_name != manager_name:
            continue
        if selected and target.cid not in selected:
            continue
        if target.cid in seen_cids:
            continue
        seen_cids.add(target.cid)
        filtered.append(target)
        if limit and len(filtered) >= limit:
            break

    unresolved = sorted(selected - seen_cids) if selected else []
    for cid in unresolved:
        filtered.append(ContentTarget(cid=cid, content_name="", manager_name=manager_name))
        if limit and len(filtered) >= limit:
            break

    still_missing = sorted(selected - {target.cid for target in filtered}) if selected else []
    return filtered, still_missing


def build_api_url(base_url: str, path: str) -> str:
    normalized_path = path if path.startswith("/") else f"/{path}"
    return f"{base_url.rstrip('/')}{normalized_path}"


def build_authenticated_session(env_file: str) -> tuple[Any, requests.Session]:
    site = get_site("kipm")
    env_path = resolve_env_path(env_file)
    credentials = load_site_credentials(
        env_path,
        username_keys=site.username_env_keys,
        password_keys=site.password_env_keys,
        company_code=site.company_code,
    )
    auth_result = perform_api_login(
        api_base_url=site.api_base_url,
        login_path=site.api_login_path,
        user_info_path=site.api_user_info_path,
        credentials=credentials,
    )

    session = requests.Session()
    session.headers.update(
        {
            "Authorization": f"Bearer {auth_result.token}",
            "Accept": "application/json, text/plain, */*",
            "X-Requested-With": "XMLHttpRequest",
        },
    )
    return site, session


def extract_response_data(response: requests.Response, *, context: str) -> Any:
    if not response.ok:
        raise RuntimeError(f"{context} 실패: HTTP {response.status_code} {response.text[:300]}")

    payload = response.json()
    if not isinstance(payload, dict):
        raise RuntimeError(f"{context} 응답 형식이 예상과 다릅니다: {type(payload).__name__}")

    code = str(payload.get("code") or "").strip()
    if code and code not in {"REQUEST_SUCCESS", "SUCCESS"}:
        raise RuntimeError(
            f"{context} 실패: code={code} message={payload.get('message')!r}",
        )
    return payload.get("data")


def fetch_contract_rows(
    session: requests.Session,
    *,
    api_base_url: str,
    cid: str,
    page_size: int,
) -> tuple[list[dict[str, Any]], int]:
    response = session.get(
        build_api_url(api_base_url, "cntsd/cntslt/ctns-list/cntr-list"),
        params={
            "srcCtnsId": cid,
            "srchCntrId": "",
            "pageNum": 1,
            "pageSize": page_size,
        },
        timeout=30,
    )
    data = extract_response_data(response, context=f"계약목록 조회 cid={cid}")
    if not isinstance(data, dict):
        return [], 0

    rows = list(data.get("list") or [])
    total = int(data.get("total") or len(rows))
    if total > len(rows):
        retry_response = session.get(
            build_api_url(api_base_url, "cntsd/cntslt/ctns-list/cntr-list"),
            params={
                "srcCtnsId": cid,
                "srchCntrId": "",
                "pageNum": 1,
                "pageSize": total,
            },
            timeout=30,
        )
        retry_data = extract_response_data(retry_response, context=f"계약목록 재조회 cid={cid}")
        if isinstance(retry_data, dict):
            rows = list(retry_data.get("list") or [])
            total = int(retry_data.get("total") or len(rows))
    return rows, total


def fetch_main_contract_id(
    session: requests.Session,
    *,
    api_base_url: str,
    contract_id: str,
    cache: dict[str, str],
) -> str:
    if not contract_id:
        return ""
    cached = cache.get(contract_id)
    if cached is not None:
        return cached

    response = session.get(
        build_api_url(api_base_url, f"cntsd/cntslt/ctns-list/get-main-cntrId/{contract_id}"),
        timeout=30,
    )
    data = extract_response_data(response, context=f"대표 계약ID 조회 cntrId={contract_id}")
    main_contract_id = str(data or "").strip()
    cache[contract_id] = main_contract_id
    return main_contract_id


def contract_change_url(site: Any, contract_id: str) -> str:
    if not contract_id:
        return ""
    return site.resolve_url(f"/ip/cntr/cntrchg/cntr-chg-reg?cntrId={contract_id}")


def build_flat_row(
    site: Any,
    target: ContentTarget,
    *,
    contract_count: int,
    lookup_status: str,
    contract: dict[str, Any] | None,
    lookup_error: str = "",
    main_contract_id: str = "",
) -> dict[str, Any]:
    contract = contract or {}
    visible_contract_id = str(contract.get("cntrId") or "").strip()
    unity_contract_id = str(contract.get("unityCntrId") or "").strip()
    lookup_contract_id = main_contract_id or visible_contract_id
    return {
        "cid": target.cid,
        "content_name": target.content_name,
        "manager_name": target.manager_name,
        "contract_count": contract_count,
        "lookup_status": lookup_status,
        "lookup_error": lookup_error,
        "visible_contract_id": visible_contract_id,
        "unity_contract_id": unity_contract_id,
        "main_contract_id": main_contract_id,
        "lookup_contract_id": lookup_contract_id,
        "contract_change_url": contract_change_url(site, lookup_contract_id),
        "contract_name": str(contract.get("cntrNm") or "").strip(),
        "contract_kind": str(contract.get("cntrKndNm") or "").strip(),
        "contract_kind_code": str(contract.get("cntrKndCd") or "").strip(),
        "is_primary_contract": str(contract.get("ocntrYn") or "").strip(),
        "counterparty_name": str(contract.get("cntrPtnNm") or "").strip(),
        "related_contract_name": str(contract.get("relCntrNm") or "").strip(),
        "contract_date": str(contract.get("cntrCclsDt") or "").strip(),
        "contract_status": str(contract.get("cntrStsNm") or "").strip(),
        "attachment_name": str(contract.get("atchfOrgNm") or "").strip(),
        "contract_manager_name": str(contract.get("chgerNm") or "").strip(),
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key in seen:
                continue
            seen.add(key)
            fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_report(args: argparse.Namespace) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    targets = read_workbook_targets(Path(args.ips_export), sheet_name=args.sheet_name)
    selected_targets, missing_cids = filter_targets(
        targets,
        manager_name=args.manager,
        selected_cids=args.cid,
        limit=args.limit,
    )

    site, session = build_authenticated_session(args.env_file)
    main_contract_cache: dict[str, str] = {}

    flat_rows: list[dict[str, Any]] = []
    nested_rows: list[dict[str, Any]] = []
    with_contracts = 0
    multi_contracts = 0
    no_contracts = 0
    errors = 0

    for index, target in enumerate(selected_targets, start=1):
        if index == 1 or index % 25 == 0:
            print(f"[lookup] {index}/{len(selected_targets)} cid={target.cid}")
        try:
            contracts, total = fetch_contract_rows(
                session,
                api_base_url=site.api_base_url,
                cid=target.cid,
                page_size=args.page_size,
            )
        except Exception as exc:  # noqa: BLE001
            error_text = str(exc)
            errors += 1
            flat_rows.append(
                build_flat_row(
                    site,
                    target,
                    contract_count=0,
                    lookup_status="error",
                    contract=None,
                    lookup_error=error_text,
                ),
            )
            nested_rows.append(
                {
                    "cid": target.cid,
                    "content_name": target.content_name,
                    "manager_name": target.manager_name,
                    "contract_count": 0,
                    "lookup_status": "error",
                    "lookup_error": error_text,
                    "contracts": [],
                },
            )
            continue

        if not contracts:
            no_contracts += 1
            flat_rows.append(
                build_flat_row(
                    site,
                    target,
                    contract_count=total,
                    lookup_status="no_contract",
                    contract=None,
                ),
            )
            nested_rows.append(
                {
                    "cid": target.cid,
                    "content_name": target.content_name,
                    "manager_name": target.manager_name,
                    "contract_count": total,
                    "lookup_status": "no_contract",
                    "lookup_error": "",
                    "contracts": [],
                },
            )
            continue

        with_contracts += 1
        if len(contracts) > 1:
            multi_contracts += 1

        nested_contracts: list[dict[str, Any]] = []
        for contract in contracts:
            visible_contract_id = str(contract.get("cntrId") or "").strip()
            kind_code = str(contract.get("cntrKndCd") or "").strip()
            main_contract_id = ""
            if kind_code == "002" and visible_contract_id:
                main_contract_id = fetch_main_contract_id(
                    session,
                    api_base_url=site.api_base_url,
                    contract_id=visible_contract_id,
                    cache=main_contract_cache,
                )

            row = build_flat_row(
                site,
                target,
                contract_count=total,
                lookup_status="matched",
                contract=contract,
                main_contract_id=main_contract_id,
            )
            flat_rows.append(row)
            nested_contracts.append(
                {
                    "visible_contract_id": row["visible_contract_id"],
                    "unity_contract_id": row["unity_contract_id"],
                    "main_contract_id": row["main_contract_id"],
                    "lookup_contract_id": row["lookup_contract_id"],
                    "contract_change_url": row["contract_change_url"],
                    "contract_name": row["contract_name"],
                    "contract_kind": row["contract_kind"],
                    "contract_kind_code": row["contract_kind_code"],
                    "is_primary_contract": row["is_primary_contract"],
                    "counterparty_name": row["counterparty_name"],
                    "related_contract_name": row["related_contract_name"],
                    "contract_date": row["contract_date"],
                    "contract_status": row["contract_status"],
                    "attachment_name": row["attachment_name"],
                    "contract_manager_name": row["contract_manager_name"],
                },
            )

        nested_rows.append(
            {
                "cid": target.cid,
                "content_name": target.content_name,
                "manager_name": target.manager_name,
                "contract_count": total,
                "lookup_status": "matched",
                "lookup_error": "",
                "contracts": nested_contracts,
            },
        )

    summary = {
        "ips_export": str(Path(args.ips_export).resolve()),
        "sheet_name": args.sheet_name,
        "manager_filter": args.manager,
        "selected_cids": [target.cid for target in selected_targets],
        "missing_requested_cids": missing_cids,
        "target_count": len(selected_targets),
        "with_contracts": with_contracts,
        "multi_contract_cids": multi_contracts,
        "no_contracts": no_contracts,
        "errors": errors,
    }
    payload = {"summary": summary, "rows": nested_rows}
    return flat_rows, payload


def main() -> None:
    args = parse_args()
    csv_output, json_output = build_output_paths(args)
    flat_rows, payload = build_report(args)

    if not args.dry_run:
        if csv_output is not None:
            write_csv(csv_output, flat_rows)
        if json_output is not None:
            json_output.parent.mkdir(parents=True, exist_ok=True)
            json_output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    result = {
        **payload["summary"],
        "csv_output": str(csv_output) if csv_output is not None and not args.dry_run else "",
        "json_output": str(json_output) if json_output is not None and not args.dry_run else "",
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
