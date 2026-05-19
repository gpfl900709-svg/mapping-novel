from __future__ import annotations

import argparse
import csv
import json
import threading
import time
import warnings
from collections import Counter, OrderedDict, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

from build_ips_contract_id_map import (
    build_api_url,
    build_authenticated_session,
    extract_response_data,
    fetch_main_contract_id,
    safe_suffix,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1] / "SIAAN Project"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output" / "ips_contract_links"
DEFAULT_IPS_EXPORT = (
    Path(__file__).resolve().parents[1]
    / "SIAANE_v2"
    / "담당자없는작품_재정리"
    / "20260518_ips.xlsx"
)


@dataclass(frozen=True)
class ContentTarget:
    cid: str
    content_name: str
    content_type: str
    owning_company: str
    department: str
    manager_name: str


_thread_local = threading.local()
_main_contract_cache_lock = threading.Lock()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Collect CID -> contract ID links from KIPM 콘텐츠상세 정산정보/계약목록 APIs. "
            "This is read-only and does not mutate IPS."
        ),
    )
    parser.add_argument("--ips-export", default=str(DEFAULT_IPS_EXPORT))
    parser.add_argument("--sheet-name", default="Sheet1")
    parser.add_argument("--env-file", default="")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--cid", action="append", default=[])
    parser.add_argument("--department", default="")
    parser.add_argument("--manager", default="")
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument("--page-size", type=int, default=1000)
    parser.add_argument("--output", default="")
    parser.add_argument("--edges-output", default="")
    parser.add_argument("--json-output", default="")
    parser.add_argument("--progress-every", type=int, default=250)
    return parser.parse_args()


def read_targets(path: Path, *, sheet_name: str) -> list[ContentTarget]:
    warnings.filterwarnings("ignore", message="Workbook contains no default style")
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    header_index = {header: index for index, header in enumerate(headers)}
    required = ("콘텐츠ID", "콘텐츠명", "콘텐츠형태", "귀속법인", "담당부서", "담당자명")
    missing = [column for column in required if column not in header_index]
    if missing:
        raise KeyError(f"필수 컬럼이 없습니다: {', '.join(missing)}")

    seen: set[str] = set()
    targets: list[ContentTarget] = []
    for values in rows:
        cid = str(values[header_index["콘텐츠ID"]] or "").strip()
        if not cid or cid in seen:
            continue
        seen.add(cid)
        targets.append(
            ContentTarget(
                cid=cid,
                content_name=str(values[header_index["콘텐츠명"]] or "").strip(),
                content_type=str(values[header_index["콘텐츠형태"]] or "").strip(),
                owning_company=str(values[header_index["귀속법인"]] or "").strip(),
                department=str(values[header_index["담당부서"]] or "").strip(),
                manager_name=str(values[header_index["담당자명"]] or "").strip(),
            ),
        )
    return targets


def filter_targets(
    targets: list[ContentTarget],
    *,
    selected_cids: list[str],
    department: str,
    manager: str,
    limit: int,
) -> tuple[list[ContentTarget], list[str]]:
    selected = {str(cid or "").strip() for cid in selected_cids if str(cid or "").strip()}
    department = str(department or "").strip()
    manager = str(manager or "").strip()

    filtered: list[ContentTarget] = []
    seen: set[str] = set()
    for target in targets:
        if selected and target.cid not in selected:
            continue
        if department and target.department != department:
            continue
        if manager and target.manager_name != manager:
            continue
        if target.cid in seen:
            continue
        seen.add(target.cid)
        filtered.append(target)
        if limit and len(filtered) >= limit:
            break

    missing = sorted(selected - {target.cid for target in filtered}) if selected else []
    return filtered, missing


def build_output_paths(args: argparse.Namespace) -> tuple[Path, Path, Path]:
    if args.output:
        summary_path = Path(args.output)
    else:
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix_bits = []
        if args.department:
            suffix_bits.append(args.department)
        if args.manager:
            suffix_bits.append(args.manager)
        if args.limit:
            suffix_bits.append(f"limit{args.limit}")
        suffix = safe_suffix("_".join(suffix_bits) or "all")
        summary_path = DEFAULT_OUTPUT_DIR / f"{stamp}__ips_contract_links_{suffix}.csv"

    edge_path = Path(args.edges_output) if args.edges_output else summary_path.with_name(
        f"{summary_path.stem}__edges.csv",
    )
    json_path = Path(args.json_output) if args.json_output else summary_path.with_suffix(".json")
    return summary_path, edge_path, json_path


def get_thread_session(base_headers: dict[str, str]) -> requests.Session:
    session = getattr(_thread_local, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update(base_headers)
        _thread_local.session = session
    return session


def fetch_list(
    session: requests.Session,
    *,
    api_base_url: str,
    path: str,
    params: dict[str, Any],
    page_size: int,
    context: str,
) -> tuple[list[dict[str, Any]], int]:
    first_params = {**params, "pageNum": 1, "pageSize": page_size}
    response = session.get(build_api_url(api_base_url, path), params=first_params, timeout=40)
    data = extract_response_data(response, context=context)
    if isinstance(data, list):
        return [row for row in data if isinstance(row, dict)], len(data)
    if not isinstance(data, dict):
        return [], 0

    rows = [row for row in (data.get("list") or []) if isinstance(row, dict)]
    total = int(data.get("total") or len(rows))
    if total <= len(rows):
        return rows, total

    if total <= page_size * 5:
        retry_params = {**params, "pageNum": 1, "pageSize": total}
        retry = session.get(build_api_url(api_base_url, path), params=retry_params, timeout=60)
        retry_data = extract_response_data(retry, context=f"{context} 재조회")
        if isinstance(retry_data, dict):
            retry_rows = [row for row in (retry_data.get("list") or []) if isinstance(row, dict)]
            return retry_rows, int(retry_data.get("total") or len(retry_rows))

    pages = (total + page_size - 1) // page_size
    all_rows = list(rows)
    for page_num in range(2, pages + 1):
        page_params = {**params, "pageNum": page_num, "pageSize": page_size}
        page = session.get(build_api_url(api_base_url, path), params=page_params, timeout=40)
        page_data = extract_response_data(page, context=f"{context} page={page_num}")
        if isinstance(page_data, dict):
            all_rows.extend(row for row in (page_data.get("list") or []) if isinstance(row, dict))
    return all_rows, total


def unique_strings(values: list[Any]) -> list[str]:
    result: OrderedDict[str, None] = OrderedDict()
    for value in values:
        cleaned = str(value or "").strip()
        if cleaned and cleaned != "0":
            result[cleaned] = None
    return list(result.keys())


def summarize_settlement_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]], int]:
    grouped: dict[str, dict[str, Any]] = {}
    blank_rows = 0
    for row in rows:
        contract_id = str(row.get("cntrId") or "").strip()
        if not contract_id or contract_id == "0":
            blank_rows += 1
            continue
        bucket = grouped.setdefault(
            contract_id,
            {
                "contract_id": contract_id,
                "settlement_row_count": 0,
                "payment_settlement_ids": OrderedDict(),
                "payment_detail_ids": OrderedDict(),
                "channels": OrderedDict(),
                "counterparties": OrderedDict(),
            },
        )
        bucket["settlement_row_count"] += 1
        for source_key, target_key in (
            ("pymtSetlId", "payment_settlement_ids"),
            ("pymtSetlDtlId", "payment_detail_ids"),
            ("schnNm", "channels"),
            ("bcncNm", "counterparties"),
        ):
            value = str(row.get(source_key) or "").strip()
            if value:
                bucket[target_key][value] = None

    summaries: list[dict[str, Any]] = []
    for contract_id, bucket in grouped.items():
        summaries.append(
            {
                "contract_id": contract_id,
                "settlement_row_count": bucket["settlement_row_count"],
                "payment_settlement_ids": list(bucket["payment_settlement_ids"].keys()),
                "payment_detail_ids": list(bucket["payment_detail_ids"].keys()),
                "channels": list(bucket["channels"].keys()),
                "counterparties": list(bucket["counterparties"].keys()),
            },
        )
    return list(grouped.keys()), summaries, blank_rows


def collect_one(
    target: ContentTarget,
    *,
    api_base_url: str,
    base_headers: dict[str, str],
    page_size: int,
    main_contract_cache: dict[str, str],
) -> dict[str, Any]:
    session = get_thread_session(base_headers)

    contract_rows, contract_total = fetch_list(
        session,
        api_base_url=api_base_url,
        path="cntsd/cntslt/ctns-list/cntr-list",
        params={"srcCtnsId": target.cid, "srchCntrId": ""},
        page_size=page_size,
        context=f"계약목록 cid={target.cid}",
    )
    settlement_rows, settlement_total = fetch_list(
        session,
        api_base_url=api_base_url,
        path="cntsd/cntslt/ctns-list/ctns-cntr-setl-srcCntrId-list",
        params={
            "srcCprCd": "1000",
            "srcCtnsId": target.cid,
            "srchCntrId": "",
            "srcBcncCd": "",
            "pymtSetlSetmId": "",
        },
        page_size=page_size,
        context=f"정산정보 cid={target.cid}",
    )

    contract_summaries: list[dict[str, Any]] = []
    visible_ids: list[Any] = []
    unity_ids: list[Any] = []
    lookup_ids: list[Any] = []
    for contract in contract_rows:
        visible_id = str(contract.get("cntrId") or "").strip()
        unity_id = str(contract.get("unityCntrId") or "").strip()
        main_id = ""
        if str(contract.get("cntrKndCd") or "").strip() == "002" and visible_id:
            with _main_contract_cache_lock:
                cached = main_contract_cache.get(visible_id)
            if cached is None:
                main_id = fetch_main_contract_id(
                    session,
                    api_base_url=api_base_url,
                    contract_id=visible_id,
                    cache=main_contract_cache,
                )
            else:
                main_id = cached
        lookup_id = main_id or unity_id or visible_id
        visible_ids.append(visible_id)
        unity_ids.append(unity_id)
        lookup_ids.append(lookup_id)
        contract_summaries.append(
            {
                "visible_contract_id": visible_id,
                "unity_contract_id": unity_id,
                "main_contract_id": main_id,
                "lookup_contract_id": lookup_id,
                "contract_name": str(contract.get("cntrNm") or "").strip(),
                "contract_kind": str(contract.get("cntrKndNm") or "").strip(),
                "contract_kind_code": str(contract.get("cntrKndCd") or "").strip(),
                "is_primary_contract": str(contract.get("ocntrYn") or "").strip(),
                "contract_status": str(contract.get("cntrStsNm") or "").strip(),
                "counterparty_name": str(contract.get("cntrPtnNm") or contract.get("bcncNm") or "").strip(),
                "contract_date": str(contract.get("cntrCclsDt") or "").strip(),
                "attachment_name": str(contract.get("atchfOrgNm") or "").strip(),
                "contract_manager_name": str(contract.get("chgerNm") or "").strip(),
            },
        )

    settlement_ids, settlement_summaries, settlement_blank_rows = summarize_settlement_rows(settlement_rows)
    visible_ids = unique_strings(visible_ids)
    unity_ids = unique_strings(unity_ids)
    lookup_ids = unique_strings(lookup_ids)
    all_ids = unique_strings(lookup_ids + visible_ids + settlement_ids)

    if contract_summaries and settlement_ids:
        status = "matched_both"
    elif contract_summaries:
        status = "contract_tab_only"
    elif settlement_ids:
        status = "settlement_only"
    elif settlement_total or settlement_blank_rows:
        status = "settlement_rows_without_contract"
    else:
        status = "no_contract"

    return {
        "cid": target.cid,
        "content_name": target.content_name,
        "content_type": target.content_type,
        "owning_company": target.owning_company,
        "department": target.department,
        "manager_name": target.manager_name,
        "status": status,
        "all_contract_ids": all_ids,
        "contract_tab_total": contract_total,
        "contract_tab_visible_ids": visible_ids,
        "contract_tab_unity_ids": unity_ids,
        "contract_tab_lookup_ids": lookup_ids,
        "settlement_total": settlement_total,
        "settlement_contract_ids": settlement_ids,
        "settlement_blank_contract_rows": settlement_blank_rows,
        "contracts": contract_summaries,
        "settlement_contracts": settlement_summaries,
        "error": "",
    }


def collect_with_error(
    target: ContentTarget,
    *,
    api_base_url: str,
    base_headers: dict[str, str],
    page_size: int,
    main_contract_cache: dict[str, str],
) -> dict[str, Any]:
    try:
        return collect_one(
            target,
            api_base_url=api_base_url,
            base_headers=base_headers,
            page_size=page_size,
            main_contract_cache=main_contract_cache,
        )
    except Exception as exc:  # noqa: BLE001
        return {
            "cid": target.cid,
            "content_name": target.content_name,
            "content_type": target.content_type,
            "owning_company": target.owning_company,
            "department": target.department,
            "manager_name": target.manager_name,
            "status": "error",
            "all_contract_ids": [],
            "contract_tab_total": 0,
            "contract_tab_visible_ids": [],
            "contract_tab_unity_ids": [],
            "contract_tab_lookup_ids": [],
            "settlement_total": 0,
            "settlement_contract_ids": [],
            "settlement_blank_contract_rows": 0,
            "contracts": [],
            "settlement_contracts": [],
            "error": str(exc),
        }


def join_values(values: list[Any]) -> str:
    return ";".join(str(value) for value in values if str(value or "").strip())


def build_summary_csv_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in rows:
        output.append(
            {
                "cid": row["cid"],
                "content_name": row["content_name"],
                "content_type": row["content_type"],
                "owning_company": row["owning_company"],
                "department": row["department"],
                "manager_name": row["manager_name"],
                "status": row["status"],
                "all_contract_ids": join_values(row["all_contract_ids"]),
                "contract_tab_total": row["contract_tab_total"],
                "contract_tab_visible_ids": join_values(row["contract_tab_visible_ids"]),
                "contract_tab_unity_ids": join_values(row["contract_tab_unity_ids"]),
                "contract_tab_lookup_ids": join_values(row["contract_tab_lookup_ids"]),
                "settlement_total": row["settlement_total"],
                "settlement_contract_ids": join_values(row["settlement_contract_ids"]),
                "settlement_blank_contract_rows": row["settlement_blank_contract_rows"],
                "error": row["error"],
            },
        )
    return output


def build_edge_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    edge_rows: list[dict[str, Any]] = []
    for row in rows:
        base = {
            "cid": row["cid"],
            "content_name": row["content_name"],
            "department": row["department"],
            "manager_name": row["manager_name"],
        }
        for contract in row["contracts"]:
            edge_rows.append(
                {
                    **base,
                    "source": "contract_tab",
                    "contract_id": contract["lookup_contract_id"],
                    "visible_contract_id": contract["visible_contract_id"],
                    "unity_contract_id": contract["unity_contract_id"],
                    "main_contract_id": contract["main_contract_id"],
                    "contract_kind": contract["contract_kind"],
                    "contract_status": contract["contract_status"],
                    "counterparty_name": contract["counterparty_name"],
                    "contract_name": contract["contract_name"],
                    "settlement_row_count": "",
                    "channels": "",
                    "payment_settlement_ids": "",
                },
            )
        for settlement in row["settlement_contracts"]:
            edge_rows.append(
                {
                    **base,
                    "source": "settlement_table",
                    "contract_id": settlement["contract_id"],
                    "visible_contract_id": "",
                    "unity_contract_id": "",
                    "main_contract_id": "",
                    "contract_kind": "",
                    "contract_status": "",
                    "counterparty_name": join_values(settlement["counterparties"]),
                    "contract_name": "",
                    "settlement_row_count": settlement["settlement_row_count"],
                    "channels": join_values(settlement["channels"]),
                    "payment_settlement_ids": join_values(settlement["payment_settlement_ids"]),
                },
            )
        if not row["contracts"] and not row["settlement_contracts"]:
            edge_rows.append(
                {
                    **base,
                    "source": row["status"],
                    "contract_id": "",
                    "visible_contract_id": "",
                    "unity_contract_id": "",
                    "main_contract_id": "",
                    "contract_kind": "",
                    "contract_status": "",
                    "counterparty_name": "",
                    "contract_name": "",
                    "settlement_row_count": "",
                    "channels": "",
                    "payment_settlement_ids": "",
                },
            )
    return edge_rows


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                fieldnames.append(key)
                seen.add(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    all_targets = read_targets(Path(args.ips_export), sheet_name=args.sheet_name)
    targets, missing = filter_targets(
        all_targets,
        selected_cids=args.cid,
        department=args.department,
        manager=args.manager,
        limit=args.limit,
    )
    summary_path, edge_path, json_path = build_output_paths(args)

    site, login_session = build_authenticated_session(args.env_file)
    base_headers = dict(login_session.headers)
    main_contract_cache: dict[str, str] = {}

    started = time.monotonic()
    rows_by_cid: dict[str, dict[str, Any]] = {}
    workers = max(1, int(args.workers or 1))
    print(f"[start] targets={len(targets)} workers={workers} page_size={args.page_size}", flush=True)
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(
                collect_with_error,
                target,
                api_base_url=site.api_base_url,
                base_headers=base_headers,
                page_size=args.page_size,
                main_contract_cache=main_contract_cache,
            ): target
            for target in targets
        }
        for index, future in enumerate(as_completed(futures), start=1):
            target = futures[future]
            rows_by_cid[target.cid] = future.result()
            if index == 1 or index % args.progress_every == 0 or index == len(targets):
                elapsed = time.monotonic() - started
                rate = index / elapsed if elapsed else 0
                print(f"[progress] {index}/{len(targets)} elapsed={elapsed:.1f}s rate={rate:.2f}/s", flush=True)

    rows = [rows_by_cid[target.cid] for target in targets]
    summary_rows = build_summary_csv_rows(rows)
    edge_rows = build_edge_rows(rows)
    write_csv(summary_path, summary_rows)
    write_csv(edge_path, edge_rows)

    status_counts = Counter(row["status"] for row in rows)
    summary = {
        "ips_export": str(Path(args.ips_export).resolve()),
        "sheet_name": args.sheet_name,
        "target_count": len(targets),
        "missing_requested_cids": missing,
        "workers": workers,
        "page_size": args.page_size,
        "status_counts": dict(status_counts),
        "rows_with_any_contract_id": sum(1 for row in rows if row["all_contract_ids"]),
        "rows_without_contract_id": sum(1 for row in rows if not row["all_contract_ids"]),
        "edge_count": len(edge_rows),
        "elapsed_seconds": round(time.monotonic() - started, 3),
        "summary_csv": str(summary_path),
        "edges_csv": str(edge_path),
        "json_output": str(json_path),
    }
    payload = {"summary": summary, "rows": rows, "edges": edge_rows}
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
