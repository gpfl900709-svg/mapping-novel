from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from build_work_index import build_title_keys, clean_title_for_display, normalize_text, split_folder_name
from ips_title_rules import (
    build_title_catalog_from_overrides,
    canonicalize_full_remark,
    canonicalize_ips_category,
    compute_explicit_target_title,
)
from ips.core.auth import resolve_env_path
from ips.core.browser import BrowserSettings
from ips.core.harness import IPSHarness
from ips.sites import get_site
from probe_ips_content_detail import FETCH_DETAIL_SCRIPT
from work_cid_utils import DEFAULT_WORK_CID_REGISTRY_PATH


PROJECT_ROOT = Path(__file__).resolve().parents[1] / "SIAAN Project"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output" / "ips_sales_channel_harness"
DEFAULT_MANUAL_OVERRIDES_PATH = PROJECT_ROOT / "config" / "manual_overrides.local.json"
DEFAULT_LOGIN_PATH = "/ip/cntsd/cntschg/ctns-chg-list?pageNum=1&pageSize=10"
DETAIL_PATH_TEMPLATE = "/cntsd/cntschg/ctns-chg-list/detail/{cid}"
EDIT_PATH_TEMPLATE = "/ip/cntsd/cntschg/ctns-chg-edit?parCtnsId={cid}"

DEFAULT_TITLE_COLUMNS = (
    "미매핑_콘텐츠마스터명",
    "콘텐츠마스터명",
    "작품명",
    "title",
    "raw_title",
)
DEFAULT_PLATFORM_COLUMNS = (
    "채널",
    "플랫폼",
    "판매채널",
    "platform",
    "channel",
)
DEFAULT_MANAGER_COLUMNS = (
    "담당자(없을 시 공란)",
    "담당자명",
    "담당자",
    "manager",
)
DEFAULT_ROW_ID_COLUMNS = (
    "row_id",
    "행번호",
    "id",
)


@dataclass
class RegistryEntry:
    row: dict[str, str]
    folder_name: str
    folder_title: str
    author: str
    manager: str
    work_cid: str
    remark: str
    ips_name: str
    title_keys: set[str]


@dataclass
class CandidateScore:
    entry: RegistryEntry
    score: int
    reason: str


@dataclass
class ExplicitTarget:
    work_cid: str
    target_title: str
    author_name: str
    remark: str
    note: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Resolve raw titles to IPS canonical CID and fetch 판매채널콘텐츠ID from IPS detail API.",
    )
    parser.add_argument("--title", default="", help="Single raw title to resolve.")
    parser.add_argument("--platform", default="", help="Single platform / 판매채널명.")
    parser.add_argument("--manager", default="", help="Optional manager filter for single lookup.")
    parser.add_argument("--cid", default="", help="Skip title resolution and query this CID directly.")
    parser.add_argument("--input", default="", help="CSV/TSV/XLSX input file for batch lookup.")
    parser.add_argument("--sheet", default="", help="Optional sheet name for XLSX input. Default is first sheet.")
    parser.add_argument("--title-column", default="", help="Override title column name for batch input.")
    parser.add_argument("--platform-column", default="", help="Override platform column name for batch input.")
    parser.add_argument("--manager-column", default="", help="Override manager column name for batch input.")
    parser.add_argument("--row-id-column", default="", help="Optional row id column to carry through.")
    parser.add_argument("--default-manager", default="", help="Use this manager when input rows omit manager.")
    parser.add_argument("--filter-manager", default="", help="Only keep rows whose manager column matches this name.")
    parser.add_argument(
        "--only-empty-column",
        default="",
        help="Only keep rows where this source column is empty. Useful for 생성 ID blank-only runs.",
    )
    parser.add_argument("--registry", default=str(DEFAULT_WORK_CID_REGISTRY_PATH))
    parser.add_argument("--manual-overrides", default=str(DEFAULT_MANUAL_OVERRIDES_PATH))
    parser.add_argument("--env-file", default="")
    parser.add_argument("--output", default="", help="Optional CSV/JSON output path. Suffix decides format.")
    parser.add_argument("--json-output", default="", help="Optional JSON output path when main output is CSV.")
    parser.add_argument("--headless", action="store_true")
    parser.add_argument("--timeout-ms", type=int, default=20_000)
    parser.add_argument("--slow-mo-ms", type=int, default=0)
    parser.add_argument("--max-candidates", type=int, default=5)
    return parser.parse_args()


def load_registry_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_manual_overrides(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def load_explicit_targets(path: Path) -> list[ExplicitTarget]:
    overrides = load_manual_overrides(path)
    titles_by_cid, titles_by_author = build_title_catalog_from_overrides(overrides)
    rows = overrides.get("ips_explicit_targets", [])
    results: list[ExplicitTarget] = []
    for row in rows:
        work_cid = str(row.get("work_cid") or "").strip()
        target_title = compute_explicit_target_title(row, titles_by_cid, titles_by_author)
        if not work_cid or not target_title:
            continue
        results.append(
            ExplicitTarget(
                work_cid=work_cid,
                target_title=target_title,
                author_name=str(row.get("author_name") or "").strip(),
                remark=canonicalize_full_remark(str(row.get("remark") or "").strip()),
                note=str(row.get("note") or "").strip(),
            )
        )
    return results


def collect_entry_keys(row: dict[str, str], folder_title: str) -> set[str]:
    variants: list[str] = []
    for value in (
        row.get("title_final", ""),
        folder_title,
        row.get("matched_ips_title", ""),
        row.get("matched_s2_title", ""),
    ):
        cleaned = clean_title_for_display(value)
        if cleaned:
            variants.append(cleaned)

    key_set: set[str] = set()
    for value in variants:
        key_set.update(build_title_keys(value))
    return {key for key in key_set if key}


def build_registry_entries(rows: list[dict[str, str]]) -> list[RegistryEntry]:
    entries: list[RegistryEntry] = []
    for row in rows:
        work_cid = (row.get("work_cid") or "").strip()
        title_final = clean_title_for_display(row.get("title_final", ""))
        if not work_cid or not title_final:
            continue
        folder_path = Path(row.get("folder_path") or "")
        folder_name = folder_path.name if str(folder_path) else ""
        folder_title, folder_author = split_folder_name(folder_name) if folder_name else ("", "")
        author = clean_title_for_display(row.get("author_final", "") or folder_author)
        title_keys = collect_entry_keys(row, folder_title)
        if not title_keys:
            continue
        entries.append(
            RegistryEntry(
                row=row,
                folder_name=folder_name,
                folder_title=folder_title or title_final,
                author=author,
                manager=(row.get("manager") or "").strip(),
                work_cid=work_cid,
                remark=(row.get("remark") or "").strip(),
                ips_name=(row.get("ips_name") or "").strip(),
                title_keys=title_keys,
            )
        )
    return entries


def detect_delimiter(path: Path) -> str:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
    if "\t" in sample and sample.count("\t") >= sample.count(","):
        return "\t"
    return ","


def read_tabular_input(path: Path, *, sheet_name: str = "") -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows_iter)]
        rows: list[dict[str, Any]] = []
        for values in rows_iter:
            row = {}
            for header, value in zip(headers, values):
                row[header] = "" if value is None else str(value)
            rows.append(row)
        return rows

    delimiter = detect_delimiter(path)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


def resolve_column_name(headers: list[str], explicit_name: str, candidates: tuple[str, ...]) -> str:
    if explicit_name:
        column_index = excel_column_ref_to_index(explicit_name)
        if column_index is not None and column_index < len(headers):
            return headers[column_index]
        return explicit_name
    header_map = {normalize_text(header): header for header in headers}
    for candidate in candidates:
        match = header_map.get(normalize_text(candidate))
        if match:
            return match
    return ""


def excel_column_ref_to_index(value: str) -> int | None:
    stripped = value.strip().upper()
    if not stripped or not stripped.isalpha() or len(stripped) > 3:
        return None
    index = 0
    for char in stripped:
        index = index * 26 + (ord(char) - 64)
    return index - 1


def infer_platform_column(headers: list[str], resolved_column: str) -> str:
    if resolved_column:
        return resolved_column
    if not headers:
        return ""
    first_header = headers[0]
    # Google Sheets exports sometimes return the A1 header as a blank/space
    # even when the visible column is the platform/channel column.
    if not first_header.strip() or normalize_text(first_header) == normalize_text("채널"):
        return first_header
    return ""


def query_title_keys(raw_title: str) -> set[str]:
    cleaned = clean_title_for_display(raw_title)
    return {key for key in build_title_keys(cleaned) if key}


def title_similarity_score(raw_title: str, entry: RegistryEntry) -> CandidateScore:
    query_keys = query_title_keys(raw_title)
    if not query_keys:
        return CandidateScore(entry=entry, score=0, reason="no_query_key")

    best_score = 0
    best_reason = "no_match"
    for query_key in query_keys:
        for entry_key in entry.title_keys:
            if query_key == entry_key:
                return CandidateScore(entry=entry, score=1000, reason="exact_key")
            if query_key in entry_key or entry_key in query_key:
                score = 930 - abs(len(entry_key) - len(query_key))
                if score > best_score:
                    best_score = score
                    best_reason = "substring_key"
                continue
            ratio = SequenceMatcher(None, query_key, entry_key).ratio()
            score = int(round(ratio * 1000))
            if score > best_score:
                best_score = score
                best_reason = "fuzzy_key"
    return CandidateScore(entry=entry, score=best_score, reason=best_reason)


def pick_candidates(
    raw_title: str,
    manager_name: str,
    entries: list[RegistryEntry],
    *,
    max_candidates: int,
) -> list[CandidateScore]:
    manager_key = normalize_text(manager_name)
    filtered = [
        entry for entry in entries
        if not manager_key or normalize_text(entry.manager) == manager_key
    ]
    if not filtered:
        filtered = entries

    scored = [title_similarity_score(raw_title, entry) for entry in filtered]
    scored = [item for item in scored if item.score > 0]
    scored.sort(
        key=lambda item: (
            item.score,
            normalize_text(item.entry.manager) == manager_key if manager_key else False,
            item.entry.work_cid,
            item.entry.folder_title,
        ),
        reverse=True,
    )
    return scored[:max_candidates]


def summarize_candidates(candidates: list[CandidateScore]) -> str:
    parts: list[str] = []
    for item in candidates[:3]:
        parts.append(
            f"{item.entry.folder_title}({item.entry.author})|cid={item.entry.work_cid}|score={item.score}"
        )
    return " || ".join(parts)


def axios_get_detail(page: Any, cid: str) -> dict[str, Any]:
    detail_path = DETAIL_PATH_TEMPLATE.format(cid=cid)
    return page.evaluate(FETCH_DETAIL_SCRIPT, detail_path)


def platform_key(value: str) -> str:
    return normalize_text(clean_title_for_display(value))


def explicit_title_key(value: str) -> str:
    return normalize_text(clean_title_for_display(value)).replace("_", "")


def platform_route_family(platform_name: str) -> dict[str, tuple[str, ...]] | None:
    platform_normalized = platform_key(platform_name)
    if platform_normalized in {
        platform_key("네이버_장르(광고수익)"),
        platform_key("네이버_연재(광고수익)"),
    }:
        return {
            "route_name": ("naver_ads",),
            "preferred_tokens": ("네이버광고수익", "네이버 광고수익"),
            "fallback_tokens": (),
            "exclude_tokens": ("카카오", "카카오광고수익", "카카오선투자", "(사용안함)", "중복"),
        }
    if platform_normalized == platform_key("카카오페이지(창작지원금)"):
        return {
            "route_name": ("kakao_ads",),
            "preferred_tokens": ("카카오광고수익", "카카오 광고수익"),
            "fallback_tokens": ("카카오창작지원금", "카카오 창작지원금"),
            "exclude_tokens": ("네이버", "네이버광고수익", "네이버mg", "네이버 mg", "카카오선투자", "(사용안함)", "중복"),
        }
    if "카카오페이지(선투자)" in platform_normalized or "카카오선투자" in platform_normalized:
        return {
            "route_name": ("kakao_investment",),
            "preferred_tokens": ("카카오선투자", "카카오 선투자", "카카오mg", "카카오 mg"),
            "fallback_tokens": ("_카카오", "카카오"),
            "exclude_tokens": ("광고수익", "카카오광고수익", "창작지원금", "카카오창작지원금", "(사용안함)", "중복"),
        }
    if "네이버" in platform_normalized:
        if "광고수익" in platform_normalized:
            return {
                "route_name": ("naver_ads",),
                "preferred_tokens": ("네이버광고수익", "네이버 광고수익"),
                "fallback_tokens": ("네이버mg", "네이버 mg", "_네이버", "네이버"),
                "exclude_tokens": ("카카오", "카카오광고수익", "카카오선투자", "(사용안함)", "중복"),
            }
        return {
            "route_name": ("naver_mg",),
            "preferred_tokens": ("네이버mg", "네이버 mg"),
            "fallback_tokens": ("원작", "_네이버", "네이버"),
            "exclude_tokens": ("카카오", "카카오광고수익", "카카오선투자", "(사용안함)", "중복"),
        }
    return None


def entry_matches_route_family(entry: RegistryEntry, platform_name: str) -> tuple[bool, str]:
    family = platform_route_family(platform_name)
    if family is None:
        return False, ""
    combined_key = explicit_title_key(f"{entry.ips_name} {entry.remark}")
    preferred = tuple(explicit_title_key(token) for token in family["preferred_tokens"])
    fallback = tuple(explicit_title_key(token) for token in family["fallback_tokens"])
    if any(token and token in combined_key for token in preferred + fallback):
        return True, family["route_name"][0]
    return False, family["route_name"][0]


def choose_platform_specific_target(
    entry: RegistryEntry,
    platform_name: str,
    explicit_targets: list[ExplicitTarget],
) -> tuple[ExplicitTarget | None, str]:
    family = platform_route_family(platform_name)
    if family is None:
        return None, ""

    entry_title_key = explicit_title_key(entry.folder_title or entry.row.get("title_final", ""))
    entry_author_key = normalize_text(entry.author)
    ranked: list[tuple[int, ExplicitTarget]] = []
    for target in explicit_targets:
        combined_text = f"{target.target_title} {target.remark}".strip()
        combined_key = explicit_title_key(combined_text)
        author_key = normalize_text(target.author_name)
        canonical_remark = canonicalize_ips_category(target.remark)

        if entry_title_key and entry_title_key not in combined_key:
            continue
        if author_key and entry_author_key and author_key != entry_author_key:
            continue
        if any(token in combined_key for token in (explicit_title_key(value) for value in family["exclude_tokens"])):
            continue

        score = 0
        if normalize_text(canonical_remark) == normalize_text("카카오MG"):
            score += 400
        if normalize_text(canonical_remark) == normalize_text("네이버MG"):
            score += 450
        if normalize_text(canonical_remark).startswith(normalize_text("네이버광고수익")):
            score += 430
        if normalize_text(canonical_remark) == normalize_text("원작"):
            score += 250
        if normalize_text(canonical_remark).startswith(normalize_text("카카오광고수익")):
            score += 430
        if normalize_text(canonical_remark) == normalize_text("카카오창작지원금"):
            score += 330
        if any(explicit_title_key(token) in combined_key for token in family["preferred_tokens"]):
            score += 300
        if any(explicit_title_key(token) in combined_key for token in family["fallback_tokens"]):
            score += 100
        if target.work_cid == entry.work_cid:
            score -= 500
        if score > 0:
            ranked.append((score, target))

    if not ranked:
        return None, family["route_name"][0]
    ranked.sort(key=lambda item: (item[0], item[1].work_cid), reverse=True)
    return ranked[0][1], family["route_name"][0]


def match_platform_rows(detail_data: dict[str, Any], platform_name: str) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    requested_key = platform_key(platform_name)
    requested_keys = platform_match_keys(platform_name)
    info_list = detail_data.get("schnCtnsInfoList") or []
    rows = [item for item in info_list if isinstance(item, dict)]
    if not requested_key:
        return None, rows

    exact: list[dict[str, Any]] = []
    partial: list[dict[str, Any]] = []
    for row in rows:
        row_name = str(row.get("lwerSchnNm") or "").strip()
        row_key = platform_key(row_name)
        if not row_key:
            continue
        if row_key in requested_keys:
            exact.append(row)
        elif any(requested in row_key or row_key in requested for requested in requested_keys):
            partial.append(row)
    if exact:
        return exact[0], rows
    if partial:
        return partial[0], rows
    return None, rows


def platform_match_keys(platform_name: str) -> set[str]:
    requested_key = platform_key(platform_name)
    if not requested_key:
        return set()
    keys = {requested_key}
    aliases = {
        platform_key("예스24(소설)"): ("Yes24(서점)",),
        platform_key("Yes24(소설)"): ("Yes24(서점)",),
    }
    for alias in aliases.get(requested_key, ()):
        alias_key = platform_key(alias)
        if alias_key:
            keys.add(alias_key)
    return keys


def format_platform_snapshot(rows: list[dict[str, Any]]) -> str:
    pairs: list[str] = []
    for row in rows:
        name = str(row.get("lwerSchnNm") or "").strip()
        schn_ctns_id = str(row.get("schnCtnsId") or "").strip()
        if not name:
            continue
        if schn_ctns_id:
            pairs.append(f"{name}:{schn_ctns_id}")
        else:
            pairs.append(name)
    return " | ".join(pairs)


def build_edit_url(cid: str) -> str:
    site = get_site("kipm")
    return site.resolve_url(EDIT_PATH_TEMPLATE.format(cid=cid))


def build_output_paths(args: argparse.Namespace) -> tuple[Path | None, Path | None]:
    main_output = Path(args.output) if args.output else None
    json_output = Path(args.json_output) if args.json_output else None
    if main_output is None and json_output is None and args.input:
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        main_output = DEFAULT_OUTPUT_DIR / f"{stamp}__ips_sales_channel_harness.csv"
        json_output = DEFAULT_OUTPUT_DIR / f"{stamp}__ips_sales_channel_harness.json"
    elif main_output is not None and main_output.suffix.lower() == ".csv" and json_output is None:
        json_output = main_output.with_suffix(".json")
    return main_output, json_output


def single_input_row(args: argparse.Namespace) -> list[dict[str, Any]]:
    if not args.title and not args.cid:
        return []
    return [{
        "__row_id": "1",
        "__raw_title": args.title,
        "__platform": args.platform,
        "__manager": args.manager or args.default_manager,
        "__cid": args.cid,
    }]


def normalize_input_rows(args: argparse.Namespace) -> list[dict[str, Any]]:
    if args.input:
        input_path = Path(args.input)
        raw_rows = read_tabular_input(input_path, sheet_name=args.sheet)
        if not raw_rows:
            return []
        headers = list(raw_rows[0].keys())
        title_column = resolve_column_name(headers, args.title_column, DEFAULT_TITLE_COLUMNS)
        platform_column = infer_platform_column(
            headers,
            resolve_column_name(headers, args.platform_column, DEFAULT_PLATFORM_COLUMNS),
        )
        manager_column = resolve_column_name(headers, args.manager_column, DEFAULT_MANAGER_COLUMNS)
        row_id_column = resolve_column_name(headers, args.row_id_column, DEFAULT_ROW_ID_COLUMNS)
        only_empty_column = ""
        if args.only_empty_column:
            only_empty_column = resolve_column_name(headers, args.only_empty_column, (args.only_empty_column,))
            if not only_empty_column:
                only_empty_column = args.only_empty_column

        rows: list[dict[str, Any]] = []
        for index, raw_row in enumerate(raw_rows, start=2):
            row_manager = (raw_row.get(manager_column) or args.default_manager or "").strip() if manager_column or args.default_manager else ""
            if args.filter_manager and normalize_text(row_manager) != normalize_text(args.filter_manager):
                continue
            if only_empty_column and str(raw_row.get(only_empty_column) or "").strip():
                continue
            rows.append(
                {
                    **raw_row,
                    "__row_id": (raw_row.get(row_id_column) or str(index)) if row_id_column else str(index),
                    "__raw_title": (raw_row.get(title_column) or "").strip() if title_column else "",
                    "__platform": (raw_row.get(platform_column) or "").strip() if platform_column else "",
                    "__manager": row_manager,
                    "__cid": "",
                }
            )
        return rows

    return single_input_row(args)


def resolve_registry_match(
    raw_title: str,
    manager_name: str,
    entries: list[RegistryEntry],
    *,
    max_candidates: int,
) -> dict[str, Any]:
    candidates = pick_candidates(raw_title, manager_name, entries, max_candidates=max_candidates)
    if not candidates:
        return {
            "status": "title_unresolved",
            "match_reason": "no_candidate",
            "candidate_summary": "",
            "entry": None,
        }

    top = candidates[0]
    second = candidates[1] if len(candidates) > 1 else None
    if top.score < 780:
        status = "title_review_needed"
    elif second and second.score == top.score and second.entry.work_cid != top.entry.work_cid:
        status = "title_review_needed"
    else:
        status = "resolved"

    return {
        "status": status,
        "match_reason": top.reason,
        "candidate_summary": summarize_candidates(candidates),
        "entry": top.entry,
        "score": top.score,
        "candidates": candidates,
    }


def run_lookup(args: argparse.Namespace) -> list[dict[str, Any]]:
    input_rows = normalize_input_rows(args)
    if not input_rows:
        return []

    registry_rows = load_registry_rows(Path(args.registry))
    entries = build_registry_entries(registry_rows)
    explicit_targets = load_explicit_targets(Path(args.manual_overrides))

    settings = BrowserSettings(
        headless=args.headless,
        slow_mo_ms=args.slow_mo_ms,
        timeout_ms=args.timeout_ms,
        artifacts_root=PROJECT_ROOT / "output" / "ips_harness",
    )
    env_path = resolve_env_path(args.env_file)
    site = get_site("kipm")

    detail_cache: dict[str, dict[str, Any]] = {}
    results: list[dict[str, Any]] = []

    with IPSHarness(site, settings=settings, env_path=env_path) as harness:
        harness.ensure_logged_in(path=DEFAULT_LOGIN_PATH)

        for raw_row in input_rows:
            result_row = dict(raw_row)
            raw_title = str(raw_row.get("__raw_title") or "").strip()
            platform_name = str(raw_row.get("__platform") or "").strip()
            manager_name = str(raw_row.get("__manager") or "").strip()
            direct_cid = str(raw_row.get("__cid") or "").strip()

            entry: RegistryEntry | None = None
            resolution: dict[str, Any] = {}
            if direct_cid:
                direct_entries = [item for item in entries if item.work_cid == direct_cid]
                entry = direct_entries[0] if direct_entries else None
                resolution = {
                    "status": "resolved" if entry else "cid_only",
                    "match_reason": "direct_cid",
                    "candidate_summary": "",
                    "score": 1000 if entry else 0,
                }
            else:
                resolution = resolve_registry_match(raw_title, manager_name, entries, max_candidates=args.max_candidates)
                entry = resolution.get("entry")

            result_row.update(
                {
                    "lookup_status": resolution.get("status") or "",
                    "lookup_reason": resolution.get("match_reason") or "",
                    "lookup_score": resolution.get("score") or "",
                    "candidate_summary": resolution.get("candidate_summary") or "",
                    "input_title": raw_title,
                    "input_platform": platform_name,
                    "input_manager": manager_name,
                    "canonical_work_cid": entry.work_cid if entry else direct_cid,
                    "work_cid": entry.work_cid if entry else direct_cid,
                    "resolved_title": entry.folder_title if entry else "",
                    "resolved_author": entry.author if entry else "",
                    "resolved_manager": entry.manager if entry else manager_name,
                    "canonical_remark": entry.remark if entry else "",
                    "remark": entry.remark if entry else "",
                    "canonical_ips_name": entry.ips_name if entry else "",
                    "ips_name": entry.ips_name if entry else "",
                    "folder_name": entry.folder_name if entry else "",
                    "folder_path": entry.row.get("folder_path", "") if entry else "",
                    "platform_route_status": "",
                    "platform_route_reason": "",
                    "platform_route_target_title": "",
                    "edit_url": build_edit_url(entry.work_cid if entry else direct_cid) if (entry or direct_cid) else "",
                    "detail_status": "",
                    "platform_match_status": "",
                    "matched_platform_name": "",
                    "sales_channel_content_id": "",
                    "next_action": "",
                    "existing_platform_count": "",
                    "existing_platform_snapshot": "",
                }
            )

            if entry and not direct_cid:
                routed_target, route_name = choose_platform_specific_target(entry, platform_name, explicit_targets)
                if routed_target is not None:
                    result_row["work_cid"] = routed_target.work_cid
                    result_row["remark"] = routed_target.remark or result_row["remark"]
                    result_row["ips_name"] = f"IPS - {routed_target.target_title}"
                    result_row["platform_route_status"] = "routed_special_cid"
                    result_row["platform_route_reason"] = f"platform_family:{route_name}"
                    result_row["platform_route_target_title"] = routed_target.target_title
                    result_row["edit_url"] = build_edit_url(routed_target.work_cid)
                else:
                    already_special, existing_route_name = entry_matches_route_family(entry, platform_name)
                    result_row["platform_route_status"] = "already_special_cid" if already_special else "canonical"
                    result_row["platform_route_reason"] = (
                        f"platform_family:{existing_route_name or route_name}" if (existing_route_name or route_name) else ""
                    )
            elif direct_cid:
                result_row["platform_route_status"] = "direct_cid"

            cid_to_query = result_row["work_cid"]
            if not cid_to_query:
                result_row["detail_status"] = "skipped_no_cid"
                result_row["next_action"] = "review_title_match"
                results.append(result_row)
                continue

            if resolution.get("status") == "title_review_needed":
                result_row["detail_status"] = "skipped_title_review"
                result_row["next_action"] = "review_title_match"
                results.append(result_row)
                continue

            detail_payload = detail_cache.get(cid_to_query)
            if detail_payload is None:
                detail_resp = axios_get_detail(harness.page, cid_to_query)
                if not detail_resp.get("ok"):
                    result_row["detail_status"] = f"detail_failed:{detail_resp.get('status')}"
                    result_row["lookup_reason"] = str(detail_resp.get("error") or detail_resp.get("data") or "")[:300]
                    result_row["next_action"] = "check_ips_login_or_cid"
                    results.append(result_row)
                    continue
                detail_payload = detail_resp.get("data") if isinstance(detail_resp.get("data"), dict) else {}
                detail_cache[cid_to_query] = detail_payload

            result_row["detail_status"] = "loaded"
            matched_platform, all_platform_rows = match_platform_rows(detail_payload, platform_name)
            result_row["existing_platform_count"] = len(all_platform_rows)
            result_row["existing_platform_snapshot"] = format_platform_snapshot(all_platform_rows)
            if not platform_name:
                result_row["platform_match_status"] = "platform_not_requested"
                result_row["next_action"] = "inspect_existing_platforms"
            elif matched_platform:
                result_row["platform_match_status"] = "found"
                result_row["matched_platform_name"] = str(matched_platform.get("lwerSchnNm") or "").strip()
                result_row["sales_channel_content_id"] = str(matched_platform.get("schnCtnsId") or "").strip()
                result_row["next_action"] = "paste_sales_channel_content_id"
            else:
                result_row["platform_match_status"] = "missing_platform"
                result_row["next_action"] = "add_platform_in_ips"
            results.append(result_row)

    return results


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key in seen:
                continue
            seen.add(key)
            fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    rows = run_lookup(args)
    main_output, json_output = build_output_paths(args)

    if main_output is not None:
        if main_output.suffix.lower() == ".json":
            main_output.parent.mkdir(parents=True, exist_ok=True)
            main_output.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
        else:
            write_csv(main_output, rows)
    if json_output is not None:
        json_output.parent.mkdir(parents=True, exist_ok=True)
        json_output.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(
        {
            "rows": rows,
            "output": str(main_output) if main_output else "",
            "json_output": str(json_output) if json_output else "",
        },
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()
