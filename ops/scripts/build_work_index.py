from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from work_cid_utils import DEFAULT_WORK_CID_REGISTRY_PATH, load_work_cid_registry, split_pipe_list


AUTHOR_SUFFIX_PATTERN = re.compile(r"^(?P<title>.+)\((?P<author>[^()]*)\)$")
BRACKET_PREFIX_PATTERN = re.compile(r"^(?:\[[^\]]+\])+\s*")
SQUARE_BRACKET_PATTERN = re.compile(r"\[[^\]]*\]")
PAREN_PATTERN = re.compile(r"\([^)]*\)")
MULTISPACE_PATTERN = re.compile(r"\s+")
AUTHOR_HINT_NOISE_PATTERN = re.compile(r"^(?:지은이\s*[_:]\s*|작가\s*[_:]\s*)", re.IGNORECASE)
TITLE_NOISE_TOKENS = ("개정판", "완결", "외전", "단행")


@dataclass
class TitleResolution:
    author: str
    source: str
    issues: list[str]


@dataclass
class MatchCandidate:
    folder_name: str
    folder_path: str
    source_root: str
    confidence: float
    reason: str
    epub_count: int
    title: str


def collapse_spaces(value: str) -> str:
    return MULTISPACE_PATTERN.sub(" ", value or "").strip()


def normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value or "").strip().lower()
    normalized = re.sub(r"\s+", "", normalized)
    return normalized


def clean_author_hint(value: str) -> str:
    cleaned = unicodedata.normalize("NFKC", value or "").strip()
    cleaned = cleaned.replace("_", " ")
    cleaned = AUTHOR_HINT_NOISE_PATTERN.sub("", cleaned)
    return collapse_spaces(cleaned)


def clean_title_for_display(value: str) -> str:
    cleaned = unicodedata.normalize("NFKC", value or "").strip()
    while True:
        updated = BRACKET_PREFIX_PATTERN.sub("", cleaned).strip()
        if updated == cleaned:
            break
        cleaned = updated
    cleaned = cleaned.replace("_", " ")
    cleaned = cleaned.replace(" - ", " ")
    cleaned = collapse_spaces(cleaned)
    return cleaned


def strip_parenthetical_parts(value: str) -> str:
    return collapse_spaces(PAREN_PATTERN.sub(" ", value or ""))


def simplify_title(value: str) -> str:
    cleaned = clean_title_for_display(value)
    cleaned = SQUARE_BRACKET_PATTERN.sub(" ", cleaned)
    cleaned = PAREN_PATTERN.sub(" ", cleaned)
    for token in TITLE_NOISE_TOKENS:
        cleaned = cleaned.replace(token, " ")
    return collapse_spaces(cleaned)


def normalize_title_key(value: str) -> str:
    compact = normalize_text(simplify_title(value))
    return re.sub(r"[^0-9a-zA-Z가-힣]", "", compact)


def build_title_keys(value: str) -> list[str]:
    variants = [
        clean_title_for_display(value),
        strip_parenthetical_parts(clean_title_for_display(value)),
        simplify_title(value),
    ]
    keys: list[str] = []
    seen: set[str] = set()
    for variant in variants:
        key = re.sub(r"[^0-9a-zA-Z가-힣]", "", normalize_text(variant))
        if key and key not in seen:
            keys.append(key)
            seen.add(key)
    return keys


def split_folder_name(folder_name: str) -> tuple[str, str]:
    match = AUTHOR_SUFFIX_PATTERN.match(folder_name.strip())
    if not match:
        return clean_title_for_display(folder_name.strip()), ""
    return clean_title_for_display(match.group("title").strip()), clean_author_hint(match.group("author").strip())


def load_config(config_path: Path) -> dict[str, Any]:
    return json.loads(config_path.read_text(encoding="utf-8"))


def load_manual_overrides(config: dict[str, Any]) -> dict[str, Any]:
    override_path_value = config.get("manual_override_path", "")
    if not override_path_value:
        return {"ebook_matches": [], "ebook_exclusions": [], "manager_assignments": [], "work_status_overrides": []}

    override_path = Path(override_path_value)
    if not override_path.exists():
        return {"ebook_matches": [], "ebook_exclusions": [], "manager_assignments": [], "work_status_overrides": []}

    return json.loads(override_path.read_text(encoding="utf-8"))


def build_manager_override_map(manual_overrides: dict[str, Any]) -> dict[str, dict[str, str]]:
    override_map: dict[str, dict[str, str]] = {}
    for item in manual_overrides.get("manager_assignments", []):
        folder_path = item.get("work_folder_path", "")
        manager_name = item.get("manager_name", "")
        if not folder_path or not manager_name:
            continue
        override_map[folder_path] = {
            "manager_name": manager_name,
            "note": item.get("note", ""),
        }
    return override_map


def build_ebook_exclusion_map(manual_overrides: dict[str, Any]) -> dict[str, dict[str, str]]:
    exclusion_map: dict[str, dict[str, str]] = {}
    for item in manual_overrides.get("ebook_exclusions", []):
        folder_path = item.get("work_folder_path", "")
        if not folder_path:
            continue
        exclusion_map[folder_path] = {
            "status": item.get("status", "미정"),
            "note": item.get("note", ""),
            "reason": item.get("reason", "manual_exclusion"),
        }
    return exclusion_map


def build_work_status_override_map(manual_overrides: dict[str, Any]) -> dict[str, dict[str, str]]:
    override_map: dict[str, dict[str, str]] = {}
    for item in manual_overrides.get("work_status_overrides", []):
        folder_path = item.get("work_folder_path", "")
        status = item.get("status", "")
        if not folder_path or not status:
            continue
        override_map[folder_path] = {
            "status": status,
            "note": item.get("note", ""),
        }
    return override_map


def load_workbook_mappings(config: dict[str, Any]) -> tuple[dict[str, str], dict[str, set[str]]]:
    workbook_path = config["mapping_workbook"]["path"]
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    author_sheet_name = config["mapping_workbook"]["author_manager_sheet"]
    work_sheet_name = config["mapping_workbook"]["work_author_sheet"]

    author_manager_map: dict[str, str] = {}
    author_ws = wb[author_sheet_name]
    for row in author_ws.iter_rows(min_row=2, values_only=True):
        author = (row[0] or "").strip() if len(row) > 0 and row[0] else ""
        manager = (row[1] or "").strip() if len(row) > 1 and row[1] else ""
        if not author or not manager:
            continue
        author_manager_map[normalize_text(author)] = manager

    work_author_map: dict[str, set[str]] = defaultdict(set)
    work_ws = wb[work_sheet_name]
    for row in work_ws.iter_rows(min_row=2, values_only=True):
        title = (row[0] or "").strip() if len(row) > 0 and row[0] else ""
        author = (row[1] or "").strip() if len(row) > 1 and row[1] else ""
        if not title or not author:
            continue
        for key in build_title_keys(title):
            work_author_map[key].add(author)

    return author_manager_map, work_author_map


def resolve_author(
    title: str,
    author_hint: str,
    work_author_map: dict[str, set[str]],
) -> TitleResolution:
    issues: list[str] = []
    author_hint_key = normalize_text(author_hint)

    for index, key in enumerate(build_title_keys(title)):
        candidates = sorted(work_author_map.get(key, set()))
        if not candidates:
            continue
        source = "work_author_sheet" if index == 0 else "work_author_sheet_variant"
        if len(candidates) == 1:
            return TitleResolution(author=candidates[0], source=source, issues=issues)
        if author_hint:
            hinted = [candidate for candidate in candidates if normalize_text(candidate) == author_hint_key]
            if len(hinted) == 1:
                return TitleResolution(author=hinted[0], source=f"{source}_with_folder_hint", issues=issues)
        issues.append("work_author_ambiguous")

    if author_hint:
        return TitleResolution(author=author_hint, source="folder_suffix", issues=issues)

    issues.append("author_unresolved")
    return TitleResolution(author="", source="unresolved", issues=issues)


def should_skip_folder(folder_name: str, source: dict[str, Any]) -> bool:
    if folder_name in set(source.get("exclude_names", [])):
        return True
    for prefix in source.get("exclude_prefixes", []):
        if folder_name.startswith(prefix):
            return True
    return False


def count_epubs(folder: Path) -> int:
    return sum(1 for _ in iter_epub_files(folder))


def episode_file_sort_key(path: Path) -> tuple[int, int, str]:
    numbers = re.findall(r"\d+", path.stem)
    if numbers:
        return (1, int(numbers[-1]), normalize_text(path.stem))
    return (0, -1, normalize_text(path.stem))


def get_epub_search_roots(folder: Path) -> list[Path]:
    preferred_roots: list[Path] = []
    for child in folder.iterdir():
        if not child.is_dir():
            continue
        if "원고" in child.name:
            preferred_roots.append(child)
    return preferred_roots or [folder]


def iter_epub_files(folder: Path):
    for root in get_epub_search_roots(folder):
        yield from (path for path in root.rglob("*.epub") if path.is_file())


def get_epub_metrics(folder: Path) -> dict[str, Any]:
    epub_paths = list(iter_epub_files(folder))
    epub_count = len(epub_paths)
    latest_mtime = 0.0
    highest_episode_path: Path | None = None

    if epub_paths:
        highest_episode_path = max(epub_paths, key=episode_file_sort_key)
        latest_mtime = highest_episode_path.stat().st_mtime

    last_updated_at = ""
    if latest_mtime:
        last_updated_at = datetime.fromtimestamp(latest_mtime).date().isoformat()

    return {
        "epub_count": epub_count,
        "last_updated_at": last_updated_at,
        "highest_episode_file": highest_episode_path.name if highest_episode_path else "",
    }


def canonical_status(value: str) -> str:
    status = (value or "").strip()
    if status in {"연재", "연재중", "serializing"}:
        return "연재중"
    if status in {"완결", "completed"}:
        return "완결"
    if status in {"미런칭", "prelaunch"}:
        return "미런칭"
    return "미정"


def ebook_status_for_canonical(value: str) -> str:
    if value in {"있음", "없음", "미정"}:
        return value
    return "미정"


def has_recent_update(last_updated_at: str, recent_window_days: int) -> bool:
    if not last_updated_at:
        return False
    cutoff = date.today() - timedelta(days=recent_window_days)
    return date.fromisoformat(last_updated_at) >= cutoff


def scan_work_records(config: dict[str, Any], manual_overrides: dict[str, Any]) -> list[dict[str, Any]]:
    author_manager_map, work_author_map = load_workbook_mappings(config)
    manager_override_map = build_manager_override_map(manual_overrides)
    work_status_override_map = build_work_status_override_map(manual_overrides)
    recent_window_days = int(config.get("recent_update_window_days", 20))
    records: list[dict[str, Any]] = []

    for source in config["work_sources"]:
        source_path = Path(source["path"])
        default_status = source.get("default_status", "미정")
        source_group = source.get("group", "general")

        for folder in sorted(source_path.iterdir(), key=lambda item: item.name):
            if not folder.is_dir():
                continue
            if should_skip_folder(folder.name, source):
                continue

            serial_metrics = get_epub_metrics(folder)
            title, author_hint = split_folder_name(folder.name)
            resolution = resolve_author(title=title, author_hint=author_hint, work_author_map=work_author_map)
            resolved_manager = author_manager_map.get(normalize_text(resolution.author), "")
            manager_name = resolved_manager or "미정"
            manager_source = "담당자찾기" if resolved_manager else "미정"
            manager_status = "confirmed" if resolved_manager else "pending"
            manager_override = manager_override_map.get(str(folder))
            manager_override_note = ""
            if manager_override:
                manager_name = manager_override["manager_name"]
                manager_source = "수동오버라이드"
                manager_status = "confirmed"
                manager_override_note = manager_override.get("note", "")

            work_status_override = work_status_override_map.get(str(folder))
            platform_status = canonical_status(default_status)
            if work_status_override:
                platform_status = canonical_status(work_status_override["status"])

            issues = list(dict.fromkeys(resolution.issues))
            if not resolved_manager and not manager_override:
                issues.append("manager_unassigned")

            readj_path = folder / "readj.md"
            if not readj_path.exists():
                issues.append("metadata_missing")

            note_parts: list[str] = []
            if resolution.source == "folder_suffix":
                note_parts.append("폴더명 작가 표기 기준 매칭")
            elif resolution.source == "unresolved":
                note_parts.append("작가 확인 필요")
            if "work_author_ambiguous" in issues:
                note_parts.append("작품명 기준 작가 후보 복수")
            if "metadata_missing" in issues:
                note_parts.append("readj.md 없음")
            if "manager_unassigned" in issues:
                note_parts.append("담당자 미정")
            if manager_override:
                note_parts.append("담당자 수동확정")
                if manager_override_note:
                    note_parts.append(f"담당자 오버라이드 메모: {manager_override_note}")
            if work_status_override:
                note_parts.append("작품상태 수동확정")
                if work_status_override.get("note"):
                    note_parts.append(f"작품상태 오버라이드 메모: {work_status_override['note']}")

            record = {
                "source_root": source["name"],
                "source_group": source_group,
                "folder_name": folder.name,
                "folder_path": str(folder),
                "readj_path": str(readj_path),
                "readj_exists": readj_path.exists(),
                "title": title,
                "title_keys": build_title_keys(title),
                "title_normalized": normalize_title_key(title),
                "author": {
                    "name": resolution.author,
                    "source": resolution.source,
                },
                "manager": {
                    "name": manager_name,
                    "status": manager_status,
                    "source": manager_source,
                },
                "platform": {
                    "name": "",
                    "serialization_state": platform_status,
                    "episode_count": serial_metrics["epub_count"],
                },
                "serial": {
                    "epub_count": serial_metrics["epub_count"],
                    "last_updated_at": serial_metrics["last_updated_at"],
                    "recent_update_window_days": recent_window_days,
                    "recent_update_within_window": has_recent_update(serial_metrics["last_updated_at"], recent_window_days),
                },
                "ebook": {
                    "status": "없음",
                    "folder_name": "",
                    "folder_path": "",
                    "source_root": "",
                    "epub_count": 0,
                    "match_confidence": 0.0,
                    "match_reason": "",
                    "review_candidates": [],
                },
                "flags": [],
                "notes": note_parts,
                "issues": issues,
            }
            records.append(record)

    records.sort(key=lambda item: (item["manager"]["name"], item["title"]))
    return records


def scan_ebook_records(config: dict[str, Any]) -> list[dict[str, Any]]:
    ebook_records: list[dict[str, Any]] = []
    for source in config.get("ebook_sources", []):
        source_path = Path(source["path"])
        source_group = source.get("group", "general")

        for folder in sorted(source_path.iterdir(), key=lambda item: item.name):
            if not folder.is_dir():
                continue
            if should_skip_folder(folder.name, source):
                continue

            title, author_hint = split_folder_name(folder.name)
            ebook_records.append(
                {
                    "source_root": source["name"],
                    "source_group": source_group,
                    "folder_name": folder.name,
                    "folder_path": str(folder),
                    "title": title,
                    "title_keys": build_title_keys(title),
                    "title_normalized": normalize_title_key(title),
                    "author": {
                        "name": author_hint,
                    },
                    "epub_count": count_epubs(folder),
                }
            )

    return ebook_records


def author_bonus(work_author: str, ebook_author: str) -> float:
    if not work_author or not ebook_author:
        return 0.0
    return 0.02 if normalize_text(work_author) == normalize_text(ebook_author) else 0.0


def build_ebook_indices(ebook_records: list[dict[str, Any]]) -> tuple[dict[tuple[str, str], list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    by_group_key: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    by_group: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in ebook_records:
        by_group[record["source_group"]].append(record)
        for key in record["title_keys"]:
            by_group_key[(record["source_group"], key)].append(record)
    return by_group_key, by_group


def collect_match_candidates(
    work_record: dict[str, Any],
    ebook_by_group_key: dict[tuple[str, str], list[dict[str, Any]]],
    ebook_by_group: dict[str, list[dict[str, Any]]],
) -> list[MatchCandidate]:
    candidate_map: dict[str, MatchCandidate] = {}
    source_group = work_record["source_group"]

    for key in work_record["title_keys"]:
        for ebook_record in ebook_by_group_key.get((source_group, key), []):
            confidence = min(1.0, 0.98 + author_bonus(work_record["author"]["name"], ebook_record["author"]["name"]))
            candidate = MatchCandidate(
                folder_name=ebook_record["folder_name"],
                folder_path=ebook_record["folder_path"],
                source_root=ebook_record["source_root"],
                confidence=confidence,
                reason="exact_title_key",
                epub_count=ebook_record["epub_count"],
                title=ebook_record["title"],
            )
            current = candidate_map.get(candidate.folder_path)
            if current is None or current.confidence < candidate.confidence:
                candidate_map[candidate.folder_path] = candidate

    if not candidate_map:
        for ebook_record in ebook_by_group.get(source_group, []):
            work_key = work_record["title_normalized"]
            ebook_key = ebook_record["title_normalized"]
            if not work_key or not ebook_key:
                continue
            confidence = SequenceMatcher(None, work_key, ebook_key).ratio()
            confidence = min(0.99, confidence + author_bonus(work_record["author"]["name"], ebook_record["author"]["name"]))
            if confidence < 0.78:
                continue
            candidate = MatchCandidate(
                folder_name=ebook_record["folder_name"],
                folder_path=ebook_record["folder_path"],
                source_root=ebook_record["source_root"],
                confidence=confidence,
                reason="fuzzy_title",
                epub_count=ebook_record["epub_count"],
                title=ebook_record["title"],
            )
            current = candidate_map.get(candidate.folder_path)
            if current is None or current.confidence < candidate.confidence:
                candidate_map[candidate.folder_path] = candidate

    return sorted(candidate_map.values(), key=lambda item: (-item.confidence, item.folder_name))


def should_auto_assign(candidates: list[MatchCandidate], chosen: MatchCandidate | None) -> bool:
    if chosen is None:
        return False
    if chosen.confidence < 0.90:
        return False
    remaining = [candidate for candidate in candidates if candidate.folder_path != chosen.folder_path]
    second_confidence = remaining[0].confidence if remaining else 0.0
    if chosen.reason == "exact_title_key" and chosen.confidence >= 0.98:
        return chosen.confidence - second_confidence >= 0.01
    return chosen.confidence - second_confidence >= 0.03


def resolve_manual_ebook_record(
    ebook_folder_path: str,
    ebook_by_path: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    existing = ebook_by_path.get(ebook_folder_path)
    if existing is not None:
        return existing

    path = Path(ebook_folder_path)
    if not path.exists() or not path.is_dir():
        return None

    title, _ = split_folder_name(path.name)
    return {
        "folder_name": path.name,
        "folder_path": str(path),
        "source_root": str(path.parent),
        "epub_count": count_epubs(path),
        "title": title,
    }


def attach_ebook_matches(
    records: list[dict[str, Any]],
    ebook_records: list[dict[str, Any]],
    manual_overrides: dict[str, Any],
) -> tuple[int, int, int, int, int]:
    ebook_by_group_key, ebook_by_group = build_ebook_indices(ebook_records)
    ebook_by_path = {record["folder_path"]: record for record in ebook_records}
    ebook_exclusion_map = build_ebook_exclusion_map(manual_overrides)
    candidates_by_index = {
        index: collect_match_candidates(record, ebook_by_group_key, ebook_by_group)
        for index, record in enumerate(records)
    }

    used_ebook_paths: set[str] = set()
    assignments: dict[int, MatchCandidate] = {}
    manual_override_count = 0
    manual_exclusion_count = 0
    excluded_indices: set[int] = set()

    record_index_by_path = {record["folder_path"]: index for index, record in enumerate(records)}

    for work_folder_path, exclusion in ebook_exclusion_map.items():
        record_index = record_index_by_path.get(work_folder_path)
        if record_index is None:
            continue
        record = records[record_index]
        record["ebook"] = {
            "status": exclusion["status"],
            "folder_name": "",
            "folder_path": "",
            "source_root": "",
            "epub_count": 0,
            "match_confidence": 0.0,
            "match_reason": exclusion["reason"],
            "review_candidates": [],
        }
        record["notes"].append(f"단행본 {exclusion['status']} 처리")
        record["notes"].append("하네스 수동 제외 반영")
        if exclusion["note"]:
            record["notes"].append(f"단행본 제외 메모: {exclusion['note']}")
        excluded_indices.add(record_index)
        manual_exclusion_count += 1

    for item in manual_overrides.get("ebook_matches", []):
        work_folder_path = item.get("work_folder_path", "")
        ebook_folder_path = item.get("ebook_folder_path", "")
        if not work_folder_path or not ebook_folder_path:
            continue
        record_index = record_index_by_path.get(work_folder_path)
        ebook_record = resolve_manual_ebook_record(ebook_folder_path, ebook_by_path)
        if record_index is None or ebook_record is None or record_index in excluded_indices:
            continue
        note = item.get("note", "")
        assignments[record_index] = MatchCandidate(
            folder_name=ebook_record["folder_name"],
            folder_path=ebook_record["folder_path"],
            source_root=ebook_record["source_root"],
            confidence=1.0,
            reason="manual_override",
            epub_count=ebook_record["epub_count"],
            title=ebook_record["title"],
        )
        if note:
            records[record_index]["notes"].append(f"단행본 수동 메모: {note}")
        used_ebook_paths.add(ebook_folder_path)
        manual_override_count += 1

    sorted_indices = sorted(
        candidates_by_index,
        key=lambda index: candidates_by_index[index][0].confidence if candidates_by_index[index] else -1.0,
        reverse=True,
    )

    for index in sorted_indices:
        if index in assignments or index in excluded_indices:
            continue
        candidates = candidates_by_index[index]
        chosen = next((candidate for candidate in candidates if candidate.folder_path not in used_ebook_paths), None)
        if not should_auto_assign(candidates, chosen):
            continue
        assignments[index] = chosen
        used_ebook_paths.add(chosen.folder_path)

    auto_matched = 0
    review_needed = 0
    missing = 0

    for index, record in enumerate(records):
        if index in excluded_indices:
            continue
        if index in assignments:
            match = assignments[index]
            record["ebook"] = {
                "status": "있음",
                "folder_name": match.folder_name,
                "folder_path": match.folder_path,
                "source_root": match.source_root,
                "epub_count": match.epub_count,
                "match_confidence": round(match.confidence, 4),
                "match_reason": match.reason,
                "review_candidates": [],
            }
            if match.reason == "manual_override":
                record["notes"].append(f"단행본 수동확정 / epub {match.epub_count}개")
                record["notes"].append("하네스 수동 오버라이드 반영")
            else:
                record["notes"].append(f"단행본 자동매칭 {match.confidence:.0%} / epub {match.epub_count}개")
                auto_matched += 1
            continue

        candidates = candidates_by_index[index]
        if candidates:
            top_candidates = candidates[:5]
            record["ebook"] = {
                "status": "검토필요",
                "folder_name": "",
                "folder_path": "",
                "source_root": "",
                "epub_count": 0,
                "match_confidence": round(top_candidates[0].confidence, 4),
                "match_reason": "review_needed",
                "review_candidates": [
                    {
                        "folder_name": candidate.folder_name,
                        "folder_path": candidate.folder_path,
                        "source_root": candidate.source_root,
                        "confidence": round(candidate.confidence, 4),
                        "reason": candidate.reason,
                        "epub_count": candidate.epub_count,
                    }
                    for candidate in top_candidates
                ],
            }
            record["notes"].append("단행본 후보 검토 필요")
            review_needed += 1
        else:
            record["ebook"] = {
                "status": "없음",
                "folder_name": "",
                "folder_path": "",
                "source_root": "",
                "epub_count": 0,
                "match_confidence": 0.0,
                "match_reason": "not_found",
                "review_candidates": [],
            }
            missing += 1

    return auto_matched, review_needed, missing, manual_override_count, manual_exclusion_count


def export_json(records: list[dict[str, Any]], output_path: Path) -> None:
    output_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def export_pending_manager(records: list[dict[str, Any]], output_path: Path) -> None:
    pending: list[dict[str, Any]] = []
    for record in records:
        review_reasons = [issue for issue in record["issues"] if issue in {"author_unresolved", "manager_unassigned", "work_author_ambiguous"}]
        if not review_reasons:
            continue
        pending.append(
            {
                "title": record["title"],
                "folder_name": record["folder_name"],
                "folder_path": record["folder_path"],
                "source_group": record["source_group"],
                "author": record["author"]["name"],
                "manager": record["manager"]["name"],
                "manager_status": "미정",
                "reasons": review_reasons,
                "suggested_action": "담당자 미정으로 유지 후 추후 확정",
            }
        )

    output_path.write_text(json.dumps(pending, ensure_ascii=False, indent=2), encoding="utf-8")


def export_pending_readj(records: list[dict[str, Any]], output_path: Path, manager_filter: str = "") -> None:
    missing: list[dict[str, Any]] = []
    for record in records:
        if record["readj_exists"]:
            continue
        if manager_filter and record["manager"]["name"] != manager_filter:
            continue
        missing.append(
            {
                "title": record["title"],
                "folder_name": record["folder_name"],
                "folder_path": record["folder_path"],
                "source_group": record["source_group"],
                "author": record["author"]["name"],
                "manager": record["manager"]["name"],
                "platform_state": record["platform"]["serialization_state"],
                "ebook_status": record["ebook"]["status"],
                "readj_path": record["readj_path"],
            }
        )

    output_path.write_text(json.dumps(missing, ensure_ascii=False, indent=2), encoding="utf-8")


def export_pending_ebook_review(records: list[dict[str, Any]], output_path: Path) -> None:
    pending: list[dict[str, Any]] = []
    for record in records:
        if record["ebook"]["status"] != "검토필요":
            continue
        pending.append(
            {
                "title": record["title"],
                "folder_name": record["folder_name"],
                "folder_path": record["folder_path"],
                "source_group": record["source_group"],
                "manager": record["manager"]["name"],
                "ebook_status": record["ebook"]["status"],
                "suggested_action": "후보 확인 후 수동 확정",
                "top_candidates": record["ebook"]["review_candidates"],
            }
        )

    output_path.write_text(json.dumps(pending, ensure_ascii=False, indent=2), encoding="utf-8")


def export_missing_ebook(records: list[dict[str, Any]], output_path: Path) -> None:
    missing: list[dict[str, Any]] = []
    for record in records:
        if record["ebook"]["status"] != "없음":
            continue
        missing.append(
            {
                "title": record["title"],
                "folder_name": record["folder_name"],
                "folder_path": record["folder_path"],
                "source_group": record["source_group"],
                "manager": record["manager"]["name"],
                "suggested_action": "단행본 없음 여부 확인",
            }
        )

    output_path.write_text(json.dumps(missing, ensure_ascii=False, indent=2), encoding="utf-8")


def issue_notes(record: dict[str, Any]) -> str:
    if record["notes"]:
        return " / ".join(record["notes"])
    if record["issues"]:
        return " / ".join(record["issues"])
    return ""


def attach_canonical_fields(
    records: list[dict[str, Any]],
    work_cid_registry: dict[str, dict[str, str]] | None = None,
) -> None:
    work_cid_registry = work_cid_registry or {}
    for record in records:
        cid_entry = work_cid_registry.get(record["folder_path"], {})
        sales_channel_cids = split_pipe_list(cid_entry.get("sales_channel_cids", ""))
        sales_channel_names = split_pipe_list(cid_entry.get("sales_channel_names", ""))
        contract_ids = split_pipe_list(cid_entry.get("contract_ids", ""))

        record["cid"] = {
            "work_cid": (cid_entry.get("work_cid") or "").strip(),
            "legacy_work_id": (cid_entry.get("legacy_work_id") or "").strip(),
            "sales_channel_cids": sales_channel_cids,
            "sales_channel_names": sales_channel_names,
            "contract_ids": contract_ids,
            "remark": (cid_entry.get("remark") or "").strip(),
            "ips_name": (cid_entry.get("ips_name") or "").strip(),
            "match_status": (cid_entry.get("match_status") or "").strip(),
            "match_reason": (cid_entry.get("match_reason") or "").strip(),
        }
        record["canonical"] = {
            "work_id": record["cid"]["legacy_work_id"],
            "work_cid": record["cid"]["work_cid"],
            "title": record["title"],
            "author": record["author"]["name"],
            "manager": record["manager"]["name"],
            "status": canonical_status(record["platform"]["serialization_state"]),
            "episode_count": record["platform"]["episode_count"],
            "recent_update_within_20d": record["serial"]["recent_update_within_window"],
            "last_update_at": record["serial"]["last_updated_at"],
            "serial_epub_count": record["serial"]["epub_count"],
            "ebook_status": ebook_status_for_canonical(record["ebook"]["status"]),
            "ebook_epub_count": record["ebook"]["epub_count"],
            "sales_channel_cids": sales_channel_cids,
            "sales_channel_names": sales_channel_names,
            "contract_ids": contract_ids,
            "remark": record["cid"]["remark"],
            "ips_name": record["cid"]["ips_name"],
            "note": issue_notes(record),
        }


def export_canonical_json(records: list[dict[str, Any]], output_path: Path) -> None:
    payload = [record["canonical"] for record in records]
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export_csv(records: list[dict[str, Any]], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["작품", "작가", "담당자", "상태", "화수", "최근20일업데이트", "최근업데이트일", "단행본", "연재EPUB수", "단행본EPUB수", "특이사항"],
        )
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "작품": record["title"],
                    "작가": record["author"]["name"],
                    "담당자": record["manager"]["name"],
                    "상태": record["canonical"]["status"],
                    "화수": record["canonical"]["episode_count"] or "",
                    "최근20일업데이트": "Y" if record["canonical"]["recent_update_within_20d"] else "N",
                    "최근업데이트일": record["canonical"]["last_update_at"],
                    "단행본": record["canonical"]["ebook_status"],
                    "연재EPUB수": record["canonical"]["serial_epub_count"] or "",
                    "단행본EPUB수": record["canonical"]["ebook_epub_count"] or "",
                    "특이사항": issue_notes(record),
                }
            )


def export_detailed_csv(records: list[dict[str, Any]], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["담당자", "작품", "구분", "상태", "화수", "단행본상태", "단행본EPUB수", "단행본매칭신뢰도", "특이사항"],
        )
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "담당자": record["manager"]["name"],
                    "작품": record["title"],
                    "구분": "성인" if record["source_group"] == "adult" else "일반",
                    "상태": record["canonical"]["status"],
                    "화수": record["canonical"]["episode_count"] or "",
                    "단행본상태": record["canonical"]["ebook_status"],
                    "단행본EPUB수": record["canonical"]["ebook_epub_count"] or "",
                    "단행본매칭신뢰도": record["ebook"]["match_confidence"] or "",
                    "특이사항": (
                        f"최근20일업데이트={'Y' if record['canonical']['recent_update_within_20d'] else 'N'}"
                        + (f" / 최근업데이트일={record['canonical']['last_update_at']}" if record["canonical"]["last_update_at"] else "")
                        + (f" / {issue_notes(record)}" if issue_notes(record) else "")
                    ),
                }
            )


def filter_records_by_manager(records: list[dict[str, Any]], manager_name: str) -> list[dict[str, Any]]:
    manager_key = normalize_text(manager_name)
    return [record for record in records if normalize_text(record["manager"]["name"]) == manager_key]


def main() -> None:
    parser = argparse.ArgumentParser(description="Build writer-management work index from NAS folders and Excel mappings.")
    parser.add_argument(
        "--config",
        default="config/paths.local.json",
        help="Path to the local config JSON file.",
    )
    parser.add_argument(
        "--manager",
        default="",
        help="Optional manager name to build a filtered my-works export.",
    )
    args = parser.parse_args()

    config_path = Path(args.config)
    config = load_config(config_path)
    manual_overrides = load_manual_overrides(config)
    work_cid_registry = load_work_cid_registry(DEFAULT_WORK_CID_REGISTRY_PATH)

    records = scan_work_records(config, manual_overrides)
    ebook_records = scan_ebook_records(config)
    auto_matched, review_needed, missing, manual_override_count, manual_exclusion_count = attach_ebook_matches(records, ebook_records, manual_overrides)
    attach_canonical_fields(records, work_cid_registry)

    catalog_dir = Path("data/catalog")
    exports_dir = Path("data/exports")
    catalog_dir.mkdir(parents=True, exist_ok=True)
    exports_dir.mkdir(parents=True, exist_ok=True)

    export_json(records, catalog_dir / "work_index.json")
    export_canonical_json(records, catalog_dir / "canonical_work_index.json")
    export_json(ebook_records, catalog_dir / "ebook_index.json")
    export_pending_manager(records, catalog_dir / "pending_manager_review.json")
    export_pending_ebook_review(records, catalog_dir / "pending_ebook_match_review.json")
    export_missing_ebook(records, catalog_dir / "ebook_missing.json")
    export_pending_readj(records, catalog_dir / "pending_readj_seed.json")

    target_manager = config.get("readj_seed_target_manager", "")
    if target_manager:
        export_pending_readj(records, catalog_dir / "pending_readj_seed_target_manager.json", manager_filter=target_manager)

    export_csv(records, exports_dir / "management_table.csv")
    export_detailed_csv(records, exports_dir / "management_table_detailed.csv")

    if args.manager:
        filtered = filter_records_by_manager(records, args.manager)
        safe_name = normalize_title_key(args.manager) or "manager"
        export_json(filtered, exports_dir / f"my_works_{safe_name}.json")
        export_csv(filtered, exports_dir / f"my_works_{safe_name}.csv")
        export_detailed_csv(filtered, exports_dir / f"my_works_{safe_name}_detailed.csv")
        print(f"filtered_manager={args.manager}")
        print(f"filtered_records={len(filtered)}")

    print(f"total_records={len(records)}")
    print(f"ebook_records={len(ebook_records)}")
    print(
        "pending_manager_reviews="
        f"{sum(1 for record in records if any(issue in {'author_unresolved', 'manager_unassigned', 'work_author_ambiguous'} for issue in record['issues']))}"
    )
    print(f"missing_readj={sum(1 for record in records if not record['readj_exists'])}")
    print(f"ebook_auto_matched={auto_matched}")
    print(f"ebook_manual_overrides={manual_override_count}")
    print(f"ebook_manual_exclusions={manual_exclusion_count}")
    print(f"ebook_review_needed={review_needed}")
    print(f"ebook_missing={missing}")


if __name__ == "__main__":
    main()
