from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

from build_account_observation_bundle import (
    MANAGER,
    TARGET_ROOT,
    _GOOD_FILL,
    _WARN_FILL,
    apply_body_style,
    apply_header_style,
    compact_key,
    latest_path,
    load_frame,
    normalize_text,
    normalize_matchable_work_title,
    set_column_widths,
    write_df_sheet,
)


INPUT_ACTUAL_CSV = TARGET_ROOT / f"latest__account_actual_inventory_{MANAGER}.csv"
INPUT_DECISION_CSV = TARGET_ROOT / f"latest__account_decision_queue_{MANAGER}.csv"
OUTPUT_XLSX = TARGET_ROOT / f"latest__account_work_gap_queue_{MANAGER}.xlsx"
OUTPUT_CSV = TARGET_ROOT / f"latest__account_work_gap_queue_{MANAGER}.csv"
OUTPUT_JSON = TARGET_ROOT / f"latest__account_work_gap_queue_{MANAGER}.json"

_LEADING_TAG_RE = re.compile(r"^[\[(]([^\])]+)[\])]\s*")
_TRAILING_VOLUME_RE = re.compile(r"[\s_]+\d+(?:-\d+)?\s*(?:화|권|회)(?:\s*[(（]?(?:완|완결|사용x|사용금지)[)）]?)?\s*$")
_TRAILING_COMPLETION_RE = re.compile(r"\s*[(（]?(?:완|완결)[)）]?\s*$")
_TRAILING_SET_RE = re.compile(r"\s*(?:합본|세트)\s*$")
_TRAILING_SET_PAREN_RE = re.compile(r"\s*[(（]세트[)）]\s*$")
_TRAILING_FULLSET_RE = re.compile(r"\s*[(（]전\d+권\s*완결[)）]\s*$")
_TRAILING_ASCII_PAREN_RE = re.compile(r"\s*[(（][A-Za-z0-9 .,'&_-]+[)）]\s*$")
_LEADING_PREFIXES = (
    "카카오_",
    "카카오창작지원금_",
    "단행본_",
    "단행_",
)
_TAG_KEYWORDS = (
    "카카오",
    "네이버",
    "윌라",
    "봄툰",
    "문피아",
    "리디",
    "세트",
    "단행본",
    "광고수익",
    "웹툰",
    "원작",
    "타플",
    "연재",
    "특별",
    "오디오북",
)


def parse_int(value: object) -> int:
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return 0
    return int(parsed)


def normalize_gap_title(value: object) -> str:
    current = normalize_matchable_work_title(value).replace("_", " ")
    for _ in range(3):
        before = current
        current = _TRAILING_ASCII_PAREN_RE.sub("", current)
        current = normalize_text(current).strip(" -_?[]()")
        if current == before:
            break
    return normalize_text(current)


def build_candidate_lookup() -> dict[tuple[str, str], list[dict[str, str]]]:
    mapping_path = latest_path(f"*__NAS_IPS_ACCOUNT_예금주_{MANAGER}.csv")
    mapping_df = load_frame(mapping_path)
    candidate_rows: dict[tuple[str, str], dict[str, dict[str, str]]] = {}

    for row in mapping_df.to_dict("records"):
        author_name = normalize_text(row.get("author_final"))
        title = normalize_text(row.get("title_final"))
        folder_path = normalize_text(row.get("folder_path"))
        normalized_title = normalize_gap_title(title)
        if not author_name or not title or not folder_path or not normalized_title:
            continue

        key = (author_name, normalized_title)
        bucket = candidate_rows.setdefault(key, {})
        bucket_key = f"{folder_path}||{title}"
        bucket[bucket_key] = {
            "folder_path": folder_path,
            "title_final": title,
            "matched_copyright_codes": normalize_text(row.get("matched_copyright_codes")),
        }

    return {key: list(value.values()) for key, value in candidate_rows.items()}


def summarize_candidate_codes(
    current_code: str,
    candidate_codes_raw: str,
    candidate_folder_count: str,
) -> tuple[str, str, str]:
    current_code = normalize_text(current_code)
    candidate_codes = [normalize_text(code) for code in str(candidate_codes_raw or "").split("|")]
    candidate_codes = [code for code in candidate_codes if code]
    unique_codes = sorted(set(candidate_codes))
    same_code = current_code != "" and current_code in unique_codes
    has_folder_candidate = parse_int(candidate_folder_count) > 0

    if has_folder_candidate and not unique_codes:
        return "코드없음후보", "N", "0"
    if not unique_codes:
        return "없음", "N", "0"
    if same_code:
        return "동일코드포함", "Y", str(len(unique_codes))
    return "타코드전용", "N", str(len(unique_codes))


def build_work_gap_df() -> pd.DataFrame:
    actual_df = pd.read_csv(INPUT_ACTUAL_CSV, dtype=str).fillna("")
    decision_df = pd.read_csv(INPUT_DECISION_CSV, dtype=str).fillna("")
    candidate_lookup = build_candidate_lookup()

    filtered = actual_df[actual_df["관측_미매핑_작품수"].map(parse_int).gt(0)].copy()
    if filtered.empty:
        return pd.DataFrame(
            columns=[
                "대표작가명",
                "account_작가코드",
                "account_저작권코드",
                "현재_저작권명",
                "action_제안",
                "특수",
                "관측_작품수",
                "관측_작품_폴더수",
                "관측_미매핑_작품수",
                "관측_미매핑_작품목록",
                "추정_폴더후보수",
                "추정_작품후보목록",
                "추정_저작권코드후보목록",
                "추정_폴더후보목록",
                "관측_작품목록",
                "관측_작품_폴더",
                "대표_제목",
                "연결_선인세명",
                "주의플래그",
            ]
        )

    decision_lookup = decision_df[
        [
            "account_작가코드",
            "account_저작권코드",
            "대표작가명",
            "현재_저작권명",
            "action_제안",
            "특수",
            "주의플래그",
        ]
    ].drop_duplicates(subset=["account_작가코드", "account_저작권코드"])

    merged = filtered.merge(
        decision_lookup,
        on=["account_작가코드", "account_저작권코드"],
        how="left",
    ).fillna("")

    candidate_counts: list[str] = []
    candidate_folders: list[str] = []
    candidate_titles: list[str] = []
    candidate_codes: list[str] = []
    for row in merged.to_dict("records"):
        author_name = normalize_text(row.get("대표작가명")) or normalize_text(row.get("scope_author_primary"))
        seen_titles: set[str] = set()
        seen_folders: set[str] = set()
        seen_codes: set[str] = set()
        for raw_title in str(row.get("관측_미매핑_작품목록") or "").split("|"):
            title = normalize_text(raw_title)
            normalized_title = normalize_gap_title(title)
            if not author_name or not normalized_title:
                continue
            for candidate in candidate_lookup.get((author_name, normalized_title), []):
                title_final = normalize_text(candidate.get("title_final"))
                folder_path = normalize_text(candidate.get("folder_path"))
                matched_codes = normalize_text(candidate.get("matched_copyright_codes"))
                if title_final:
                    seen_titles.add(title_final)
                if folder_path:
                    seen_folders.add(folder_path)
                for code in matched_codes.split("|"):
                    code_clean = normalize_text(code)
                    if code_clean:
                        seen_codes.add(code_clean)
        candidate_counts.append(str(len(seen_folders)))
        candidate_folders.append(" | ".join(sorted(seen_folders)))
        candidate_titles.append(" | ".join(sorted(seen_titles)))
        candidate_codes.append(" | ".join(sorted(seen_codes)))

    merged["추정_폴더후보수"] = candidate_counts
    merged["추정_폴더후보목록"] = candidate_folders
    merged["추정_작품후보목록"] = candidate_titles
    merged["추정_저작권코드후보목록"] = candidate_codes
    candidate_meta = merged.apply(
        lambda row: summarize_candidate_codes(
            normalize_text(row.get("account_저작권코드")),
            normalize_text(row.get("추정_저작권코드후보목록")),
            normalize_text(row.get("추정_폴더후보수")),
        ),
        axis=1,
        result_type="expand",
    )
    merged["추정_후보유형"] = candidate_meta[0]
    merged["후보_동일저작권코드포함(Y/N)"] = candidate_meta[1]
    merged["추정_저작권코드후보수"] = candidate_meta[2]

    merged["정렬_미매핑수"] = merged["관측_미매핑_작품수"].map(parse_int)
    merged["정렬_후보수"] = merged["추정_폴더후보수"].map(parse_int)
    merged["정렬_후보유형"] = merged["추정_후보유형"].map(
        {
            "동일코드포함": 0,
            "타코드전용": 1,
            "코드없음후보": 2,
            "없음": 3,
        }
    ).fillna(9)
    merged["정렬_관측수"] = merged["관측_작품수"].map(parse_int)
    merged["정렬_대표작가명"] = merged["대표작가명"].map(normalize_text)
    merged["정렬_저작권코드"] = pd.to_numeric(merged["account_저작권코드"], errors="coerce").fillna(10**18)
    merged = (
        merged.sort_values(
            by=["정렬_미매핑수", "정렬_후보수", "정렬_후보유형", "정렬_관측수", "정렬_대표작가명", "정렬_저작권코드"],
            ascending=[False, False, True, False, True, True],
            kind="stable",
        )
        .drop(columns=["정렬_미매핑수", "정렬_후보수", "정렬_후보유형", "정렬_관측수", "정렬_대표작가명", "정렬_저작권코드"])
        .reset_index(drop=True)
    )

    columns = [
        "대표작가명",
        "account_작가코드",
        "account_저작권코드",
        "현재_저작권명",
        "action_제안",
        "특수",
        "관측_작품수",
        "관측_작품_폴더수",
        "관측_미매핑_작품수",
        "관측_미매핑_작품목록",
        "추정_폴더후보수",
        "추정_후보유형",
        "후보_동일저작권코드포함(Y/N)",
        "추정_저작권코드후보수",
        "추정_작품후보목록",
        "추정_저작권코드후보목록",
        "추정_폴더후보목록",
        "관측_작품목록",
        "관측_작품_폴더",
        "대표_제목",
        "연결_선인세명",
        "주의플래그",
    ]
    return merged[columns].copy()


def build_gap_cluster_df(gap_df: pd.DataFrame) -> pd.DataFrame:
    exploded_rows: list[dict[str, str]] = []
    for row in gap_df.to_dict("records"):
        for raw_title in str(row.get("관측_미매핑_작품목록") or "").split("|"):
            title = normalize_text(raw_title)
            if not title:
                continue
            normalized_title = normalize_gap_title(title)
            exploded_rows.append(
                {
                    "대표작가명": normalize_text(row.get("대표작가명")),
                    "account_저작권코드": normalize_text(row.get("account_저작권코드")),
                    "action_제안": normalize_text(row.get("action_제안")),
                    "원본_작품명": title,
                    "정규화_작품명": normalized_title or title,
                    "정규화_작품키": compact_key(normalized_title or title),
                }
            )

    if not exploded_rows:
        return pd.DataFrame(
            columns=[
                "정규화_작품명",
                "원본변형수",
                "관련_권리행수",
                "관련_대표작가명목록",
                "관련_저작권코드목록",
                "action_제안목록",
                "원본변형목록",
            ]
        )

    exploded_df = pd.DataFrame(exploded_rows)
    grouped_rows: list[dict[str, str]] = []
    for _, group in exploded_df.groupby(["정규화_작품키", "정규화_작품명"], dropna=False, sort=False):
        grouped_rows.append(
            {
                "정규화_작품명": normalize_text(group["정규화_작품명"].iloc[0]),
                "원본변형수": str(group["원본_작품명"].nunique()),
                "관련_권리행수": str(group[["대표작가명", "account_저작권코드"]].drop_duplicates().shape[0]),
                "관련_대표작가명목록": " | ".join(sorted({normalize_text(v) for v in group["대표작가명"] if normalize_text(v)})),
                "관련_저작권코드목록": " | ".join(sorted({normalize_text(v) for v in group["account_저작권코드"] if normalize_text(v)})),
                "action_제안목록": " | ".join(sorted({normalize_text(v) for v in group["action_제안"] if normalize_text(v)})),
                "원본변형목록": " | ".join(sorted({normalize_text(v) for v in group["원본_작품명"] if normalize_text(v)})),
            }
        )

    clustered_df = pd.DataFrame(grouped_rows)
    clustered_df["정렬_변형수"] = clustered_df["원본변형수"].map(parse_int)
    clustered_df["정렬_권리행수"] = clustered_df["관련_권리행수"].map(parse_int)
    clustered_df = (
        clustered_df.sort_values(
            by=["정렬_변형수", "정렬_권리행수", "정규화_작품명"],
            ascending=[False, False, True],
            kind="stable",
        )
        .drop(columns=["정렬_변형수", "정렬_권리행수"])
        .reset_index(drop=True)
    )
    return clustered_df


def write_summary_sheet(wb: Workbook, gap_df: pd.DataFrame, cluster_df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{MANAGER} account work gap queue"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "지표"
    ws["B3"] = "값"
    apply_header_style(ws["A3"])
    apply_header_style(ws["B3"])

    metrics = [
        ("raw 미매핑 권리 행 수", len(gap_df)),
        ("raw 미매핑 작품 수 합계", int(pd.to_numeric(gap_df["관측_미매핑_작품수"], errors="coerce").fillna(0).sum())),
        ("정규화 클러스터 수", len(cluster_df)),
        ("추정 폴더 후보 보유 행 수", int(pd.to_numeric(gap_df["추정_폴더후보수"], errors="coerce").fillna(0).gt(0).sum())),
        ("동일코드 후보 포함 행 수", int(gap_df["후보_동일저작권코드포함(Y/N)"].eq("Y").sum())),
        ("타코드 전용 후보 행 수", int(gap_df["추정_후보유형"].eq("타코드전용").sum())),
        ("코드없음 후보 행 수", int(gap_df["추정_후보유형"].eq("코드없음후보").sum())),
        ("CID분해필요 행 수", int(gap_df["action_제안"].eq("CID분해필요").sum())),
        ("이름수정검토 행 수", int(gap_df["action_제안"].eq("이름수정검토").sum())),
        ("작품확인 행 수", int(gap_df["action_제안"].eq("작품확인").sum())),
    ]

    for row_idx, (label, value) in enumerate(metrics, start=4):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2))

    start_row = 12
    ws.cell(row=start_row, column=1, value="action_제안")
    ws.cell(row=start_row, column=2, value="행 수")
    apply_header_style(ws.cell(row=start_row, column=1))
    apply_header_style(ws.cell(row=start_row, column=2))

    action_counter = Counter(gap_df["action_제안"].tolist())
    labels = ["CID분해필요", "이름수정검토", "작품확인", "작가확인", "유지"]
    for idx, label in enumerate(labels, start=start_row + 1):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=action_counter.get(label, 0))
        apply_body_style(ws.cell(row=idx, column=1))
        apply_body_style(ws.cell(row=idx, column=2))

    chart = BarChart()
    chart.title = "work gap action 분포"
    chart.y_axis.title = "행 수"
    chart.x_axis.title = "action"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(labels))
    categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(labels))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "D4")

    ws["A20"] = "해석"
    ws["A21"] = "raw popup에는 있지만 registry/NAS 폴더가 아직 안 잡힌 작품 목록이다."
    ws["A22"] = "판단용 큐가 아니라 stale/누락 관측용 리포트다."
    ws["A23"] = "관측_작품목록은 raw 기준, 관측_작품_폴더는 registry 기준이다."
    ws["A24"] = "normalized_gap_clusters 시트는 플랫폼 태그, 세트, 권수 표기를 자동으로 접은 보조 요약이다."
    ws["A25"] = "추정_폴더후보* 컬럼은 같은 작가의 다른 저작권코드에서 이미 잡힌 폴더 후보를 보여준다."
    ws["A26"] = "추정_후보유형=동일코드포함 이면 현재 저작권코드와도 이미 연결 흔적이 있는 강한 후보다."
    ws["A27"] = "추정_후보유형=코드없음후보 이면 폴더 후보는 있지만 mapping summary에 저작권코드 연결 흔적은 비어 있다."
    apply_header_style(ws["A20"])
    apply_body_style(ws["A21"], fill=_GOOD_FILL, wrap=True)
    apply_body_style(ws["A22"], fill=_WARN_FILL, wrap=True)
    apply_body_style(ws["A23"], wrap=True)
    apply_body_style(ws["A24"], wrap=True)
    apply_body_style(ws["A25"], wrap=True)
    apply_body_style(ws["A26"], wrap=True)
    apply_body_style(ws["A27"], wrap=True)

    set_column_widths(ws)


def write_outputs(gap_df: pd.DataFrame, cluster_df: pd.DataFrame) -> dict[str, Path]:
    TARGET_ROOT.mkdir(parents=True, exist_ok=True)
    gap_df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    wb = Workbook()
    write_summary_sheet(wb, gap_df, cluster_df)
    write_df_sheet(wb, "work_gap_queue", gap_df)
    candidate_df = gap_df[pd.to_numeric(gap_df["추정_폴더후보수"], errors="coerce").fillna(0).gt(0)].copy()
    write_df_sheet(wb, "candidate_gap_queue", candidate_df)
    write_df_sheet(wb, "normalized_gap_clusters", cluster_df)

    try:
        wb.save(OUTPUT_XLSX)
    except PermissionError as exc:
        raise RuntimeError(
            f"최신판 1개 정책을 유지하려면 '{OUTPUT_XLSX.name}' 파일을 닫은 뒤 다시 실행해야 합니다."
        ) from exc

    summary = {
        "manager": MANAGER,
        "gap_rows": int(len(gap_df)),
        "unmapped_work_total": int(pd.to_numeric(gap_df["관측_미매핑_작품수"], errors="coerce").fillna(0).sum()),
        "normalized_clusters": int(len(cluster_df)),
        "candidate_rows": int(pd.to_numeric(gap_df["추정_폴더후보수"], errors="coerce").fillna(0).gt(0).sum()),
        "same_code_candidate_rows": int(gap_df["후보_동일저작권코드포함(Y/N)"].eq("Y").sum()),
        "code_less_candidate_rows": int(gap_df["추정_후보유형"].eq("코드없음후보").sum()),
        "workbook": str(OUTPUT_XLSX),
    }
    OUTPUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "csv": OUTPUT_CSV,
        "workbook": OUTPUT_XLSX,
        "summary": OUTPUT_JSON,
    }


def main() -> None:
    gap_df = build_work_gap_df()
    cluster_df = build_gap_cluster_df(gap_df)
    outputs = write_outputs(gap_df, cluster_df)
    print("=== account work gap queue built ===")
    print(
        json.dumps(
            {
                "manager": MANAGER,
                "gap_rows": len(gap_df),
                "normalized_clusters": len(cluster_df),
                "candidate_rows": int(pd.to_numeric(gap_df["추정_폴더후보수"], errors="coerce").fillna(0).gt(0).sum()),
                "same_code_candidate_rows": int(gap_df["후보_동일저작권코드포함(Y/N)"].eq("Y").sum()),
                "code_less_candidate_rows": int(gap_df["추정_후보유형"].eq("코드없음후보").sum()),
                "workbook": str(outputs["workbook"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
