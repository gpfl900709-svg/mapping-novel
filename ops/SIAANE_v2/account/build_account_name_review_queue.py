from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

from build_account_observation_bundle import (
    MANAGER,
    TARGET_ROOT,
    _GOOD_FILL,
    _WARN_FILL,
    apply_body_style,
    apply_header_style,
    normalize_text,
    set_column_widths,
    write_df_sheet,
)


INPUT_CSV = TARGET_ROOT / f"latest__account_decision_queue_{MANAGER}.csv"
OUTPUT_XLSX = TARGET_ROOT / f"latest__account_name_review_queue_{MANAGER}.xlsx"
OUTPUT_CSV = TARGET_ROOT / f"latest__account_name_review_queue_{MANAGER}.csv"
OUTPUT_JSON = TARGET_ROOT / f"latest__account_name_review_queue_{MANAGER}.json"


def review_question(row: pd.Series) -> str:
    suggested_name = normalize_text(row.get("canonical_저작권명_제안"))
    special = normalize_text(row.get("특수"))
    if special in {"", "일반"}:
        return "대표작품명으로 확정해도 되는가?"
    return f"'{suggested_name}'로 확정해도 되는가?"


def review_priority(row: pd.Series) -> str:
    special = normalize_text(row.get("특수"))
    if special in {"카카오MG", "네이버MG", "원작"}:
        return "특수먼저"
    if normalize_text(row.get("작가특수RS요약")):
        return "RS확인"
    return "일반"


def build_name_review_df() -> pd.DataFrame:
    df = pd.read_csv(INPUT_CSV, dtype=str).fillna("")
    filtered = df[
        df["action_제안"].eq("이름수정검토")
        & df["처리완료(Y/N)"].map(normalize_text).ne("Y")
    ].copy()
    filtered["권장_수동_최종저작권명"] = filtered["canonical_저작권명_제안"].map(normalize_text)
    filtered["검토질문"] = filtered.apply(review_question, axis=1)
    filtered["검토우선순위"] = filtered.apply(review_priority, axis=1)
    filtered["정렬_대표작품"] = filtered["대표작품"].map(normalize_text)
    filtered["정렬_대표작가명"] = filtered["대표작가명"].map(normalize_text)
    filtered["정렬_특수우선"] = filtered["검토우선순위"].map({"특수먼저": 0, "RS확인": 1, "일반": 2}).fillna(9)
    filtered["정렬_저작권코드"] = pd.to_numeric(filtered["account_저작권코드"], errors="coerce").fillna(10**18)
    filtered = (
        filtered.sort_values(
            by=["정렬_특수우선", "정렬_대표작가명", "정렬_대표작품", "정렬_저작권코드"],
            kind="stable",
        )
        .drop(columns=["정렬_대표작품", "정렬_대표작가명", "정렬_특수우선", "정렬_저작권코드"])
        .reset_index(drop=True)
    )

    columns = [
        "검토우선순위",
        "대표작가명",
        "account_저작권코드",
        "현재_저작권명",
        "대표작품",
        "특수",
        "권장_수동_최종저작권명",
        "검토질문",
        "판정근거",
        "관측_작품목록",
        "연결_선인세명",
        "작가특수RS요약",
        "수동_최종저작권명",
        "수동_action",
        "수동_메모",
        "처리완료(Y/N)",
    ]
    return filtered[columns].copy()


def write_summary_sheet(wb: Workbook, review_df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{MANAGER} account name review queue"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "지표"
    ws["B3"] = "값"
    apply_header_style(ws["A3"])
    apply_header_style(ws["B3"])

    metrics = [
        ("미처리 이름수정검토 행 수", len(review_df)),
        ("특수먼저 행 수", int(review_df["검토우선순위"].eq("특수먼저").sum())),
        ("RS확인 행 수", int(review_df["검토우선순위"].eq("RS확인").sum())),
        ("일반 행 수", int(review_df["검토우선순위"].eq("일반").sum())),
        ("수동입력 완료 행 수", int(review_df["수동_최종저작권명"].astype(str).str.strip().ne("").sum())),
    ]

    for row_idx, (label, value) in enumerate(metrics, start=4):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2))

    start_row = 12
    ws.cell(row=start_row, column=1, value="검토우선순위")
    ws.cell(row=start_row, column=2, value="행 수")
    apply_header_style(ws.cell(row=start_row, column=1))
    apply_header_style(ws.cell(row=start_row, column=2))

    priority_counter = Counter(review_df["검토우선순위"].tolist())
    labels = ["특수먼저", "RS확인", "일반"]
    for idx, label in enumerate(labels, start=start_row + 1):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=priority_counter.get(label, 0))
        apply_body_style(ws.cell(row=idx, column=1))
        apply_body_style(ws.cell(row=idx, column=2))

    chart = BarChart()
    chart.title = "name review 우선순위 분포"
    chart.y_axis.title = "행 수"
    chart.x_axis.title = "우선순위"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(labels))
    categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(labels))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "D4")

    ws["A18"] = "작업 기준"
    ws["A19"] = "대표작품명으로 닫을 수 있으면 수동_최종저작권명 입력"
    ws["A20"] = "애매하면 수동_최종저작권명은 비우고 수동_action만 입력"
    ws["A21"] = "이 파일은 처리완료(Y/N) != Y 인 미처리 건만 보여준다"
    apply_header_style(ws["A18"])
    apply_body_style(ws["A19"], fill=_GOOD_FILL, wrap=True)
    apply_body_style(ws["A20"], fill=_WARN_FILL, wrap=True)
    apply_body_style(ws["A21"], wrap=True)

    set_column_widths(ws)


def write_outputs(review_df: pd.DataFrame) -> dict[str, Path]:
    TARGET_ROOT.mkdir(parents=True, exist_ok=True)
    review_df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    wb = Workbook()
    write_summary_sheet(wb, review_df)
    write_df_sheet(wb, "name_review_queue", review_df)

    try:
        wb.save(OUTPUT_XLSX)
    except PermissionError as exc:
        raise RuntimeError(
            f"최신판 1개 정책을 유지하려면 '{OUTPUT_XLSX.name}' 파일을 닫은 뒤 다시 실행해야 합니다."
        ) from exc

    summary = {
        "manager": MANAGER,
        "pending_name_review_rows": int(len(review_df)),
        "priority_special_rows": int(review_df["검토우선순위"].eq("특수먼저").sum()),
        "priority_rs_rows": int(review_df["검토우선순위"].eq("RS확인").sum()),
        "priority_general_rows": int(review_df["검토우선순위"].eq("일반").sum()),
        "workbook": str(OUTPUT_XLSX),
    }
    OUTPUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "csv": OUTPUT_CSV,
        "workbook": OUTPUT_XLSX,
        "summary": OUTPUT_JSON,
    }


def main() -> None:
    review_df = build_name_review_df()
    outputs = write_outputs(review_df)
    print("=== account name review queue built ===")
    print(
        json.dumps(
            {
                "manager": MANAGER,
                "pending_name_review_rows": len(review_df),
                "workbook": str(outputs["workbook"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
