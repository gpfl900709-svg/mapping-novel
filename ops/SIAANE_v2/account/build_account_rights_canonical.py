from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from build_account_observation_bundle import (
    MANAGER,
    ROOT,
    TARGET_ROOT,
    _GOOD_FILL,
    _WARN_FILL,
    apply_body_style,
    apply_header_style,
    load_frame,
    normalize_text,
    pipe_join,
    set_column_widths,
    split_pipe_values,
    write_df_sheet,
)


DECISION_PATH = TARGET_ROOT / f"latest__account_decision_queue_{MANAGER}.csv"
CID_SEED_PATH = ROOT / "SIAANE_v2" / "crosswalk" / "exports" / f"latest__account_ips_cid_seed_{MANAGER}.csv"
OUTPUT_XLSX = TARGET_ROOT / f"latest__account_rights_canonical_{MANAGER}.xlsx"
OUTPUT_CSV = TARGET_ROOT / f"latest__account_rights_canonical_{MANAGER}.csv"
OUTPUT_JSON = TARGET_ROOT / f"latest__account_rights_canonical_{MANAGER}.json"

STATUS_ORDER = {
    "담당제외": 0,
    "CID분해": 1,
    "이름확정": 2,
    "유지": 3,
    "IPS_CID생성보류": 4,
    "작가확정보류": 5,
    "검토필요": 9,
}


def unique(values: list[Any]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        cleaned = normalize_text(value)
        if not cleaned or cleaned in seen:
            continue
        ordered.append(cleaned)
        seen.add(cleaned)
    return ordered


def parse_int(value: Any) -> int:
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return 0
    return int(parsed)


def resolve_canonical_status(row: dict[str, Any]) -> str:
    action = normalize_text(row.get("action_제안"))
    manual_action = normalize_text(row.get("수동_action"))
    processed = normalize_text(row.get("처리완료(Y/N)"))

    if manual_action == "담당제외":
        return "담당제외"
    if manual_action == "IPS_CID생성보류":
        return "IPS_CID생성보류"
    if manual_action == "작가확정보류":
        return "작가확정보류"
    if action == "CID분해필요":
        return "CID분해"
    if action == "이름수정검토" and normalize_text(row.get("수동_최종저작권명")):
        return "이름확정"
    if action == "유지":
        return "유지"
    if processed == "Y":
        return f"{action or '처리완료'}"
    return "검토필요"


def resolve_final_rights_name(row: dict[str, Any], status: str) -> tuple[str, str]:
    manual_name = normalize_text(row.get("수동_최종저작권명"))
    current_name = normalize_text(row.get("현재_저작권명"))
    suggested_name = normalize_text(row.get("canonical_저작권명_제안"))

    if manual_name:
        return manual_name, "수동_최종저작권명"
    if status == "이름확정" and suggested_name:
        return suggested_name, "canonical_저작권명_제안"
    if status == "CID분해":
        return current_name, "CID분해 권리행명 유지"
    if status == "담당제외":
        return current_name, "담당제외 원권리명 보존"
    if status in {"IPS_CID생성보류", "작가확정보류"}:
        return current_name, "보류 원권리명 보존"
    if current_name:
        return current_name, "현재_저작권명"
    return suggested_name, "canonical_저작권명_제안"


def build_seed_summary(seed_df: pd.DataFrame) -> pd.DataFrame:
    if seed_df.empty:
        return pd.DataFrame(columns=[
            "account_저작권코드",
            "CID_seed_행수",
            "CID_작품목록",
            "CID_대표Y_작품목록",
            "CID_작가목록",
            "CID_정산자목록",
            "CID_seed_id_목록",
            "CID_주의플래그목록",
        ])

    grouped_rows: list[dict[str, Any]] = []
    for rights_code, group in seed_df.groupby("account_저작권코드", sort=False):
        primary_works = group.loc[group["정산대표Y/N"].map(normalize_text).eq("Y"), "작품명"].tolist()
        grouped_rows.append(
            {
                "account_저작권코드": normalize_text(rights_code),
                "CID_seed_행수": str(len(group)),
                "CID_작품목록": pipe_join(group["작품명"].tolist()),
                "CID_대표Y_작품목록": pipe_join(primary_works),
                "CID_작가목록": pipe_join(group["대표작가명"].tolist()),
                "CID_정산자목록": pipe_join(group["정산자"].tolist()),
                "CID_seed_id_목록": pipe_join(group["CID_seed_id"].tolist()),
                "CID_주의플래그목록": pipe_join(sum([split_pipe_values(value) for value in group["주의플래그"].tolist()], [])),
            }
        )
    return pd.DataFrame(grouped_rows)


def build_canonical_df() -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Path]]:
    decision_df = load_frame(DECISION_PATH)
    seed_df = load_frame(CID_SEED_PATH)
    seed_summary_df = build_seed_summary(seed_df)

    merged = decision_df.merge(seed_summary_df, on="account_저작권코드", how="left").fillna("")

    canonical_rows: list[dict[str, Any]] = []
    for row in merged.to_dict("records"):
        status = resolve_canonical_status(row)
        final_name, final_name_basis = resolve_final_rights_name(row, status)
        seed_count = parse_int(row.get("CID_seed_행수"))
        action = normalize_text(row.get("action_제안"))
        processed = normalize_text(row.get("처리완료(Y/N)"))
        manual_action = normalize_text(row.get("수동_action"))

        canonical_rows.append(
            {
                "canonical_status": status,
                "canonical_사용(Y/N)": "N" if status == "담당제외" else "Y",
                "account_저작권코드": normalize_text(row.get("account_저작권코드")),
                "account_최종저작권명": final_name,
                "최종저작권명_근거": final_name_basis,
                "현재_저작권명": normalize_text(row.get("현재_저작권명")),
                "canonical_저작권명_제안": normalize_text(row.get("canonical_저작권명_제안")),
                "수동_최종저작권명": normalize_text(row.get("수동_최종저작권명")),
                "action_제안": action,
                "수동_action": manual_action,
                "처리완료(Y/N)": processed,
                "대표작가명": normalize_text(row.get("대표작가명")),
                "관련작가명": normalize_text(row.get("관련작가명")),
                "account_작가코드": normalize_text(row.get("account_작가코드")),
                "특수": normalize_text(row.get("특수")) or "일반",
                "특수_판정근거": normalize_text(row.get("특수_판정근거")),
                "관측_작품수": normalize_text(row.get("관측_작품수")),
                "관측_작품목록": normalize_text(row.get("관측_작품목록")),
                "대표작품": normalize_text(row.get("대표작품")),
                "CID_seed_필요(Y/N)": "Y" if action == "CID분해필요" and status != "담당제외" else "N",
                "CID_seed_상태": (
                    "담당제외"
                    if status == "담당제외"
                    else ("seed생성" if seed_count > 0 else ("대상아님" if action != "CID분해필요" else "seed없음"))
                ),
                "CID_seed_행수": str(seed_count),
                "CID_작품목록": normalize_text(row.get("CID_작품목록")),
                "CID_대표Y_작품목록": normalize_text(row.get("CID_대표Y_작품목록")),
                "CID_작가목록": normalize_text(row.get("CID_작가목록")),
                "CID_정산자목록": normalize_text(row.get("CID_정산자목록")),
                "CID_seed_id_목록": normalize_text(row.get("CID_seed_id_목록")),
                "연결_선인세코드": normalize_text(row.get("연결_선인세코드")),
                "연결_선인세명": normalize_text(row.get("연결_선인세명")),
                "rights_선인세명": normalize_text(row.get("rights_선인세명")),
                "B2C_정산율(%)": normalize_text(row.get("B2C_정산율(%)")),
                "B2BC_정산율(%)": normalize_text(row.get("B2BC_정산율(%)")),
                "B2B_정산율(%)": normalize_text(row.get("B2B_정산율(%)")),
                "account_정산기준": normalize_text(row.get("account_정산기준")),
                "account_계약담당자": normalize_text(row.get("account_계약담당자")),
                "대표_상품번호": normalize_text(row.get("대표_상품번호")),
                "대표_제목": normalize_text(row.get("대표_제목")),
                "fetched_상품유형": normalize_text(row.get("fetched_상품유형")),
                "admin_이용등급": normalize_text(row.get("admin_이용등급")),
                "작가특수RS건수": normalize_text(row.get("작가특수RS건수")),
                "작가특수RS요약": normalize_text(row.get("작가특수RS요약")),
                "scope_assignment_basis": normalize_text(row.get("scope_assignment_basis")),
                "scope_assignment_note": normalize_text(row.get("scope_assignment_note")),
                "관측_match_status": normalize_text(row.get("관측_match_status")),
                "관측_match_reason": normalize_text(row.get("관측_match_reason")),
                "주의플래그": normalize_text(row.get("주의플래그")),
                "수동_메모": normalize_text(row.get("수동_메모")),
                "decision_signature": normalize_text(row.get("decision_signature")),
            }
        )

    canonical_df = pd.DataFrame(canonical_rows)
    canonical_df["정렬_status"] = canonical_df["canonical_status"].map(lambda value: STATUS_ORDER.get(normalize_text(value), 99))
    canonical_df["정렬_저작권코드"] = pd.to_numeric(canonical_df["account_저작권코드"], errors="coerce").fillna(10**18)
    canonical_df = (
        canonical_df.sort_values(
            by=["정렬_status", "대표작가명", "account_최종저작권명", "정렬_저작권코드"],
            kind="stable",
        )
        .drop(columns=["정렬_status", "정렬_저작권코드"])
        .reset_index(drop=True)
    )

    source_paths = {
        "account_decision_queue": DECISION_PATH,
        "account_ips_cid_seed": CID_SEED_PATH,
    }
    return canonical_df, seed_summary_df, source_paths


def write_summary_sheet(
    wb: Workbook,
    *,
    canonical_df: pd.DataFrame,
    seed_summary_df: pd.DataFrame,
    source_paths: dict[str, Path],
) -> None:
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{MANAGER} account rights canonical"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "지표"
    ws["B3"] = "값"
    apply_header_style(ws["A3"])
    apply_header_style(ws["B3"])

    metrics = [
        ("canonical 행 수", len(canonical_df)),
        ("canonical 사용 행 수", int(canonical_df["canonical_사용(Y/N)"].eq("Y").sum())),
        ("담당제외 행 수", int(canonical_df["canonical_status"].eq("담당제외").sum())),
        ("CID분해 행 수", int(canonical_df["canonical_status"].eq("CID분해").sum())),
        ("이름확정 행 수", int(canonical_df["canonical_status"].eq("이름확정").sum())),
        ("유지 행 수", int(canonical_df["canonical_status"].eq("유지").sum())),
        ("미처리 행 수", int(canonical_df["처리완료(Y/N)"].map(normalize_text).ne("Y").sum())),
        ("CID seed 연결 권리코드 수", len(seed_summary_df)),
        ("CID seed 총 행 수", int(canonical_df["CID_seed_행수"].map(parse_int).sum())),
    ]
    for row_idx, (label, value) in enumerate(metrics, start=4):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2))

    start_row = 16
    ws.cell(row=start_row, column=1, value="canonical_status")
    ws.cell(row=start_row, column=2, value="행 수")
    apply_header_style(ws.cell(row=start_row, column=1))
    apply_header_style(ws.cell(row=start_row, column=2))
    status_counter = Counter(canonical_df["canonical_status"].tolist())
    for idx, (status, count) in enumerate(status_counter.items(), start=start_row + 1):
        ws.cell(row=idx, column=1, value=status)
        ws.cell(row=idx, column=2, value=count)
        apply_body_style(ws.cell(row=idx, column=1))
        apply_body_style(ws.cell(row=idx, column=2))

    source_row = start_row + len(status_counter) + 3
    ws.cell(row=source_row, column=1, value="source")
    ws.cell(row=source_row, column=2, value="path")
    apply_header_style(ws.cell(row=source_row, column=1))
    apply_header_style(ws.cell(row=source_row, column=2))
    for offset, (label, path) in enumerate(source_paths.items(), start=1):
        ws.cell(row=source_row + offset, column=1, value=label)
        ws.cell(row=source_row + offset, column=2, value=str(path))
        apply_body_style(ws.cell(row=source_row + offset, column=1))
        apply_body_style(ws.cell(row=source_row + offset, column=2), wrap=True)

    set_column_widths(ws)


def write_canonical_sheet(wb: Workbook, canonical_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("account_rights_canonical")
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    rows = [list(canonical_df.columns)] + canonical_df.astype(object).where(pd.notnull(canonical_df), "").values.tolist()

    status_idx = canonical_df.columns.get_loc("canonical_status") + 1
    seed_status_idx = canonical_df.columns.get_loc("CID_seed_상태") + 1
    for row_idx, row_values in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                apply_header_style(cell)
                continue

            status = normalize_text(ws.cell(row=row_idx, column=status_idx).value)
            seed_status = normalize_text(ws.cell(row=row_idx, column=seed_status_idx).value)
            fill = None
            if status in {"담당제외", "검토필요"} or seed_status == "seed없음":
                fill = _WARN_FILL
            elif status in {"유지", "이름확정"}:
                fill = _GOOD_FILL
            apply_body_style(cell, fill=fill, wrap=True)

    set_column_widths(ws)


def write_outputs(
    *,
    canonical_df: pd.DataFrame,
    seed_summary_df: pd.DataFrame,
    source_paths: dict[str, Path],
) -> dict[str, Path]:
    TARGET_ROOT.mkdir(parents=True, exist_ok=True)
    canonical_df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    wb = Workbook()
    write_summary_sheet(wb, canonical_df=canonical_df, seed_summary_df=seed_summary_df, source_paths=source_paths)
    write_canonical_sheet(wb, canonical_df)
    write_df_sheet(wb, "cid_seed_summary", seed_summary_df)
    try:
        wb.save(OUTPUT_XLSX)
    finally:
        wb.close()

    OUTPUT_JSON.write_text(
        json.dumps(
            {
                "manager": MANAGER,
                "summary": {
                    "canonical_rows": int(len(canonical_df)),
                    "usable_rows": int(canonical_df["canonical_사용(Y/N)"].eq("Y").sum()),
                    "excluded_rows": int(canonical_df["canonical_status"].eq("담당제외").sum()),
                    "cid_split_rows": int(canonical_df["canonical_status"].eq("CID분해").sum()),
                    "pending_rows": int(canonical_df["처리완료(Y/N)"].map(normalize_text).ne("Y").sum()),
                    "cid_seed_rows": int(canonical_df["CID_seed_행수"].map(parse_int).sum()),
                },
                "sources": {label: str(path) for label, path in source_paths.items()},
                "rows": canonical_df.to_dict("records"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    return {
        "xlsx": OUTPUT_XLSX,
        "csv": OUTPUT_CSV,
        "json": OUTPUT_JSON,
    }


def main() -> None:
    canonical_df, seed_summary_df, source_paths = build_canonical_df()
    outputs = write_outputs(
        canonical_df=canonical_df,
        seed_summary_df=seed_summary_df,
        source_paths=source_paths,
    )
    print("=== account rights canonical built ===")
    print(
        json.dumps(
            {
                "manager": MANAGER,
                "canonical_rows": int(len(canonical_df)),
                "usable_rows": int(canonical_df["canonical_사용(Y/N)"].eq("Y").sum()),
                "excluded_rows": int(canonical_df["canonical_status"].eq("담당제외").sum()),
                "cid_split_rows": int(canonical_df["canonical_status"].eq("CID분해").sum()),
                "pending_rows": int(canonical_df["처리완료(Y/N)"].map(normalize_text).ne("Y").sum()),
                "workbook": str(outputs["xlsx"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
