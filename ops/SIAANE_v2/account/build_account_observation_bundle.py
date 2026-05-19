from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = ROOT / "SIAAN Project"
TARGET_ROOT = ROOT / "SIAANE_v2" / "account" / "exports"
NOTES_ROOT = ROOT / "SIAANE_v2" / "account" / "notes"
MANAGER = "조원재"
BAROBOOK_DIR = PROJECT_ROOT / "data" / "exports" / "barobook"
SSOT_PATH = PROJECT_ROOT / "data" / "exports" / "ssot" / f"latest__siaan_ssot_{MANAGER}.csv"
MANAGER_SCOPE_PATH = ROOT / "SIAANE_v2" / "담당작가_ssot" / "manager_author_ssot.csv"
MANUAL_SPECIAL_OVERRIDES_PATH = NOTES_ROOT / "manual_special_overrides.json"
MANUAL_WORK_SCOPE_OVERRIDES_PATH = NOTES_ROOT / "manual_work_scope_overrides.json"

_NON_WORD_RE = re.compile(r"[^0-9A-Za-z가-힣]+")
_PAREN_RE = re.compile(r"^(.*?)\((.*?)\)\s*$")
_SERIES_PREFIX_RE = re.compile(r"^[(\[]\s*연재\s*[)\]]\s*")
_SERIES_TRAIL_NUM_UNIT_RE = re.compile(r"\s+\d+(?:-\d+)?\s*(?:화|권)(?:\s+완결)?\s*$")
_SERIES_TRAIL_EPISODE_RE = re.compile(r"\s+\d+\s*회(?:\s+.*)?\s*$")
_SERIES_TRAIL_COMPLETION_RE = re.compile(r"\s+완결\s*$")
_MATCH_LEADING_TAG_RE = re.compile(r"^[\[(]([^\])]+)[\])]\s*")
_MATCH_TRAILING_VOLUME_RE = re.compile(r"[\s_]+\d+(?:-\d+)?\s*(?:화|권|회)(?:\s*[(（]?(?:완|완결|사용x|사용금지)[)）]?)?\s*$")
_MATCH_TRAILING_COMPLETION_RE = re.compile(r"\s*[(（]?(?:완|완결)[)）]?\s*$")
_MATCH_TRAILING_SET_RE = re.compile(r"\s*(?:합본|세트)\s*$")
_MATCH_TRAILING_SET_PAREN_RE = re.compile(r"\s*[(（]세트[)）]\s*$")
_MATCH_TRAILING_FULLSET_RE = re.compile(r"\s*[(（]전\d+권\s*완결[)）]\s*$")
_MATCH_TRAILING_TAG_RE = re.compile(
    r"\s*[(（\[]\s*(?:연재|세트|카카오\s*제외|카카오\s*외\s*연재|타플)\s*[)）\]]\s*$"
)
_MATCH_TRAILING_ENGLISH_PAREN_RE = re.compile(r"\s*[(（][A-Za-z0-9 .,'-]+[)）]\s*$")
_IGNORED_WORK_TITLE_RE = re.compile(r"사용금지|사용\s*[xX]")
_THIN = Side(style="thin", color="D9DEE6")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SUBHEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
_GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")

TARGET_SPECIALS = [
    "일반",
    "카카오MG",
    "네이버MG",
    "원작",
]

MATCHABLE_TAG_KEYWORDS = (
    "카카오",
    "네이버",
    "윌라",
    "봄툰",
    "문피아",
    "리디",
    "세트",
    "연재",
    "단행본",
    "단행",
    "오디오북",
    "창작지원금",
    "광고수익",
    "타플",
)


def normalize_text(value: Any) -> str:
    return str(value or "").replace("\u00a0", " ").strip()


def compact_key(value: Any) -> str:
    return _NON_WORD_RE.sub("", normalize_text(value)).lower()


def normalize_raw_series_name(value: Any) -> str:
    current = normalize_text(value)
    for _ in range(3):
        before = current
        current = _SERIES_PREFIX_RE.sub("", current)
        current = _SERIES_TRAIL_NUM_UNIT_RE.sub("", current)
        current = _SERIES_TRAIL_EPISODE_RE.sub("", current)
        current = _SERIES_TRAIL_COMPLETION_RE.sub("", current)
        current = normalize_text(current)
        if current == before:
            break
    return current


def normalize_matchable_work_title(value: Any) -> str:
    current = normalize_text(value).replace("_", " ")
    for _ in range(5):
        before = current

        tag_match = _MATCH_LEADING_TAG_RE.match(current)
        while tag_match:
            tag = normalize_text(tag_match.group(1))
            if not any(keyword in tag for keyword in MATCHABLE_TAG_KEYWORDS):
                break
            current = normalize_text(current[tag_match.end():])
            tag_match = _MATCH_LEADING_TAG_RE.match(current)

        for prefix in (
            "카카오창작지원금 ",
            "네이버MG ",
            "네이버광고수익 ",
            "광고수익 ",
            "카카오 ",
            "카카오_",
            "단행본 ",
            "단행본_",
            "단행 ",
            "단행_",
        ):
            if current.startswith(prefix):
                current = normalize_text(current[len(prefix):])

        current = _MATCH_TRAILING_FULLSET_RE.sub("", current)
        current = _MATCH_TRAILING_SET_PAREN_RE.sub("", current)
        current = _MATCH_TRAILING_SET_RE.sub("", current)
        current = _MATCH_TRAILING_TAG_RE.sub("", current)
        current = _MATCH_TRAILING_ENGLISH_PAREN_RE.sub("", current)
        current = _MATCH_TRAILING_VOLUME_RE.sub("", current)
        current = _MATCH_TRAILING_COMPLETION_RE.sub("", current)
        current = normalize_text(current).strip(" -_?")

        if current == before:
            break

    return normalize_text(current)


def should_ignore_work_title(value: Any) -> bool:
    raw = normalize_text(value)
    if not raw:
        return True
    normalized = normalize_matchable_work_title(raw)
    return bool(
        _IGNORED_WORK_TITLE_RE.search(raw)
        or (normalized and _IGNORED_WORK_TITLE_RE.search(normalized))
    )


def unique(values: list[Any]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        cleaned = normalize_text(value)
        if not cleaned or cleaned in seen:
            continue
        ordered.append(cleaned)
        seen.add(cleaned)
    return ordered


def pipe_join(values: list[Any]) -> str:
    return " | ".join(unique(values))


def split_pipe_values(raw: Any) -> list[str]:
    return unique(str(raw or "").split("|"))


def first_pipe_value(raw: Any) -> str:
    values = split_pipe_values(raw)
    return values[0] if values else ""


def contains_pipe_token(raw: Any, token: str) -> bool:
    token_clean = normalize_text(token)
    if not token_clean:
        return False
    return token_clean in split_pipe_values(raw)


def latest_path(pattern: str) -> Path:
    hits = sorted(BAROBOOK_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not hits:
        raise FileNotFoundError(pattern)
    return hits[0]


def load_frame(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, dtype=str).fillna("")
    return pd.read_excel(path, dtype=str).fillna("")


def load_manual_special_overrides(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError(f"manual special overrides must be a list: {path}")
    overrides: list[dict[str, str]] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        normalized = {str(key): normalize_text(value) for key, value in item.items()}
        if normalized.get("특수") not in TARGET_SPECIALS:
            continue
        overrides.append(normalized)
    return overrides


def load_manual_work_scope_overrides(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError(f"manual work scope overrides must be a list: {path}")

    overrides: list[dict[str, str]] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        normalized = {str(key): normalize_text(value) for key, value in item.items()}
        if not normalized.get("작품명") or not normalized.get("대표작가명"):
            continue
        overrides.append(normalized)
    return overrides


def match_manual_special_override(row: dict[str, Any], overrides: list[dict[str, str]]) -> tuple[str, str] | None:
    author_code = normalize_text(row.get("account_작가코드"))
    rights_code = normalize_text(row.get("account_저작권코드"))

    matched: tuple[str, str] | None = None
    for override in overrides:
        override_author_code = normalize_text(override.get("account_작가코드"))
        override_rights_code = normalize_text(override.get("account_저작권코드"))

        if override_author_code and override_author_code != author_code:
            continue
        if override_rights_code and override_rights_code != rights_code:
            continue

        reason = normalize_text(override.get("특수_판정근거")) or "수동override"
        matched = (normalize_text(override.get("특수")), reason)

    return matched


def resolve_manual_work_scope_override(
    *,
    author_code: str,
    rights_code: str = "",
    observed_titles: list[str],
    overrides: list[dict[str, str]],
) -> tuple[list[str], str]:
    normalized_titles = {
        normalize_matchable_work_title(title) or normalize_text(title)
        for title in observed_titles
        if normalize_text(title)
    }
    matched: list[tuple[bool, str, str]] = []
    rights_code = normalize_text(rights_code)

    for override in overrides:
        override_author_code = normalize_text(override.get("account_작가코드"))
        if override_author_code and override_author_code != author_code:
            continue
        override_rights_code = normalize_text(override.get("account_저작권코드"))
        if override_rights_code and override_rights_code != rights_code:
            continue

        target_title = normalize_matchable_work_title(override.get("작품명")) or normalize_text(override.get("작품명"))
        if not target_title or target_title not in normalized_titles:
            continue

        author_name = normalize_text(override.get("대표작가명"))
        reason = normalize_text(override.get("판정근거")) or f"{target_title}->{normalize_text(override.get('대표작가명'))}"
        matched.append((bool(override_rights_code), author_name, reason))

    if any(is_rights_specific for is_rights_specific, _, _ in matched):
        matched = [item for item in matched if item[0]]

    matched_authors = [author for _, author, _ in matched]
    matched_notes = [reason for _, _, reason in matched]
    return unique(matched_authors), " | ".join(unique(matched_notes))


def explode_name_tokens(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []

    pieces: list[str] = [raw]
    for pipe_piece in split_pipe_values(raw):
        pieces.append(pipe_piece)

    match = _PAREN_RE.match(raw)
    if match:
        pieces.append(match.group(1))
        for inner_piece in re.split(r"[|/,]", match.group(2)):
            pieces.append(inner_piece)

    return unique(pieces)


def build_scope_alias_indexes(scope_df: pd.DataFrame) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    exact: dict[str, set[str]] = defaultdict(set)
    compact: dict[str, set[str]] = defaultdict(set)
    alias_columns = [
        "대표작가명",
        "index_f_작가표기",
        "NAS_작가표기",
        "account_pen_name",
        "account_저작권자명_필명",
        "관련이름묶음",
        "수동동치묶음",
    ]

    for row in scope_df.to_dict("records"):
        scope_author = normalize_text(row.get("대표작가명"))
        if not scope_author:
            continue

        aliases: list[str] = [scope_author]
        for column in alias_columns:
            aliases.extend(explode_name_tokens(row.get(column)))

        for alias in unique(aliases):
            exact[alias].add(scope_author)
            compact[compact_key(alias)].add(scope_author)

    return (
        {key: sorted(value) for key, value in exact.items()},
        {key: sorted(value) for key, value in compact.items()},
    )


def build_scope_author_code_owners(scope_df: pd.DataFrame) -> dict[str, list[str]]:
    owners: dict[str, set[str]] = defaultdict(set)
    for row in scope_df.to_dict("records"):
        scope_author = normalize_text(row.get("대표작가명"))
        if not scope_author:
            continue
        for author_code in split_pipe_values(row.get("account_작가코드")):
            owners[author_code].add(scope_author)
    return {code: sorted(names) for code, names in owners.items()}


def resolve_scope_candidates(
    values: list[Any],
    *,
    alias_exact: dict[str, list[str]],
    alias_compact: dict[str, list[str]],
) -> list[str]:
    candidates: list[str] = []
    for value in values:
        for token in explode_name_tokens(value):
            candidates.extend(alias_exact.get(token, []))
            candidates.extend(alias_compact.get(compact_key(token), []))
    return unique(candidates)


def load_scope_df() -> pd.DataFrame:
    df = load_frame(MANAGER_SCOPE_PATH)
    return df[df["판정"].astype(str).str.contains("확정", regex=False)].copy()


def classify_candidate_specials(text: Any, *, source_hint: str = "") -> list[str]:
    raw = normalize_text(text)
    lowered = raw.lower()
    source_hint = normalize_text(source_hint)
    candidates: list[str] = []

    if source_hint == "observed_special":
        mapping = {
            "일반": "일반",
            "네이버 광고수익": "광고수익",
        }
        mapped = mapping.get(raw)
        return [mapped] if mapped else []

    if not raw:
        return []

    if "광고수익" in raw:
        candidates.append("광고수익")
    if any(keyword in raw for keyword in ["기본정산율", " 일반", "일반 ", "_일반", "일반도서"]):
        candidates.append("일반")
    if raw == "일반":
        candidates.append("일반")

    return unique(candidates)


def sort_key_frame(df: pd.DataFrame, column: str) -> pd.DataFrame:
    sortable = df.copy()
    sortable[f"{column}__sort_num"] = pd.to_numeric(sortable[column], errors="coerce")
    sortable[f"{column}__sort_num"] = sortable[f"{column}__sort_num"].fillna(10**18)
    return sortable


def build_raw_series_index(series_map_df: pd.DataFrame) -> dict[tuple[str, str], pd.DataFrame]:
    if series_map_df.empty:
        return {}

    required_columns = [
        "작가코드",
        "저작권코드",
        "시리즈명",
        "대표_상품번호",
        "대표_제목",
    ]
    sortable = series_map_df.copy()
    for column in required_columns:
        if column not in sortable.columns:
            sortable[column] = ""
        sortable[column] = sortable[column].map(normalize_text)

    sortable["정규화_시리즈명"] = sortable.apply(
        lambda row: normalize_raw_series_name(row.get("시리즈명"))
        or normalize_raw_series_name(row.get("대표_제목")),
        axis=1,
    )
    sortable = sortable[sortable["정규화_시리즈명"].ne("")].copy()
    sortable["대표_상품번호__sort_num"] = pd.to_numeric(sortable["대표_상품번호"], errors="coerce")
    sortable["대표_상품번호__sort_num"] = sortable["대표_상품번호__sort_num"].fillna(10**18)
    sortable = sortable.sort_values(
        by=["작가코드", "저작권코드", "정규화_시리즈명", "대표_상품번호__sort_num", "시리즈명", "대표_제목"],
        kind="stable",
    )

    collapsed_rows: list[dict[str, Any]] = []
    for _, group in sortable.groupby(["작가코드", "저작권코드", "정규화_시리즈명"], dropna=False, sort=False):
        first = group.iloc[0].to_dict()
        first["시리즈명"] = normalize_text(first.get("정규화_시리즈명"))
        collapsed_rows.append(first)

    collapsed_df = pd.DataFrame(collapsed_rows)
    grouped_rows: dict[tuple[str, str], pd.DataFrame] = {}
    for (author_code, copyright_code), group in collapsed_df.groupby(["작가코드", "저작권코드"], dropna=False, sort=False):
        author_code = normalize_text(author_code)
        copyright_code = normalize_text(copyright_code)
        if not author_code or not copyright_code:
            continue
        grouped_rows[(author_code, copyright_code)] = group.drop(
            columns=["대표_상품번호__sort_num", "정규화_시리즈명"]
        ).reset_index(drop=True)

    return grouped_rows


def derive_observed_work_values(
    *,
    raw_series_rows: pd.DataFrame,
    mapping_rows: pd.DataFrame,
) -> tuple[str, str, str, str, str, str]:
    if not raw_series_rows.empty:
        observed_titles = unique(
            [
                normalize_text(row.get("시리즈명")) or normalize_text(row.get("대표_제목"))
                for row in raw_series_rows.to_dict("records")
                if not should_ignore_work_title(
                    normalize_text(row.get("시리즈명")) or normalize_text(row.get("대표_제목"))
                )
            ]
        )
    else:
        observed_titles = unique(
            [
                title
                for title in (mapping_rows["title_final"].tolist() if not mapping_rows.empty else [])
                if not should_ignore_work_title(title)
            ]
        )

    folder_by_title: dict[str, str] = {}
    folder_by_compact_title: dict[str, str] = {}
    folder_by_match_title: dict[str, str] = {}
    folder_by_match_compact_title: dict[str, str] = {}
    if not mapping_rows.empty:
        for item in mapping_rows.to_dict("records"):
            title = normalize_text(item.get("title_final"))
            folder = normalize_text(item.get("folder_path"))
            if not title or not folder:
                continue
            folder_by_title.setdefault(title, folder)
            folder_by_compact_title.setdefault(compact_key(title), folder)
            match_title = normalize_matchable_work_title(title)
            if match_title:
                folder_by_match_title.setdefault(match_title, folder)
                folder_by_match_compact_title.setdefault(compact_key(match_title), folder)

    observed_folders: list[str] = []
    seen_folders: set[str] = set()
    missing_titles: list[str] = []
    for title in observed_titles:
        match_title = normalize_matchable_work_title(title)
        folder = (
            folder_by_title.get(title)
            or folder_by_compact_title.get(compact_key(title), "")
            or folder_by_match_title.get(match_title, "")
            or folder_by_match_compact_title.get(compact_key(match_title), "")
        )
        if not folder:
            missing_titles.append(title)
            continue
        if folder in seen_folders:
            continue
        observed_folders.append(folder)
        seen_folders.add(folder)

    return (
        str(len(observed_titles)),
        pipe_join(observed_titles),
        pipe_join(observed_folders),
        str(len(observed_folders)),
        str(len(missing_titles)),
        pipe_join(missing_titles),
    )


def derive_special_from_advances(row: dict[str, Any]) -> tuple[str, str]:
    advance_names = split_pipe_values(row.get("연결_선인세명"))
    joined = pipe_join(advance_names)

    if joined:
        if "웹툰" in joined:
            return "원작", f"연결선인세_웹툰:{joined}"
        if "네이버" in joined:
            return "네이버MG", f"연결선인세:{joined}"
        if any(keyword in joined for keyword in ["카카오", "선투자", "콘텐츠 특별 공급"]):
            return "카카오MG", f"연결선인세:{joined}"
        return "일반", f"연결선인세_비MG:{joined}"

    return "일반", "연결선인세없음"


def parse_int(value: Any) -> int:
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return 0
    return int(parsed)


def apply_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def apply_body_style(cell, fill: PatternFill | None = None, wrap: bool = False) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    if fill is not None:
        cell.fill = fill


def set_column_widths(ws) -> None:
    max_widths: dict[int, int] = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            value = normalize_text(cell.value)
            if not value:
                continue
            width = min(max(len(value), 8), 60)
            max_widths[cell.column] = max(max_widths[cell.column], width)
    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def write_df_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, *, freeze_cell: str = "A2") -> None:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = freeze_cell
    ws.sheet_view.showGridLines = False
    rows = [list(df.columns)] + df.astype(object).where(pd.notnull(df), "").values.tolist()
    for row_idx, row_values in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                apply_header_style(cell)
            else:
                fill = None
                if sheet_name == "account_actual_inventory":
                    if normalize_text(ws.cell(row=row_idx, column=1).value):
                        fill = _GOOD_FILL
                elif sheet_name == "special_evidence_inventory":
                    if normalize_text(ws.cell(row=row_idx, column=6).value) in {"카카오MG", "네이버MG", "원작"}:
                        fill = _WARN_FILL
                apply_body_style(cell, fill=fill, wrap=True)
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24
    set_column_widths(ws)


def write_summary_sheet(
    wb: Workbook,
    *,
    manager: str,
    actual_df: pd.DataFrame,
    evidence_df: pd.DataFrame,
    source_paths: dict[str, Path],
) -> None:
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{manager} account observation"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "지표"
    ws["B3"] = "값"
    apply_header_style(ws["A3"])
    apply_header_style(ws["B3"])

    metrics = [
        ("권리 실황 행 수", len(actual_df)),
        ("고유 작가코드 수", int(actual_df["account_작가코드"].nunique())),
        ("고유 저작권코드 수", int(actual_df["account_저작권코드"].nunique())),
        ("선인세 연결 행 수", int(actual_df["연결_선인세건수"].astype(int).gt(0).sum())),
        ("관측된 작품 연결 행 수", int(actual_df["관측_작품수"].astype(int).gt(0).sum())),
        ("미매핑 작품 보유 행 수", int(actual_df["관측_미매핑_작품수"].astype(int).gt(0).sum())),
        ("미매핑 작품 수 합계", int(pd.to_numeric(actual_df["관측_미매핑_작품수"], errors="coerce").fillna(0).sum())),
        ("특수 근거 행 수", len(evidence_df)),
        ("특수 후보 종류 수", int(evidence_df["특수"].nunique()) if not evidence_df.empty else 0),
    ]

    for row_idx, (label, value) in enumerate(metrics, start=4):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2))

    start_row = 13
    ws.cell(row=start_row, column=1, value="후보특수")
    ws.cell(row=start_row, column=2, value="근거행수")
    apply_header_style(ws.cell(row=start_row, column=1))
    apply_header_style(ws.cell(row=start_row, column=2))

    counter = Counter(evidence_df["특수"].tolist()) if not evidence_df.empty else Counter()
    for idx, special in enumerate(TARGET_SPECIALS, start=start_row + 1):
        ws.cell(row=idx, column=1, value=special)
        ws.cell(row=idx, column=2, value=counter.get(special, 0))
        apply_body_style(ws.cell(row=idx, column=1))
        apply_body_style(ws.cell(row=idx, column=2))

    chart = BarChart()
    chart.title = "특수 후보 관측 건수"
    chart.y_axis.title = "근거행 수"
    chart.x_axis.title = "후보특수"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(TARGET_SPECIALS))
    categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(TARGET_SPECIALS))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "D4")

    source_row = start_row + len(TARGET_SPECIALS) + 3
    ws.cell(row=source_row, column=1, value="입력 소스")
    ws.cell(row=source_row, column=2, value="경로")
    apply_header_style(ws.cell(row=source_row, column=1))
    apply_header_style(ws.cell(row=source_row, column=2))
    for offset, (label, path) in enumerate(source_paths.items(), start=1):
        ws.cell(row=source_row + offset, column=1, value=label)
        ws.cell(row=source_row + offset, column=2, value=str(path))
        apply_body_style(ws.cell(row=source_row + offset, column=1))
        apply_body_style(ws.cell(row=source_row + offset, column=2), wrap=True)

    set_column_widths(ws)


def build_observation_data(manager: str) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Path]]:
    scope_df = load_scope_df()
    alias_exact, alias_compact = build_scope_alias_indexes(scope_df)
    author_code_owners = build_scope_author_code_owners(scope_df)

    source_paths = {
        "manager_scope": MANAGER_SCOPE_PATH,
        "ssot": SSOT_PATH,
        "author_details": latest_path(f"*__NAS_IPS_ACCOUNT_예금주_{manager}_author_details.csv"),
        "mapping_summary": latest_path(f"*__NAS_IPS_ACCOUNT_예금주_{manager}.csv"),
        "raw_series_map": latest_path(f"*__저작권_시리즈매핑_{manager}*.xlsx"),
        "copyright_codes": latest_path(f"*__저작권코드_{manager}.xlsx"),
        "advance_balances": latest_path(f"*__선인세잔액_{manager}.xlsx"),
        "admin_meta": latest_path(f"*__어드민_상품메타_{manager}.xlsx"),
        "manual_special_overrides": MANUAL_SPECIAL_OVERRIDES_PATH,
        "manual_work_scope_overrides": MANUAL_WORK_SCOPE_OVERRIDES_PATH,
    }

    ssot_df = load_frame(source_paths["ssot"])
    author_details_df = load_frame(source_paths["author_details"])
    mapping_summary_df = load_frame(source_paths["mapping_summary"])
    raw_series_map_df = load_frame(source_paths["raw_series_map"])
    copyright_df = load_frame(source_paths["copyright_codes"])
    advance_df = load_frame(source_paths["advance_balances"])
    admin_df = load_frame(source_paths["admin_meta"])
    manual_special_overrides = load_manual_special_overrides(source_paths["manual_special_overrides"])
    manual_work_scope_overrides = load_manual_work_scope_overrides(source_paths["manual_work_scope_overrides"])
    raw_series_by_rights = build_raw_series_index(raw_series_map_df)

    # Canonical author map for ssot rows.
    ssot_df["대표작가명"] = ssot_df["작가"].map(
        lambda value: pipe_join(
            resolve_scope_candidates(
                [value],
                alias_exact=alias_exact,
                alias_compact=alias_compact,
            )
        )
    )
    ssot_scope_df = ssot_df[ssot_df["대표작가명"].astype(str).str.strip().ne("")].copy()

    author_details_by_code = {
        normalize_text(row["작가코드"]): row
        for row in author_details_df.to_dict("records")
        if normalize_text(row.get("작가코드"))
    }
    admin_grade_by_product = {
        normalize_text(row["상품번호"]): normalize_text(row.get("이용등급"))
        for row in admin_df.to_dict("records")
        if normalize_text(row.get("상품번호"))
    }

    actual_rows: list[dict[str, Any]] = []
    for row in copyright_df.to_dict("records"):
        author_code = normalize_text(row.get("작가코드"))
        copyright_code = normalize_text(row.get("저작권코드"))
        if not author_code or not copyright_code:
            continue

        author_detail = author_details_by_code.get(author_code, {})
        scope_candidates = resolve_scope_candidates(
            [
                row.get("저작권자명_필명"),
                author_detail.get("pen_name"),
                author_detail.get("real_name"),
                author_detail.get("예금주"),
            ],
            alias_exact=alias_exact,
            alias_compact=alias_compact,
        )

        mapping_rows = mapping_summary_df[
            mapping_summary_df["작가코드"].map(normalize_text).eq(author_code)
            & mapping_summary_df["matched_copyright_codes"].map(lambda raw: contains_pipe_token(raw, copyright_code))
        ].copy()
        raw_series_rows = raw_series_by_rights.get((author_code, copyright_code), pd.DataFrame())
        observed_titles = unique(
            [
                title
                for title in [
                    normalize_text(row.get("시리즈명")) or normalize_text(row.get("대표_제목"))
                    for row in raw_series_rows.to_dict("records")
                ]
                if not should_ignore_work_title(title)
            ]
        ) if not raw_series_rows.empty else unique(
            [
                title
                for title in (mapping_rows["title_final"].tolist() if not mapping_rows.empty else [])
                if not should_ignore_work_title(title)
            ]
        )
        mapping_scope_authors = unique(
            sum(
                [
                    resolve_scope_candidates(
                        [
                            item.get("author_final"),
                            item.get("matched_s2_author"),
                            item.get("matched_ips_author"),
                        ],
                        alias_exact=alias_exact,
                        alias_compact=alias_compact,
                    )
                    for item in mapping_rows.to_dict("records")
                ],
                [],
            )
        )
        owner_candidates = author_code_owners.get(author_code, [])
        manual_scope_authors, manual_scope_note = resolve_manual_work_scope_override(
            author_code=author_code,
            rights_code=copyright_code,
            observed_titles=observed_titles,
            overrides=manual_work_scope_overrides,
        )

        if not scope_candidates and not mapping_scope_authors and not owner_candidates and not manual_scope_authors:
            continue

        ssot_rows = ssot_scope_df[
            ssot_scope_df["어카운트_저작권코드"].map(normalize_text).eq(copyright_code)
            & ssot_scope_df["어카운트_작가코드"].map(lambda raw: contains_pipe_token(raw, author_code))
        ].copy()

        work_scope_authors = unique(mapping_scope_authors + sum(
            [
                resolve_scope_candidates(
                    [item.get("대표작가명"), item.get("작가")],
                    alias_exact=alias_exact,
                    alias_compact=alias_compact,
                )
                for item in ssot_rows.to_dict("records")
            ],
            [],
        ))
        if len(manual_scope_authors) == 1:
            scope_author_primary = manual_scope_authors[0]
            scope_authors_related = manual_scope_authors
            scope_assignment_basis = "manual_work_scope_override"
        elif len(owner_candidates) == 1:
            scope_author_primary = owner_candidates[0]
            scope_authors_related = owner_candidates
            scope_assignment_basis = "author_code_owner"
        else:
            scope_authors_related = unique(manual_scope_authors + work_scope_authors + scope_candidates)
            scope_author_primary = (
                manual_scope_authors[0]
                if len(manual_scope_authors) == 1
                else (
                    work_scope_authors[0]
                    if len(work_scope_authors) == 1
                    else (scope_candidates[0] if len(scope_candidates) == 1 else "")
                )
            )
            scope_assignment_basis = (
                "observed_work_author"
                if len(work_scope_authors) == 1
                else (
                    "alias_signal"
                    if len(scope_candidates) == 1
                    else ("manual_work_scope_conflict" if manual_scope_authors else "ambiguous_scope")
                )
            )

        rights_advance_name = normalize_text(row.get("선인세명"))
        rights_advance_mapping_raw = normalize_text(row.get("선인세_매핑원문"))
        linked_advance_rows = advance_df[
            advance_df["작가코드"].map(normalize_text).eq(author_code)
            & advance_df.apply(
                lambda record: (
                    bool(rights_advance_name)
                    and normalize_text(record.get("선인세명")) == rights_advance_name
                )
                or (
                    bool(rights_advance_mapping_raw)
                    and normalize_text(record.get("선인세명"))
                    and normalize_text(record.get("선인세명")) in rights_advance_mapping_raw
                ),
                axis=1,
            )
        ].copy()

        representative_product_no = normalize_text(row.get("대표_상품번호"))
        (
            observed_work_count,
            observed_work_list,
            observed_work_folders,
            observed_folder_count,
            unmapped_work_count,
            unmapped_work_list,
        ) = derive_observed_work_values(
            raw_series_rows=raw_series_rows,
            mapping_rows=mapping_rows,
        )
        actual_rows.append(
            {
                "scope_author_primary": scope_author_primary,
                "scope_authors_related": pipe_join(scope_authors_related),
                "scope_assignment_basis": scope_assignment_basis,
                "scope_assignment_note": manual_scope_note,
                "account_작가코드": author_code,
                "account_pen_name": normalize_text(author_detail.get("pen_name")),
                "account_real_name": normalize_text(author_detail.get("real_name")),
                "account_예금주": normalize_text(author_detail.get("예금주")),
                "account_계약담당자": normalize_text(author_detail.get("contract_manager")),
                "account_정산기준": normalize_text(author_detail.get("settlement_basis")),
                "account_status": normalize_text(author_detail.get("status")),
                "account_등록일": normalize_text(author_detail.get("registered_at")),
                "account_추가메모": normalize_text(author_detail.get("additional_note")),
                "account_저작권코드": copyright_code,
                "account_저작권명": normalize_text(row.get("저작권명")),
                "기본정산율여부": normalize_text(row.get("기본정산율여부")),
                "B2C_정산율(%)": normalize_text(row.get("B2C_정산율(%)")),
                "B2BC_정산율(%)": normalize_text(row.get("B2BC_정산율(%)")),
                "B2B_정산율(%)": normalize_text(row.get("B2B_정산율(%)")),
                "rights_선인세명": rights_advance_name,
                "rights_선인세매핑원문": rights_advance_mapping_raw,
                "rights_선인세차감비율(%)": normalize_text(row.get("선인세차감비율(%)")),
                "연결_선인세코드": pipe_join(linked_advance_rows["선인세코드"].tolist() if not linked_advance_rows.empty else []),
                "연결_선인세명": pipe_join(linked_advance_rows["선인세명"].tolist() if not linked_advance_rows.empty else []),
                "연결_선인세잔액": pipe_join(linked_advance_rows["선인세잔액"].tolist() if not linked_advance_rows.empty else []),
                "연결_선인세건수": str(len(linked_advance_rows)),
                "대표_상품번호": representative_product_no,
                "대표_제목": normalize_text(row.get("대표_제목")),
                "fetched_상품유형": normalize_text(row.get("fetched_상품유형")),
                "fetched_시리즈번호": normalize_text(row.get("fetched_시리즈번호")),
                "admin_이용등급": admin_grade_by_product.get(representative_product_no, ""),
                "관측_작품수": observed_work_count,
                "관측_작품목록": observed_work_list,
                "관측_작품_폴더": observed_work_folders,
                "관측_작품_폴더수": observed_folder_count,
                "관측_미매핑_작품수": unmapped_work_count,
                "관측_미매핑_작품목록": unmapped_work_list,
                "관측_작품_scope_authors": pipe_join(work_scope_authors),
                "관측_specials_from_ssot": pipe_join(ssot_rows["특수"].tolist()),
                "관측_채널명": pipe_join(
                    ssot_rows["지급정산_판매채널명"].tolist()
                    + ssot_rows["외부매출_판매채널명"].tolist()
                    + mapping_rows["sales_channel_names"].tolist()
                ),
                "관측_match_status": pipe_join(mapping_rows["match_status"].tolist()),
                "관측_match_reason": pipe_join(mapping_rows["match_reason"].tolist()),
            }
        )

    actual_df = pd.DataFrame(actual_rows).drop_duplicates(
        subset=["account_작가코드", "account_저작권코드"],
        keep="first",
    )
    if not actual_df.empty:
        actual_df["정렬_대표작품"] = actual_df["관측_작품목록"].map(first_pipe_value)
        actual_df["정렬_대표작품"] = actual_df["정렬_대표작품"].mask(
            actual_df["정렬_대표작품"].eq(""),
            actual_df["대표_제목"].map(normalize_text),
        )
        actual_df["정렬_대표작품없음"] = actual_df["정렬_대표작품"].map(normalize_text).eq("")
        actual_df = sort_key_frame(actual_df, "account_저작권코드").sort_values(
            by=["정렬_대표작품없음", "정렬_대표작품", "account_저작권코드__sort_num", "account_저작권코드", "account_작가코드"],
            kind="stable",
        ).drop(columns=["정렬_대표작품없음", "정렬_대표작품", "account_저작권코드__sort_num"]).reset_index(drop=True)

    evidence_rows: list[dict[str, Any]] = []
    for row in actual_df.to_dict("records"):
        special, reason = derive_special_from_advances(row)
        manual_override = match_manual_special_override(row, manual_special_overrides)
        if manual_override is not None:
            special, manual_reason = manual_override
            reason = f"{manual_reason} | 기존={reason}"
        evidence_rows.append(
            {
                "scope_author_primary": row.get("scope_author_primary"),
                "scope_authors_related": row.get("scope_authors_related"),
                "account_작가코드": row.get("account_작가코드"),
                "account_저작권코드": row.get("account_저작권코드"),
                "account_저작권명": row.get("account_저작권명"),
                "특수": special,
                "특수_판정근거": reason,
                "연결_선인세코드": row.get("연결_선인세코드"),
                "연결_선인세명": row.get("연결_선인세명"),
                "대표_상품번호": row.get("대표_상품번호"),
                "관련_작품수": parse_int(row.get("관측_작품수")),
                "관련_작품목록": row.get("관측_작품목록"),
            }
        )

    evidence_df = pd.DataFrame(evidence_rows)
    if not evidence_df.empty:
        evidence_df = evidence_df[evidence_df["특수"].isin(TARGET_SPECIALS)].copy()
        evidence_df["정렬_대표작품"] = evidence_df["관련_작품목록"].map(first_pipe_value)
        evidence_df["정렬_대표작품없음"] = evidence_df["정렬_대표작품"].map(normalize_text).eq("")
        evidence_df = sort_key_frame(evidence_df, "account_저작권코드").sort_values(
            by=["정렬_대표작품없음", "정렬_대표작품", "account_저작권코드__sort_num", "account_저작권코드", "특수"],
            kind="stable",
        ).drop(columns=["정렬_대표작품없음", "정렬_대표작품", "account_저작권코드__sort_num"]).reset_index(drop=True)

    return actual_df, evidence_df, source_paths


def write_outputs(
    *,
    actual_df: pd.DataFrame,
    evidence_df: pd.DataFrame,
    source_paths: dict[str, Path],
) -> dict[str, Path]:
    TARGET_ROOT.mkdir(parents=True, exist_ok=True)
    output_paths = {
        "actual_csv": TARGET_ROOT / f"latest__account_actual_inventory_{MANAGER}.csv",
        "evidence_csv": TARGET_ROOT / f"latest__special_evidence_inventory_{MANAGER}.csv",
        "workbook": TARGET_ROOT / f"latest__account_observation_{MANAGER}.xlsx",
        "summary": TARGET_ROOT / f"latest__account_observation_{MANAGER}.json",
    }
    actual_df.to_csv(output_paths["actual_csv"], index=False, encoding="utf-8-sig")
    evidence_df.to_csv(output_paths["evidence_csv"], index=False, encoding="utf-8-sig")

    wb = Workbook()
    write_summary_sheet(
        wb,
        manager=MANAGER,
        actual_df=actual_df,
        evidence_df=evidence_df,
        source_paths=source_paths,
    )
    write_df_sheet(wb, "account_actual_inventory", actual_df)
    write_df_sheet(wb, "special_evidence_inventory", evidence_df if not evidence_df.empty else pd.DataFrame(columns=["특수"]))

    try:
        wb.save(output_paths["workbook"])
    except PermissionError as exc:
        raise RuntimeError(
            f"최신판 1개 정책을 유지하려면 '{output_paths['workbook'].name}' 파일을 닫은 뒤 다시 실행해야 합니다."
        ) from exc

    summary = {
        "manager": MANAGER,
        "actual_inventory_rows": int(len(actual_df)),
        "special_evidence_rows": int(len(evidence_df)),
        "distinct_author_codes": int(actual_df["account_작가코드"].nunique()),
        "distinct_copyright_codes": int(actual_df["account_저작권코드"].nunique()),
        "workbook": str(output_paths["workbook"]),
    }
    output_paths["summary"].write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_paths


def main() -> None:
    actual_df, evidence_df, source_paths = build_observation_data(MANAGER)
    outputs = write_outputs(actual_df=actual_df, evidence_df=evidence_df, source_paths=source_paths)
    print("=== account observation bundle built ===")
    print(
        json.dumps(
            {
                "manager": MANAGER,
                "actual_inventory_rows": len(actual_df),
                "special_evidence_rows": len(evidence_df),
                "workbook": str(outputs["workbook"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
