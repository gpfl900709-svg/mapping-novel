from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

from build_account_observation_bundle import (
    MANAGER,
    TARGET_ROOT,
    _GOOD_FILL,
    _WARN_FILL,
    apply_body_style,
    apply_header_style,
    first_pipe_value,
    latest_path,
    load_frame,
    normalize_text,
    pipe_join,
    set_column_widths,
    split_pipe_values,
)


ACTUAL_PATH = TARGET_ROOT / f"latest__account_actual_inventory_{MANAGER}.csv"
EVIDENCE_PATH = TARGET_ROOT / f"latest__special_evidence_inventory_{MANAGER}.csv"
OUTPUT_XLSX = TARGET_ROOT / f"latest__account_decision_queue_{MANAGER}.xlsx"
OUTPUT_CSV = TARGET_ROOT / f"latest__account_decision_queue_{MANAGER}.csv"
OUTPUT_JSON = TARGET_ROOT / f"latest__account_decision_queue_{MANAGER}.json"
MANUAL_COLUMNS = [
    "수동_최종저작권명",
    "수동_action",
    "수동_메모",
    "처리완료(Y/N)",
]
DECISION_SIGNATURE_COLUMNS = [
    "대표작가명",
    "관련작가명",
    "account_저작권코드",
    "현재_저작권명",
    "대표작품",
    "관측_작품수",
    "관측_작품목록",
    "특수",
    "canonical_저작권명_제안",
    "action_제안",
    "판정근거",
    "주의플래그",
    "연결_선인세코드",
    "연결_선인세명",
    "scope_assignment_basis",
    "scope_assignment_note",
]
_AUTO_CID_MEMO_RE = re.compile(r"^관측_작품수=\d+, 수동_최종저작권명 비움$")
AUTO_BULK_APPROVAL_MEMO = "권장안 일괄 승인"
AUTO_KEEP_MEMO = "자동 유지"

GENERIC_RIGHTS_NAMES = {
    "기본정산율",
    "구간",
    "선인세 정산 도서",
    "윌라 정산",
    "[카카오창작지원금]",
    "카카오창작지원금",
}
GENERIC_RIGHTS_TOKENS = [
    "기본정산율",
    "구간",
    "신작",
    "카카오창작지원금",
    "창작지원금",
    "선인세 정산 도서",
    "윌라 정산",
]
SPECIAL_SORT_ORDER = {
    "일반": 0,
    "카카오MG": 1,
    "네이버MG": 2,
    "원작": 3,
}


def parse_int(value: Any) -> int:
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return 0
    return int(parsed)


def unique(values: list[Any]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        cleaned = normalize_text(value)
        if not cleaned or cleaned in seen:
            continue
        ordered.append(cleaned)
        seen.add(cleaned)
    return ordered


def observed_works(raw: Any, *, fallback_title: Any = "") -> list[str]:
    works = split_pipe_values(raw)
    if works:
        return works
    fallback = normalize_text(fallback_title)
    return [fallback] if fallback else []


def summarize_author_special_rs(df: pd.DataFrame) -> pd.DataFrame:
    grouped_rows: list[dict[str, Any]] = []
    for author_code, group in df.groupby("작가코드", sort=False):
        service_summaries = unique(
            [
                " ".join(
                    part
                    for part in [
                        normalize_text(row.get("서비스")),
                        f"{normalize_text(row.get('정산율(%)'))}%" if normalize_text(row.get("정산율(%)")) else "",
                        f"[{normalize_text(row.get('정산기준'))}]" if normalize_text(row.get("정산기준")) else "",
                    ]
                    if part
                )
                for row in group.to_dict("records")
            ]
        )
        grouped_rows.append(
            {
                "account_작가코드": normalize_text(author_code),
                "작가특수RS건수": str(len(group)),
                "작가특수RS서비스목록": pipe_join(group["서비스"].tolist()),
                "작가특수RS관리제목목록": pipe_join(group["관리제목"].tolist()),
                "작가특수RS정산기준목록": pipe_join(group["정산기준"].tolist()),
                "작가특수RS요약": " / ".join(service_summaries),
            }
        )
    return pd.DataFrame(grouped_rows)


def is_generic_rights_name(value: Any) -> bool:
    raw = normalize_text(value)
    if not raw:
        return True
    if raw in GENERIC_RIGHTS_NAMES:
        return True
    return any(token in raw for token in GENERIC_RIGHTS_TOKENS)


def suggest_canonical_name(*, representative_work: str, special: str) -> str:
    representative_work = normalize_text(representative_work)
    special = normalize_text(special)
    if not representative_work:
        return ""
    if special in {"", "일반"}:
        return representative_work
    return f"{representative_work} {special}"


def suggest_action(row: dict[str, Any]) -> tuple[str, str, str]:
    assignment_basis = normalize_text(row.get("scope_assignment_basis"))
    assignment_note = normalize_text(row.get("scope_assignment_note"))
    representative_work = normalize_text(row.get("대표작품"))
    rights_name = normalize_text(row.get("현재_저작권명"))
    canonical_name = normalize_text(row.get("canonical_저작권명_제안"))
    special = normalize_text(row.get("특수"))
    generic_name = is_generic_rights_name(rights_name)
    author_special_rs_count = parse_int(row.get("작가특수RS건수"))
    observed_work_count = parse_int(row.get("관측_작품수"))
    unmapped_work_count = parse_int(row.get("관측_미매핑_작품수"))

    flags: list[str] = []
    reasons: list[str] = []

    if assignment_basis == "ambiguous_scope":
        flags.append("작가범위애매")
        reasons.append("scope_assignment_basis=ambiguous_scope")
    if assignment_basis == "manual_work_scope_conflict":
        flags.append("수동작가충돌")
        reasons.append("scope_assignment_basis=manual_work_scope_conflict")
    if assignment_basis in {"manual_work_scope_override", "manual_work_scope_conflict"} and assignment_note:
        reasons.append(f"manual_scope={assignment_note}")
    if not representative_work:
        flags.append("대표작품없음")
        reasons.append("관측_작품목록 비어있음")
    if observed_work_count > 1:
        flags.append("복수작품권리행")
        reasons.append(f"관측_작품수={observed_work_count}")
    if unmapped_work_count > 0:
        flags.append("raw_미매핑작품")
        reasons.append(f"관측_미매핑_작품수={unmapped_work_count}")
    if generic_name:
        flags.append("범용저작권명")
        reasons.append(f"현재 저작권명='{rights_name or '(빈값)'}'")
    if author_special_rs_count > 0:
        flags.append("작가특수RS")
        reasons.append(f"작가특수RS {author_special_rs_count}건")
    if special in {"카카오MG", "네이버MG", "원작"}:
        flags.append("특수권리")
        reasons.append(f"특수={special}")

    if assignment_basis == "ambiguous_scope":
        return "작가확인", pipe_join(reasons), pipe_join(flags)
    if assignment_basis == "manual_work_scope_conflict" and observed_work_count <= 1:
        return "작가확인", pipe_join(reasons), pipe_join(flags)
    if not representative_work:
        return "작품확인", pipe_join(reasons), pipe_join(flags)
    if observed_work_count > 1:
        reasons.append("작품별 CID seed로 분해 필요")
        return "CID분해필요", pipe_join(reasons), pipe_join(flags)
    if generic_name:
        return "이름수정검토", pipe_join(reasons), pipe_join(flags)
    if canonical_name and rights_name == canonical_name:
        return "유지", pipe_join(reasons or ["현재 저작권명과 제안명이 일치"]), pipe_join(flags)
    if canonical_name and rights_name != canonical_name:
        reasons.append(f"제안명='{canonical_name}'")
        return "이름수정검토", pipe_join(reasons), pipe_join(flags)

    return "유지", pipe_join(reasons or ["추가 이슈 없음"]), pipe_join(flags)


def default_manual_fields(*, action: str, observed_work_count: int) -> tuple[str, str, str]:
    if action != "CID분해필요":
        return "", "", ""
    return (
        "",
        "CID분해필요",
        f"관측_작품수={observed_work_count}, 수동_최종저작권명 비움",
    )


def decision_signature(row: dict[str, Any]) -> str:
    payload = "\u241f".join(
        normalize_text(row.get(column))
        for column in DECISION_SIGNATURE_COLUMNS
    )
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def attach_decision_signatures(df: pd.DataFrame) -> pd.DataFrame:
    signed_df = df.copy()
    signed_df["decision_signature"] = [
        decision_signature(row)
        for row in signed_df.to_dict("records")
    ]
    return signed_df


def preserve_existing_manual_fields(decision_df: pd.DataFrame) -> pd.DataFrame:
    if not OUTPUT_CSV.exists():
        return decision_df

    existing_df = load_frame(OUTPUT_CSV)
    if existing_df.empty or "account_저작권코드" not in existing_df.columns:
        return decision_df

    if "decision_signature" not in existing_df.columns:
        existing_df = attach_decision_signatures(existing_df)

    available_columns = [
        "account_저작권코드",
        "decision_signature",
    ] + [column for column in MANUAL_COLUMNS if column in existing_df.columns]
    existing_manual_df = existing_df[available_columns].drop_duplicates(
        subset=["account_저작권코드", "decision_signature"],
        keep="last",
    )

    merged = decision_df.merge(
        existing_manual_df,
        on=["account_저작권코드", "decision_signature"],
        how="left",
        suffixes=("", "__existing"),
    )

    for column in MANUAL_COLUMNS:
        existing_column = f"{column}__existing"
        if existing_column not in merged.columns:
            continue
        merged[column] = merged[existing_column].where(
            merged[existing_column].map(normalize_text).ne(""),
            merged[column],
        )
        merged = merged.drop(columns=[existing_column])

    return merged


def refresh_autogenerated_manual_fields(decision_df: pd.DataFrame) -> pd.DataFrame:
    refreshed_rows: list[dict[str, Any]] = []
    for row in decision_df.to_dict("records"):
        action = normalize_text(row.get("action_제안"))
        manual_name = normalize_text(row.get("수동_최종저작권명"))
        manual_action = normalize_text(row.get("수동_action"))
        manual_note = normalize_text(row.get("수동_메모"))
        processed = normalize_text(row.get("처리완료(Y/N)"))
        observed_work_count = parse_int(row.get("관측_작품수"))

        _, default_action, default_note = default_manual_fields(
            action=action,
            observed_work_count=observed_work_count,
        )
        auto_cid_note = bool(_AUTO_CID_MEMO_RE.match(manual_note))
        generated_name_approval = manual_note == AUTO_BULK_APPROVAL_MEMO

        if action in {"작가확인", "작품확인"} and (generated_name_approval or auto_cid_note):
            row["수동_최종저작권명"] = ""
            row["수동_action"] = ""
            row["수동_메모"] = ""
            row["처리완료(Y/N)"] = ""
            refreshed_rows.append(row)
            continue

        if action == "CID분해필요":
            if manual_name and (generated_name_approval or not manual_action):
                row["수동_최종저작권명"] = ""
                row["수동_action"] = default_action
                row["수동_메모"] = default_note
                row["처리완료(Y/N)"] = ""
                refreshed_rows.append(row)
                continue
            if manual_action in {"", "CID분해필요"} and (not manual_note or auto_cid_note):
                if manual_note != default_note:
                    row["처리완료(Y/N)"] = ""
                row["수동_action"] = default_action
                row["수동_메모"] = default_note
        elif manual_action == "CID분해필요" and (auto_cid_note or generated_name_approval):
            row["수동_action"] = ""
            if auto_cid_note:
                row["수동_메모"] = ""
                row["처리완료(Y/N)"] = ""
            elif not manual_name:
                row["처리완료(Y/N)"] = ""

        if processed != "Y" and not manual_name:
            if action == "CID분해필요":
                if manual_action in {"", "CID분해필요"}:
                    row["수동_action"] = default_action
                if not manual_note or auto_cid_note:
                    row["수동_메모"] = default_note
            elif manual_action == "CID분해필요" and auto_cid_note:
                row["수동_action"] = ""
                row["수동_메모"] = ""

        refreshed_rows.append(row)

    return pd.DataFrame(refreshed_rows)


def build_decision_df() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Path]]:
    actual_df = load_frame(ACTUAL_PATH)
    evidence_df = load_frame(EVIDENCE_PATH)
    special_rs_path = latest_path(f"*__작가특수RS_{MANAGER}.xlsx")
    special_rs_raw = load_frame(special_rs_path)

    evidence_cols = [
        "account_작가코드",
        "account_저작권코드",
        "특수",
        "특수_판정근거",
    ]
    merged = actual_df.merge(
        evidence_df[evidence_cols],
        on=["account_작가코드", "account_저작권코드"],
        how="left",
    )

    rs_summary_df = summarize_author_special_rs(special_rs_raw)
    if not rs_summary_df.empty:
        merged = merged.merge(rs_summary_df, on="account_작가코드", how="left")
    else:
        for column in [
            "작가특수RS건수",
            "작가특수RS서비스목록",
            "작가특수RS관리제목목록",
            "작가특수RS정산기준목록",
            "작가특수RS요약",
        ]:
            merged[column] = ""

    merged = merged.fillna("")

    decision_rows: list[dict[str, Any]] = []
    for row in merged.to_dict("records"):
        works = observed_works(
            row.get("관측_작품목록"),
            fallback_title=row.get("대표_제목"),
        )
        representative_work = works[0] if works else ""
        observed_work_count = len(works)
        canonical_name = suggest_canonical_name(
            representative_work=representative_work,
            special=normalize_text(row.get("특수")) or "일반",
        )
        if observed_work_count > 1:
            canonical_name = ""
        current_row = {
            "대표작가명": normalize_text(row.get("scope_author_primary")),
            "관련작가명": normalize_text(row.get("scope_authors_related")),
            "account_작가코드": normalize_text(row.get("account_작가코드")),
            "account_저작권코드": normalize_text(row.get("account_저작권코드")),
            "현재_저작권명": normalize_text(row.get("account_저작권명")),
            "대표작품": representative_work,
            "관측_작품수": str(observed_work_count),
            "관측_작품목록": normalize_text(row.get("관측_작품목록")),
            "관측_작품_폴더수": normalize_text(row.get("관측_작품_폴더수")),
            "관측_미매핑_작품수": normalize_text(row.get("관측_미매핑_작품수")),
            "관측_미매핑_작품목록": normalize_text(row.get("관측_미매핑_작품목록")),
            "특수": normalize_text(row.get("특수")) or "일반",
            "특수_판정근거": normalize_text(row.get("특수_판정근거")),
            "canonical_저작권명_제안": canonical_name,
            "수동_최종저작권명": "",
            "action_제안": "",
            "수동_action": "",
            "판정근거": "",
            "수동_메모": "",
            "처리완료(Y/N)": "",
            "주의플래그": "",
            "연결_선인세코드": normalize_text(row.get("연결_선인세코드")),
            "연결_선인세명": normalize_text(row.get("연결_선인세명")),
            "rights_선인세명": normalize_text(row.get("rights_선인세명")),
            "B2C_정산율(%)": normalize_text(row.get("B2C_정산율(%)")),
            "B2BC_정산율(%)": normalize_text(row.get("B2BC_정산율(%)")),
            "B2B_정산율(%)": normalize_text(row.get("B2B_정산율(%)")),
            "작가특수RS건수": normalize_text(row.get("작가특수RS건수")),
            "작가특수RS서비스목록": normalize_text(row.get("작가특수RS서비스목록")),
            "작가특수RS관리제목목록": normalize_text(row.get("작가특수RS관리제목목록")),
            "작가특수RS정산기준목록": normalize_text(row.get("작가특수RS정산기준목록")),
            "작가특수RS요약": normalize_text(row.get("작가특수RS요약")),
            "account_정산기준": normalize_text(row.get("account_정산기준")),
            "account_계약담당자": normalize_text(row.get("account_계약담당자")),
            "대표_상품번호": normalize_text(row.get("대표_상품번호")),
            "대표_제목": normalize_text(row.get("대표_제목")),
            "fetched_상품유형": normalize_text(row.get("fetched_상품유형")),
            "admin_이용등급": normalize_text(row.get("admin_이용등급")),
            "scope_assignment_basis": normalize_text(row.get("scope_assignment_basis")),
            "scope_assignment_note": normalize_text(row.get("scope_assignment_note")),
            "관측_match_status": normalize_text(row.get("관측_match_status")),
            "관측_match_reason": normalize_text(row.get("관측_match_reason")),
        }
        action, reason, flags = suggest_action(current_row)
        manual_name, manual_action, manual_note = default_manual_fields(
            action=action,
            observed_work_count=observed_work_count,
        )
        current_row["action_제안"] = action
        current_row["판정근거"] = reason
        current_row["주의플래그"] = flags
        current_row["수동_최종저작권명"] = manual_name
        current_row["수동_action"] = manual_action
        current_row["수동_메모"] = manual_note
        if action == "유지":
            current_row["수동_메모"] = AUTO_KEEP_MEMO
            current_row["처리완료(Y/N)"] = "Y"
        decision_rows.append(current_row)

    decision_df = pd.DataFrame(decision_rows)
    decision_df["정렬_대표작품"] = decision_df["대표작품"].map(normalize_text)
    decision_df["정렬_대표작품없음"] = decision_df["정렬_대표작품"].eq("")
    decision_df["정렬_저작권코드"] = pd.to_numeric(decision_df["account_저작권코드"], errors="coerce").fillna(10**18)
    decision_df["정렬_특수"] = decision_df["특수"].map(lambda value: SPECIAL_SORT_ORDER.get(normalize_text(value), 99))
    decision_df = (
        decision_df.sort_values(
            by=[
                "정렬_대표작품없음",
                "정렬_대표작품",
                "대표작가명",
                "정렬_저작권코드",
                "정렬_특수",
                "account_저작권코드",
            ],
            kind="stable",
        )
        .drop(columns=["정렬_대표작품", "정렬_대표작품없음", "정렬_저작권코드", "정렬_특수"])
        .reset_index(drop=True)
    )
    decision_df = attach_decision_signatures(decision_df)
    decision_df = preserve_existing_manual_fields(decision_df)
    decision_df = refresh_autogenerated_manual_fields(decision_df)

    special_rs_raw = special_rs_raw.sort_values(
        by=["작가코드", "서비스", "작가특수RS코드"],
        kind="stable",
    ).reset_index(drop=True)

    if not rs_summary_df.empty:
        author_names = (
            decision_df.groupby("account_작가코드")["대표작가명"]
            .agg(lambda series: pipe_join(series.tolist()))
            .reset_index()
        )
        rs_summary_df = (
            author_names.merge(rs_summary_df, on="account_작가코드", how="right")
            .rename(columns={"대표작가명": "scope_대표작가명"})
            .sort_values(by=["scope_대표작가명", "account_작가코드"], kind="stable")
            .reset_index(drop=True)
        )

    source_paths = {
        "account_actual_inventory": ACTUAL_PATH,
        "special_evidence_inventory": EVIDENCE_PATH,
        "author_special_rs": special_rs_path,
    }
    return decision_df, rs_summary_df, special_rs_raw, source_paths


def write_summary_sheet(
    wb: Workbook,
    *,
    decision_df: pd.DataFrame,
    rs_summary_df: pd.DataFrame,
    source_paths: dict[str, Path],
) -> None:
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{MANAGER} account decision queue"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "지표"
    ws["B3"] = "값"
    apply_header_style(ws["A3"])
    apply_header_style(ws["B3"])

    metrics = [
        ("decision 행 수", len(decision_df)),
        ("고유 작가코드 수", int(decision_df["account_작가코드"].nunique())),
        ("고유 저작권코드 수", int(decision_df["account_저작권코드"].nunique())),
        ("이름수정검토 행 수", int(decision_df["action_제안"].eq("이름수정검토").sum())),
        ("CID분해필요 행 수", int(decision_df["action_제안"].eq("CID분해필요").sum())),
        ("작가확인 행 수", int(decision_df["action_제안"].eq("작가확인").sum())),
        ("작품확인 행 수", int(decision_df["action_제안"].eq("작품확인").sum())),
        ("raw 미매핑 작품 보유 권리 행 수", int(decision_df["관측_미매핑_작품수"].map(parse_int).gt(0).sum())),
        ("작가특수RS 보유 작가 수", int(rs_summary_df["account_작가코드"].nunique()) if not rs_summary_df.empty else 0),
        ("작가특수RS 보유 권리 행 수", int(decision_df["작가특수RS건수"].map(parse_int).gt(0).sum())),
    ]

    for row_idx, (label, value) in enumerate(metrics, start=4):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2))

    action_counter = Counter(decision_df["action_제안"].tolist())
    start_row = 14
    ws.cell(row=start_row, column=1, value="action_제안")
    ws.cell(row=start_row, column=2, value="행 수")
    apply_header_style(ws.cell(row=start_row, column=1))
    apply_header_style(ws.cell(row=start_row, column=2))

    action_labels = ["유지", "이름수정검토", "CID분해필요", "작가확인", "작품확인"]
    for idx, label in enumerate(action_labels, start=start_row + 1):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=action_counter.get(label, 0))
        apply_body_style(ws.cell(row=idx, column=1))
        apply_body_style(ws.cell(row=idx, column=2))

    chart = BarChart()
    chart.title = "decision action 분포"
    chart.y_axis.title = "행 수"
    chart.x_axis.title = "action"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(action_labels))
    categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(action_labels))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "D4")

    source_row = start_row + len(action_labels) + 3
    ws.cell(row=source_row, column=1, value="입력 소스")
    ws.cell(row=source_row, column=2, value="경로")
    apply_header_style(ws.cell(row=source_row, column=1))
    apply_header_style(ws.cell(row=source_row, column=2))

    for offset, (label, path) in enumerate(source_paths.items(), start=1):
        ws.cell(row=source_row + offset, column=1, value=label)
        ws.cell(row=source_row + offset, column=2, value=str(path))
        apply_body_style(ws.cell(row=source_row + offset, column=1))
        apply_body_style(ws.cell(row=source_row + offset, column=2), wrap=True)

    set_column_widths(ws)


def write_decision_sheet(wb: Workbook, decision_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("account_decision_queue")
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    rows = [list(decision_df.columns)] + decision_df.astype(object).where(pd.notnull(decision_df), "").values.tolist()

    for row_idx, row_values in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                apply_header_style(cell)
                continue

            action = normalize_text(ws.cell(row=row_idx, column=13).value)
            special = normalize_text(ws.cell(row=row_idx, column=9).value)
            flags = normalize_text(ws.cell(row=row_idx, column=18).value)

            fill = None
            if action in {"이름수정검토", "CID분해필요", "작가확인", "작품확인"}:
                fill = _WARN_FILL
            elif special in {"카카오MG", "네이버MG", "원작"} or "작가특수RS" in flags:
                fill = _GOOD_FILL

            apply_body_style(cell, fill=fill, wrap=True)

    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24
    set_column_widths(ws)


def write_outputs(
    *,
    decision_df: pd.DataFrame,
    rs_summary_df: pd.DataFrame,
    special_rs_raw: pd.DataFrame,
    source_paths: dict[str, Path],
) -> dict[str, Path]:
    TARGET_ROOT.mkdir(parents=True, exist_ok=True)
    decision_df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    wb = Workbook()
    write_summary_sheet(wb, decision_df=decision_df, rs_summary_df=rs_summary_df, source_paths=source_paths)
    write_decision_sheet(wb, decision_df)

    if not rs_summary_df.empty:
        from build_account_observation_bundle import write_df_sheet

        write_df_sheet(wb, "author_special_rs_summary", rs_summary_df)
    else:
        from build_account_observation_bundle import write_df_sheet

        write_df_sheet(wb, "author_special_rs_summary", pd.DataFrame(columns=["account_작가코드"]))

    write_df_sheet(wb, "author_special_rs_raw", special_rs_raw if not special_rs_raw.empty else pd.DataFrame(columns=["작가코드"]))

    try:
        wb.save(OUTPUT_XLSX)
    except PermissionError as exc:
        raise RuntimeError(
            f"최신판 1개 정책을 유지하려면 '{OUTPUT_XLSX.name}' 파일을 닫은 뒤 다시 실행해야 합니다."
        ) from exc

    summary = {
        "manager": MANAGER,
        "decision_rows": int(len(decision_df)),
        "rename_review_rows": int(decision_df["action_제안"].eq("이름수정검토").sum()),
        "cid_split_rows": int(decision_df["action_제안"].eq("CID분해필요").sum()),
        "author_review_rows": int(decision_df["action_제안"].eq("작가확인").sum()),
        "work_review_rows": int(decision_df["action_제안"].eq("작품확인").sum()),
        "author_special_rs_rows": int(len(special_rs_raw)),
        "workbook": str(OUTPUT_XLSX),
    }
    OUTPUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "csv": OUTPUT_CSV,
        "workbook": OUTPUT_XLSX,
        "summary": OUTPUT_JSON,
    }


def main() -> None:
    decision_df, rs_summary_df, special_rs_raw, source_paths = build_decision_df()
    outputs = write_outputs(
        decision_df=decision_df,
        rs_summary_df=rs_summary_df,
        special_rs_raw=special_rs_raw,
        source_paths=source_paths,
    )
    print("=== account decision queue built ===")
    print(
        json.dumps(
            {
                "manager": MANAGER,
                "decision_rows": len(decision_df),
                "author_special_rs_rows": len(special_rs_raw),
                "workbook": str(outputs["workbook"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
