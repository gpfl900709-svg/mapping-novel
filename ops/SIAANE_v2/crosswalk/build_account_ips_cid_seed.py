from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parents[1]
ACCOUNT_DIR = ROOT / "SIAANE_v2" / "account"
if str(ACCOUNT_DIR) not in sys.path:
    sys.path.insert(0, str(ACCOUNT_DIR))

from build_account_observation_bundle import (  # noqa: E402
    MANAGER,
    MANUAL_WORK_SCOPE_OVERRIDES_PATH,
    _GOOD_FILL,
    _WARN_FILL,
    apply_body_style,
    apply_header_style,
    compact_key,
    load_manual_work_scope_overrides,
    load_frame,
    normalize_matchable_work_title,
    normalize_text,
    pipe_join,
    resolve_manual_work_scope_override,
    set_column_widths,
    should_ignore_work_title,
    split_pipe_values,
    write_df_sheet,
)


OUTPUT_ROOT = ROOT / "SIAANE_v2" / "crosswalk" / "exports"
ACTUAL_PATH = ROOT / "SIAANE_v2" / "account" / "exports" / f"latest__account_actual_inventory_{MANAGER}.csv"
EVIDENCE_PATH = ROOT / "SIAANE_v2" / "account" / "exports" / f"latest__special_evidence_inventory_{MANAGER}.csv"
DECISION_PATH = ROOT / "SIAANE_v2" / "account" / "exports" / f"latest__account_decision_queue_{MANAGER}.csv"

OUTPUT_XLSX = OUTPUT_ROOT / f"latest__account_ips_cid_seed_{MANAGER}.xlsx"
OUTPUT_CSV = OUTPUT_ROOT / f"latest__account_ips_cid_seed_{MANAGER}.csv"
OUTPUT_JSON = OUTPUT_ROOT / f"latest__account_ips_cid_seed_{MANAGER}.json"
MANUAL_CID_SEED_WORK_OVERRIDES_PATH = (
    ROOT / "SIAANE_v2" / "account" / "notes" / "manual_cid_seed_work_overrides.json"
)
_SEED_TRAILING_CHANNEL_TAG_RE = re.compile(r"\s+(?:카카오\s*제외|카카오\s*외\s*연재|타플)\s*$")
_SEED_TRAILING_CHANNEL_TAG_PAREN_RE = re.compile(r"\s*[(（](?:카카오\s*제외|카카오\s*외\s*연재|타플)[)）]\s*$")
_SEED_VOLUME_BEFORE_CHANNEL_TAG_RE = re.compile(
    r"\s+\d+(?:-\d+)?\s*(?:화|권)(?=\s+(?:카카오\s*제외|카카오\s*외\s*연재|타플)\s*$)"
)
_SEED_TRAILING_PUNCT_RE = re.compile(r"[!?.,~]+$")


def unique(values: list[Any]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        cleaned = normalize_text(value)
        if not cleaned or cleaned in seen:
            continue
        ordered.append(cleaned)
        seen.add(cleaned)
    return ordered


def parse_int(value: Any) -> int:
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return 0
    return int(parsed)


def non_empty_nunique(series: pd.Series) -> int:
    cleaned = series.map(normalize_text)
    return int(cleaned[cleaned.ne("")].nunique())


def load_manual_cid_seed_work_overrides(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError(f"manual CID seed work overrides must be a list: {path}")

    overrides: list[dict[str, str]] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        normalized = {str(key): normalize_text(value) for key, value in item.items()}
        if not normalized.get("작품명"):
            continue
        overrides.append(normalized)
    return overrides


def observed_works(raw: Any, *, fallback_title: Any = "") -> list[str]:
    return unique([normalized for normalized, _ in observed_work_pairs(raw, fallback_title=fallback_title)])


def normalize_seed_work_title(value: Any) -> str:
    raw = normalize_text(value)
    if should_ignore_work_title(raw):
        return ""

    normalized = normalize_matchable_work_title(raw) or raw
    normalized = _SEED_TRAILING_CHANNEL_TAG_PAREN_RE.sub("", normalized)
    normalized = _SEED_VOLUME_BEFORE_CHANNEL_TAG_RE.sub("", normalized)
    normalized = _SEED_TRAILING_CHANNEL_TAG_RE.sub("", normalized)
    normalized = normalize_text(normalized).strip(" -_")
    normalized = _SEED_TRAILING_PUNCT_RE.sub("", normalized)
    return normalize_text(normalized)


def observed_work_pairs(raw: Any, *, fallback_title: Any = "") -> list[tuple[str, str]]:
    raw_titles = split_pipe_values(raw)
    if not raw_titles:
        fallback = normalize_text(fallback_title)
        raw_titles = [fallback] if fallback else []

    pairs: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    seen_compact_titles: set[str] = set()
    for raw_title in raw_titles:
        cleaned_raw = normalize_text(raw_title)
        normalized_title = normalize_seed_work_title(cleaned_raw)
        if not normalized_title:
            continue
        compact_title = compact_key(normalize_matchable_work_title(normalized_title) or normalized_title)
        if compact_title and compact_title in seen_compact_titles:
            continue
        pair = (normalized_title, cleaned_raw)
        if pair in seen:
            continue
        pairs.append(pair)
        seen.add(pair)
        if compact_title:
            seen_compact_titles.add(compact_title)
    return pairs


def choose_settlement_actor(row: dict[str, Any]) -> tuple[str, str]:
    candidates = [
        ("예금주", row.get("account_예금주")),
        ("실명", row.get("account_real_name")),
        ("필명", row.get("account_pen_name")),
    ]
    for source, value in candidates:
        cleaned = normalize_text(value)
        if cleaned:
            return cleaned, source
    return "", ""


def derive_representative_work(row: dict[str, Any], works: list[str]) -> str:
    representative_title = normalize_seed_work_title(row.get("대표_제목"))
    if representative_title and representative_title in works:
        return representative_title
    return works[0] if works else representative_title


def match_manual_cid_seed_work_override(
    *,
    row: dict[str, Any],
    work_title: str,
    raw_work_title: str,
    overrides: list[dict[str, str]],
) -> dict[str, str] | None:
    rights_code = normalize_text(row.get("account_저작권코드"))
    author_code = normalize_text(row.get("account_작가코드"))
    candidate_titles = unique([work_title, raw_work_title])
    candidate_keys = {
        compact_key(normalize_matchable_work_title(title) or title)
        for title in candidate_titles
        if normalize_text(title)
    }

    matched: dict[str, str] | None = None
    for override in overrides:
        override_rights_code = normalize_text(override.get("account_저작권코드"))
        override_author_code = normalize_text(override.get("account_작가코드"))
        if override_rights_code and override_rights_code != rights_code:
            continue
        if override_author_code and override_author_code != author_code:
            continue

        override_title = normalize_text(override.get("작품명"))
        override_key = compact_key(normalize_matchable_work_title(override_title) or override_title)
        if override_key and override_key in candidate_keys:
            matched = override

    return matched


def apply_manual_cid_seed_work_overrides(
    *,
    row: dict[str, Any],
    work_pairs: list[tuple[str, str]],
    overrides: list[dict[str, str]],
) -> list[tuple[str, str]]:
    resolved_pairs: list[tuple[str, str]] = []
    seen_compact_titles: set[str] = set()

    for work_title, raw_work_title in work_pairs:
        override = match_manual_cid_seed_work_override(
            row=row,
            work_title=work_title,
            raw_work_title=raw_work_title,
            overrides=overrides,
        )
        if override and normalize_text(override.get("action")).lower() in {"exclude", "제외"}:
            continue

        resolved_work_title = work_title
        if override and normalize_text(override.get("CID_작품명")):
            resolved_work_title = normalize_seed_work_title(override.get("CID_작품명"))

        compact_title = compact_key(normalize_matchable_work_title(resolved_work_title) or resolved_work_title)
        if compact_title and compact_title in seen_compact_titles:
            continue
        resolved_pairs.append((resolved_work_title, raw_work_title))
        if compact_title:
            seen_compact_titles.add(compact_title)

    return resolved_pairs


def resolve_seed_work_author(
    *,
    row: dict[str, Any],
    work_title: str,
    raw_work_title: str,
    manual_work_scope_overrides: list[dict[str, str]],
) -> tuple[str, str]:
    manual_authors, manual_note = resolve_manual_work_scope_override(
        author_code=normalize_text(row.get("account_작가코드")),
        rights_code=normalize_text(row.get("account_저작권코드")),
        observed_titles=unique([raw_work_title, work_title]),
        overrides=manual_work_scope_overrides,
    )
    if len(manual_authors) == 1:
        return manual_authors[0], manual_note

    return normalize_text(row.get("scope_author_primary")), ""


def build_cid_seed_id(
    *,
    work_title: str,
    author_name: str,
    advance_code: str,
    special: str,
    rights_code: str,
    settlement_actor: str,
    is_primary: str,
) -> str:
    parts = [
        normalize_text(work_title) or "미상작품",
        normalize_text(author_name) or "미상작가",
        normalize_text(advance_code) or "선인세없음",
        normalize_text(special) or "일반",
        normalize_text(rights_code) or "미상권리",
        normalize_text(settlement_actor) or "미상",
        normalize_text(is_primary) or "N",
    ]
    return "_".join(parts)


def build_seed_df() -> tuple[pd.DataFrame, dict[str, Path]]:
    actual_df = load_frame(ACTUAL_PATH)
    evidence_df = load_frame(EVIDENCE_PATH)
    decision_df = load_frame(DECISION_PATH)
    manual_work_scope_overrides = load_manual_work_scope_overrides(MANUAL_WORK_SCOPE_OVERRIDES_PATH)
    manual_cid_seed_work_overrides = load_manual_cid_seed_work_overrides(
        MANUAL_CID_SEED_WORK_OVERRIDES_PATH
    )

    evidence_cols = [
        "account_작가코드",
        "account_저작권코드",
        "특수",
        "특수_판정근거",
    ]
    merged = actual_df.merge(
        evidence_df[evidence_cols],
        on=["account_작가코드", "account_저작권코드"],
        how="left",
    ).fillna("")

    decision_cols = [
        "account_저작권코드",
        "action_제안",
        "수동_최종저작권명",
        "수동_action",
        "수동_메모",
        "처리완료(Y/N)",
        "주의플래그",
    ]
    merged = merged.merge(
        decision_df[decision_cols],
        on="account_저작권코드",
        how="left",
    ).fillna("")

    seed_rows: list[dict[str, Any]] = []
    for row in merged.to_dict("records"):
        decision_action = normalize_text(row.get("action_제안"))
        if decision_action in {"작가확인", "작품확인"}:
            continue
        if normalize_text(row.get("수동_action")) == "담당제외":
            continue

        work_pairs = observed_work_pairs(
            row.get("관측_작품목록"),
            fallback_title=row.get("대표_제목"),
        )
        work_pairs = apply_manual_cid_seed_work_overrides(
            row=row,
            work_pairs=work_pairs,
            overrides=manual_cid_seed_work_overrides,
        )
        if not work_pairs:
            continue
        works = unique([normalized for normalized, _ in work_pairs])
        representative_work = derive_representative_work(row, works)
        settlement_actor, settlement_source = choose_settlement_actor(row)
        special = normalize_text(row.get("특수")) or "일반"
        advance_code = normalize_text(row.get("연결_선인세코드"))
        advance_name = normalize_text(row.get("연결_선인세명"))
        primary_author = normalize_text(row.get("scope_author_primary"))
        related_authors = normalize_text(row.get("scope_authors_related"))
        observed_work_count = len(works)

        if not work_pairs:
            work_pairs = [("", "")]

        for work, raw_work in work_pairs:
            work_author, work_author_note = resolve_seed_work_author(
                row=row,
                work_title=work,
                raw_work_title=raw_work,
                manual_work_scope_overrides=manual_work_scope_overrides,
            )
            related_author_list = unique([work_author] + split_pipe_values(related_authors))
            is_primary = "Y" if work and work == representative_work else "N"
            if len(works) == 1 and work:
                is_primary = "Y"

            flags = []
            if observed_work_count > 1:
                flags.append("복수작품권리행")
            if not advance_code:
                flags.append("선인세미연결")
            if not settlement_actor:
                flags.append("정산자미상")
            if not work_author:
                flags.append("작가미확정")
            if work_author_note:
                flags.append("작품별수동작가")

            seed_rows.append(
                {
                    "작품명": normalize_text(work),
                    "원본_작품명_목록": normalize_text(raw_work),
                    "대표작가명": work_author,
                    "관련작가명": pipe_join(related_author_list),
                    "정산자": settlement_actor,
                    "정산자_근거": settlement_source,
                    "연결_선인세코드": advance_code,
                    "연결_선인세명": advance_name,
                    "특수": special,
                    "account_저작권코드": normalize_text(row.get("account_저작권코드")),
                    "account_작가코드": normalize_text(row.get("account_작가코드")),
                    "account_저작권명": normalize_text(row.get("account_저작권명")),
                    "관측_작품수": str(observed_work_count),
                    "정산대표Y/N": is_primary,
                    "CID_seed_id": build_cid_seed_id(
                        work_title=normalize_text(work),
                        author_name=work_author,
                        advance_code=advance_code,
                        special=special,
                        rights_code=normalize_text(row.get("account_저작권코드")),
                        settlement_actor=settlement_actor,
                        is_primary=is_primary,
                    ),
                    "decision_action_제안": decision_action,
                    "decision_수동_최종저작권명": normalize_text(row.get("수동_최종저작권명")),
                    "decision_수동_action": normalize_text(row.get("수동_action")),
                    "decision_수동_메모": normalize_text(row.get("수동_메모")),
                    "decision_처리완료(Y/N)": normalize_text(row.get("처리완료(Y/N)")),
                    "주의플래그": pipe_join(flags + split_pipe_values(row.get("주의플래그"))),
                    "대표_상품번호": normalize_text(row.get("대표_상품번호")),
                    "대표_제목": normalize_text(row.get("대표_제목")),
                    "account_정산기준": normalize_text(row.get("account_정산기준")),
                    "account_계약담당자": normalize_text(row.get("account_계약담당자")),
                }
            )

    seed_df = pd.DataFrame(seed_rows)
    if seed_df.empty:
        return seed_df, {
            "account_actual_inventory": ACTUAL_PATH,
            "special_evidence_inventory": EVIDENCE_PATH,
            "account_decision_queue": DECISION_PATH,
            "manual_work_scope_overrides": MANUAL_WORK_SCOPE_OVERRIDES_PATH,
            "manual_cid_seed_work_overrides": MANUAL_CID_SEED_WORK_OVERRIDES_PATH,
        }

    group_cols = [
        "작품명",
        "대표작가명",
        "연결_선인세코드",
        "특수",
        "account_저작권코드",
        "정산자",
    ]
    aggregate_cols = [
        "관련작가명",
        "원본_작품명_목록",
        "정산자_근거",
        "연결_선인세명",
        "account_작가코드",
        "account_저작권명",
        "관측_작품수",
        "CID_seed_id",
        "decision_action_제안",
        "decision_수동_최종저작권명",
        "decision_수동_action",
        "decision_수동_메모",
        "decision_처리완료(Y/N)",
        "주의플래그",
        "대표_상품번호",
        "대표_제목",
        "account_정산기준",
        "account_계약담당자",
    ]

    aggregated_rows: list[dict[str, Any]] = []
    for key, group in seed_df.groupby(group_cols, dropna=False, sort=False):
        first_row = group.iloc[0].to_dict()
        row = {column: value for column, value in zip(group_cols, key)}
        row["정산대표Y/N"] = "Y" if group["정산대표Y/N"].astype(str).eq("Y").any() else "N"
        for column in aggregate_cols:
            if column == "CID_seed_id":
                row[column] = build_cid_seed_id(
                    work_title=row["작품명"],
                    author_name=row["대표작가명"],
                    advance_code=row["연결_선인세코드"],
                    special=row["특수"],
                    rights_code=row["account_저작권코드"],
                    settlement_actor=row["정산자"],
                    is_primary=row["정산대표Y/N"],
                )
                continue
            row[column] = pipe_join(group[column].tolist())
        row["대표_상품번호"] = first_row.get("대표_상품번호", "")
        row["대표_제목"] = first_row.get("대표_제목", "")
        aggregated_rows.append(row)

    seed_df = pd.DataFrame(aggregated_rows)
    seed_df["정렬_작품명"] = seed_df["작품명"].map(normalize_text)
    seed_df["정렬_작가명"] = seed_df["대표작가명"].map(normalize_text)
    seed_df["정렬_저작권코드"] = pd.to_numeric(seed_df["account_저작권코드"], errors="coerce").fillna(10**18)
    seed_df = (
        seed_df.sort_values(
            by=["정렬_작품명", "정렬_작가명", "정렬_저작권코드", "정산대표Y/N"],
            ascending=[True, True, True, False],
            kind="stable",
        )
        .drop(columns=["정렬_작품명", "정렬_작가명", "정렬_저작권코드"])
        .reset_index(drop=True)
    )

    return seed_df, {
        "account_actual_inventory": ACTUAL_PATH,
        "special_evidence_inventory": EVIDENCE_PATH,
        "account_decision_queue": DECISION_PATH,
        "manual_work_scope_overrides": MANUAL_WORK_SCOPE_OVERRIDES_PATH,
        "manual_cid_seed_work_overrides": MANUAL_CID_SEED_WORK_OVERRIDES_PATH,
    }


def write_summary_sheet(wb: Workbook, *, seed_df: pd.DataFrame, source_paths: dict[str, Path]) -> None:
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{MANAGER} account→IPS CID seed"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "지표"
    ws["B3"] = "값"
    apply_header_style(ws["A3"])
    apply_header_style(ws["B3"])

    metrics = [
        ("seed 행 수", len(seed_df)),
        ("고유 작품 수", non_empty_nunique(seed_df["작품명"])),
        ("고유 저작권코드 수", int(seed_df["account_저작권코드"].nunique())),
        ("고유 선인세코드 수", int(seed_df["연결_선인세코드"].astype(str).str.strip().replace("", pd.NA).dropna().nunique())),
        ("복수작품권리행 seed 수", int(seed_df["주의플래그"].astype(str).str.contains("복수작품권리행", regex=False).sum())),
        ("선인세미연결 seed 수", int(seed_df["주의플래그"].astype(str).str.contains("선인세미연결", regex=False).sum())),
    ]

    for row_idx, (label, value) in enumerate(metrics, start=4):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2))

    action_counter = Counter(seed_df["decision_action_제안"].tolist())
    start_row = 12
    ws.cell(row=start_row, column=1, value="decision_action_제안")
    ws.cell(row=start_row, column=2, value="행 수")
    apply_header_style(ws.cell(row=start_row, column=1))
    apply_header_style(ws.cell(row=start_row, column=2))

    action_labels = ["유지", "이름수정검토", "CID분해필요", "작가확인", "작품확인"]
    for idx, label in enumerate(action_labels, start=start_row + 1):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=action_counter.get(label, 0))
        apply_body_style(ws.cell(row=idx, column=1))
        apply_body_style(ws.cell(row=idx, column=2))

    chart = BarChart()
    chart.title = "CID seed source action 분포"
    chart.y_axis.title = "행 수"
    chart.x_axis.title = "action"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(action_labels))
    categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(action_labels))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "D4")

    source_row = start_row + len(action_labels) + 3
    ws.cell(row=source_row, column=1, value="입력 소스")
    ws.cell(row=source_row, column=2, value="경로")
    apply_header_style(ws.cell(row=source_row, column=1))
    apply_header_style(ws.cell(row=source_row, column=2))

    for offset, (label, path) in enumerate(source_paths.items(), start=1):
        ws.cell(row=source_row + offset, column=1, value=label)
        ws.cell(row=source_row + offset, column=2, value=str(path))
        apply_body_style(ws.cell(row=source_row + offset, column=1))
        apply_body_style(ws.cell(row=source_row + offset, column=2), wrap=True)

    set_column_widths(ws)


def write_seed_sheet(wb: Workbook, seed_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("account_ips_cid_seed")
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    rows = [list(seed_df.columns)] + seed_df.astype(object).where(pd.notnull(seed_df), "").values.tolist()

    decision_action_idx = seed_df.columns.get_loc("decision_action_제안") + 1
    special_idx = seed_df.columns.get_loc("특수") + 1
    flags_idx = seed_df.columns.get_loc("주의플래그") + 1

    for row_idx, row_values in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                apply_header_style(cell)
                continue

            action = normalize_text(ws.cell(row=row_idx, column=decision_action_idx).value)
            special = normalize_text(ws.cell(row=row_idx, column=special_idx).value)
            flags = normalize_text(ws.cell(row=row_idx, column=flags_idx).value)

            fill = None
            if "복수작품권리행" in flags or "선인세미연결" in flags or action in {"CID분해필요", "작가확인", "작품확인"}:
                fill = _WARN_FILL
            elif special in {"카카오MG", "네이버MG", "원작"}:
                fill = _GOOD_FILL

            apply_body_style(cell, fill=fill, wrap=True)

    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24
    set_column_widths(ws)


def write_outputs(*, seed_df: pd.DataFrame, source_paths: dict[str, Path]) -> dict[str, Path]:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    seed_df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    wb = Workbook()
    write_summary_sheet(wb, seed_df=seed_df, source_paths=source_paths)
    write_seed_sheet(wb, seed_df)
    write_df_sheet(wb, "raw_seed", seed_df)

    try:
        wb.save(OUTPUT_XLSX)
    except PermissionError as exc:
        raise RuntimeError(
            f"최신판 1개 정책을 유지하려면 '{OUTPUT_XLSX.name}' 파일을 닫은 뒤 다시 실행해야 합니다."
        ) from exc

    summary = {
        "manager": MANAGER,
        "seed_rows": int(len(seed_df)),
        "distinct_works": non_empty_nunique(seed_df["작품명"]),
        "distinct_rights_codes": int(seed_df["account_저작권코드"].nunique()),
        "multi_work_rows": int(seed_df["주의플래그"].astype(str).str.contains("복수작품권리행", regex=False).sum()),
        "missing_advance_rows": int(seed_df["주의플래그"].astype(str).str.contains("선인세미연결", regex=False).sum()),
        "workbook": str(OUTPUT_XLSX),
    }
    OUTPUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "csv": OUTPUT_CSV,
        "workbook": OUTPUT_XLSX,
        "summary": OUTPUT_JSON,
    }


def main() -> None:
    seed_df, source_paths = build_seed_df()
    outputs = write_outputs(seed_df=seed_df, source_paths=source_paths)
    print("=== account→IPS CID seed built ===")
    print(
        json.dumps(
            {
                "manager": MANAGER,
                "seed_rows": len(seed_df),
                "workbook": str(outputs["workbook"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
