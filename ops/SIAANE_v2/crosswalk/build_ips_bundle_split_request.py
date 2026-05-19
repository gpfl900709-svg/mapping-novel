from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"
WORK_ORDER = EXPORT_ROOT / f"latest__ips_apply_work_order_{MANAGER}.csv"
OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_bundle_split_request_{MANAGER}.xlsx"
OUTPUT_CSV = EXPORT_ROOT / f"latest__ips_bundle_split_request_{MANAGER}.csv"


def text(value: Any) -> str:
    return str(value or "").strip()


def split_target_name(name: str) -> list[str]:
    parts = text(name).split("_")
    if len(parts) != 7:
        return ["", "", "", "", "", "", ""]
    return parts


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    input_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = ws.cell(row=1, column=cell.column).value
                if header in {"신규_콘텐츠ID_입력", "분해결과", "등록메모", "분해후_원묶음CID_처리"}:
                    cell.fill = input_fill
                if ws.title in {"검증", "확인필요"}:
                    cell.fill = warn_fill
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(text(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 90)
    wb.save(path)


def build_rows(source: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for _, row in source.iterrows():
        parts = split_target_name(row["변경후IPS명"])
        rows.append(
            {
                "분해순번": len(rows) + 1,
                "원묶음CID": row["대상CID"],
                "원묶음명": row["대상현재명"],
                "분해후_원묶음CID_처리": "유지/종료 판단",
                "신규_콘텐츠ID_입력": "",
                "신규_콘텐츠명": row["변경후IPS명"],
                "콘텐츠형태": "소설",
                "귀속법인": "키다리스튜디오",
                "서비스유형": "연재",
                "담당부서": "소설편집팀",
                "담당자명": MANAGER,
                "작품명": parts[0] or row["작품명"],
                "대표작가명": parts[1] or row["대표작가명"],
                "선인세코드": parts[2] or row["선인세코드"],
                "특수": parts[3] or row["특수"],
                "account_저작권코드": parts[4] or row["account_저작권코드"],
                "정산자": parts[5] or row["정산자"],
                "정산대표Y/N": parts[6] or row["정산대표Y/N"],
                "분해결과": "",
                "등록메모": "",
                "처리메모": row["처리메모"],
                "seed_row_index": row["seed_row_index"],
            }
        )
    return pd.DataFrame(rows)


def main() -> None:
    work = pd.read_csv(WORK_ORDER, dtype=str).fillna("")
    split = work[work["작업구분"].eq("04_묶음분해")].copy()
    ready = split[split["작업상태"].eq("대기")].copy()
    hold = split[~split["작업상태"].eq("대기")].copy()

    request = build_rows(ready)
    request = request.sort_values(["원묶음CID", "대표작가명", "작품명", "신규_콘텐츠명"], kind="stable")
    hold = hold.sort_values(["대상CID", "대표작가명", "작품명", "변경후IPS명"], kind="stable")

    validation = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("입력 작업지시서", str(WORK_ORDER)),
            ("분해 요청 건수", len(request)),
            ("확인필요 보류 건수", len(hold)),
            ("원묶음CID 수", int(request["원묶음CID"].nunique()) if not request.empty else 0),
            ("신규_콘텐츠명 중복", int(request.duplicated(["신규_콘텐츠명"], keep=False).sum()) if not request.empty else 0),
            ("7파트 포맷 오류", int((request["신규_콘텐츠명"].map(lambda v: len(text(v).split("_"))) != 7).sum()) if not request.empty else 0),
            ("특수값 오류", int((~request["특수"].isin(["일반", "카카오MG", "네이버MG", "원작"])).sum()) if not request.empty else 0),
            ("원칙", "원묶음 CID를 로컬에서 덮어쓰지 않음. IPS 등록/분해 후 발급 CID를 입력."),
        ],
        columns=["항목", "값"],
    )
    summary = (
        request.groupby(["원묶음CID", "특수"], dropna=False).size().reset_index(name="건수")
        if not request.empty
        else pd.DataFrame(columns=["원묶음CID", "특수", "건수"])
    )

    request.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        validation.to_excel(writer, sheet_name="검증", index=False)
        summary.to_excel(writer, sheet_name="요약", index=False)
        request.to_excel(writer, sheet_name="묶음분해_요청", index=False)
        if not hold.empty:
            hold.to_excel(writer, sheet_name="확인필요", index=False)
    autosize(OUTPUT_XLSX)

    print("=== IPS bundle split request built ===")
    print(f"ready={len(request)} hold={len(hold)}")
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
