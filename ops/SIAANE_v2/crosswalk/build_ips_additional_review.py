from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"
WORK_ORDER = EXPORT_ROOT / f"latest__ips_apply_work_order_{MANAGER}.csv"
CANDIDATES = EXPORT_ROOT / f"latest__account_ips_match_candidates_{MANAGER}.csv"
OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_additional_review_{MANAGER}.xlsx"
OUTPUT_CSV = EXPORT_ROOT / f"latest__ips_additional_review_{MANAGER}.csv"


def text(value: Any) -> str:
    return str(value or "").strip()


def truthy_y(value: Any) -> bool:
    return text(value).upper() == "Y"


def decide(row: pd.Series) -> tuple[str, str]:
    action = text(row.get("적용액션"))
    current_name = text(row.get("현재IPS명"))
    target_name = text(row.get("변경후IPS명"))
    candidate_score = pd.to_numeric(row.get("best_candidate_score", ""), errors="coerce")
    disabled = truthy_y(row.get("best_is_disabled(Y/N)"))
    weak = truthy_y(row.get("best_weak_match(Y/N)"))
    reason = text(row.get("best_match_reason"))

    if action == "사용안함검토":
        if disabled and not weak and candidate_score >= 80 and "title+author" in reason:
            return "사용안함_승격가능", "사용안함 사유만 확인 후 목표 IPS명으로 승격 가능."
        return "사용안함_보류", "사용안함 후보이지만 약매칭/근거부족. 수동 확인 필요."

    if action == "매칭검토":
        if not disabled and not weak and candidate_score >= 90 and current_name == target_name:
            return "매칭_유지가능", "현재명과 목표명이 같고 강매칭. 유지 처리 가능."
        if not disabled and not weak and candidate_score >= 85:
            return "매칭_이름수정가능", "강매칭 후보. 현재 CID 유지 후 목표 IPS명으로 이름수정 검토."
        return "매칭_보류", "후보 점수/근거가 약하거나 다중 후보. 수동 확인 필요."

    return "기타_보류", "예상하지 못한 추가검토 유형."


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if ws.title not in {"검증", "요약"}:
            headers = [text(cell.value) for cell in ws[1]]
            cls_idx = headers.index("검토판정") + 1 if "검토판정" in headers else None
            if cls_idx:
                for row in ws.iter_rows(min_row=2):
                    fill = ok_fill if "가능" in text(row[cls_idx - 1].value) else warn_fill
                    for cell in row:
                        cell.fill = fill
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(text(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 90)
    wb.save(path)


def main() -> None:
    work = pd.read_csv(WORK_ORDER, dtype=str).fillna("")
    candidates = pd.read_csv(CANDIDATES, dtype=str).fillna("")
    review = work[work["작업구분"].eq("06_추가검토")].copy()

    best_rows: list[dict[str, Any]] = []
    for _, row in review.iterrows():
        seed_idx = text(row.get("seed_row_index"))
        cand = candidates[candidates["seed_row_index"].map(text).eq(seed_idx)].copy()
        if cand.empty:
            best_rows.append({})
            continue
        cand["_score"] = pd.to_numeric(cand["candidate_rank_score"], errors="coerce").fillna(0)
        cand = cand.sort_values(["_score"], ascending=False, kind="stable")
        best = cand.iloc[0]
        best_rows.append(
            {
                "best_candidate_count": len(cand),
                "best_candidate_CID": best.get("current_콘텐츠ID", ""),
                "best_candidate_name": best.get("current_콘텐츠명", ""),
                "best_candidate_manager": best.get("current_담당자명", ""),
                "best_candidate_score": best.get("candidate_rank_score", ""),
                "best_match_tier": best.get("match_tier", ""),
                "best_match_reason": best.get("match_reason", ""),
                "best_is_disabled(Y/N)": best.get("current_is_disabled(Y/N)", ""),
                "best_is_bundle(Y/N)": best.get("current_is_bundle(Y/N)", ""),
                "best_weak_match(Y/N)": best.get("weak_match(Y/N)", ""),
            }
        )

    enriched = pd.concat([review.reset_index(drop=True), pd.DataFrame(best_rows)], axis=1)
    decisions = enriched.apply(decide, axis=1, result_type="expand")
    enriched["검토판정"] = decisions[0]
    enriched["권장처리"] = decisions[1]

    output_cols = [
        "검토판정",
        "권장처리",
        "적용액션",
        "작품명",
        "대표작가명",
        "선인세코드",
        "특수",
        "account_저작권코드",
        "정산자",
        "정산대표Y/N",
        "현재CID",
        "현재IPS명",
        "변경후IPS명",
        "best_candidate_count",
        "best_candidate_CID",
        "best_candidate_name",
        "best_candidate_manager",
        "best_candidate_score",
        "best_match_tier",
        "best_match_reason",
        "best_is_disabled(Y/N)",
        "best_is_bundle(Y/N)",
        "best_weak_match(Y/N)",
        "위험플래그",
        "seed_row_index",
    ]
    enriched = enriched[[col for col in output_cols if col in enriched.columns]]
    enriched = enriched.sort_values(["검토판정", "대표작가명", "작품명"], kind="stable")

    summary = enriched.groupby(["적용액션", "검토판정"], dropna=False).size().reset_index(name="건수")
    validation = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("입력 작업지시서", str(WORK_ORDER)),
            ("입력 후보", str(CANDIDATES)),
            ("추가검토 행 수", len(enriched)),
            ("가능 판정 수", int(enriched["검토판정"].str.contains("가능", regex=False).sum())),
            ("보류 판정 수", int(enriched["검토판정"].str.contains("보류", regex=False).sum())),
        ],
        columns=["항목", "값"],
    )

    enriched.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        validation.to_excel(writer, sheet_name="검증", index=False)
        summary.to_excel(writer, sheet_name="요약", index=False)
        for name in ["사용안함_승격가능", "매칭_이름수정가능", "매칭_유지가능", "사용안함_보류", "매칭_보류"]:
            subset = enriched[enriched["검토판정"].eq(name)]
            if not subset.empty:
                subset.to_excel(writer, sheet_name=name[:31], index=False)
        enriched.to_excel(writer, sheet_name="전체", index=False)
    autosize(OUTPUT_XLSX)

    print("=== IPS additional review built ===")
    print(f"rows={len(enriched)}")
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
