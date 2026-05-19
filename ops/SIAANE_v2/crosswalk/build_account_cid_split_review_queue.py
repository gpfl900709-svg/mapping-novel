from __future__ import annotations

import hashlib
import json
import sys
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parents[1]
ACCOUNT_DIR = ROOT / "SIAANE_v2" / "account"
if str(ACCOUNT_DIR) not in sys.path:
    sys.path.insert(0, str(ACCOUNT_DIR))

from build_account_observation_bundle import (  # noqa: E402
    MANAGER,
    _GOOD_FILL,
    _WARN_FILL,
    apply_body_style,
    apply_header_style,
    normalize_text,
    pipe_join,
    set_column_widths,
    split_pipe_values,
    write_df_sheet,
)


OUTPUT_ROOT = ROOT / "SIAANE_v2" / "crosswalk" / "exports"
DECISION_PATH = ROOT / "SIAANE_v2" / "account" / "exports" / f"latest__account_decision_queue_{MANAGER}.csv"
SEED_PATH = ROOT / "SIAANE_v2" / "crosswalk" / "exports" / f"latest__account_ips_cid_seed_{MANAGER}.csv"

OUTPUT_XLSX = OUTPUT_ROOT / f"latest__account_cid_split_review_queue_{MANAGER}.xlsx"
OUTPUT_RIGHTS_CSV = OUTPUT_ROOT / f"latest__account_cid_split_rights_review_{MANAGER}.csv"
OUTPUT_WORK_CSV = OUTPUT_ROOT / f"latest__account_cid_split_work_review_{MANAGER}.csv"
OUTPUT_JSON = OUTPUT_ROOT / f"latest__account_cid_split_review_queue_{MANAGER}.json"

RIGHTS_MANUAL_COLUMNS = ["수동_메모", "처리완료(Y/N)"]
WORK_MANUAL_COLUMNS = ["수동_유지(Y/N)", "수동_대표보정", "수동_CID_병합그룹", "수동_메모", "처리완료(Y/N)"]


def unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        cleaned = normalize_text(value)
        if not cleaned or cleaned in seen:
            continue
        ordered.append(cleaned)
        seen.add(cleaned)
    return ordered


def preserve_manual_columns(df: pd.DataFrame, *, path: Path, key_columns: list[str], manual_columns: list[str]) -> pd.DataFrame:
    if not path.exists():
        return df

    existing_df = pd.read_csv(path, dtype=str).fillna("")
    available_columns = key_columns + [column for column in manual_columns if column in existing_df.columns]
    if not available_columns:
        return df

    existing_df = existing_df[available_columns].drop_duplicates(subset=key_columns, keep="last")
    merged = df.merge(existing_df, on=key_columns, how="left", suffixes=("", "__existing"))

    for column in manual_columns:
        existing_column = f"{column}__existing"
        if existing_column not in merged.columns:
            continue
        merged[column] = merged[existing_column].where(
            merged[existing_column].map(normalize_text).ne(""),
            merged[column],
        )
        merged = merged.drop(columns=[existing_column])

    return merged


def compute_seed_list_signature(raw: str) -> str:
    seed_ids = sorted(unique(split_pipe_values(raw)))
    if not seed_ids:
        return ""
    payload = "||".join(seed_ids).encode("utf-8")
    return hashlib.sha1(payload).hexdigest()


def preserve_rights_manual_columns(df: pd.DataFrame) -> pd.DataFrame:
    if not OUTPUT_RIGHTS_CSV.exists():
        return df

    current_df = df.copy()
    current_df["CID_split_signature"] = current_df["CID_seed_id_목록"].map(compute_seed_list_signature)

    existing_df = pd.read_csv(OUTPUT_RIGHTS_CSV, dtype=str).fillna("")
    if "CID_split_signature" not in existing_df.columns:
        if "CID_seed_id_목록" in existing_df.columns:
            existing_df["CID_split_signature"] = existing_df["CID_seed_id_목록"].map(compute_seed_list_signature)
        else:
            existing_df["CID_split_signature"] = ""

    available_columns = [
        "account_저작권코드",
        "CID_split_signature",
    ] + [column for column in RIGHTS_MANUAL_COLUMNS if column in existing_df.columns]
    existing_df = existing_df[available_columns].drop_duplicates(
        subset=["account_저작권코드", "CID_split_signature"],
        keep="last",
    )

    merged = current_df.merge(
        existing_df,
        on=["account_저작권코드", "CID_split_signature"],
        how="left",
        suffixes=("", "__existing"),
    )
    for column in RIGHTS_MANUAL_COLUMNS:
        existing_column = f"{column}__existing"
        if existing_column not in merged.columns:
            continue
        merged[column] = merged[existing_column].where(
            merged[existing_column].map(normalize_text).ne(""),
            merged[column],
        )
        merged = merged.drop(columns=[existing_column])

    return merged.drop(columns=["CID_split_signature"])


def build_rights_review_question(raw_count: str, seed_count: str) -> str:
    raw_count = normalize_text(raw_count) or "0"
    seed_count = normalize_text(seed_count) or "0"
    if raw_count != seed_count:
        return f"raw {raw_count}개 -> seed {seed_count}개 정규화 분해를 그대로 유지할까?"
    return f"{seed_count}개 작품 seed를 그대로 유지할까?"


def build_rights_review_df(decision_df: pd.DataFrame, seed_df: pd.DataFrame) -> pd.DataFrame:
    cid_rights_df = decision_df[
        decision_df["action_제안"].eq("CID분해필요")
        & decision_df["수동_action"].map(normalize_text).ne("담당제외")
    ].copy()
    if cid_rights_df.empty:
        return cid_rights_df

    cid_seed_df = seed_df[seed_df["decision_action_제안"].eq("CID분해필요")].copy()
    aggregated_seed = (
        cid_seed_df.groupby("account_저작권코드", sort=False)
        .agg(
            CID_seed_행수=("CID_seed_id", "count"),
            정산자_목록=("정산자", lambda s: pipe_join(s.tolist())),
            연결_선인세코드_목록=("연결_선인세코드", lambda s: pipe_join(s.tolist())),
            연결_선인세명_목록=("연결_선인세명", lambda s: pipe_join(s.tolist())),
            대표Y_작품=("작품명", lambda s: pipe_join(cid_seed_df.loc[s.index][cid_seed_df.loc[s.index, "정산대표Y/N"].eq("Y")]["작품명"].tolist())),
            CID_seed_id_목록=("CID_seed_id", lambda s: pipe_join(s.tolist())),
        )
        .reset_index()
    )

    review_df = cid_rights_df.merge(aggregated_seed, on="account_저작권코드", how="left").fillna("")
    review_df["검토질문"] = review_df.apply(
        lambda row: build_rights_review_question(
            row.get("관측_작품수", ""),
            row.get("CID_seed_행수", ""),
        ),
        axis=1,
    )
    review_df["수동_메모"] = ""
    review_df["처리완료(Y/N)"] = ""
    review_df["검토우선순위"] = review_df["특수"].map(lambda value: "특수권리" if normalize_text(value) in {"카카오MG", "네이버MG", "원작"} else "일반")
    review_df = review_df.sort_values(
        by=["검토우선순위", "대표작가명", "대표작품", "account_저작권코드"],
        ascending=[True, True, True, True],
        kind="stable",
    ).reset_index(drop=True)

    columns = [
        "검토우선순위",
        "대표작가명",
        "account_저작권코드",
        "특수",
        "대표작품",
        "관측_작품수",
        "관측_작품목록",
        "정산자_목록",
        "연결_선인세코드_목록",
        "연결_선인세명_목록",
        "대표Y_작품",
        "검토질문",
        "판정근거",
        "CID_seed_행수",
        "CID_seed_id_목록",
        "수동_메모",
        "처리완료(Y/N)",
    ]
    review_df = review_df[columns].copy()
    return preserve_rights_manual_columns(review_df)


def build_work_review_df(seed_df: pd.DataFrame) -> pd.DataFrame:
    work_df = seed_df[seed_df["decision_action_제안"].eq("CID분해필요")].copy()
    if work_df.empty:
        return work_df

    work_df["수동_유지(Y/N)"] = ""
    work_df["수동_대표보정"] = ""
    work_df["수동_CID_병합그룹"] = ""
    work_df["수동_메모"] = ""
    work_df["처리완료(Y/N)"] = ""
    work_df = work_df.sort_values(
        by=["대표작가명", "account_저작권코드", "정산대표Y/N", "작품명"],
        ascending=[True, True, False, True],
        kind="stable",
    ).reset_index(drop=True)

    columns = [
        "대표작가명",
        "작품명",
        "account_저작권코드",
        "특수",
        "정산자",
        "연결_선인세코드",
        "연결_선인세명",
        "정산대표Y/N",
        "CID_seed_id",
        "주의플래그",
        "수동_유지(Y/N)",
        "수동_대표보정",
        "수동_CID_병합그룹",
        "수동_메모",
        "처리완료(Y/N)",
    ]
    work_df = work_df[columns].copy()
    return preserve_manual_columns(
        work_df,
        path=OUTPUT_WORK_CSV,
        key_columns=["CID_seed_id"],
        manual_columns=WORK_MANUAL_COLUMNS,
    )


def write_summary_sheet(wb: Workbook, *, rights_df: pd.DataFrame, work_df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{MANAGER} account CID split review"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "지표"
    ws["B3"] = "값"
    apply_header_style(ws["A3"])
    apply_header_style(ws["B3"])

    metrics = [
        ("CID분해필요 권리행 수", len(rights_df)),
        ("CID분해필요 seed 행 수", len(work_df)),
        ("특수권리 권리행 수", int(rights_df["검토우선순위"].eq("특수권리").sum()) if not rights_df.empty else 0),
        ("대표Y seed 행 수", int(work_df["정산대표Y/N"].eq("Y").sum()) if not work_df.empty else 0),
        ("선인세미연결 seed 행 수", int(work_df["주의플래그"].astype(str).str.contains("선인세미연결", regex=False).sum()) if not work_df.empty else 0),
    ]

    for row_idx, (label, value) in enumerate(metrics, start=4):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2))

    start_row = 12
    ws.cell(row=start_row, column=1, value="검토우선순위")
    ws.cell(row=start_row, column=2, value="행 수")
    apply_header_style(ws.cell(row=start_row, column=1))
    apply_header_style(ws.cell(row=start_row, column=2))

    priority_counter = Counter(rights_df["검토우선순위"].tolist()) if not rights_df.empty else Counter()
    labels = ["특수권리", "일반"]
    for idx, label in enumerate(labels, start=start_row + 1):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=priority_counter.get(label, 0))
        apply_body_style(ws.cell(row=idx, column=1))
        apply_body_style(ws.cell(row=idx, column=2))

    chart = BarChart()
    chart.title = "CID split 검토 우선순위"
    chart.y_axis.title = "행 수"
    chart.x_axis.title = "우선순위"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(labels))
    categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(labels))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "D4")

    ws["A18"] = "작업 기준"
    ws["A19"] = "rights_review: 이 권리행을 현재 seed 분해대로 가져갈지 메모"
    ws["A20"] = "work_seed_review: 작품별 seed 유지/대표Y/N/병합그룹만 표시"
    ws["A21"] = "애매하면 수동_메모만 남기고 다음 턴에 정리"
    apply_header_style(ws["A18"])
    apply_body_style(ws["A19"], fill=_GOOD_FILL, wrap=True)
    apply_body_style(ws["A20"], fill=_WARN_FILL, wrap=True)
    apply_body_style(ws["A21"], wrap=True)

    set_column_widths(ws)


def write_review_sheet(wb: Workbook, *, sheet_name: str, df: pd.DataFrame, kind: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    rows = [list(df.columns)] + df.astype(object).where(pd.notnull(df), "").values.tolist()

    for row_idx, row_values in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                apply_header_style(cell)
                continue

            fill = None
            if kind == "rights":
                priority = normalize_text(ws.cell(row=row_idx, column=1).value)
                if priority == "특수권리":
                    fill = _WARN_FILL
            else:
                flags = normalize_text(ws.cell(row=row_idx, column=10).value)
                if "선인세미연결" in flags or "복수작품권리행" in flags:
                    fill = _WARN_FILL
                elif normalize_text(ws.cell(row=row_idx, column=4).value) in {"카카오MG", "네이버MG", "원작"}:
                    fill = _GOOD_FILL

            apply_body_style(cell, fill=fill, wrap=True)

    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24
    set_column_widths(ws)


def write_outputs(*, rights_df: pd.DataFrame, work_df: pd.DataFrame) -> dict[str, Path]:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    rights_df.to_csv(OUTPUT_RIGHTS_CSV, index=False, encoding="utf-8-sig")
    work_df.to_csv(OUTPUT_WORK_CSV, index=False, encoding="utf-8-sig")

    wb = Workbook()
    write_summary_sheet(wb, rights_df=rights_df, work_df=work_df)
    write_review_sheet(wb, sheet_name="cid_rights_review", df=rights_df, kind="rights")
    write_review_sheet(wb, sheet_name="cid_work_seed_review", df=work_df, kind="work")
    write_df_sheet(wb, "raw_cid_seed", work_df)

    try:
        wb.save(OUTPUT_XLSX)
    except PermissionError as exc:
        raise RuntimeError(
            f"최신판 1개 정책을 유지하려면 '{OUTPUT_XLSX.name}' 파일을 닫은 뒤 다시 실행해야 합니다."
        ) from exc

    summary = {
        "manager": MANAGER,
        "cid_rights_rows": int(len(rights_df)),
        "cid_seed_rows": int(len(work_df)),
        "priority_special_rows": int(rights_df["검토우선순위"].eq("특수권리").sum()) if not rights_df.empty else 0,
        "workbook": str(OUTPUT_XLSX),
    }
    OUTPUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "rights_csv": OUTPUT_RIGHTS_CSV,
        "work_csv": OUTPUT_WORK_CSV,
        "workbook": OUTPUT_XLSX,
        "summary": OUTPUT_JSON,
    }


def main() -> None:
    decision_df = pd.read_csv(DECISION_PATH, dtype=str).fillna("")
    seed_df = pd.read_csv(SEED_PATH, dtype=str).fillna("")

    rights_df = build_rights_review_df(decision_df, seed_df)
    work_df = build_work_review_df(seed_df)
    outputs = write_outputs(rights_df=rights_df, work_df=work_df)

    print("=== account CID split review queue built ===")
    print(
        json.dumps(
            {
                "manager": MANAGER,
                "cid_rights_rows": len(rights_df),
                "cid_seed_rows": len(work_df),
                "workbook": str(outputs["workbook"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
