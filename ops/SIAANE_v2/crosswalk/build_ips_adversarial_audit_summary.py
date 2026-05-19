from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from pandas.errors import EmptyDataError
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"
CREATE = EXPORT_ROOT / f"latest__ips_create_request_{MANAGER}.csv"
BUNDLE = EXPORT_ROOT / f"latest__ips_bundle_split_request_{MANAGER}.csv"
EXISTING_AUDIT = EXPORT_ROOT / f"latest__ips_existing_candidate_audit_{MANAGER}.csv"
ADDITIONAL = EXPORT_ROOT / f"latest__ips_additional_review_{MANAGER}.csv"
QUEUE = EXPORT_ROOT / f"latest__account_ips_action_queue_{MANAGER}.csv"
OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_adversarial_audit_summary_{MANAGER}.xlsx"


def text(value: Any) -> str:
    return str(value or "").strip()


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            first = text(row[0].value)
            fill = ok_fill if "통과" in first or "0" == text(row[-1].value) else warn_fill
            if ws.title != "요약":
                for cell in row:
                    cell.fill = fill
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(text(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 90)
    wb.save(path)


def main() -> None:
    create = pd.read_csv(CREATE, dtype=str).fillna("")
    bundle = pd.read_csv(BUNDLE, dtype=str).fillna("")
    try:
        existing = pd.read_csv(EXISTING_AUDIT, dtype=str).fillna("")
    except EmptyDataError:
        existing = pd.DataFrame(columns=["판정"])
    additional = pd.read_csv(ADDITIONAL, dtype=str).fillna("")
    queue = pd.read_csv(QUEUE, dtype=str).fillna("")

    allowed = {"일반", "카카오MG", "네이버MG", "원작"}
    create_bad_format = int((create["신규_콘텐츠명"].map(lambda v: len(text(v).split("_"))) != 7).sum())
    bundle_bad_format = int((bundle["신규_콘텐츠명"].map(lambda v: len(text(v).split("_"))) != 7).sum())
    create_bundle_overlap = len(set(create["신규_콘텐츠명"]) & set(bundle["신규_콘텐츠명"]))

    checks = pd.DataFrame(
        [
            ("1회차_요청서중복", "신규생성명/분해명 중복 및 상호 중복 검사", int(create.duplicated(["신규_콘텐츠명"], keep=False).sum()) + int(bundle.duplicated(["신규_콘텐츠명"], keep=False).sum()) + create_bundle_overlap),
            ("1회차_포맷", "7파트 포맷 및 특수값 검사", create_bad_format + bundle_bad_format + int((~create["특수"].isin(allowed)).sum()) + int((~bundle["특수"].isin(allowed)).sum())),
            ("2회차_기존후보", "신규생성 33건에서 기존 CID 후보 재탐색", int(existing["판정"].str.startswith("자동후보").sum()) if not existing.empty else 0),
            ("3회차_묶음분해", "묶음분해 원 CID 단일성 및 대상명 중복 검사", 0 if int(bundle["원묶음CID"].nunique()) == 1 and not bundle.duplicated(["신규_콘텐츠명"], keep=False).any() else 1),
            ("3회차_추가검토", "추가검토 자동 가능 잔여 검사", int(additional["검토판정"].str.contains("가능", regex=False).sum()) if not additional.empty else 0),
        ],
        columns=["감리항목", "검사내용", "문제건수"],
    )
    checks["결론"] = checks["문제건수"].map(lambda n: "통과" if int(n) == 0 else "확인필요")

    summary = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("최신 IPS_유지", int(queue["IPS_action_제안"].eq("IPS_유지").sum())),
            ("최신 IPS_신규생성", int(queue["IPS_action_제안"].eq("IPS_신규생성").sum())),
            ("신규생성 요청 건수", len(create)),
            ("묶음분해 요청 건수", len(bundle)),
            ("추가검토 보류 건수", len(additional)),
            ("최종 받아올 CID 수", len(create) + len(bundle)),
        ],
        columns=["항목", "값"],
    )

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="요약", index=False)
        checks.to_excel(writer, sheet_name="감리결과", index=False)
        existing.to_excel(writer, sheet_name="기존후보_보류", index=False)
        additional.to_excel(writer, sheet_name="추가검토_보류", index=False)
    autosize(OUTPUT_XLSX)
    print("=== IPS adversarial audit summary built ===")
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
