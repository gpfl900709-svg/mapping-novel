from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"

ACTION_QUEUE = EXPORT_ROOT / f"latest__account_ips_action_queue_{MANAGER}.csv"
PROMOTION_REVIEW = EXPORT_ROOT / f"latest__ips_disabled_promotion_review_{MANAGER}.csv"
OWNER_PLAN = EXPORT_ROOT / f"latest__ips_owner_decision_plan_{MANAGER}.csv"
IPS_INVENTORY = PROJECT_ROOT / "ips" / "stage" / f"latest__current_ips_inventory_{MANAGER}.csv"
ABSORBED_SEED_INDEXES = {
    "244",  # 핵무기도 만들어 드릴까요 외전: 본편 카카오MG CID에 흡수
    "245",  # 핵무기도 만들어 드림 804/N: 드림 일반 대표 CID(1005653/Y)로 흡수
    "229",  # 타락의 검은실 2부: 기존 타락 CID 중 하나로 흡수
    "4",  # 4인 청춘 레포트: 담당 작품 아님
    "37",  # 나선미로: 담당 작품 아님
    "59",  # 드래곤 스튜던트: 담당 작품 아님
    "63",  # 라이오스의 불량기사: 담당 작품 아님
    "107",  # 사립루레인학원 윤리선생: 담당 작품 아님
    "108",  # 사립루레인학원 윤리선생(개정판): 담당 작품 아님
    "213",  # 짐승들의 만찬: 담당 작품 아님
    "230",  # 투 브라더스: 담당 작품 아님
    "231",  # 투 시스터즈: 담당 작품 아님
    "234",  # 페르기온의 황제: 담당 작품 아님
    "252",  # 회귀로 나 혼자 독식 네이버MG: 기존 107850에 흡수
    "45",  # 노출광 여대생: 런칭 전이라 생성 제외
    "76",  # 마왕으로 산다 카카오MG: 카카오MG 미수령이라 생성 제외
    "92",  # 무림누나 능욕게임: 2026-04-27 KIPM 라이브 신규 CID 327814 생성 완료
    "162",  # 여자들이 집착하는 축구 선수가 되었다: 사용자 확인. 쓰고또쓰고 CID 320890이 맞아 IPS 변경 없음
    "50",  # 달인: 2026-04-27 더미 계약서로 KIPM 라이브 신규 CID 327817 생성 완료
}

MANUAL_ACTION_QUEUE_DECISIONS = {
    "64": {
        "work_type": "01_이름수정",
        "action": "이름수정",
        "target_cid": "319394",
        "target_current_name": "레벨원",
        "new_name": "레벨원_박천웅_선인세없음_일반_1004809_박천웅_N",
        "memo": "사용자 확인: 레벨원은 박천웅이 맞고 하늘곰과 동일인물. 기존 CID 319394를 목표 IPS명으로 rename.",
    },
}

OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_apply_work_order_{MANAGER}.xlsx"
OUTPUT_CSV = EXPORT_ROOT / f"latest__ips_apply_work_order_{MANAGER}.csv"

AD_REVENUE_TOKENS = ("네이버광고수익", "카카오광고수익")


def text(value: Any) -> str:
    return str(value or "").strip()


def normalize_target_name(value: Any) -> str:
    return text(value).replace("_미연결_", "_선인세없음_")


def common_record(row: pd.Series, work_type: str, action: str, source: str) -> dict[str, Any]:
    return {
        "작업구분": work_type,
        "적용액션": action,
        "출처": source,
        "작업상태": "대기",
        "확인메모": "",
        "작품명": row.get("작품명", ""),
        "대표작가명": row.get("대표작가명", ""),
        "선인세코드": row.get("연결_선인세코드", ""),
        "특수": row.get("특수", ""),
        "account_저작권코드": row.get("account_저작권코드", ""),
        "정산자": row.get("정산자", ""),
        "정산대표Y/N": row.get("정산대표Y/N", ""),
        "목표IPS명": normalize_target_name(row.get("desired_ips_name", "") or row.get("신규생성명", "")),
        "현재CID": row.get("current_콘텐츠ID", "") or row.get("현재CID", ""),
        "현재IPS명": row.get("current_콘텐츠명", "") or row.get("현재IPS명", ""),
        "대상CID": "",
        "대상현재명": "",
        "변경후IPS명": "",
        "처리메모": "",
        "위험플래그": row.get("주의플래그", "") or row.get("위험플래그", ""),
        "seed_row_index": row.get("seed_row_index", ""),
    }


def build_from_action_queue(queue: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for _, row in queue.iterrows():
        manual = MANUAL_ACTION_QUEUE_DECISIONS.get(text(row.get("seed_row_index")))
        if manual is not None:
            rec = common_record(row, manual["work_type"], manual["action"], "manual_action_queue_decision")
            rec["대상CID"] = manual["target_cid"]
            rec["대상현재명"] = manual["target_current_name"]
            rec["변경후IPS명"] = manual["new_name"]
            rec["처리메모"] = manual["memo"]
            rows.append(rec)
            continue

        action = text(row.get("IPS_action_제안"))
        if action == "IPS_이름수정":
            rec = common_record(row, "01_이름수정", "이름수정", "action_queue")
            rec["대상CID"] = row.get("current_콘텐츠ID", "")
            rec["대상현재명"] = row.get("current_콘텐츠명", "")
            rec["변경후IPS명"] = normalize_target_name(row.get("desired_ips_name", ""))
            rec["처리메모"] = "현재 CID는 유지하고 IPS명만 목표 포맷으로 수정."
            rows.append(rec)
        elif action == "IPS_신규생성":
            rec = common_record(row, "03_신규생성", "신규생성", "action_queue")
            rec["변경후IPS명"] = normalize_target_name(row.get("desired_ips_name", ""))
            rec["처리메모"] = "매칭 CID 없음. 목표 IPS명으로 신규 생성."
            rows.append(rec)
        elif action in {"IPS_매칭검토", "IPS_사용안함검토"}:
            rec = common_record(row, "06_추가검토", action.replace("IPS_", ""), "action_queue")
            rec["대상CID"] = row.get("current_콘텐츠ID", "")
            rec["대상현재명"] = row.get("current_콘텐츠명", "")
            rec["변경후IPS명"] = normalize_target_name(row.get("desired_ips_name", ""))
            rec["처리메모"] = "자동 적용 전 후보/사용안함 사유 확인 필요."
            rows.append(rec)
    return rows


def build_from_promotion(review: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    source = review[
        review["세부분류"].map(text).isin({"일반_사용안함CID승격", "일반_사용안함CID승격후보"})
    ].copy()
    if source.empty:
        return rows

    source["_has_advance"] = source["연결_선인세코드"].map(lambda value: 1 if text(value) else 0)
    source["_is_primary"] = source["정산대표Y/N"].map(lambda value: 1 if text(value).upper() == "Y" else 0)
    source["_is_confirm"] = source["세부분류"].map(lambda value: 0 if text(value).endswith("후보") else 1)
    source = source.sort_values(
        ["사용안함_승격후보CID", "_is_confirm", "_is_primary", "_has_advance", "account_저작권코드"],
        ascending=[True, False, False, False, True],
        kind="stable",
    )
    source["_promotion_rank"] = source.groupby("사용안함_승격후보CID").cumcount()

    for _, row in source.iterrows():
        detail = text(row.get("세부분류"))
        rank = int(row.get("_promotion_rank", 0))
        if rank == 0:
            rec = common_record(row, "02_사용안함승격", "사용안함승격", "disabled_promotion_review")
            rec["대상CID"] = row.get("사용안함_승격후보CID", "")
            rec["대상현재명"] = row.get("사용안함_승격후보명", "")
            rec["변경후IPS명"] = normalize_target_name(row.get("desired_ips_name", ""))
            rec["처리메모"] = "사용안함 CID를 일반 CID로 승격하고 목표 IPS명으로 수정."
        else:
            rec = common_record(row, "03_신규생성", "신규생성", "disabled_promotion_review_duplicate")
            rec["변경후IPS명"] = normalize_target_name(row.get("desired_ips_name", ""))
            rec["처리메모"] = (
                f"동일 사용안함 CID({row.get('사용안함_승격후보CID', '')})는 다른 seed에 승격 배정됨. "
                "한 CID를 여러 이름으로 변경할 수 없으므로 이 seed는 신규 생성."
            )

        if detail.endswith("후보"):
            rec["작업상태"] = "확인필요"
            rec["처리메모"] += " 보류 해소 후 적용."
        rows.append(rec)
    return rows


def build_from_owner_plan(plan: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for _, row in plan.iterrows():
        decision = text(row.get("내판단"))
        apply_action = text(row.get("적용액션"))
        if apply_action == "유지":
            rec = common_record(row, "05_유지확인", "유지", "owner_decision_plan")
            rec["대상CID"] = row.get("유지CID", "") or row.get("현재CID", "")
            rec["대상현재명"] = row.get("현재IPS명", "")
            rec["변경후IPS명"] = normalize_target_name(row.get("목표IPS명", ""))
            rec["처리메모"] = row.get("처리메모", "")
            rows.append(rec)
        elif apply_action in {"신규/분해", "신규생성/후보재탐색"}:
            rec = common_record(row, "04_묶음분해", apply_action, "owner_decision_plan")
            rec["대상CID"] = row.get("현재CID", "")
            rec["대상현재명"] = row.get("현재IPS명", "")
            rec["변경후IPS명"] = normalize_target_name(row.get("신규생성명", ""))
            rec["처리메모"] = row.get("처리메모", "") or decision
            if apply_action == "신규생성/후보재탐색":
                rec["작업상태"] = "확인필요"
            rows.append(rec)
        elif apply_action == "이름수정":
            rec = common_record(row, "01_이름수정", "이름수정", "owner_decision_plan")
            rec["대상CID"] = row.get("유지CID", "") or row.get("현재CID", "")
            rec["대상현재명"] = row.get("대상현재명", "") or row.get("현재IPS명", "")
            rec["변경후IPS명"] = normalize_target_name(row.get("신규생성명", "") or row.get("목표IPS명", ""))
            rec["처리메모"] = row.get("처리메모", "")
            rows.append(rec)
        elif apply_action == "신규생성":
            rec = common_record(row, "03_신규생성", "신규생성", "owner_decision_plan")
            rec["대상CID"] = ""
            rec["대상현재명"] = row.get("현재IPS명", "")
            rec["변경후IPS명"] = normalize_target_name(row.get("신규생성명", ""))
            rec["처리메모"] = row.get("처리메모", "") or decision
            rows.append(rec)
    return rows


def build_from_ad_revenue_inventory(inventory: pd.DataFrame, protected_cids: set[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if inventory.empty:
        return rows

    for _, row in inventory.iterrows():
        content_id = text(row.get("콘텐츠ID"))
        content_name = text(row.get("콘텐츠명"))
        if not content_id or content_id in protected_cids:
            continue
        if not any(token in content_name for token in AD_REVENUE_TOKENS):
            continue

        rows.append(
            {
                "작업구분": "07_광고수익폐기",
                "적용액션": "사용안함처리",
                "출처": "ips_inventory_ad_revenue_sweep",
                "작업상태": "대기",
                "확인메모": "",
                "작품명": "",
                "대표작가명": "",
                "선인세코드": "",
                "특수": "광고수익",
                "account_저작권코드": "",
                "정산자": "",
                "정산대표Y/N": "",
                "목표IPS명": "",
                "현재CID": content_id,
                "현재IPS명": content_name,
                "대상CID": content_id,
                "대상현재명": content_name,
                "변경후IPS명": f"(사용안함)_{content_name}",
                "처리메모": "광고수익은 별도 canonical CID로 유지하지 않음. 네이버MG/카카오MG 수익 배분은 IPS 내 플랫폼 비율에서 처리하므로 이 CID는 사용안함 처리.",
                "위험플래그": "광고수익CID_폐기",
                "seed_row_index": "",
            }
        )
    return rows


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    status_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            if ws.title != "요약" and text(row[3].value) == "확인필요":
                for cell in row:
                    cell.fill = status_fill
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(text(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 70)
    wb.save(path)


def main() -> None:
    queue = pd.read_csv(ACTION_QUEUE, dtype=str).fillna("")
    review = pd.read_csv(PROMOTION_REVIEW, dtype=str).fillna("")
    plan = pd.read_csv(OWNER_PLAN, dtype=str).fillna("")
    inventory = pd.read_csv(IPS_INVENTORY, dtype=str).fillna("") if IPS_INVENTORY.exists() else pd.DataFrame()

    rows = []
    rows.extend(build_from_action_queue(queue))
    rows.extend(build_from_promotion(review))
    rows.extend(build_from_owner_plan(plan))
    protected_cids = {
        text(row.get("대상CID"))
        for row in rows
        if text(row.get("적용액션")) == "이름수정" and text(row.get("대상CID"))
    }
    rows.extend(build_from_ad_revenue_inventory(inventory, protected_cids))
    work = pd.DataFrame(rows)
    resolved_seed_indexes = set(
        queue[queue["IPS_action_제안"].eq("IPS_유지")]["seed_row_index"].map(text)
    )
    if not work.empty and "seed_row_index" in work.columns:
        work = work[~work["seed_row_index"].map(text).isin(resolved_seed_indexes)].copy()
        work = work[~work["seed_row_index"].map(text).isin(ABSORBED_SEED_INDEXES)].copy()
    work = work.sort_values(
        ["작업구분", "작업상태", "대표작가명", "작품명", "특수", "변경후IPS명"],
        kind="stable",
    )

    summary = (
        work.groupby(["작업구분", "적용액션", "작업상태"], dropna=False)
        .size()
        .reset_index(name="건수")
        .sort_values(["작업구분", "적용액션", "작업상태"], kind="stable")
    )
    meta = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("action_queue", str(ACTION_QUEUE)),
            ("promotion_review", str(PROMOTION_REVIEW)),
            ("owner_decision_plan", str(OWNER_PLAN)),
            ("총 작업행", len(work)),
            ("작업구분별", dict(Counter(work["작업구분"]))),
        ],
        columns=["항목", "값"],
    )

    work.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        meta.to_excel(writer, sheet_name="요약", index=False, startrow=0)
        summary.to_excel(writer, sheet_name="요약", index=False, startrow=len(meta) + 2)
        sheet_order = [
            "01_이름수정",
            "02_사용안함승격",
            "03_신규생성",
            "04_묶음분해",
            "05_유지확인",
            "06_추가검토",
            "07_광고수익폐기",
        ]
        for sheet in sheet_order:
            subset = work[work["작업구분"].eq(sheet)]
            if not subset.empty:
                subset.to_excel(writer, sheet_name=sheet, index=False)
        work.to_excel(writer, sheet_name="전체", index=False)
    autosize(OUTPUT_XLSX)

    print("=== IPS apply work order built ===")
    print(f"rows={len(work)}")
    print(dict(Counter(work["작업구분"])))
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
