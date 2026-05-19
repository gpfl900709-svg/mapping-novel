from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = PROJECT_ROOT.parent

ACCOUNT_CID_SEED_PATH = PROJECT_ROOT / "crosswalk" / "exports" / f"latest__account_ips_cid_seed_{MANAGER}.csv"
ACCOUNT_CANONICAL_PATH = PROJECT_ROOT / "account" / "exports" / f"latest__account_rights_canonical_{MANAGER}.csv"
AUTHOR_GROUPS_PATH = PROJECT_ROOT / "manager_author_manual_groups.json"
IPS_WORKBOOK_PATHS = [
    PROJECT_ROOT / "update_ips.xlsx",
]

TITLE_OVERRIDES = {
    "하이퍼 서버 벤젠스": "하이퍼 서퍼 벤젠스",
    "골 떄리는 비제이들": "골 때리는 여자 비제이들",
}

OUTPUT_XLSX = PROJECT_ROOT / "crosswalk" / "exports" / f"latest__account_ips_action_queue_{MANAGER}.xlsx"
OUTPUT_CSV = PROJECT_ROOT / "crosswalk" / "exports" / f"latest__account_ips_action_queue_{MANAGER}.csv"
OUTPUT_JSON = PROJECT_ROOT / "crosswalk" / "exports" / f"latest__account_ips_action_queue_{MANAGER}.json"
DETAIL_CSV = PROJECT_ROOT / "crosswalk" / "exports" / f"latest__account_ips_match_candidates_{MANAGER}.csv"
INVENTORY_CSV = PROJECT_ROOT / "ips" / "stage" / f"latest__current_ips_inventory_{MANAGER}.csv"

SHEET_NAME = "콘텐츠 목록"
ACTION_ORDER = {
    "IPS_분해/중복검토": 1,
    "IPS_사용안함검토": 2,
    "IPS_매칭검토": 3,
    "IPS_신규생성": 4,
    "IPS_이름수정": 5,
    "IPS_유지": 6,
}


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[^0-9a-z가-힣]", "", text)


def join_pipe(values: list[Any]) -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = str(value or "").strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return " | ".join(result)


def load_author_alias_map(path: Path) -> dict[str, set[str]]:
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    alias_map: dict[str, set[str]] = {}
    for group in payload.get("groups", []) or []:
        keys = {normalize_text(name) for name in group if normalize_text(name)}
        for key in keys:
            alias_map[key] = keys
    return alias_map


def aliases_for(name: str, alias_map: dict[str, set[str]]) -> set[str]:
    key = normalize_text(name)
    if not key:
        return set()
    return alias_map.get(key, {key})


def strip_known_prefixes(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    text = re.sub(r"^\s*\(사용안함\)[_\s-]*", "", text)
    text = re.sub(r"^\s*사용안함[_\s-]*", "", text)
    text = re.sub(r"^\s*0_", "", text)
    for _ in range(4):
        text = re.sub(
            r"^\s*(\[[^\]]+\]|\([^)]*연재[^)]*\)|\([^)]*단행본[^)]*\)|단행[_\s-]*|"
            r"카카오창작지원금_|창작지원금|카카오_?|네이버_?|리디_?|봄툰_?|윌라_?|시리즈_?)\s*",
            "",
            text,
        )
    return re.sub(r"^[_\s-]+", "", text).strip()


def strip_known_suffixes(value: str) -> str:
    text = strip_known_prefixes(value)
    changed = True
    while changed:
        before = text
        text = re.sub(r"\s*카카오\s*제외\s*$", "", text)
        text = re.sub(
            r"\s*\([^)]*(세트|외전증보판|개정판|단행본|연재|카카오\s*선투자|보장인세)[^)]*\)\s*$",
            "",
            text,
        )
        text = re.sub(r"\s*(\d+\s*권|세트|외전|완결|단행본|연재|개정판)\s*$", "", text)
        changed = text != before
    return text.strip()


def title_key(value: str) -> str:
    return normalize_text(strip_known_suffixes(value))


def split_author_keys(value: str, alias_map: dict[str, set[str]]) -> set[str]:
    keys: set[str] = set()
    for part in re.split(r"[,/|·ㆍ]", str(value or "")):
        key = normalize_text(part)
        if key:
            keys |= alias_map.get(key, {key})
    return keys


def special_matches(seed_special: str, current_text: str) -> bool:
    seed = normalize_text(seed_special)
    current = normalize_text(current_text)
    if not seed:
        return False
    if seed == normalize_text("일반"):
        negative_tokens = ("카카오MG", "카카오선투자", "네이버MG", "원작", "웹툰")
        return normalize_text("일반") in current or not any(normalize_text(token) in current for token in negative_tokens)
    if seed == normalize_text("카카오MG"):
        return any(
            normalize_text(token) in current
            for token in ("카카오MG", "카카오선투자", "카카오", "선투자", "콘텐츠특별공급")
        )
    if seed == normalize_text("네이버MG"):
        return any(normalize_text(token) in current for token in ("네이버MG", "네이버"))
    if seed == normalize_text("원작"):
        return any(normalize_text(token) in current for token in ("원작", "웹툰"))
    return seed in current


def parse_ips_name(name: str, alias_map: dict[str, set[str]]) -> dict[str, Any]:
    raw = unicodedata.normalize("NFKC", str(name or "")).strip()
    disabled = bool(re.search(r"사용안함|사용금지|사용x|사용X", raw))
    working = strip_known_prefixes(raw)
    parts = [part.strip() for part in working.split("_") if part.strip()]

    title_parts: list[str]
    authors: list[str] = []
    remarks: list[str] = []
    if len(parts) >= 4 and re.fullmatch(r"100\d{4}", parts[-2]):
        remarks = [parts[-1]]
        authors = [parts[-3]]
        title_parts = parts[:-3]
    elif len(parts) >= 3:
        remarks = [parts[-1]]
        authors = [parts[-2]]
        title_parts = parts[:-2]
    else:
        title_parts = [working]

    if len(title_parts) == 1 and ("·" in title_parts[0] or "ㆍ" in title_parts[0]):
        dot_parts = re.split(r"[·ㆍ]", title_parts[0])
        if len(dot_parts) >= 2:
            title_parts = ["·".join(dot_parts[:-1]).strip()]
            authors.append(dot_parts[-1].strip())

    titles: list[str] = []
    for title in title_parts:
        cleaned = strip_known_suffixes(title)
        if cleaned:
            titles.append(cleaned)
    if len(title_parts) > 1:
        combined = strip_known_suffixes("_".join(title_parts))
        if combined:
            titles.append(combined)
    if not titles:
        fallback = strip_known_suffixes(working)
        if fallback:
            titles.append(fallback)

    seen: set[str] = set()
    clean_titles: list[str] = []
    for title in titles:
        key = title_key(title)
        if key and key not in seen:
            seen.add(key)
            clean_titles.append(title)

    author_keys: set[str] = set()
    for author in authors:
        author_keys |= split_author_keys(author, alias_map)

    return {
        "parsed_titles": clean_titles,
        "title_keys": {title_key(title) for title in clean_titles if title_key(title)},
        "author_keys": author_keys,
        "account_codes": set(re.findall(r"100\d{4}", raw)),
        "remarks": remarks,
        "is_disabled": disabled,
        "is_bundle": len(title_parts) > 1,
        "raw_norm": normalize_text(raw),
    }


def read_ips_workbooks(paths: list[Path], alias_map: dict[str, set[str]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for path in paths:
        if not path.exists():
            continue
        frame = pd.read_excel(path, sheet_name=SHEET_NAME, dtype=str, engine="openpyxl").fillna("")
        frame["source_file"] = path.name
        frame["source_path"] = str(path)
        frames.append(frame)
    if not frames:
        raise FileNotFoundError("No IPS workbook inputs were found.")

    inventory = pd.concat(frames, ignore_index=True).fillna("")
    parsed_rows = [parse_ips_name(name, alias_map) for name in inventory["콘텐츠명"]]
    inventory["parsed_titles"] = [join_pipe(row["parsed_titles"]) for row in parsed_rows]
    inventory["parsed_title_keys"] = [join_pipe(sorted(row["title_keys"])) for row in parsed_rows]
    inventory["parsed_author_keys"] = [join_pipe(sorted(row["author_keys"])) for row in parsed_rows]
    inventory["parsed_account_codes"] = [join_pipe(sorted(row["account_codes"])) for row in parsed_rows]
    inventory["parsed_remarks"] = [join_pipe(row["remarks"]) for row in parsed_rows]
    inventory["is_disabled"] = ["Y" if row["is_disabled"] else "N" for row in parsed_rows]
    inventory["is_bundle"] = ["Y" if row["is_bundle"] else "N" for row in parsed_rows]
    inventory["_parsed"] = parsed_rows
    return inventory


def build_indexes(inventory: pd.DataFrame) -> tuple[dict[str, list[int]], dict[str, list[int]]]:
    exact_index: dict[str, list[int]] = defaultdict(list)
    title_index: dict[str, list[int]] = defaultdict(list)
    for idx, row in inventory.iterrows():
        exact_index[normalize_text(row.get("콘텐츠명", ""))].append(idx)
        for key in row["_parsed"]["title_keys"]:
            title_index[key].append(idx)
    return exact_index, title_index


def score_candidate(
    seed_row: pd.Series,
    inventory_row: pd.Series,
    alias_map: dict[str, set[str]],
) -> tuple[int, str, list[str], bool]:
    parsed = inventory_row["_parsed"]
    current_name = str(inventory_row.get("콘텐츠명", ""))
    seed_author_aliases = aliases_for(str(seed_row.get("대표작가명", "")), alias_map)
    account_code = str(seed_row.get("account_저작권코드", "")).strip()

    reasons = ["title"]
    score = 55
    author_ok = bool(
        seed_author_aliases
        and (
            seed_author_aliases & parsed["author_keys"]
            or any(alias in parsed["raw_norm"] for alias in seed_author_aliases)
        )
    )
    if author_ok:
        score += 25
        reasons.append("author_alias")
    if account_code and account_code in parsed["account_codes"]:
        score += 20
        reasons.append("account_code")
    if special_matches(str(seed_row.get("특수", "")), current_name):
        score += 8
        reasons.append("special")
    if str(inventory_row.get("담당자명", "")).strip() == MANAGER:
        score += 5
        reasons.append("manager")
    if parsed["is_disabled"]:
        score -= 12
        reasons.append("disabled")
    if parsed["is_bundle"]:
        score -= 6
        reasons.append("bundle")

    if author_ok or (account_code and account_code in parsed["account_codes"]):
        return score, "title_parsed", reasons, False
    if str(inventory_row.get("담당자명", "")).strip() == MANAGER:
        return min(score, 68), "title_only_manager", reasons + ["weak_title_only"], True
    return score, "title_only_non_manager", reasons + ["weak_title_only"], True


def build_match_candidates(
    seed: pd.DataFrame,
    inventory: pd.DataFrame,
    exact_index: dict[str, list[int]],
    title_index: dict[str, list[int]],
    alias_map: dict[str, set[str]],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for seed_idx, seed_row in seed.iterrows():
        desired_name = str(seed_row.get("CID_seed_id", "")).strip()
        seed_title_key = title_key(str(seed_row.get("작품명", "")))
        candidates: dict[tuple[str, str], dict[str, Any]] = {}

        for inventory_idx in exact_index.get(normalize_text(desired_name), []):
            current = inventory.loc[inventory_idx]
            candidate = {
                "seed_row_index": seed_idx,
                "desired_ips_name": desired_name,
                "candidate_rank_score": 120,
                "match_tier": "exact_desired_name",
                "match_reason": "exact_desired_name",
                "weak_match(Y/N)": "N",
                "source_file": current.get("source_file", ""),
                "current_콘텐츠ID": current.get("콘텐츠ID", ""),
                "current_콘텐츠명": current.get("콘텐츠명", ""),
                "current_담당부서": current.get("담당부서", ""),
                "current_담당자명": current.get("담당자명", ""),
                "current_is_disabled(Y/N)": current.get("is_disabled", ""),
                "current_is_bundle(Y/N)": current.get("is_bundle", ""),
                "current_parsed_titles": current.get("parsed_titles", ""),
                "current_parsed_author_keys": current.get("parsed_author_keys", ""),
                "current_parsed_account_codes": current.get("parsed_account_codes", ""),
            }
            candidates[(str(current.get("source_file", "")), str(current.get("콘텐츠ID", "")))] = candidate

        for inventory_idx in title_index.get(seed_title_key, []):
            current = inventory.loc[inventory_idx]
            score, tier, reasons, weak = score_candidate(seed_row, current, alias_map)
            if score < 70 and not (weak and score >= 60):
                continue
            candidate = {
                "seed_row_index": seed_idx,
                "desired_ips_name": desired_name,
                "candidate_rank_score": score,
                "match_tier": tier,
                "match_reason": "+".join(reasons),
                "weak_match(Y/N)": "Y" if weak else "N",
                "source_file": current.get("source_file", ""),
                "current_콘텐츠ID": current.get("콘텐츠ID", ""),
                "current_콘텐츠명": current.get("콘텐츠명", ""),
                "current_담당부서": current.get("담당부서", ""),
                "current_담당자명": current.get("담당자명", ""),
                "current_is_disabled(Y/N)": current.get("is_disabled", ""),
                "current_is_bundle(Y/N)": current.get("is_bundle", ""),
                "current_parsed_titles": current.get("parsed_titles", ""),
                "current_parsed_author_keys": current.get("parsed_author_keys", ""),
                "current_parsed_account_codes": current.get("parsed_account_codes", ""),
            }
            key = (str(current.get("source_file", "")), str(current.get("콘텐츠ID", "")))
            existing = candidates.get(key)
            if existing is None or score > int(existing["candidate_rank_score"]):
                candidates[key] = candidate

        rows.extend(candidates.values())

    if not rows:
        return pd.DataFrame()
    detail = pd.DataFrame(rows)
    detail = detail.sort_values(
        ["seed_row_index", "candidate_rank_score", "current_is_disabled(Y/N)"],
        ascending=[True, False, True],
    )
    return detail


def choose_top_matches(seed: pd.DataFrame, candidates: pd.DataFrame) -> pd.DataFrame:
    selected_rows: list[dict[str, Any]] = []
    grouped = candidates.groupby("seed_row_index") if not candidates.empty else {}

    for seed_idx, seed_row in seed.iterrows():
        desired_name = str(seed_row.get("CID_seed_id", "")).strip()
        seed_candidates = grouped.get_group(seed_idx).copy() if not candidates.empty and seed_idx in grouped.groups else pd.DataFrame()
        if seed_candidates.empty:
            selected_rows.append({"seed_row_index": seed_idx, "desired_ips_name": desired_name})
            continue

        enabled = seed_candidates[seed_candidates["current_is_disabled(Y/N)"] != "Y"]
        usable = enabled if not enabled.empty else seed_candidates
        usable = usable.sort_values("candidate_rank_score", ascending=False)
        selected_rows.append(usable.iloc[0].to_dict())

    return pd.DataFrame(selected_rows).fillna("")


def classify_actions(queue: pd.DataFrame, candidates: pd.DataFrame) -> pd.DataFrame:
    if queue.empty:
        return queue

    match_counts = candidates.groupby("seed_row_index")["current_콘텐츠ID"].nunique().to_dict() if not candidates.empty else {}
    cid_seed_counts = (
        queue[queue["current_콘텐츠ID"].astype(str).str.strip() != ""]
        .groupby("current_콘텐츠ID")["desired_ips_name"]
        .nunique()
        .to_dict()
    )
    cid_desired_names = (
        queue[queue["current_콘텐츠ID"].astype(str).str.strip() != ""]
        .groupby("current_콘텐츠ID")["desired_ips_name"]
        .apply(lambda values: join_pipe(list(values)))
        .to_dict()
    )

    actions: list[str] = []
    priorities: list[int] = []
    flags_column: list[str] = []
    current_seed_counts: list[int] = []
    tied_names: list[str] = []

    for _, row in queue.iterrows():
        current_cid = str(row.get("current_콘텐츠ID", "")).strip()
        desired_name = str(row.get("desired_ips_name", "")).strip()
        current_name = str(row.get("current_콘텐츠명", "")).strip()
        match_count = int(match_counts.get(row.get("seed_row_index"), 0))
        shared_count = int(cid_seed_counts.get(current_cid, 0)) if current_cid else 0
        flags: list[str] = []

        if not current_cid:
            action = "IPS_신규생성"
            flags.append("현재IPS매칭없음")
        elif normalize_text(current_name) == normalize_text(desired_name):
            action = "IPS_유지"
        elif str(row.get("current_is_disabled(Y/N)", "")) == "Y":
            action = "IPS_사용안함검토"
            flags.append("현재IPS_사용안함")
        elif str(row.get("weak_match(Y/N)", "")) == "Y":
            action = "IPS_매칭검토"
            flags.append("제목단독매칭")
        elif shared_count > 1 or match_count > 1 or str(row.get("current_is_bundle(Y/N)", "")) == "Y":
            action = "IPS_분해/중복검토"
            if shared_count > 1:
                flags.append("현재CID가여러seed에매칭")
            if match_count > 1:
                flags.append("seed별후보CID복수")
            if str(row.get("current_is_bundle(Y/N)", "")) == "Y":
                flags.append("현재IPS_묶음명")
        else:
            action = "IPS_이름수정"

        if current_cid and str(row.get("current_담당자명", "")).strip() != MANAGER:
            flags.append("현재담당자_조원재아님")
        if "선인세없음" in desired_name or "미연결" in desired_name:
            flags.append("선인세미연결")

        actions.append(action)
        priorities.append(ACTION_ORDER.get(action, 99))
        flags_column.append(join_pipe(flags))
        current_seed_counts.append(shared_count)
        tied_names.append(cid_desired_names.get(current_cid, ""))

    queue["IPS_action_제안"] = actions
    queue["검토우선순위"] = priorities
    queue["주의플래그"] = flags_column
    queue["현재CID_매칭_seed수"] = current_seed_counts
    queue["현재CID_매칭_desired_name목록"] = tied_names
    queue["seed별_후보CID수"] = [
        int(match_counts.get(seed_idx, 0)) for seed_idx in queue["seed_row_index"].tolist()
    ]
    return queue


def build_queue() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    alias_map = load_author_alias_map(AUTHOR_GROUPS_PATH)
    seed = pd.read_csv(ACCOUNT_CID_SEED_PATH, dtype=str, encoding="utf-8-sig").fillna("")
    seed["작품명"] = seed["작품명"].replace(TITLE_OVERRIDES)
    seed["CID_seed_id"] = seed.apply(
        lambda row: "_".join(
            [
                str(row.get("작품명", "")).strip() or "미상작품",
                str(row.get("대표작가명", "")).strip() or "미상작가",
                str(row.get("연결_선인세코드", "")).strip() or "선인세없음",
                str(row.get("특수", "")).strip() or "일반",
                str(row.get("account_저작권코드", "")).strip() or "미상권리",
                str(row.get("정산자", "")).strip() or "미상",
                str(row.get("정산대표Y/N", "")).strip() or "N",
            ]
        ),
        axis=1,
    )
    canonical = pd.read_csv(ACCOUNT_CANONICAL_PATH, dtype=str, encoding="utf-8-sig").fillna("")
    inventory = read_ips_workbooks(IPS_WORKBOOK_PATHS, alias_map)
    exact_index, title_index = build_indexes(inventory)
    candidates = build_match_candidates(seed, inventory, exact_index, title_index, alias_map)
    selected = choose_top_matches(seed, candidates)

    canonical_cols = [
        "account_저작권코드",
        "canonical_status",
        "canonical_사용(Y/N)",
        "account_최종저작권명",
        "현재_저작권명",
    ]
    canonical_small = canonical[[col for col in canonical_cols if col in canonical.columns]].drop_duplicates(
        "account_저작권코드"
    )

    seed_base = seed.reset_index(names="seed_row_index").copy()
    seed_base["desired_ips_name"] = seed_base["CID_seed_id"]
    queue = seed_base.merge(
        selected,
        on=["seed_row_index", "desired_ips_name"],
        how="left",
    )
    queue = queue.merge(canonical_small, on="account_저작권코드", how="left")
    queue = classify_actions(queue.fillna(""), candidates)

    output_cols = [
        "IPS_action_제안",
        "검토우선순위",
        "주의플래그",
        "desired_ips_name",
        "작품명",
        "대표작가명",
        "연결_선인세코드",
        "특수",
        "account_저작권코드",
        "정산자",
        "정산대표Y/N",
        "current_콘텐츠ID",
        "current_콘텐츠명",
        "current_담당부서",
        "current_담당자명",
        "source_file",
        "candidate_rank_score",
        "match_tier",
        "match_reason",
        "weak_match(Y/N)",
        "seed별_후보CID수",
        "현재CID_매칭_seed수",
        "현재CID_매칭_desired_name목록",
        "current_is_disabled(Y/N)",
        "current_is_bundle(Y/N)",
        "current_parsed_titles",
        "current_parsed_author_keys",
        "current_parsed_account_codes",
        "canonical_status",
        "account_최종저작권명",
        "account_저작권명",
        "연결_선인세명",
        "관련작가명",
        "원본_작품명_목록",
        "정산자_근거",
        "주의플래그_x",
        "대표_상품번호",
        "대표_제목",
        "account_계약담당자",
    ]
    queue = queue.rename(columns={"주의플래그_x": "seed_주의플래그"})
    output_cols = ["seed_주의플래그" if col == "주의플래그_x" else col for col in output_cols]
    existing_cols = [col for col in output_cols if col in queue.columns]
    remaining_cols = [col for col in queue.columns if col not in existing_cols and not col.startswith("_")]
    queue = queue[existing_cols + remaining_cols].sort_values(
        ["검토우선순위", "account_저작권코드", "작품명", "desired_ips_name"]
    )

    action_counts = Counter(queue["IPS_action_제안"])
    summary = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "manager": MANAGER,
        "seed_rows": int(len(seed)),
        "ips_inventory_rows": int(len(inventory)),
        "match_candidate_rows": int(len(candidates)),
        "queue_rows": int(len(queue)),
        "action_counts": dict(action_counts),
        "input_account_cid_seed": str(ACCOUNT_CID_SEED_PATH),
        "input_account_canonical": str(ACCOUNT_CANONICAL_PATH),
        "input_ips_workbooks": [str(path) for path in IPS_WORKBOOK_PATHS if path.exists()],
        "output_xlsx": str(OUTPUT_XLSX),
        "output_csv": str(OUTPUT_CSV),
        "detail_csv": str(DETAIL_CSV),
        "inventory_csv": str(INVENTORY_CSV),
    }
    return queue, candidates.fillna(""), inventory.drop(columns=["_parsed"]).fillna(""), summary


def autosize_sheet(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = 0
            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def write_outputs(queue: pd.DataFrame, candidates: pd.DataFrame, inventory: pd.DataFrame, summary: dict[str, Any]) -> None:
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    INVENTORY_CSV.parent.mkdir(parents=True, exist_ok=True)

    summary_rows = [
        {"metric": key, "value": json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value}
        for key, value in summary.items()
    ]
    action_summary = (
        queue["IPS_action_제안"]
        .value_counts()
        .rename_axis("IPS_action_제안")
        .reset_index(name="rows")
        .sort_values("IPS_action_제안")
    )

    queue.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    candidates.to_csv(DETAIL_CSV, index=False, encoding="utf-8-sig")
    inventory.to_csv(INVENTORY_CSV, index=False, encoding="utf-8-sig")
    OUTPUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="요약", index=False)
        action_summary.to_excel(writer, sheet_name="action_summary", index=False)
        queue.to_excel(writer, sheet_name="ips_action_queue", index=False)
        candidates.to_excel(writer, sheet_name="match_candidates", index=False)
        inventory.to_excel(writer, sheet_name="current_ips_inventory", index=False)

    autosize_sheet(OUTPUT_XLSX)


def main() -> None:
    queue, candidates, inventory, summary = build_queue()
    write_outputs(queue, candidates, inventory, summary)
    print("=== account→IPS action queue built ===")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
