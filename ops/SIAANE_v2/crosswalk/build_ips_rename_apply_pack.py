from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"
UPDATE_IPS = PROJECT_ROOT / "update_ips.xlsx"
WORK_ORDER = EXPORT_ROOT / f"latest__ips_apply_work_order_{MANAGER}.csv"

OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_rename_apply_{MANAGER}.xlsx"
OUTPUT_CSV = EXPORT_ROOT / f"latest__ips_rename_apply_{MANAGER}.csv"
PREVIEW_XLSX = EXPORT_ROOT / f"update_ips__rename_preview_{MANAGER}.xlsx"


def text(value: Any) -> str:
    return str(value or "").strip()


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if "검증" in ws.title:
            for row in ws.iter_rows(min_row=2):
                if text(row[1].value) not in {"OK", "0"}:
                    for cell in row:
                        cell.fill = warning_fill
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(text(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 80)
    wb.save(path)


def main() -> None:
    work = pd.read_csv(WORK_ORDER, dtype=str).fillna("")
    rename = work[work["작업구분"].eq("01_이름수정")].copy()
    rename = rename.sort_values(["대표작가명", "작품명", "대상CID"], kind="stable")

    xl = pd.ExcelFile(UPDATE_IPS)
    sheet_name = xl.sheet_names[0]
    ips = pd.read_excel(UPDATE_IPS, sheet_name=sheet_name, dtype=str).fillna("")
    ips["콘텐츠ID"] = ips["콘텐츠ID"].map(text)

    rename_ids = set(rename["대상CID"].map(text))
    found_ids = set(ips[ips["콘텐츠ID"].isin(rename_ids)]["콘텐츠ID"].map(text))
    missing_ids = sorted(rename_ids - found_ids)

    current_map = ips.set_index("콘텐츠ID")["콘텐츠명"].to_dict()
    rename["원본파일현재명"] = rename["대상CID"].map(lambda cid: current_map.get(text(cid), ""))
    rename["원본현재명일치(Y/N)"] = rename.apply(
        lambda row: "Y" if text(row["대상현재명"]) == text(row["원본파일현재명"]) else "N",
        axis=1,
    )
    rename["적용가능(Y/N)"] = rename["대상CID"].map(lambda cid: "Y" if text(cid) in found_ids else "N")

    apply_cols = [
        "대상CID",
        "대상현재명",
        "원본파일현재명",
        "원본현재명일치(Y/N)",
        "변경후IPS명",
        "작품명",
        "대표작가명",
        "선인세코드",
        "특수",
        "account_저작권코드",
        "정산자",
        "정산대표Y/N",
        "처리메모",
        "seed_row_index",
        "적용가능(Y/N)",
    ]
    apply_df = rename[[col for col in apply_cols if col in rename.columns]].copy()

    preview = ips.copy()
    rename_map = dict(zip(rename["대상CID"].map(text), rename["변경후IPS명"].map(text)))
    preview["rename_적용대상(Y/N)"] = preview["콘텐츠ID"].map(lambda cid: "Y" if cid in rename_map else "N")
    preview["rename_기존콘텐츠명"] = preview["콘텐츠명"]
    preview["콘텐츠명"] = preview.apply(
        lambda row: rename_map[row["콘텐츠ID"]] if row["콘텐츠ID"] in rename_map else row["콘텐츠명"],
        axis=1,
    )
    changed_preview = preview[preview["rename_적용대상(Y/N)"].eq("Y")].copy()

    validation = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("입력 update_ips", str(UPDATE_IPS)),
            ("입력 work_order", str(WORK_ORDER)),
            ("rename 행 수", len(rename)),
            ("update_ips 내 대상 CID 발견", len(found_ids)),
            ("update_ips 내 대상 CID 누락", len(missing_ids)),
            ("현재명 불일치", int(rename["원본현재명일치(Y/N)"].ne("Y").sum())),
            ("대상CID 중복", int(rename.duplicated(["대상CID"], keep=False).sum())),
            ("변경후IPS명 중복", int(rename.duplicated(["변경후IPS명"], keep=False).sum())),
        ],
        columns=["항목", "값"],
    )
    missing = pd.DataFrame({"누락_대상CID": missing_ids})
    mismatch = rename[rename["원본현재명일치(Y/N)"].ne("Y")][
        ["대상CID", "대상현재명", "원본파일현재명", "변경후IPS명"]
    ]

    apply_df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        validation.to_excel(writer, sheet_name="검증", index=False)
        apply_df.to_excel(writer, sheet_name="rename_적용목록", index=False)
        changed_preview.to_excel(writer, sheet_name="preview_변경행", index=False)
        if not mismatch.empty:
            mismatch.to_excel(writer, sheet_name="현재명불일치", index=False)
        if not missing.empty:
            missing.to_excel(writer, sheet_name="누락CID", index=False)
    autosize(OUTPUT_XLSX)

    with pd.ExcelWriter(PREVIEW_XLSX, engine="openpyxl") as writer:
        preview.to_excel(writer, sheet_name=sheet_name, index=False)
    autosize(PREVIEW_XLSX)

    print("=== IPS rename apply pack built ===")
    print(f"rename_rows={len(rename)}")
    print(f"missing_ids={len(missing_ids)}")
    print(f"name_mismatch={int(rename['원본현재명일치(Y/N)'].ne('Y').sum())}")
    print(OUTPUT_XLSX)
    print(PREVIEW_XLSX)


if __name__ == "__main__":
    main()
