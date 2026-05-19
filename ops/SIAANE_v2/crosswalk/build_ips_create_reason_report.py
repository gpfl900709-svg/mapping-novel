from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from pandas.errors import EmptyDataError
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"
CREATE = EXPORT_ROOT / f"latest__ips_create_request_{MANAGER}.csv"
QUEUE = EXPORT_ROOT / f"latest__account_ips_action_queue_{MANAGER}.csv"
EXISTING_AUDIT = EXPORT_ROOT / f"latest__ips_existing_candidate_audit_{MANAGER}.csv"
OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_create_reason_report_{MANAGER}.xlsx"
OUTPUT_CSV = EXPORT_ROOT / f"latest__ips_create_reason_report_{MANAGER}.csv"


def text(value: Any) -> str:
    return str(value or "").strip()


def reason_for(row: pd.Series, audit_rows: pd.DataFrame) -> tuple[str, str]:
    flags = text(row.get("주의플래그"))
    if audit_rows.empty:
        return "현재IPS매칭없음", "현재 IPS inventory에서 작품+작가 기준 후보가 없음. 신규 생성 필요."

    decisions = set(audit_rows["판정"].map(text))
    if "보류_RS세부CID" in decisions:
        return "기존후보_RS세부CID", "기존 후보는 작가선인세/광고수익/창작지원금 등 세부 RS CID라 canonical 일반/MG CID로 사용 금지."
    if "보류_특수seed_현재일반" in decisions:
        return "기존후보_특수충돌", "특수 seed인데 기존 후보가 일반/윌라 등 다른 용도 CID라 신규 생성 필요."
    if "보류_담당자다름" in decisions:
        return "기존후보_담당자다름", "기존 후보는 타 담당/타 부서 CID라 자동 사용 금지."
    if "보류_사용안함묶음가능" in decisions:
        return "기존후보_사용안함묶음", "사용안함 후보가 있으나 여러 작품 묶음 가능성이 있어 신규 생성이 안전."
    if "현재IPS매칭없음" in flags:
        return "현재IPS매칭없음", "현재 IPS inventory에서 안전 후보가 없음. 신규 생성 필요."
    return "기존후보_검토보류", "후보는 있으나 자동 사용 기준 미충족. 신규 생성 요청으로 분리."


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(text(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 90)
    wb.save(path)


def main() -> None:
    create = pd.read_csv(CREATE, dtype=str).fillna("")
    queue = pd.read_csv(QUEUE, dtype=str).fillna("")
    try:
        audit = pd.read_csv(EXISTING_AUDIT, dtype=str).fillna("")
    except EmptyDataError:
        audit = pd.DataFrame(columns=["신규_콘텐츠명", "판정"])
    queue_small = queue[
        [
            "seed_row_index",
            "IPS_action_제안",
            "주의플래그",
            "current_콘텐츠ID",
            "current_콘텐츠명",
            "현재CID_매칭_seed수",
            "seed별_후보CID수",
            "desired_ips_name",
        ]
    ].copy()
    create = create.merge(queue_small, left_on="seed_row_index", right_on="seed_row_index", how="left")

    columns = [
        "신규생성사유",
        "사유설명",
        "신규_콘텐츠명",
        "작품명",
        "대표작가명",
        "선인세코드",
        "특수",
        "account_저작권코드",
        "정산자",
        "정산대표Y/N",
        "action_queue_제안",
        "주의플래그",
        "기존후보수",
        "기존후보요약",
        "seed_row_index",
    ]
    records: list[dict[str, Any]] = []
    for _, row in create.iterrows():
        target_name = text(row["신규_콘텐츠명"])
        audit_rows = audit[audit["신규_콘텐츠명"].map(text).eq(target_name)].copy()
        reason_code, reason = reason_for(row, audit_rows)
        records.append(
            {
                "신규생성사유": reason_code,
                "사유설명": reason,
                "신규_콘텐츠명": target_name,
                "작품명": row["작품명"],
                "대표작가명": row["대표작가명"],
                "선인세코드": row["선인세코드"],
                "특수": row["특수"],
                "account_저작권코드": row["account_저작권코드"],
                "정산자": row["정산자"],
                "정산대표Y/N": row["정산대표Y/N"],
                "action_queue_제안": row.get("IPS_action_제안", ""),
                "주의플래그": row.get("주의플래그", ""),
                "기존후보수": len(audit_rows),
                "기존후보요약": " | ".join(
                    f"{r['후보CID']}:{r['후보명']}({r['판정']})" for _, r in audit_rows.iterrows()
                ),
                "seed_row_index": row["seed_row_index"],
            }
        )

    report = pd.DataFrame(records, columns=columns).sort_values(
        ["신규생성사유", "대표작가명", "작품명"],
        kind="stable",
    )
    if report.empty:
        summary = pd.DataFrame(columns=["신규생성사유", "특수", "건수"])
    else:
        summary = report.groupby(["신규생성사유", "특수"], dropna=False).size().reset_index(name="건수")
    validation = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("신규생성 요청 건수", len(report)),
            ("기존 후보가 없는 건수", int(report["기존후보수"].eq(0).sum())),
            ("기존 후보가 있으나 보류된 건수", int(report["기존후보수"].gt(0).sum())),
        ],
        columns=["항목", "값"],
    )

    report.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        validation.to_excel(writer, sheet_name="검증", index=False)
        summary.to_excel(writer, sheet_name="요약", index=False)
        report.to_excel(writer, sheet_name="신규생성_사유", index=False)
    autosize(OUTPUT_XLSX)
    print("=== IPS create reason report built ===")
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
