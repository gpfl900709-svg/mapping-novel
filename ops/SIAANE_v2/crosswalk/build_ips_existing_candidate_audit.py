from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"
CREATE_REQUEST = EXPORT_ROOT / f"latest__ips_create_request_{MANAGER}.csv"
INVENTORY = PROJECT_ROOT / "ips" / "stage" / f"latest__current_ips_inventory_{MANAGER}.csv"
OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_existing_candidate_audit_{MANAGER}.xlsx"
OUTPUT_CSV = EXPORT_ROOT / f"latest__ips_existing_candidate_audit_{MANAGER}.csv"


def text(value: Any) -> str:
    return str(value or "").strip()


def norm(value: Any) -> str:
    value = text(value).lower().replace("떄", "때")
    value = re.sub(r"\s+", "", value)
    value = re.sub(r"[\-_·|,.:;!?？！\[\](){}+/]", "", value)
    return value


def classify(candidate_name: str, special: str, manager: str, score: int, reason: str) -> tuple[str, str]:
    name = text(candidate_name)
    if manager != MANAGER:
        return "보류_담당자다름", "담당자가 조원재가 아니므로 자동 사용 금지."
    if "사용안함" in name:
        return "보류_사용안함묶음가능", "사용안함 후보이지만 묶음/분리 여부 확인 필요."
    if any(token in name for token in ["선투자", "광고수익", "창작지원금", "작가선인세", "작품선인세"]):
        return "보류_RS세부CID", "세부 RS/플랫폼성 CID라 일반/MG seed에 바로 사용 금지."
    if "·" in name or "+" in name or name.count("_") >= 5:
        return "보류_묶음CID", "여러 작품 묶음명으로 보여 분해 대상."
    if special in {"카카오MG", "네이버MG"} and "일반" in name:
        return "보류_특수seed_현재일반", "특수 seed인데 현재 후보가 일반 CID."
    if score >= 120 and "title_exact_name" in reason:
        return "자동후보_정확제목", "정확 제목 후보. 작가/특수 충돌 없으면 rename 가능."
    return "보류_검토필요", "후보는 있으나 자동 확정 기준 미달."


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if ws.title not in {"검증", "요약"}:
            headers = [text(cell.value) for cell in ws[1]]
            if "판정" in headers:
                idx = headers.index("판정") + 1
                for row in ws.iter_rows(min_row=2):
                    fill = ok_fill if text(row[idx - 1].value).startswith("자동후보") else warn_fill
                    for cell in row:
                        cell.fill = fill
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(text(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 90)
    wb.save(path)


def main() -> None:
    create = pd.read_csv(CREATE_REQUEST, dtype=str).fillna("")
    inventory = pd.read_csv(INVENTORY, dtype=str).fillna("")
    rows: list[dict[str, Any]] = []
    for _, req in create.iterrows():
        title_key = norm(req["작품명"])
        author_key = norm(req["대표작가명"])
        for _, inv in inventory.iterrows():
            name = text(inv["콘텐츠명"])
            name_key = norm(name)
            score = 0
            reasons: list[str] = []
            if title_key and title_key == name_key:
                score += 80
                reasons.append("title_exact_name")
            if title_key and title_key in name_key:
                score += 50
                reasons.append("title_in_name")
            if author_key and author_key in name_key:
                score += 30
                reasons.append("author_in_name")
            if text(inv.get("담당자명")) == MANAGER:
                score += 5
                reasons.append("manager")
            if "사용안함" in name:
                score -= 5
                reasons.append("disabled")
            if any(mark in name for mark in ["·", "+", "/"]) or name.count("_") >= 5:
                score -= 20
                reasons.append("bundleish")
            if score < 75:
                continue
            decision, note = classify(name, text(req["특수"]), text(inv.get("담당자명")), score, "+".join(reasons))
            rows.append(
                {
                    "판정": decision,
                    "권장": note,
                    "신규_콘텐츠명": req["신규_콘텐츠명"],
                    "작품명": req["작품명"],
                    "대표작가명": req["대표작가명"],
                    "특수": req["특수"],
                    "후보CID": inv["콘텐츠ID"],
                    "후보명": name,
                    "후보담당자": inv.get("담당자명", ""),
                    "후보담당부서": inv.get("담당부서", ""),
                    "score": score,
                    "reason": "+".join(reasons),
                }
            )
    audit = pd.DataFrame(rows)
    if not audit.empty:
        audit = audit.sort_values(["판정", "신규_콘텐츠명", "score"], ascending=[True, True, False], kind="stable")
    summary = audit.groupby("판정", dropna=False).size().reset_index(name="건수") if not audit.empty else pd.DataFrame(columns=["판정", "건수"])
    validation = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("신규생성 요청 건수", len(create)),
            ("기존 후보 행 수", len(audit)),
            ("자동후보 수", int(audit["판정"].str.startswith("자동후보").sum()) if not audit.empty else 0),
        ],
        columns=["항목", "값"],
    )
    audit.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        validation.to_excel(writer, sheet_name="검증", index=False)
        summary.to_excel(writer, sheet_name="요약", index=False)
        if not audit.empty:
            audit.to_excel(writer, sheet_name="전체후보", index=False)
    autosize(OUTPUT_XLSX)
    print("=== IPS existing candidate audit built ===")
    print(f"rows={len(audit)}")
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
