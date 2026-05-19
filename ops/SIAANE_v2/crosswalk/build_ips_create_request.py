from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"
WORK_ORDER = EXPORT_ROOT / f"latest__ips_apply_work_order_{MANAGER}.csv"
OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_create_request_{MANAGER}.xlsx"
OUTPUT_CSV = EXPORT_ROOT / f"latest__ips_create_request_{MANAGER}.csv"


def text(value: Any) -> str:
    return str(value or "").strip()


def split_target_name(name: str) -> list[str]:
    parts = text(name).split("_")
    if len(parts) != 7:
        return ["", "", "", "", "", "", ""]
    return parts


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    input_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if ws.title == "신규생성_요청" and ws.cell(row=1, column=cell.column).value in {
                    "신규_콘텐츠ID_입력",
                    "등록결과",
                    "등록메모",
                }:
                    cell.fill = input_fill
                if ws.title == "검증" and text(row[1].value) not in {"0", "OK"}:
                    cell.fill = warn_fill
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(text(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 80)
    wb.save(path)


def main() -> None:
    work = pd.read_csv(WORK_ORDER, dtype=str).fillna("")
    create = work[(work["작업구분"].eq("03_신규생성")) & (work["작업상태"].eq("대기"))].copy()
    create = create.sort_values(["대표작가명", "작품명", "특수", "변경후IPS명"], kind="stable")

    columns = [
        "등록순번",
        "신규_콘텐츠ID_입력",
        "신규_콘텐츠명",
        "콘텐츠형태",
        "귀속법인",
        "서비스유형",
        "담당부서",
        "담당자명",
        "작품명",
        "대표작가명",
        "선인세코드",
        "특수",
        "account_저작권코드",
        "정산자",
        "정산대표Y/N",
        "등록결과",
        "등록메모",
        "작업출처",
        "처리메모",
        "seed_row_index",
    ]
    rows: list[dict[str, Any]] = []
    for _, row in create.iterrows():
        parts = split_target_name(row["변경후IPS명"])
        rows.append(
            {
                "등록순번": len(rows) + 1,
                "신규_콘텐츠ID_입력": "",
                "신규_콘텐츠명": row["변경후IPS명"],
                "콘텐츠형태": "소설",
                "귀속법인": "키다리스튜디오",
                "서비스유형": "연재",
                "담당부서": "소설편집팀",
                "담당자명": MANAGER,
                "작품명": parts[0] or row["작품명"],
                "대표작가명": parts[1] or row["대표작가명"],
                "선인세코드": parts[2] or row["선인세코드"],
                "특수": parts[3] or row["특수"],
                "account_저작권코드": parts[4] or row["account_저작권코드"],
                "정산자": parts[5] or row["정산자"],
                "정산대표Y/N": parts[6] or row["정산대표Y/N"],
                "등록결과": "",
                "등록메모": "",
                "작업출처": row["출처"],
                "처리메모": row["처리메모"],
                "seed_row_index": row["seed_row_index"],
            }
        )

    request = pd.DataFrame(rows, columns=columns)
    validation = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("입력 작업지시서", str(WORK_ORDER)),
            ("신규생성 요청 건수", len(request)),
            ("신규 콘텐츠ID 방침", "로컬 임의 발급 금지. IPS 등록 후 발급 CID를 신규_콘텐츠ID_입력에 기입."),
            ("신규_콘텐츠명 중복", int(request.duplicated(["신규_콘텐츠명"], keep=False).sum())),
            ("7파트 포맷 오류", int((request["신규_콘텐츠명"].map(lambda v: len(text(v).split("_"))) != 7).sum())),
            (
                "특수값 오류",
                int(~request["특수"].isin(["일반", "카카오MG", "네이버MG", "원작"]).sum())
                if False
                else int((~request["특수"].isin(["일반", "카카오MG", "네이버MG", "원작"])).sum()),
            ),
        ],
        columns=["항목", "값"],
    )
    by_special = request.groupby("특수", dropna=False).size().reset_index(name="건수")

    request.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        validation.to_excel(writer, sheet_name="검증", index=False)
        by_special.to_excel(writer, sheet_name="요약", index=False)
        request.to_excel(writer, sheet_name="신규생성_요청", index=False)
    autosize(OUTPUT_XLSX)

    print("=== IPS create request built ===")
    print(f"rows={len(request)}")
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
