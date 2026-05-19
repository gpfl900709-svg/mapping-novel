from __future__ import annotations

import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"
QUEUE_CSV = EXPORT_ROOT / f"latest__account_ips_action_queue_{MANAGER}.csv"
CANDIDATES_CSV = EXPORT_ROOT / f"latest__account_ips_match_candidates_{MANAGER}.csv"
OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_disabled_promotion_review_{MANAGER}.xlsx"
OUTPUT_CSV = EXPORT_ROOT / f"latest__ips_disabled_promotion_review_{MANAGER}.csv"
OUTPUT_JSON = EXPORT_ROOT / f"latest__ips_disabled_promotion_review_{MANAGER}.json"


def text(value: Any) -> str:
    return str(value or "").strip()


def truthy_y(value: Any) -> bool:
    return text(value).upper() == "Y"


def split_pipe(value: Any) -> list[str]:
    return [part.strip() for part in text(value).split("|") if part.strip()]


def choose_disabled_candidate(candidates: pd.DataFrame) -> tuple[pd.Series | None, str]:
    disabled = candidates[candidates["current_is_disabled(Y/N)"].map(truthy_y)].copy()
    if disabled.empty:
        return None, ""

    disabled["_score_num"] = pd.to_numeric(disabled["candidate_rank_score"], errors="coerce").fillna(0)
    disabled["_weak_penalty"] = disabled["weak_match(Y/N)"].map(lambda value: 1 if truthy_y(value) else 0)
    disabled["_bundle_penalty"] = disabled["current_is_bundle(Y/N)"].map(lambda value: 1 if truthy_y(value) else 0)
    disabled = disabled.sort_values(
        ["_weak_penalty", "_bundle_penalty", "_score_num"],
        ascending=[True, True, False],
    )
    best = disabled.iloc[0]
    all_ids = split_pipe(" | ".join(disabled["current_콘텐츠ID"].map(text).tolist()))
    return best, " | ".join(all_ids)


SPECIAL_TOKENS = [
    "카카오MG",
    "네이버MG",
    "카카오선투자",
    "카카오창작지원금",
    "네이버광고수익",
    "작가선인세",
    "선투자",
    "광고수익",
    "윌라",
]


def has_any(value: Any, tokens: list[str]) -> bool:
    haystack = text(value)
    return any(token in haystack for token in tokens)


def is_general_special(value: Any) -> bool:
    special = text(value)
    return special in {"", "일반", "미연결", "선인세없음"}


def current_looks_special(current_name: Any) -> bool:
    return has_any(current_name, SPECIAL_TOKENS)


def current_looks_clean_general(row: pd.Series) -> bool:
    current_name = text(row.get("current_콘텐츠명"))
    if not current_name:
        return False
    if truthy_y(row.get("current_is_bundle(Y/N)")):
        return False
    if current_looks_special(current_name):
        return False
    return True


def current_looks_matching_special(row: pd.Series) -> bool:
    special = text(row.get("특수"))
    current_name = text(row.get("current_콘텐츠명"))
    if not special or not current_name or truthy_y(row.get("current_is_bundle(Y/N)")):
        return False
    if special in current_name:
        return True
    if special == "카카오MG" and has_any(current_name, ["카카오MG", "카카오선투자", "카카오창작지원금"]):
        return True
    if special == "네이버MG" and has_any(current_name, ["네이버MG", "네이버광고수익"]):
        return True
    return False


def classify(row: pd.Series, disabled: pd.Series | None) -> tuple[str, str]:
    current_id = text(row.get("current_콘텐츠ID"))
    current_bundle = truthy_y(row.get("current_is_bundle(Y/N)"))
    flags = text(row.get("주의플래그"))
    current_multi_seed = text(row.get("현재CID_매칭_seed수"))
    seed_candidate_count = text(row.get("seed별_후보CID수"))

    if disabled is not None:
        disabled_score = float(pd.to_numeric(disabled.get("candidate_rank_score"), errors="coerce") or 0)
        disabled_weak = truthy_y(disabled.get("weak_match(Y/N)"))
        disabled_bundle = truthy_y(disabled.get("current_is_bundle(Y/N)"))
        disabled_reason = text(disabled.get("match_reason"))
        disabled_name = text(disabled.get("current_콘텐츠명"))

        if (
            disabled_score >= 80
            and not disabled_weak
            and not disabled_bundle
            and "title+author" in disabled_reason
            and "카카오 제외" not in disabled_name
        ):
            return (
                "승격후보_사용안함CID있음",
                "사용안함 CID가 제목+작가 근거로 강하게 매칭됨. 승격 전 사용안함 사유/계약/RS만 확인.",
            )
        return (
            "승격보류_사용안함CID근거약함",
            "사용안함 CID는 있으나 약매칭/묶음/정책성 사용안함 가능성이 있어 수동 확인 필요.",
        )

    if current_bundle or "현재IPS_묶음명" in flags:
        return (
            "묶음CID수동판단",
            "사용안함 후보가 없고 현재 매칭 CID가 묶음명임. 기존 CID 분해 또는 신규 CID 생성 판단 필요.",
        )

    if current_id:
        if current_multi_seed and current_multi_seed != "1":
            return (
                "기존운영CID분해필요",
                "사용안함 후보가 없고 현재 운영 CID가 여러 seed에 걸림. 운영 CID를 어느 seed에 남길지 결정 필요.",
            )
        if seed_candidate_count and seed_candidate_count != "1":
            return (
                "기존운영CID분해필요",
                "사용안함 후보가 없고 seed별 후보 CID가 복수임. 후보 중 하나 유지/나머지 생성 판단 필요.",
            )
        return (
            "기존운영CID확인",
            "사용안함 후보는 없지만 현재 운영 CID와 매칭됨. 분해 필요 사유를 확인.",
        )

    return (
        "신규생성후보_사용안함CID없음",
        "사용안함 후보도 현재 매칭 CID도 없음. 신규 생성 후보.",
    )


def detailed_classify(row: pd.Series, disabled: pd.Series | None, bucket: str) -> tuple[str, str]:
    special = text(row.get("특수"))
    current_bundle = truthy_y(row.get("current_is_bundle(Y/N)"))

    if bucket == "승격후보_사용안함CID있음":
        if is_general_special(special):
            return (
                "일반_사용안함CID승격",
                "일반 seed는 사용안함 CID를 승격해서 커버. 승격 전 사용안함 사유/계약/RS만 확인.",
            )
        return (
            "특수_사용안함CID승격전특수성확인",
            "특수 seed는 사용안함 CID 이름에 동일 특수 근거가 있는지 확인 후 승격. 일반 CID와 섞지 않음.",
        )

    if bucket != "승격보류_사용안함CID근거약함":
        return bucket, ""

    if is_general_special(special):
        if current_looks_clean_general(row):
            return (
                "일반_기존운영CID유지",
                "현재 CID가 일반 운영 CID로 보임. 사용안함 후보는 승격하지 말고 현재 CID 유지 쪽이 우선.",
            )
        if disabled is not None:
            return (
                "일반_사용안함CID승격후보",
                "현재 CID가 특수/묶음 쪽이면 일반은 사용안함 CID 승격 후보로 분리.",
            )
        return (
            "일반_신규생성후보",
            "일반 CID 후보가 없으면 신규 생성.",
        )

    if current_looks_matching_special(row):
        return (
            "특수_기존CID유지",
            "현재 CID가 해당 특수 계약용 CID로 보임. 일반 사용안함 CID를 승격하지 않음.",
        )
    if current_bundle:
        return (
            "특수_묶음CID분해필요",
            "현재 특수 CID가 여러 작품 묶음임. 특수 CID를 작품별로 분해하거나 신규 생성.",
        )
    return (
        "특수_별도CID생성후보",
        "카카오MG/네이버MG 등 특수는 일반 사용안함 CID로 덮지 말고 별도 CID 목록으로 보냄.",
    )


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column:
                max_len = max(max_len, len(text(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)
        if ws.title == "요약":
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = note_fill
    wb.save(path)


def main() -> None:
    queue = pd.read_csv(QUEUE_CSV, dtype=str).fillna("")
    candidates = pd.read_csv(CANDIDATES_CSV, dtype=str).fillna("")
    split_queue = queue[queue["IPS_action_제안"].eq("IPS_분해/중복검토")].copy()

    rows: list[dict[str, Any]] = []
    for _, row in split_queue.iterrows():
        seed_idx = text(row.get("seed_row_index"))
        desired_name = text(row.get("desired_ips_name"))
        cand = candidates[
            candidates["seed_row_index"].map(text).eq(seed_idx)
            & candidates["desired_ips_name"].map(text).eq(desired_name)
        ].copy()
        disabled, disabled_ids = choose_disabled_candidate(cand)
        bucket, recommendation = classify(row, disabled)
        detail_bucket, detail_recommendation = detailed_classify(row, disabled, bucket)

        rows.append(
            {
                "분류": bucket,
                "세부분류": detail_bucket,
                "권장검토": recommendation,
                "세부처리방향": detail_recommendation,
                "검토우선순위": row.get("검토우선순위", ""),
                "desired_ips_name": desired_name,
                "작품명": row.get("작품명", ""),
                "대표작가명": row.get("대표작가명", ""),
                "특수": row.get("특수", ""),
                "연결_선인세코드": row.get("연결_선인세코드", ""),
                "account_저작권코드": row.get("account_저작권코드", ""),
                "정산자": row.get("정산자", ""),
                "정산대표Y/N": row.get("정산대표Y/N", ""),
                "현재CID": row.get("current_콘텐츠ID", ""),
                "현재IPS명": row.get("current_콘텐츠명", ""),
                "현재담당자": row.get("current_담당자명", ""),
                "현재CID_매칭_seed수": row.get("현재CID_매칭_seed수", ""),
                "seed별_후보CID수": row.get("seed별_후보CID수", ""),
                "주의플래그": row.get("주의플래그", ""),
                "사용안함_승격후보CID": "" if disabled is None else disabled.get("current_콘텐츠ID", ""),
                "사용안함_승격후보명": "" if disabled is None else disabled.get("current_콘텐츠명", ""),
                "사용안함_후보점수": "" if disabled is None else disabled.get("candidate_rank_score", ""),
                "사용안함_match_tier": "" if disabled is None else disabled.get("match_tier", ""),
                "사용안함_match_reason": "" if disabled is None else disabled.get("match_reason", ""),
                "사용안함_전체후보CID": disabled_ids,
                "후보CID수": len(cand),
                "seed_row_index": seed_idx,
            }
        )

    review = pd.DataFrame(rows)
    review = review.sort_values(
        ["분류", "검토우선순위", "대표작가명", "작품명", "desired_ips_name"],
        kind="stable",
    )
    summary = (
        review["분류"]
        .value_counts()
        .rename_axis("분류")
        .reset_index(name="건수")
        .sort_values("분류")
    )
    detail_summary = (
        review["세부분류"]
        .value_counts()
        .rename_axis("세부분류")
        .reset_index(name="건수")
        .sort_values("세부분류")
    )
    summary_meta = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("입력 action queue", str(QUEUE_CSV)),
            ("입력 match candidates", str(CANDIDATES_CSV)),
            ("분해/중복검토 행 수", len(review)),
            ("승격후보_사용안함CID있음", int(review["분류"].eq("승격후보_사용안함CID있음").sum())),
            ("승격보류_사용안함CID근거약함", int(review["분류"].eq("승격보류_사용안함CID근거약함").sum())),
        ],
        columns=["항목", "값"],
    )

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    review.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    payload = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "manager": MANAGER,
        "input_queue": str(QUEUE_CSV),
        "input_candidates": str(CANDIDATES_CSV),
        "rows": int(len(review)),
        "bucket_counts": dict(Counter(review["분류"])),
        "output_xlsx": str(OUTPUT_XLSX),
        "output_csv": str(OUTPUT_CSV),
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        summary_meta.to_excel(writer, sheet_name="요약", index=False, startrow=0)
        summary.to_excel(writer, sheet_name="요약", index=False, startrow=len(summary_meta) + 2)
        detail_summary.to_excel(
            writer,
            sheet_name="요약",
            index=False,
            startrow=len(summary_meta) + len(summary) + 5,
        )
        for bucket in [
            "승격후보_사용안함CID있음",
            "승격보류_사용안함CID근거약함",
            "묶음CID수동판단",
            "기존운영CID분해필요",
            "기존운영CID확인",
            "신규생성후보_사용안함CID없음",
        ]:
            subset = review[review["분류"].eq(bucket)]
            if not subset.empty:
                subset.to_excel(writer, sheet_name=bucket[:31], index=False)
        review.to_excel(writer, sheet_name="전체", index=False)
    autosize(OUTPUT_XLSX)

    print("=== IPS disabled promotion review built ===")
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
