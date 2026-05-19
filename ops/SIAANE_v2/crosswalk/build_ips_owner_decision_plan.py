from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MANAGER = "조원재"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXPORT_ROOT = PROJECT_ROOT / "crosswalk" / "exports"
INPUT_CSV = EXPORT_ROOT / f"latest__ips_owner_decision_queue_{MANAGER}.csv"
OUTPUT_CSV = EXPORT_ROOT / f"latest__ips_owner_decision_plan_{MANAGER}.csv"
OUTPUT_XLSX = EXPORT_ROOT / f"latest__ips_owner_decision_plan_{MANAGER}.xlsx"

SPECIAL_TOKENS = [
    "카카오MG",
    "네이버MG",
    "카카오선투자",
    "카카오창작지원금",
    "네이버광고수익",
    "네이버 MG",
    "작가선인세",
    "작품선인세",
    "선투자",
    "광고수익",
    "윌라",
    "카카오",
    "네이버",
]

CID_SPECIAL_VALUES = {"일반", "카카오MG", "네이버MG", "원작"}

MANUAL_DECISIONS = {
    # 네이버광고수익은 별도 canonical로 두지 않고, 핵무기도 만들어 드림 일반 대표 CID로 전환한다.
    "247": {
        "내판단": "광고수익 CID를 일반 canonical으로 전환",
        "적용액션": "이름수정",
        "유지CID": "321710",
        "승격CID": "",
        "신규생성명": "핵무기도 만들어 드림_북홀릭_선인세없음_일반_1005653_이권배_Y",
        "처리메모": "네이버광고수익은 별도 canonical 불필요. 321710을 핵무기도 만들어 드림 일반 대표 CID로 rename.",
        "위험플래그": "광고수익CID_일반전환",
        "대상현재명": "핵무기도 만들어 드림_북홀릭_네이버광고수익_4438797",
    },
    # 서오 광고수익 묶음 CID 2개를 부족한 작품별 일반 CID로 전환한다.
    "137": {
        "내판단": "광고수익 CID를 일반 canonical으로 전환",
        "적용액션": "이름수정",
        "유지CID": "325041",
        "승격CID": "",
        "신규생성명": "시간의 마에스트로_서오_1021_일반_1005190_유동우_N",
        "처리메모": "네이버광고수익은 별도 canonical 불필요. 325041을 시간의 마에스트로 일반 CID로 rename.",
        "위험플래그": "광고수익CID_일반전환 | 묶음CID",
        "대상현재명": "독식하는 재벌 3세_마법 배운 재벌집 늦둥이_미생법사_순혈의 헌터_스킬스_스킬스 - 현대편_시간의 마에스트로_연봉 1조 신입사원_현자귀환_서오_네이버광고수익_4476726",
    },
    "169": {
        "내판단": "광고수익 CID를 일반 canonical으로 전환",
        "적용액션": "이름수정",
        "유지CID": "325012",
        "승격CID": "",
        "신규생성명": "연봉 1조 신입사원_서오_선인세없음_일반_1005344_유동우_N",
        "처리메모": "카카오광고수익은 별도 canonical 불필요. 325012를 연봉 1조 신입사원 일반 CID로 rename.",
        "위험플래그": "광고수익CID_일반전환 | 묶음CID",
        "대상현재명": "독식하는 재벌 3세_마법 배운 재벌집 늦둥이_미생법사_순혈의 헌터_스킬스_스킬스 - 현대편_시간의 마에스트로_연봉 1조 신입사원_현자귀환_서오_카카오광고수익_4476726",
    },
}


def text(value: Any) -> str:
    return str(value or "").strip()


def normalize_target_name(value: Any) -> str:
    return text(value).replace("_미연결_", "_선인세없음_")


def has_any(value: Any, tokens: list[str]) -> bool:
    haystack = text(value)
    return any(token in haystack for token in tokens)


def is_general(row: pd.Series) -> bool:
    return text(row.get("특수")) in {"", "일반", "미연결", "선인세없음"}


def canonical_special(row: pd.Series) -> str:
    special = text(row.get("특수")).replace(" ", "")
    if not special or special in {"미연결", "선인세없음"}:
        return "일반"
    if special in CID_SPECIAL_VALUES:
        return special
    if special == "네이버MG":
        return "네이버MG"
    if special == "카카오MG":
        return "카카오MG"
    if special == "원작":
        return "원작"
    return "일반"


def seed_name(row: pd.Series, special: str | None = None) -> str:
    parts = [
        text(row.get("작품명")) or "미상작품",
        text(row.get("대표작가명")) or "미상작가",
        text(row.get("연결_선인세코드")) or "선인세없음",
        special or canonical_special(row),
        text(row.get("account_저작권코드")) or "미상권리",
        text(row.get("정산자")) or "미상",
        text(row.get("정산대표Y/N")) or "N",
    ]
    return "_".join(parts)


def current_is_general(row: pd.Series) -> bool:
    current_name = text(row.get("현재IPS명"))
    return "_일반" in current_name or current_name.endswith("일반")


def current_is_special(row: pd.Series) -> bool:
    return has_any(row.get("현재IPS명"), SPECIAL_TOKENS)


def current_matches_seed_author(row: pd.Series) -> bool:
    author = text(row.get("대표작가명"))
    current_name = text(row.get("현재IPS명"))
    return bool(author and author in current_name)


def current_matches_special(row: pd.Series) -> bool:
    special = text(row.get("특수"))
    current_name = text(row.get("현재IPS명"))
    if special == "카카오MG":
        return has_any(current_name, ["카카오MG", "카카오선투자", "카카오창작지원금"])
    if special == "네이버MG":
        return has_any(current_name, ["네이버MG", "네이버 MG", "네이버광고수익", "_네이버"])
    return special and special in current_name


def is_bundle(row: pd.Series) -> bool:
    try:
        return int(float(text(row.get("현재CID_매칭_seed수")) or "0")) > 1
    except ValueError:
        return False


def decide(row: pd.Series) -> dict[str, str]:
    manual = MANUAL_DECISIONS.get(text(row.get("seed_row_index")))
    if manual is not None:
        return manual

    detail = text(row.get("세부분류"))
    title = text(row.get("작품명"))
    special = text(row.get("특수"))
    current_id = text(row.get("현재CID"))
    current_name = text(row.get("현재IPS명"))
    disabled_id = text(row.get("사용안함_승격후보CID"))
    disabled_name = text(row.get("사용안함_승격후보명"))

    risk_flags: list[str] = []
    if "외전" in title:
        risk_flags.append("제목변형/외전")
    if current_is_special(row) and is_general(row) and not current_is_general(row):
        risk_flags.append("일반seed_현재특수CID")
    if not is_general(row) and current_is_general(row):
        risk_flags.append("특수seed_현재일반CID")

    if detail == "특수_묶음CID분해필요":
        return {
            "내판단": "특수 묶음CID 분해",
            "적용액션": "신규/분해",
            "유지CID": "",
            "승격CID": "",
            "신규생성명": seed_name(row),
            "처리메모": f"현재 {current_id}는 세부 RS/플랫폼 묶음명. CID명은 seed 포맷의 {canonical_special(row)}로 분해. 일반 사용안함 CID({disabled_id}) 승격 금지.",
            "위험플래그": "묶음CID | 특수계약",
        }

    if detail == "특수_별도CID생성후보":
        return {
            "내판단": "특수 CID 별도 생성",
            "적용액션": "신규생성",
            "유지CID": "",
            "승격CID": "",
            "신규생성명": seed_name(row),
            "처리메모": f"현재 {current_id}는 일반 CID. 사용안함 후보({disabled_id} {disabled_name})도 일반이라 {canonical_special(row)} 승격에 쓰지 않음.",
            "위험플래그": "특수seed_현재일반CID",
        }

    if detail == "묶음CID수동판단":
        if not is_general(row):
            return {
                "내판단": "특수 묶음CID 분해",
                "적용액션": "신규/분해",
                "유지CID": "",
                "승격CID": "",
                "신규생성명": seed_name(row),
                "처리메모": f"현재 {current_id}는 여러 작품 묶음 세부 RS/플랫폼 CID. 작품별 {canonical_special(row)} CID로 분해.",
                "위험플래그": "묶음CID | 특수계약",
            }
        return {
            "내판단": "일반 CID 별도 확보",
            "적용액션": "신규생성/후보재탐색",
            "유지CID": "",
            "승격CID": disabled_id,
            "신규생성명": seed_name(row, "일반"),
            "처리메모": f"현재 {current_id}는 특수/묶음 성격. 일반 CID로 유지하지 않음.",
            "위험플래그": "묶음CID | 일반seed_현재특수CID",
        }

    if not is_general(row):
        if current_matches_special(row) and not is_bundle(row):
            return {
                "내판단": "특수 기존CID 유지",
                "적용액션": "유지",
                "유지CID": current_id,
                "승격CID": "",
                "신규생성명": "",
                "처리메모": f"현재 CID명이 {special} 성격과 맞음. 일반 CID로 합치지 않음.",
                "위험플래그": " | ".join(risk_flags),
            }
        return {
            "내판단": "특수 CID 별도 생성",
            "적용액션": "신규생성",
            "유지CID": "",
            "승격CID": "",
            "신규생성명": seed_name(row),
            "처리메모": f"현재 CID({current_id})가 {canonical_special(row)} 용도로 보기 어려움. seed 포맷으로 별도 생성.",
            "위험플래그": " | ".join(risk_flags or ["특수CID없음"]),
        }

    if current_is_general(row) and current_matches_seed_author(row):
        return {
            "내판단": "일반 기존CID 유지",
            "적용액션": "유지",
            "유지CID": current_id,
            "승격CID": "",
            "신규생성명": "",
            "처리메모": "현재 CID가 일반명 및 작가명과 맞음. 복수 seed 충돌은 불일치/특수 쪽 신규 생성으로 해소.",
            "위험플래그": " | ".join(risk_flags),
        }

    if disabled_id:
        return {
            "내판단": "일반 사용안함CID 승격",
            "적용액션": "승격",
            "유지CID": "",
            "승격CID": disabled_id,
            "신규생성명": "",
            "처리메모": f"현재 CID({current_id})는 일반용으로 부적합. 사용안함 후보 승격.",
            "위험플래그": " | ".join(risk_flags),
        }

    return {
        "내판단": "일반 CID 신규 생성",
        "적용액션": "신규생성",
        "유지CID": "",
        "승격CID": "",
        "신규생성명": seed_name(row, "일반"),
        "처리메모": f"현재 CID({current_id})는 일반용으로 확정하기 애매함. 신규 생성이 보수적.",
        "위험플래그": " | ".join(risk_flags or ["일반CID확정불가"]),
    }


def autosize(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(text(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 70)
    wb.save(path)


def main() -> None:
    df = pd.read_csv(INPUT_CSV, dtype=str).fillna("")
    if "desired_ips_name" in df.columns:
        df["desired_ips_name"] = df["desired_ips_name"].map(normalize_target_name)
    decisions = pd.DataFrame([decide(row) for _, row in df.iterrows()])
    result = pd.concat([decisions, df.drop(columns=["내판단", "추천처리", "메모"], errors="ignore")], axis=1)
    result = result.sort_values(["적용액션", "대표작가명", "작품명", "특수"], kind="stable")

    summary = (
        result["내판단"]
        .value_counts()
        .rename_axis("내판단")
        .reset_index(name="건수")
        .sort_values("내판단")
    )
    risk_summary = (
        result["위험플래그"]
        .replace("", "없음")
        .value_counts()
        .rename_axis("위험플래그")
        .reset_index(name="건수")
        .sort_values("위험플래그")
    )
    meta = pd.DataFrame(
        [
            ("생성시각", datetime.now().isoformat(timespec="seconds")),
            ("입력", str(INPUT_CSV)),
            ("행 수", len(result)),
            ("원칙", "CID명 특수값은 일반/카카오MG/네이버MG/원작만 사용. 세부 플랫폼 RS는 IPS 내부 로직에서 조정."),
        ],
        columns=["항목", "값"],
    )

    result.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        meta.to_excel(writer, sheet_name="요약", index=False, startrow=0)
        summary.to_excel(writer, sheet_name="요약", index=False, startrow=len(meta) + 2)
        risk_summary.to_excel(writer, sheet_name="요약", index=False, startrow=len(meta) + len(summary) + 5)
        result.to_excel(writer, sheet_name="처리안", index=False)
    autosize(OUTPUT_XLSX)

    print("=== IPS owner decision plan built ===")
    print(f"rows={len(result)}")
    print(dict(Counter(result["내판단"])))
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
