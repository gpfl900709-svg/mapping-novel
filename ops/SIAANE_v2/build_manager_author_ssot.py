from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path, PureWindowsPath
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = ROOT / "SIAAN Project"
TARGET_ROOT = ROOT / "SIAANE_v2" / "담당작가_ssot"
INDEX_PATH = ROOT / "SIAANE_v2" / "index_f.xlsx"
MANUAL_GROUPS_PATH = ROOT / "SIAANE_v2" / "manager_author_manual_groups.json"
REGISTRY_PATH = PROJECT_ROOT / "config" / "work_cid_registry.local.csv"
ALIASES_PATH = PROJECT_ROOT / "config" / "cp_pen_aliases.local.json"
BAROBOOK_DIR = PROJECT_ROOT / "data" / "exports" / "barobook"
MANAGER = "조원재"

_NON_WORD_RE = re.compile(r"[^0-9A-Za-z가-힣]+")
_THIN = Side(style="thin", color="D9DEE6")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SUBHEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
_GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_BAD_FILL = PatternFill("solid", fgColor="FCE4D6")


@dataclass(frozen=True)
class ResolveResult:
    canonical_author: str
    rule: str


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def compact_key(value: Any) -> str:
    return _NON_WORD_RE.sub("", normalize_text(value)).lower()


def unique(values: list[Any]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        cleaned = normalize_text(value)
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            ordered.append(cleaned)
    return ordered


def pipe_join(values: list[Any]) -> str:
    return " | ".join(unique(values))


def latest_path(pattern: str) -> Path:
    hits = sorted(BAROBOOK_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not hits:
        raise FileNotFoundError(pattern)
    return hits[0]


def optional_latest_path(pattern: str) -> Path | None:
    hits = sorted(BAROBOOK_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    return hits[0] if hits else None


def explode_pipe_values(raw: Any) -> list[str]:
    parts = [normalize_text(piece) for piece in str(raw or "").split("|")]
    return [piece for piece in unique(parts) if piece]


def load_index_df(manager: str) -> pd.DataFrame:
    df = pd.read_excel(INDEX_PATH, sheet_name="담당자찾기", dtype=str).fillna("")
    df = df[df["담당자"].map(normalize_text) == manager].copy()
    df["작가"] = df["작가"].map(normalize_text)
    return df


def load_registry_df(manager: str) -> pd.DataFrame:
    df = pd.read_csv(REGISTRY_PATH, dtype=str).fillna("")
    df = df[df["manager"].map(normalize_text) == manager].copy()
    for column in ("author_final", "title_final", "folder_path", "match_status", "matched_s2_author", "matched_ips_author", "note"):
        if column in df.columns:
            df[column] = df[column].map(normalize_text)
    return df


def load_manual_config() -> dict[str, Any]:
    if not MANUAL_GROUPS_PATH.exists():
        return {}
    payload = json.loads(MANUAL_GROUPS_PATH.read_text(encoding="utf-8"))
    return payload if isinstance(payload, dict) else {}


def load_manual_group_maps(
    registry_names: list[str],
    payload: dict[str, Any],
) -> tuple[dict[str, str], dict[str, str], dict[str, list[str]]]:
    raw_groups = [unique(group) for group in (payload.get("groups") or [])]
    raw_groups = [group for group in raw_groups if group]
    if not raw_groups:
        return {}, {}, {}

    ordered_names = unique([name for group in raw_groups for name in group])
    ordered_index = {name: idx for idx, name in enumerate(ordered_names)}
    registry_set = set(registry_names)

    graph: dict[str, set[str]] = defaultdict(set)
    for group in raw_groups:
        for name in group:
            graph.setdefault(name, set())
        for name in group:
            for peer in group:
                if peer != name:
                    graph[name].add(peer)

    visited: set[str] = set()
    exact_map: dict[str, str] = {}
    compact_map: dict[str, str] = {}
    canonical_to_variants: dict[str, list[str]] = {}

    for seed in ordered_names:
        if seed in visited:
            continue
        stack = [seed]
        component: list[str] = []
        while stack:
            current = stack.pop()
            if current in visited:
                continue
            visited.add(current)
            component.append(current)
            stack.extend(sorted(graph.get(current, set()) - visited))

        component = sorted(unique(component), key=lambda item: ordered_index[item])
        canonical = ""
        for name in component:
            if name in registry_set:
                canonical = name
                break
        if not canonical:
            canonical = component[0]

        canonical_to_variants[canonical] = component
        for alias in component:
            exact_map[alias] = canonical
            compact_map[compact_key(alias)] = canonical

    return exact_map, compact_map, canonical_to_variants


def canonicalize_manual_name(
    value: Any,
    *,
    manual_exact_map: dict[str, str],
    manual_compact_map: dict[str, str],
) -> str:
    name = normalize_text(value)
    if not name:
        return ""
    return (
        normalize_text(manual_exact_map.get(name))
        or normalize_text(manual_compact_map.get(compact_key(name)))
        or name
    )


def build_excluded_canonical_set(
    payload: dict[str, Any],
    *,
    manual_exact_map: dict[str, str],
    manual_compact_map: dict[str, str],
) -> set[str]:
    excluded: set[str] = set()
    for raw_name in payload.get("excluded_authors") or []:
        canonical = canonicalize_manual_name(
            raw_name,
            manual_exact_map=manual_exact_map,
            manual_compact_map=manual_compact_map,
        )
        if canonical:
            excluded.add(canonical)
    return excluded


def load_manual_nas_overrides(
    payload: dict[str, Any],
    *,
    manual_exact_map: dict[str, str],
    manual_compact_map: dict[str, str],
) -> dict[str, list[str]]:
    overrides: dict[str, list[str]] = defaultdict(list)
    for entry in payload.get("force_nas_current") or []:
        if not isinstance(entry, dict):
            continue
        canonical = canonicalize_manual_name(
            entry.get("name"),
            manual_exact_map=manual_exact_map,
            manual_compact_map=manual_compact_map,
        )
        if not canonical:
            continue
        for folder_path in unique(entry.get("folder_paths") or []):
            overrides[canonical].append(folder_path)
    return {name: unique(paths) for name, paths in overrides.items()}


def load_force_confirm_authors(
    payload: dict[str, Any],
    *,
    manual_exact_map: dict[str, str],
    manual_compact_map: dict[str, str],
) -> set[str]:
    forced: set[str] = set()
    for raw_name in payload.get("force_confirm_authors") or []:
        canonical = canonicalize_manual_name(
            raw_name,
            manual_exact_map=manual_exact_map,
            manual_compact_map=manual_compact_map,
        )
        if canonical:
            forced.add(canonical)
    return forced


def folder_name_from_path(raw_path: Any) -> str:
    return normalize_text(PureWindowsPath(str(raw_path or "")).name)


def folder_title_from_path(raw_path: Any) -> str:
    folder_name = folder_name_from_path(raw_path)
    return normalize_text(re.sub(r"\([^()]+\)\s*$", "", folder_name))


def folder_author_from_path(raw_path: Any) -> str:
    folder_name = folder_name_from_path(raw_path)
    match = re.search(r"\(([^()]+)\)\s*$", folder_name)
    return normalize_text(match.group(1)) if match else ""


def load_alias_maps() -> tuple[dict[str, str], dict[str, list[str]], dict[str, list[str]]]:
    payload = json.loads(ALIASES_PATH.read_text(encoding="utf-8"))
    aliases = {
        normalize_text(registry_pen): normalize_text(cp_pen)
        for registry_pen, cp_pen in (payload.get("aliases") or {}).items()
        if normalize_text(registry_pen) and normalize_text(cp_pen)
    }
    cp_to_registry: dict[str, list[str]] = defaultdict(list)
    registry_to_cp: dict[str, list[str]] = defaultdict(list)
    for registry_pen, cp_pen in aliases.items():
        cp_to_registry[cp_pen].append(registry_pen)
        registry_to_cp[registry_pen].append(cp_pen)
    return aliases, dict(cp_to_registry), dict(registry_to_cp)


def load_cp_frames(manager: str) -> dict[str, pd.DataFrame]:
    outputs: dict[str, pd.DataFrame] = {}
    for status in ("matched", "unmatched", "excluded_by_pin", "disabled"):
        path = optional_latest_path(f"*_{manager}_{status}.xlsx")
        if path is None:
            outputs[status] = pd.DataFrame()
            continue
        frame = pd.read_excel(path, dtype=str).fillna("")
        frame["_source"] = status
        for column in ("작가코드", "저작권자명_필명", "_실명", "_필명후보", "매칭_키_필명"):
            if column in frame.columns:
                frame[column] = frame[column].map(normalize_text)
        outputs[status] = frame
    detail_path = latest_path(f"*__NAS_IPS_ACCOUNT_예금주_{manager}_author_details.csv")
    detail_df = pd.read_csv(detail_path, dtype=str).fillna("")
    for column in ("작가코드", "pen_name", "real_name", "예금주", "contract_manager", "detail_url"):
        if column in detail_df.columns:
            detail_df[column] = detail_df[column].map(normalize_text)
    outputs["author_details"] = detail_df
    return outputs


def build_registry_name_indexes(registry_df: pd.DataFrame) -> tuple[list[str], dict[str, list[str]]]:
    name_column = "canonical_author" if "canonical_author" in registry_df.columns else "author_final"
    registry_names = unique(registry_df[name_column].tolist())
    compact_index: dict[str, list[str]] = defaultdict(list)
    for name in registry_names:
        compact_index[compact_key(name)].append(name)
    return registry_names, dict(compact_index)


def resolve_index_name(
    raw_name: str,
    *,
    registry_names: list[str],
    registry_compact_index: dict[str, list[str]],
    cp_to_registry: dict[str, list[str]],
    manual_exact_map: dict[str, str],
    manual_compact_map: dict[str, str],
) -> ResolveResult:
    name = normalize_text(raw_name)
    if not name:
        return ResolveResult("", "")

    manual_exact = normalize_text(manual_exact_map.get(name))
    if manual_exact and manual_exact != name:
        return ResolveResult(manual_exact, "manual_group_exact")

    if name in registry_names:
        return ResolveResult(name, "exact_registry")

    direct_alias_hits = unique(cp_to_registry.get(name, []))
    if len(direct_alias_hits) == 1:
        return ResolveResult(direct_alias_hits[0], "alias_cp_to_registry")

    compact = compact_key(name)
    manual_compact = normalize_text(manual_compact_map.get(compact))
    if manual_compact and manual_compact != name:
        return ResolveResult(manual_compact, "manual_group_compact")

    compact_hits = unique(registry_compact_index.get(compact, []))
    if len(compact_hits) == 1:
        return ResolveResult(compact_hits[0], "compact_registry")

    compact_alias_hits: list[str] = []
    for cp_name, registry_candidates in cp_to_registry.items():
        if compact_key(cp_name) == compact:
            compact_alias_hits.extend(registry_candidates)
    compact_alias_hits = unique(compact_alias_hits)
    if len(compact_alias_hits) == 1:
        return ResolveResult(compact_alias_hits[0], "compact_alias_to_registry")

    fuzzy_scores: list[tuple[float, str]] = []
    for registry_name in registry_names:
        ratio = SequenceMatcher(None, compact, compact_key(registry_name)).ratio()
        if ratio >= 0.82:
            fuzzy_scores.append((ratio, registry_name))
    fuzzy_scores.sort(reverse=True)
    if fuzzy_scores and (len(fuzzy_scores) == 1 or fuzzy_scores[0][0] - fuzzy_scores[1][0] >= 0.06):
        return ResolveResult(fuzzy_scores[0][1], f"fuzzy_registry:{fuzzy_scores[0][0]:.2f}")

    return ResolveResult(name, "unresolved_keep")


def explode_candidate_names(row: dict[str, Any]) -> list[str]:
    values: list[str] = []
    values.extend(explode_pipe_values(row.get("_필명후보")))
    values.extend([normalize_text(row.get("_실명")), normalize_text(row.get("저작권자명_필명"))])
    return unique(values)


def row_matches_names(row: dict[str, Any], related_names: list[str]) -> bool:
    related_keys = {compact_key(name) for name in related_names if compact_key(name)}
    if not related_keys:
        return False
    for candidate_name in explode_candidate_names(row):
        if compact_key(candidate_name) in related_keys:
            return True
    return False


def detail_matches_names(row: dict[str, Any], related_names: list[str]) -> bool:
    related_keys = {compact_key(name) for name in related_names if compact_key(name)}
    if not related_keys:
        return False
    for candidate in (row.get("pen_name"), row.get("real_name"), row.get("예금주")):
        if compact_key(candidate) in related_keys:
            return True
    return False


def classify_row(*, index_flag: bool, nas_flag: bool, account_codes: list[str], contract_managers: list[str], manager: str) -> tuple[str, str, str]:
    has_account = bool(account_codes)
    has_my_manager = manager in contract_managers
    has_other_manager = bool([name for name in contract_managers if name and name != manager])

    if nas_flag and index_flag:
        if has_other_manager and not has_my_manager:
            return "Y", "확정(account 타담당 주의)", "P2"
        if not has_account:
            return "Y", "확정(account 미연결)", "P1"
        return "Y", "확정", ""
    if nas_flag and not index_flag:
        return "Y", "검토(NAS만)", "P1"
    if index_flag and not nas_flag:
        return "보류", "검토(index만)", "P1"
    return "", "기타", "P3"


def build_datasets(manager: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    index_df = load_index_df(manager)
    registry_df = load_registry_df(manager)
    manual_payload = load_manual_config()
    _, cp_to_registry, registry_to_cp = load_alias_maps()
    raw_registry_names = unique(registry_df["author_final"].tolist())
    manual_exact_map, manual_compact_map, canonical_to_manual_variants = load_manual_group_maps(
        raw_registry_names,
        manual_payload,
    )
    excluded_canonicals = build_excluded_canonical_set(
        manual_payload,
        manual_exact_map=manual_exact_map,
        manual_compact_map=manual_compact_map,
    )
    manual_nas_overrides = load_manual_nas_overrides(
        manual_payload,
        manual_exact_map=manual_exact_map,
        manual_compact_map=manual_compact_map,
    )
    force_confirm_authors = load_force_confirm_authors(
        manual_payload,
        manual_exact_map=manual_exact_map,
        manual_compact_map=manual_compact_map,
    )
    if manual_exact_map:
        registry_df["canonical_author"] = registry_df["author_final"].map(
            lambda value: canonicalize_manual_name(
                value,
                manual_exact_map=manual_exact_map,
                manual_compact_map=manual_compact_map,
            )
        )
    else:
        registry_df["canonical_author"] = registry_df["author_final"].map(normalize_text)
    if excluded_canonicals:
        index_df = index_df[
            ~index_df["작가"].map(
                lambda value: canonicalize_manual_name(
                    value,
                    manual_exact_map=manual_exact_map,
                    manual_compact_map=manual_compact_map,
                )
                in excluded_canonicals
            )
        ].copy()
        registry_df = registry_df[~registry_df["canonical_author"].isin(excluded_canonicals)].copy()

    frames = load_cp_frames(manager)
    registry_names, registry_compact_index = build_registry_name_indexes(registry_df)

    index_names = unique(index_df["작가"].tolist())
    resolve_rows: list[dict[str, str]] = []
    index_variants_by_canonical: dict[str, list[str]] = defaultdict(list)
    for raw_name in index_names:
        resolved = resolve_index_name(
            raw_name,
            registry_names=registry_names,
            registry_compact_index=registry_compact_index,
            cp_to_registry=cp_to_registry,
            manual_exact_map=manual_exact_map,
            manual_compact_map=manual_compact_map,
        )
        resolve_rows.append(
            {
                "index_f_작가표기": raw_name,
                "canonical_작가": resolved.canonical_author,
                "이름변형_규칙": resolved.rule,
            }
        )
        index_variants_by_canonical[resolved.canonical_author].append(raw_name)

    canonical_authors = [
        author
        for author in unique([row["canonical_작가"] for row in resolve_rows] + registry_names)
        if author not in excluded_canonicals
    ]

    registry_grouped = registry_df.groupby("canonical_author", dropna=False)
    cp_frames = [frame for key, frame in frames.items() if key != "author_details" and not frame.empty]
    all_cp_rows = pd.concat(cp_frames, ignore_index=True, sort=False).fillna("") if cp_frames else pd.DataFrame()
    detail_df = frames["author_details"]

    ssot_rows: list[dict[str, Any]] = []
    for canonical_author in canonical_authors:
        index_variants = unique(index_variants_by_canonical.get(canonical_author, []))
        manual_variants = unique(canonical_to_manual_variants.get(canonical_author, []))
        registry_aliases = unique(registry_to_cp.get(canonical_author, []))
        for manual_variant in manual_variants:
            registry_aliases.extend(registry_to_cp.get(manual_variant, []))
        related_names = unique([canonical_author] + index_variants + manual_variants + registry_aliases)

        index_flag = bool(index_variants)
        manual_nas_paths = unique(manual_nas_overrides.get(canonical_author, []))
        has_registry_rows = canonical_author in registry_grouped.groups
        nas_flag = has_registry_rows or bool(manual_nas_paths)

        registry_rows = registry_grouped.get_group(canonical_author).copy() if has_registry_rows else pd.DataFrame()
        nas_titles = unique(registry_rows["title_final"].tolist()) if not registry_rows.empty else []
        nas_match_statuses = unique(registry_rows["match_status"].tolist()) if not registry_rows.empty else []
        nas_aliases = unique(registry_rows["author_final"].tolist()) if not registry_rows.empty else []
        if manual_nas_paths:
            nas_titles = unique(nas_titles + [folder_title_from_path(path) for path in manual_nas_paths])
            nas_aliases = unique(nas_aliases + [folder_author_from_path(path) for path in manual_nas_paths] + [canonical_author])
            nas_match_statuses = unique(nas_match_statuses + ["manual_nas_override"])

        if all_cp_rows.empty:
            cp_rows = pd.DataFrame()
        else:
            cp_rows = all_cp_rows[
                all_cp_rows.apply(
                    lambda record: row_matches_names(record.to_dict(), related_names),
                    axis=1,
                )
            ].copy()

        account_codes = unique(cp_rows["작가코드"].tolist()) if not cp_rows.empty else []
        direct_detail_matches = detail_df[
            detail_df.apply(lambda record: detail_matches_names(record.to_dict(), related_names), axis=1)
        ].copy()
        if account_codes:
            cp_code_matches = detail_df[detail_df["작가코드"].isin(account_codes)].copy()
            detail_matches = pd.concat([cp_code_matches, direct_detail_matches], ignore_index=True, sort=False).fillna("")
        else:
            detail_matches = direct_detail_matches
        detail_matches = detail_matches.drop_duplicates(subset=["작가코드"], keep="first") if not detail_matches.empty else detail_matches

        detail_codes = unique(detail_matches["작가코드"].tolist()) if not detail_matches.empty else []
        all_account_codes = unique(account_codes + detail_codes)
        contract_managers = unique(detail_matches["contract_manager"].tolist()) if not detail_matches.empty else []
        current_operating, final_status, review_priority = classify_row(
            index_flag=index_flag,
            nas_flag=nas_flag,
            account_codes=all_account_codes,
            contract_managers=contract_managers,
            manager=manager,
        )
        if canonical_author in force_confirm_authors and nas_flag and all_account_codes:
            current_operating = "Y"
            final_status = "확정"
            review_priority = ""

        unresolved_variants = [
            row["index_f_작가표기"]
            for row in resolve_rows
            if row["canonical_작가"] == canonical_author and row["이름변형_규칙"] == "unresolved_keep"
        ]

        notes: list[str] = []
        if unresolved_variants:
            notes.append(f"index 이름 검토: {pipe_join(unresolved_variants)}")
        if not all_account_codes:
            notes.append("account 코드 없음")
        if contract_managers and manager not in contract_managers:
            notes.append(f"account 담당자: {pipe_join(contract_managers)}")
        if index_flag and not nas_flag:
            notes.append("index_f에는 있으나 현재 NAS 담당 목록에는 없음")
        if nas_flag and not index_flag:
            notes.append("현재 NAS 담당 목록에는 있으나 index_f 담당자찾기에는 없음")
        if manual_nas_paths:
            notes.append(f"수동 NAS 확인: {pipe_join([folder_name_from_path(path) for path in manual_nas_paths])}")
        if canonical_author in force_confirm_authors and nas_flag and all_account_codes:
            notes.append("수동 확정")

        account_sources = unique(cp_rows["_source"].tolist()) if not cp_rows.empty else []
        cp_names = unique(cp_rows["저작권자명_필명"].tolist()) if not cp_rows.empty else []

        ssot_rows.append(
            {
                "대표작가명": canonical_author,
                "현재운영기준(NAS)": current_operating,
                "판정": final_status,
                "검토우선순위": review_priority,
                "index_f_조원재": "Y" if index_flag else "",
                "index_f_작가표기": pipe_join(index_variants),
                "index_f_이름변형규칙": pipe_join(
                    [
                        row["이름변형_규칙"]
                        for row in resolve_rows
                        if row["canonical_작가"] == canonical_author
                    ]
                ),
                "NAS_조원재": "Y" if nas_flag else "",
                "NAS_작품수": len(nas_titles),
                "NAS_작품목록": pipe_join(nas_titles),
                "NAS_작가표기": pipe_join(nas_aliases),
                "NAS_match_status": pipe_join(nas_match_statuses),
                "account_작가코드": pipe_join(all_account_codes),
                "account_소스": pipe_join(account_sources),
                "account_저작권자명_필명": pipe_join(cp_names),
                "account_pen_name": pipe_join(detail_matches["pen_name"].tolist() if not detail_matches.empty else []),
                "account_real_name": pipe_join(detail_matches["real_name"].tolist() if not detail_matches.empty else []),
                "account_예금주": pipe_join(detail_matches["예금주"].tolist() if not detail_matches.empty else []),
                "account_contract_manager": pipe_join(contract_managers),
                "account_조원재여부": "Y" if manager in contract_managers else ("N" if contract_managers else ""),
                "수동동치묶음": pipe_join(manual_variants),
                "관련이름묶음": pipe_join(related_names),
                "비고": " / ".join(notes),
            }
        )

    ssot_df = pd.DataFrame(ssot_rows)
    operating_rank = {"Y": 0, "보류": 1, "": 2}
    status_rank = {
        "확정": 0,
        "확정(account 타담당 주의)": 1,
        "확정(account 미연결)": 2,
        "검토(NAS만)": 3,
        "검토(index만)": 4,
        "기타": 5,
    }
    ssot_df["_operating_rank"] = ssot_df["현재운영기준(NAS)"].map(lambda value: operating_rank.get(normalize_text(value), 9))
    ssot_df["_status_rank"] = ssot_df["판정"].map(lambda value: status_rank.get(normalize_text(value), 9))
    ssot_df = ssot_df.sort_values(
        by=["_operating_rank", "_status_rank", "대표작가명"],
        ascending=[True, True, True],
        kind="stable",
    ).reset_index(drop=True)
    ssot_df = ssot_df.drop(columns=["_operating_rank", "_status_rank"])

    review_df = ssot_df[
        (ssot_df["검토우선순위"] != "")
        | (~ssot_df["판정"].eq("확정"))
    ].copy()

    resolve_df = pd.DataFrame(resolve_rows).sort_values(
        by=["canonical_작가", "index_f_작가표기"],
        kind="stable",
    )
    raw_index_df = index_df[["작가", "담당자"]].copy()

    return ssot_df, review_df, resolve_df, raw_index_df


def write_support_csvs(
    *,
    target_dir: Path,
    ssot_df: pd.DataFrame,
    review_df: pd.DataFrame,
    resolve_df: pd.DataFrame,
    raw_index_df: pd.DataFrame,
) -> dict[str, Path]:
    target_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "ssot": target_dir / "manager_author_ssot.csv",
        "review": target_dir / "review_queue.csv",
        "resolve": target_dir / "name_variant_map.csv",
        "index": target_dir / "source_index_f_조원재.csv",
    }
    ssot_df.to_csv(paths["ssot"], index=False, encoding="utf-8-sig")
    review_df.to_csv(paths["review"], index=False, encoding="utf-8-sig")
    resolve_df.to_csv(paths["resolve"], index=False, encoding="utf-8-sig")
    raw_index_df.to_csv(paths["index"], index=False, encoding="utf-8-sig")
    return paths


def apply_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def apply_body_style(cell, fill: PatternFill | None = None, wrap: bool = False) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    if fill is not None:
        cell.fill = fill


def set_column_widths(ws) -> None:
    max_widths: dict[int, int] = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            value = normalize_text(cell.value)
            if not value:
                continue
            width = min(max(len(value), 8), 60)
            max_widths[cell.column] = max(max_widths[cell.column], width)
    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def write_df_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, *, freeze_cell: str = "A2") -> None:
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = freeze_cell
    ws.sheet_view.showGridLines = False
    rows = [list(df.columns)] + df.astype(object).where(pd.notnull(df), "").values.tolist()
    for row_idx, row_values in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                apply_header_style(cell)
            else:
                fill = None
                if sheet_name == "조원재_담당작가_SSOT":
                    status_value = normalize_text(ws.cell(row=row_idx, column=3).value)
                    if status_value == "확정":
                        fill = _GOOD_FILL
                    elif "검토" in status_value:
                        fill = _WARN_FILL
                    elif "주의" in status_value or "미연결" in status_value:
                        fill = _BAD_FILL
                apply_body_style(cell, fill=fill, wrap=True)
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24
    set_column_widths(ws)


def write_summary_sheet(
    wb: Workbook,
    *,
    manager: str,
    ssot_df: pd.DataFrame,
    review_df: pd.DataFrame,
    raw_index_author_count: int,
    source_paths: dict[str, Path],
) -> None:
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{manager} 담당작가 SSOT"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "지표"
    ws["B3"] = "값"
    apply_header_style(ws["A3"])
    apply_header_style(ws["B3"])

    metrics = [
        ("index_f 조원재 원표기 수", raw_index_author_count),
        ("index_f 조원재 canonical 수", int((ssot_df["index_f_조원재"] == "Y").sum())),
        ("현재 NAS 조원재 작가 수", int((ssot_df["NAS_조원재"] == "Y").sum())),
        ("확정 작가 수", int((ssot_df["판정"] == "확정").sum())),
        ("검토 필요 작가 수", len(review_df)),
        ("account 연결 작가 수", int(ssot_df["account_작가코드"].astype(str).str.strip().ne("").sum())),
        ("account 담당자=조원재 작가 수", int((ssot_df["account_조원재여부"] == "Y").sum())),
    ]
    for row_idx, (label, value) in enumerate(metrics, start=4):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2))

    start_row = 12
    ws.cell(row=start_row, column=1, value="판정")
    ws.cell(row=start_row, column=2, value="건수")
    apply_header_style(ws.cell(row=start_row, column=1))
    apply_header_style(ws.cell(row=start_row, column=2))

    status_counts = Counter(ssot_df["판정"].tolist())
    ordered_statuses = [
        "확정",
        "확정(account 타담당 주의)",
        "확정(account 미연결)",
        "검토(NAS만)",
        "검토(index만)",
    ]
    status_rows: list[tuple[str, int]] = []
    for status in ordered_statuses:
        if status_counts.get(status):
            status_rows.append((status, status_counts[status]))
    for status, count in status_counts.items():
        if status not in ordered_statuses:
            status_rows.append((status, count))

    for offset, (status, count) in enumerate(status_rows, start=1):
        row_idx = start_row + offset
        ws.cell(row=row_idx, column=1, value=status)
        ws.cell(row=row_idx, column=2, value=count)
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2))

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "판정 분포"
    chart.y_axis.title = "작가 수"
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(status_rows))
    labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(status_rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, "D3")

    path_row = start_row + len(status_rows) + 3
    ws.cell(row=path_row, column=1, value="주요 입력 경로")
    apply_header_style(ws.cell(row=path_row, column=1))
    apply_header_style(ws.cell(row=path_row, column=2))
    ws.cell(row=path_row, column=2, value="path")
    for offset, (label, path) in enumerate(source_paths.items(), start=1):
        row_idx = path_row + offset
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=str(path))
        apply_body_style(ws.cell(row=row_idx, column=1))
        apply_body_style(ws.cell(row=row_idx, column=2), wrap=True)

    ws.freeze_panes = "A4"
    set_column_widths(ws)


def build_workbook(
    *,
    manager: str,
    ssot_df: pd.DataFrame,
    review_df: pd.DataFrame,
    resolve_df: pd.DataFrame,
    raw_index_df: pd.DataFrame,
    target_path: Path,
    source_paths: dict[str, Path],
) -> Path:
    wb = Workbook()
    write_summary_sheet(
        wb,
        manager=manager,
        ssot_df=ssot_df,
        review_df=review_df,
        raw_index_author_count=int(raw_index_df["작가"].nunique()),
        source_paths=source_paths,
    )
    write_df_sheet(wb, f"{manager}_담당작가_SSOT", ssot_df)
    write_df_sheet(wb, "검토필요", review_df)
    write_df_sheet(wb, "이름변형맵", resolve_df)
    write_df_sheet(wb, "원본_담당자찾기", raw_index_df)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(target_path)
    except PermissionError as exc:
        raise RuntimeError(
            f"최신판 1개 정책을 유지하려면 '{target_path.name}' 파일을 닫은 뒤 다시 실행해야 합니다."
        ) from exc

    for stale_path in target_path.parent.glob(f"{target_path.stem}__*{target_path.suffix}"):
        stale_path.unlink(missing_ok=True)
    return target_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build manager-author SSOT inputs used by the account/crosswalk pipeline. "
            "Requires index_f.xlsx and Barobook export files under ops/SIAAN Project."
        ),
    )
    parser.add_argument("--manager", default=MANAGER)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    manager = args.manager
    TARGET_ROOT.mkdir(parents=True, exist_ok=True)
    ssot_df, review_df, resolve_df, raw_index_df = build_datasets(manager)

    support_paths = write_support_csvs(
        target_dir=TARGET_ROOT,
        ssot_df=ssot_df,
        review_df=review_df,
        resolve_df=resolve_df,
        raw_index_df=raw_index_df,
    )

    workbook_path = TARGET_ROOT / f"{manager}_담당작가_ssot.xlsx"
    source_paths = {
        "index_f.xlsx": INDEX_PATH,
        "manual_groups.json": MANUAL_GROUPS_PATH,
        "work_cid_registry.local.csv": REGISTRY_PATH,
        "matched.xlsx": latest_path(f"*_{manager}_matched.xlsx"),
        "author_details.csv": latest_path(f"*__NAS_IPS_ACCOUNT_예금주_{manager}_author_details.csv"),
        "support_csv_ssot": support_paths["ssot"],
    }
    saved_workbook_path = build_workbook(
        manager=manager,
        ssot_df=ssot_df,
        review_df=review_df,
        resolve_df=resolve_df,
        raw_index_df=raw_index_df,
        target_path=workbook_path,
        source_paths=source_paths,
    )

    summary = {
        "manager": manager,
        "index_author_labels": int(raw_index_df["작가"].nunique()),
        "index_authors_canonical": int((ssot_df["index_f_조원재"] == "Y").sum()),
        "nas_authors": int((ssot_df["NAS_조원재"] == "Y").sum()),
        "confirmed_authors": int((ssot_df["판정"] == "확정").sum()),
        "review_rows": int(len(review_df)),
        "workbook": str(saved_workbook_path),
    }
    (TARGET_ROOT / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print("=== manager author SSOT built ===")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
