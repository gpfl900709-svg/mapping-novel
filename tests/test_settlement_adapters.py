from __future__ import annotations

import csv
import io
import os
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from settlement_adapters import (
    REGISTRY,
    _file_status,
    adapter_audit_dataframe,
    adapter_blocking_messages,
    normalize_settlement,
    summarize_normalization,
)


DEFAULT_SOURCE_ROOT = Path(r"\\172.16.10.120\소설사업부\판무팀_ssot\100_계산서_매출등록_자료")
DOC_DIR = Path(__file__).resolve().parents[1] / "doc" / "2026-05-07"


def source_root() -> Path:
    return Path(os.environ.get("SETTLEMENT_SOURCE_ROOT", str(DEFAULT_SOURCE_ROOT)))


class SettlementAdapterRegistryTest(unittest.TestCase):
    def test_registry_covers_all_survey_platforms(self) -> None:
        self.assertEqual(len(REGISTRY), 37)

        blocked = [spec.platform for spec in REGISTRY.values() if spec.blocks_default_feed]
        self.assertEqual(sorted(blocked), ["보인&국립장애인도서관", "알라딘 종이책"])

    def test_amount_policy_lock_is_explicit(self) -> None:
        locked = [spec.platform for spec in REGISTRY.values() if spec.s2_amount_policy_locked]

        self.assertIn("북큐브", locked)
        self.assertIn("피우리(누온)", locked)
        self.assertNotIn("무툰", locked)
        self.assertNotIn("카카오", locked)

        self.assertGreaterEqual(len(locked), 10)
        self.assertLess(len(locked), len(REGISTRY) - 2)

    def test_human_processed_filename_is_not_blocked_by_name_alone(self) -> None:
        spec = REGISTRY["미소설"]

        self.assertEqual(_file_status(spec, "2026년 2월 미소설 사람가공 정산상세.xlsx"), "include")

    def test_romantique_does_not_promote_isbn_to_external_id(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "styleB(바로북)fixture"
        sheet.append(["도서명", "저자", "판매액", "정산액", "isbn"])
        sheet.append(["헌터 외 사망 금지 1권", "파인애플덤플링", 3300, 1815, "979-11-7530-702-5"])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = normalize_settlement(payload, platform="로망띠끄", source_name="로망띠끄_fixture.xlsx")
        feed = result.to_mapping_feed()

        self.assertEqual(len(feed), 1)
        self.assertEqual(feed["외부콘텐츠ID"].iloc[0], "")
        self.assertNotIn("979-11-7530-702-5", feed.to_csv(index=False))

    def test_naver_repairs_24_7_when_excel_coerces_title_to_date(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "contentsSelling_fixture"
        sheet.append(["컨텐츠", "작가명", "공급자코드", "합계", "정산금액", "마켓수수료"])
        sheet.append([datetime(2026, 7, 24), "이내리", "NV-247", 1000, 700, 300])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = normalize_settlement(payload, platform="네이버", source_name="네이버_연재_fixture.xlsx")
        feed = result.to_mapping_feed()

        self.assertEqual(len(feed), 1)
        self.assertEqual(feed["상품명"].iloc[0], "24/7")
        self.assertEqual(result.rows["정제_상품명"].iloc[0], "24/7")

    def test_naver_uses_header_fallback_when_sheet_name_changes(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "2026년 5월 네이버 일반 정산상세"
        sheet.append(["이용권 포함 컨텐츠별 매출 통계"])
        sheet.append([])
        sheet.append(["컨텐츠", "컨텐츠No", "공급자코드", "CP명", "작가명", "합계", "마켓수수료(추정치)"])
        sheet.append([None, None, None, None, None, None, None])
        sheet.append(["테스트 작품", 113935, "87997", "바로북_일반도서", "테스트 작가", 7200, 0])
        sheet.append(["합계", None, None, None, None, 7200, 0])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = normalize_settlement(payload, platform="네이버", source_name="2026년 5월 네이버 일반 정산상세.xlsx")
        feed = result.to_mapping_feed()
        audit = adapter_audit_dataframe(result)

        self.assertEqual(len(feed), 1)
        self.assertEqual(feed["상품명"].iloc[0], "테스트 작품")
        self.assertEqual(feed["외부콘텐츠ID"].iloc[0], "113935")
        self.assertEqual(feed["판매금액_후보"].iloc[0], 7200)
        self.assertEqual(adapter_blocking_messages(result), [])
        self.assertIn("parsed_sheet_rule_fallback", audit["status"].tolist())

    def test_hana_areum_summary_rows_are_excluded(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "한아름_fixture"
        sheet.append(["작품명", "작가명", "BOOK NO", "건당 로그 금액"])
        sheet.append(["한아름 성인 합계", "", "", 1000])
        sheet.append(["한아름 web", "", "", 1000])
        sheet.append(["아름북스 inapp", "", "", 1000])
        sheet.append(["진짜 작품", "작가", "BOOK-1", 100])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = normalize_settlement(payload, platform="한아름", source_name="한아름_fixture.xlsx")
        feed = result.to_mapping_feed()

        self.assertEqual(len(feed), 1)
        self.assertEqual(feed["상품명"].iloc[0], "진짜 작품")
        self.assertEqual(result.rows["정제_상품명"].iloc[0], "진짜작품")

    def test_alltoon_uses_header_fallback_when_sheet_name_changes(self) -> None:
        workbook = Workbook()
        invoice = workbook.active
        invoice.title = "계산서 발행"
        invoice.append(["■ 전자(세금)계산서 발행 정보"])
        invoice.append(["웹소설", "발행금액", 168033.6])
        detail = workbook.create_sheet("전체 매출 내역")
        detail.append(
            [
                "매출 기간(월)",
                "작품명",
                "정산 수수료율(%)",
                "코인 사용수량(개)",
                "총 매출액(원)",
                "앱스토어 수수료(원)",
                "올웨이즈 수수료(원)",
                "순매출액",
                "정산대상금액",
            ]
        )
        detail.append(["2026년 5월", "[소설]테스트 작품", 60, 100, 120, 36, 33.6, 84, 50.4])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = normalize_settlement(payload, platform="올툰", source_name="2026년 5월 올툰 정산상세.xlsx")
        feed = result.to_mapping_feed()
        audit = adapter_audit_dataframe(result)

        self.assertEqual(len(feed), 1)
        self.assertEqual(feed["상품명"].iloc[0], "[소설]테스트 작품")
        self.assertEqual(feed["판매금액_후보"].iloc[0], 120)
        self.assertAlmostEqual(feed["정산기준액_후보"].iloc[0], 50.4)
        self.assertAlmostEqual(feed["상계금액_후보"].iloc[0], 69.6)
        self.assertEqual(adapter_blocking_messages(result), [])
        self.assertTrue(summarize_normalization(result)["s2_amount_policy_locked"])
        self.assertIn("parsed_sheet_rule_fallback", audit["status"].tolist())
        self.assertIn("excluded_sheet", audit["status"].tolist())

    def test_alltoon_accepts_april_settlement_amount_header(self) -> None:
        workbook = Workbook()
        detail = workbook.active
        detail.title = "전체 매출 내역"
        detail.append(["CP사명: 키다리_소설"])
        detail.append(
            [
                "작품명",
                "정산 수수료율(%)",
                "코인 사용수량(개)",
                "총 매출액(원)",
                "앱스토어 수수료(원)",
                "올웨이즈 수수료(원)",
                "순매출액(원)",
                "정산 대상 금액(수수료 제)",
            ]
        )
        detail.append(["[소설]테스트 작품", 60, 100, 120, 36, 33.6, 84, 50.4])
        detail.append(["합계", "", 100, 120, 36, 33.6, 84, 50.4])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = normalize_settlement(payload, platform="올툰", source_name="2026년 4월 올툰 정산상세.xlsx")
        feed = result.to_mapping_feed()

        self.assertEqual(len(feed), 1)
        self.assertEqual(feed["상품명"].iloc[0], "[소설]테스트 작품")
        self.assertEqual(feed["판매금액_후보"].iloc[0], 120)
        self.assertAlmostEqual(feed["정산기준액_후보"].iloc[0], 50.4)
        self.assertAlmostEqual(feed["상계금액_후보"].iloc[0], 69.6)


class SettlementAdapterFixtureTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.root = source_root()
        if not cls.root.exists():
            raise unittest.SkipTest(f"정산상세 원본 루트가 없습니다: {cls.root}")

    def test_latest_fixture_files_normalize_to_expected_data_rows(self) -> None:
        manifest = DOC_DIR / "latest_origin_substitution_materialization_test_files.csv"
        with manifest.open(encoding="utf-8-sig", newline="") as handle:
            fixtures = list(csv.DictReader(handle))

        self.assertEqual(len(fixtures), 53)
        for fixture in fixtures:
            with self.subTest(platform=fixture["platform"], relative_path=fixture["relative_path"]):
                path = self.root / fixture["relative_path"]
                if not path.exists():
                    self.skipTest(f"정산상세 fixture 파일이 없습니다: {path}")
                result = normalize_settlement(path, platform=fixture["platform"], source_name=fixture["relative_path"])
                summary = summarize_normalization(result)

                expected_rows = int(fixture["parsed_rows"] or 0)
                if fixture["platform"] == "네이버" and fixture["counted_in_default_feed"] == "True":
                    # 기존 survey parsed_rows는 네이버 합계 행을 포함했다. 운영 어댑터는 feed에서 합계 행을 제외한다.
                    expected_rows -= 1

                self.assertEqual(summary["parsed_rows"], expected_rows)
                self.assertEqual(summary["title_present_rows"], expected_rows)

                if fixture["counted_in_default_feed"] == "True":
                    self.assertEqual(summary["default_feed_rows"], expected_rows)
                    self.assertEqual(adapter_blocking_messages(result), [])
                else:
                    self.assertEqual(summary["default_feed_rows"], 0)
                    self.assertTrue(adapter_blocking_messages(result))

                if expected_rows:
                    bad_titles = result.rows["상품명"].astype(str).str.strip().isin(["합계", "총 합계", "총 액"])
                    self.assertFalse(bad_titles.any())

    def test_bookcube_invalid_style_files_use_value_only_fallback(self) -> None:
        base = Path(__file__).resolve().parents[1] / "igignore" / "2026-02_정산상세_초기원형" / "북큐브" / "2월"
        fixtures = [
            base / "2026년 2월 북큐브(로맨스) 정산상세.xlsx",
            base / "2026년 북큐브(판무) 정산상세.xlsx",
        ]
        if not all(path.exists() for path in fixtures):
            raise unittest.SkipTest("북큐브 2월 스타일 오류 fixture가 없습니다.")

        for path in fixtures:
            with self.subTest(path=path.name):
                result = normalize_settlement(path, platform="북큐브", source_name=str(path))
                summary = summarize_normalization(result)

                self.assertGreater(summary["default_feed_rows"], 0)
                self.assertEqual(adapter_blocking_messages(result), [])
                self.assertTrue(summary["s2_amount_policy_locked"])


if __name__ == "__main__":
    unittest.main()
