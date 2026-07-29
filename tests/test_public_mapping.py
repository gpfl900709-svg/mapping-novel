from __future__ import annotations

import ast
import io
import unittest
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from mapping_core import build_mapping
from public_mapping import (
    PUBLIC_OUTPUT_COLUMNS,
    PUBLIC_REFERENCE_COLUMNS,
    PUBLIC_WORKBOOK_SHEETS,
    PublicMappingSecurityError,
    build_public_reference_frame,
    build_public_zip,
    export_public_mapping,
    project_public_mapping_result,
    upload_signature,
    validate_public_reference_frame,
    validate_xlsx_archive,
)
from scripts.build_public_deploy_bundle import (
    FORBIDDEN_BUNDLE_NAMES,
    PUBLIC_DEPLOY_FILES,
    resolve_safe_output,
)


REPO_ROOT = Path(__file__).resolve().parents[1]


class PublicMappingSecurityTest(unittest.TestCase):
    def test_public_reference_keeps_only_mapping_columns(self) -> None:
        source = pd.DataFrame(
            {
                "콘텐츠명": ["테스트 작품"],
                "판매채널콘텐츠ID": ["1001"],
                "콘텐츠ID": ["2001"],
                "판매채널명": ["리디북스(소설)"],
                "담당자명": ["내부 담당자"],
                "담당부서명": ["내부 부서"],
                "통합계약ID": ["CONTRACT-1"],
            }
        )

        public = build_public_reference_frame(source)

        self.assertEqual(tuple(public.columns), PUBLIC_REFERENCE_COLUMNS)
        self.assertNotIn("담당자명", public.columns)
        self.assertNotIn("통합계약ID", public.columns)

    def test_public_reference_rejects_schema_drift_and_direct_contacts(self) -> None:
        bad_schema = pd.DataFrame(
            [["작품", "1001", "2001", "리디북스(소설)", "담당자"]],
            columns=[*PUBLIC_REFERENCE_COLUMNS, "담당자명"],
        )
        with self.assertRaises(PublicMappingSecurityError):
            validate_public_reference_frame(bad_schema)

        direct_contact = pd.DataFrame(
            [["문의 test@example.com", "1001", "2001", "리디북스(소설)"]],
            columns=PUBLIC_REFERENCE_COLUMNS,
        )
        with self.assertRaises(PublicMappingSecurityError):
            validate_public_reference_frame(direct_contact)

        disguised_phone = pd.DataFrame(
            [["작품", "01012345678", "2001", "리디북스(소설)"]],
            columns=PUBLIC_REFERENCE_COLUMNS,
        )
        with self.assertRaises(PublicMappingSecurityError):
            validate_public_reference_frame(disguised_phone)

    def test_removing_assignee_and_department_does_not_change_match(self) -> None:
        full_reference = pd.DataFrame(
            {
                "콘텐츠명": ["그 남자의 비밀"],
                "판매채널콘텐츠ID": ["1001"],
                "콘텐츠ID": ["2001"],
                "판매채널명": ["리디북스(소설)"],
                "담당자명": ["내부 담당자"],
                "담당부서명": ["내부 부서"],
            }
        )
        settlement = pd.DataFrame({"상품명": ["그 남자의 비밀 1화"]})

        full_rows = build_mapping(full_reference, settlement, None).rows
        public_rows = build_mapping(build_public_reference_frame(full_reference), settlement, None).rows

        comparable = ["S2_매칭상태", "S2_판매채널콘텐츠ID", "S2_콘텐츠ID", "S2_콘텐츠명"]
        self.assertEqual(full_rows.loc[0, comparable].tolist(), public_rows.loc[0, comparable].tolist())

    def test_public_projection_is_an_exact_allowlist(self) -> None:
        reference = pd.DataFrame(
            {
                "콘텐츠명": ["테스트 작품"],
                "판매채널콘텐츠ID": ["1001"],
                "콘텐츠ID": ["2001"],
                "판매채널명": ["리디북스(소설)"],
                "담당자명": ["내부 담당자"],
                "담당부서명": ["내부 부서"],
            }
        )
        settlement = pd.DataFrame(
            {
                "상품명": ["테스트 작품"],
                "정산금액": [1000],
                "연락처": ["010-1234-5678"],
            }
        )

        projected = project_public_mapping_result(build_mapping(reference, settlement, None))

        self.assertEqual(tuple(projected.rows.columns), PUBLIC_OUTPUT_COLUMNS)
        self.assertNotIn("S2_담당자명", projected.rows.columns)
        self.assertNotIn("정산서원본_연락처", projected.rows.columns)
        self.assertNotIn("S2_후보ID목록", projected.rows.columns)

    def test_public_workbook_and_zip_have_only_safe_outputs(self) -> None:
        reference = pd.DataFrame(
            {
                "콘텐츠명": ["테스트 작품"],
                "판매채널콘텐츠ID": ["1001"],
                "콘텐츠ID": ["2001"],
                "판매채널명": ["리디북스(소설)"],
                "담당자명": ["내부 담당자"],
                "담당부서명": ["내부 부서"],
            }
        )
        settlement = pd.DataFrame({"상품명": ["테스트 작품"], "정산금액": [1000]})
        payload = export_public_mapping(build_mapping(reference, settlement, None))

        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        self.assertEqual(tuple(workbook.sheetnames), PUBLIC_WORKBOOK_SHEETS)
        result_headers = [cell.value for cell in next(workbook["행별매핑결과"].iter_rows())]
        self.assertEqual(tuple(result_headers), PUBLIC_OUTPUT_COLUMNS)
        workbook.close()

        archive_payload = build_public_zip([("테스트_매핑.xlsx", payload)])
        with zipfile.ZipFile(io.BytesIO(archive_payload)) as archive:
            self.assertEqual(archive.namelist(), ["테스트_매핑.xlsx"])

    def test_public_export_blocks_direct_contact_in_title(self) -> None:
        reference = pd.DataFrame(
            {
                "콘텐츠명": ["문의 test@example.com"],
                "판매채널콘텐츠ID": ["1001"],
                "콘텐츠ID": ["2001"],
                "판매채널명": ["리디북스(소설)"],
            }
        )
        settlement = pd.DataFrame({"상품명": ["문의 test@example.com"]})

        with self.assertRaises(PublicMappingSecurityError):
            export_public_mapping(build_mapping(reference, settlement, None))

    def test_public_export_neutralizes_excel_formula_prefixes(self) -> None:
        for title in ("=1+1", "+SUM(1,1)", "-2+3", "@SUM(1,1)"):
            with self.subTest(title=title):
                reference = pd.DataFrame(
                    {
                        "콘텐츠명": [title],
                        "판매채널콘텐츠ID": ["1001"],
                        "콘텐츠ID": ["2001"],
                        "판매채널명": ["리디북스(소설)"],
                    }
                )
                payload = export_public_mapping(
                    build_mapping(reference, pd.DataFrame({"상품명": [title]}), None)
                )
                workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
                cells = [
                    cell
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                ]
                self.assertFalse(any(cell.data_type == "f" for cell in cells))
                workbook.close()

    def test_xlsx_archive_guard_accepts_normal_and_rejects_high_ratio_member(self) -> None:
        normal = io.BytesIO()
        pd.DataFrame({"상품명": ["작품"]}).to_excel(normal, index=False)
        validate_xlsx_archive(normal.getvalue())

        crafted = io.BytesIO()
        with zipfile.ZipFile(crafted, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", "<Types/>")
            archive.writestr("xl/workbook.xml", "<workbook/>")
            archive.writestr("xl/sharedStrings.xml", "A" * (2 * 1024 * 1024))
        with self.assertRaises(PublicMappingSecurityError):
            validate_xlsx_archive(crafted.getvalue())

    def test_upload_signature_hashes_file_contents(self) -> None:
        first = upload_signature("채널", [("same.xlsx", b"AAAA")])
        second = upload_signature("채널", [("same.xlsx", b"BBBB")])
        self.assertNotEqual(first, second)

    def test_bundle_output_is_restricted_to_codex_tmp_child(self) -> None:
        safe = resolve_safe_output(REPO_ROOT / ".codex_tmp" / "public-test")
        self.assertEqual(safe.name, "public-test")
        for unsafe in (REPO_ROOT / "data", REPO_ROOT / "scripts", REPO_ROOT / ".git", REPO_ROOT / ".codex_tmp"):
            with self.subTest(path=unsafe):
                with self.assertRaises(RuntimeError):
                    resolve_safe_output(unsafe)

    def test_public_entrypoint_has_no_internal_connectors_or_secret_reads(self) -> None:
        source = (REPO_ROOT / "app.py").read_text(encoding="utf-8")
        tree = ast.parse(source)
        imported: set[str] = set()
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                imported.update(alias.name for alias in node.names)
            elif isinstance(node, ast.ImportFrom) and node.module:
                imported.add(node.module)

        forbidden = {
            "notion_tasks",
            "github_notifications",
            "s2_auth",
            "s2_direct_refresh",
            "s2_reference_guards",
            "kiss_payment_settlement",
        }
        self.assertTrue(forbidden.isdisjoint(imported))
        self.assertNotIn("st.secrets", source)
        self.assertNotIn("create_mapping_run_audit_request", source)

    def test_public_deploy_bundle_is_an_exact_file_allowlist(self) -> None:
        names = {Path(value).name.casefold() for value in PUBLIC_DEPLOY_FILES}

        self.assertNotIn("internal_app.py", names)
        self.assertTrue({name.casefold() for name in FORBIDDEN_BUNDLE_NAMES}.isdisjoint(names))
        self.assertIn("public_s2_mapping_reference.csv", names)
        self.assertNotIn("work_order_reports.py", names)

    def test_public_entrypoint_reduces_mapping_feed_to_title_only(self) -> None:
        source = (REPO_ROOT / "app.py").read_text(encoding="utf-8")
        self.assertIn('normalized.to_mapping_feed().loc[:, ["상품명"]].copy()', source)


if __name__ == "__main__":
    unittest.main()
