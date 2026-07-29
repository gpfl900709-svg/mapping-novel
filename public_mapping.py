from __future__ import annotations

import io
import hashlib
import re
import zipfile
from dataclasses import replace
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

from cleaning_rules import text
from mapping_core import MappingResult


PUBLIC_REFERENCE_COLUMNS = (
    "콘텐츠명",
    "판매채널콘텐츠ID",
    "콘텐츠ID",
    "판매채널명",
)

PUBLIC_OUTPUT_COLUMNS = (
    "정산서_원본행번호",
    "정산서_콘텐츠명",
    "정제_상품명",
    "S2_매칭상태",
    "S2_판매채널콘텐츠ID",
    "S2_콘텐츠ID",
    "S2_콘텐츠명",
    "검토필요사유",
    "검토필요(Y/N)",
)

PUBLIC_WORKBOOK_SHEETS = (
    "요약",
    "입력검증",
    "행별매핑결과",
    "검토필요",
)

FORBIDDEN_COLUMN_MARKERS = (
    "담당자",
    "담당부서",
    "이메일",
    "email",
    "메일",
    "전화",
    "휴대폰",
    "연락처",
    "계약",
    "정산마스터",
    "정산상세",
    "후보id목록",
    "후보콘텐츠명목록",
    "근거",
    "거래처",
)

EMAIL_PATTERN = re.compile(r"(?i)\b[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}\b")
PHONE_PATTERN = re.compile(
    r"(?<!\d)(?:\+?82[-.\s]?)?0(?:1[016789]|2|[3-6][1-5])(?:[-.\s]?\d{3,4})[-.\s]?\d{4}(?!\d)"
)
FORMULA_PREFIXES = ("=", "+", "-", "@")
PUBLIC_IDENTIFIER_COLUMNS = ("판매채널콘텐츠ID", "콘텐츠ID")
MAX_XLSX_MEMBERS = 5_000
MAX_XLSX_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 128 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200
MIN_RATIO_CHECK_BYTES = 1024 * 1024


class PublicMappingSecurityError(ValueError):
    pass


def _normalized_column(column: object) -> str:
    return re.sub(r"[\s_\-]+", "", text(column)).casefold()


def forbidden_columns(columns: Iterable[object]) -> list[str]:
    blocked: list[str] = []
    for column in columns:
        normalized = _normalized_column(column)
        if any(marker in normalized for marker in FORBIDDEN_COLUMN_MARKERS):
            blocked.append(text(column))
    return blocked


def find_direct_contact_values(frame: pd.DataFrame, *, limit: int = 10) -> list[str]:
    hits: list[str] = []
    for column in frame.columns:
        normalized_column = _normalized_column(column)
        phone_scan_allowed = "id" not in normalized_column and "번호" not in normalized_column
        for value in frame[column].dropna():
            value_text = text(value)
            if not value_text:
                continue
            if EMAIL_PATTERN.search(value_text) or (phone_scan_allowed and PHONE_PATTERN.search(value_text)):
                hits.append(f"{column}")
                if len(hits) >= limit:
                    return hits
    return hits


def build_public_reference_frame(source: pd.DataFrame) -> pd.DataFrame:
    missing = [column for column in PUBLIC_REFERENCE_COLUMNS if column not in source.columns]
    if missing:
        raise PublicMappingSecurityError(f"공개 기준 필수 컬럼이 없습니다: {', '.join(missing)}")
    public = source.loc[:, list(PUBLIC_REFERENCE_COLUMNS)].copy()
    public = public.fillna("")
    public = public.drop_duplicates().reset_index(drop=True)
    validate_public_reference_frame(public)
    return public


def validate_public_reference_frame(frame: pd.DataFrame) -> None:
    actual = tuple(map(str, frame.columns))
    if actual != PUBLIC_REFERENCE_COLUMNS:
        raise PublicMappingSecurityError(
            f"공개 기준 스키마가 허용목록과 다릅니다. expected={PUBLIC_REFERENCE_COLUMNS}, actual={actual}"
        )
    blocked = forbidden_columns(frame.columns)
    if blocked:
        raise PublicMappingSecurityError(f"공개 기준에 금지 컬럼이 있습니다: {', '.join(blocked)}")
    for column in PUBLIC_IDENTIFIER_COLUMNS:
        values = frame[column].map(text)
        invalid = values.loc[~values.str.fullmatch(r"\d+")]
        if not invalid.empty:
            raise PublicMappingSecurityError(f"공개 기준의 {column}는 숫자 ID만 허용됩니다.")
        phone_like = values.loc[values.str.fullmatch(r"0\d{9,10}")]
        if not phone_like.empty:
            raise PublicMappingSecurityError(f"공개 기준의 {column}에서 전화번호 형태 값을 찾았습니다.")
    contact_hits = find_direct_contact_values(frame)
    if contact_hits:
        raise PublicMappingSecurityError(
            f"공개 기준에서 이메일 또는 전화번호 패턴을 찾았습니다. 컬럼: {', '.join(contact_hits)}"
        )


def load_public_reference(path: str | Path) -> pd.DataFrame:
    frame = pd.read_csv(path, dtype=object, keep_default_na=False)
    validate_public_reference_frame(frame)
    return frame


def project_public_mapping_result(result: MappingResult) -> MappingResult:
    missing = [column for column in PUBLIC_OUTPUT_COLUMNS if column not in result.rows.columns]
    if missing:
        raise PublicMappingSecurityError(f"매핑 결과에 공개 필수 컬럼이 없습니다: {', '.join(missing)}")
    rows = result.rows.loc[:, list(PUBLIC_OUTPUT_COLUMNS)].copy()
    blocked = forbidden_columns(rows.columns)
    if blocked:
        raise PublicMappingSecurityError(f"공개 결과에 금지 컬럼이 있습니다: {', '.join(blocked)}")
    review_rows = rows.loc[rows["검토필요(Y/N)"].eq("Y")].copy()
    input_validation = result.input_validation.copy()
    if forbidden_columns(input_validation.columns):
        raise PublicMappingSecurityError("입력검증 결과의 컬럼이 공개 스키마를 위반했습니다.")
    return replace(
        result,
        rows=rows,
        review_rows=review_rows,
        duplicate_candidates=pd.DataFrame(),
        input_validation=input_validation,
    )


def _validate_public_frame_values(frame: pd.DataFrame, *, label: str) -> None:
    blocked = forbidden_columns(frame.columns)
    if blocked:
        raise PublicMappingSecurityError(f"{label}에 금지 컬럼이 있습니다: {', '.join(blocked)}")
    contact_hits = find_direct_contact_values(frame)
    if contact_hits:
        raise PublicMappingSecurityError(
            f"{label}에서 이메일 또는 전화번호 패턴을 찾았습니다. 컬럼: {', '.join(contact_hits)}"
        )


def _neutralize_formula_value(value: object) -> object:
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _neutralize_formula_frame(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.map(_neutralize_formula_value)


def export_public_mapping(result: MappingResult) -> bytes:
    public_result = project_public_mapping_result(result)
    _validate_public_frame_values(public_result.rows, label="공개 매핑 결과")
    _validate_public_frame_values(public_result.review_rows, label="공개 검토 결과")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _neutralize_formula_frame(public_result.summary).to_excel(writer, sheet_name="요약", index=False)
        _neutralize_formula_frame(public_result.input_validation).to_excel(
            writer, sheet_name="입력검증", index=False
        )
        _neutralize_formula_frame(public_result.rows).to_excel(writer, sheet_name="행별매핑결과", index=False)
        _neutralize_formula_frame(public_result.review_rows).to_excel(
            writer, sheet_name="검토필요", index=False
        )
        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            if worksheet.max_row and worksheet.max_column:
                worksheet.auto_filter.ref = worksheet.dimensions
    payload = buffer.getvalue()
    validate_public_workbook(payload)
    return payload


def validate_public_workbook(payload: bytes) -> None:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    try:
        if tuple(workbook.sheetnames) != PUBLIC_WORKBOOK_SHEETS:
            raise PublicMappingSecurityError(
                f"공개 엑셀 시트가 허용목록과 다릅니다: {tuple(workbook.sheetnames)}"
            )
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows()
            headers = tuple(text(cell.value) for cell in next(rows, ()))
            blocked = forbidden_columns(headers)
            if blocked:
                raise PublicMappingSecurityError(
                    f"{worksheet.title} 시트에 금지 컬럼이 있습니다: {', '.join(blocked)}"
                )
            for row in rows:
                for cell in row:
                    if cell.data_type == "f":
                        raise PublicMappingSecurityError(
                            f"{worksheet.title} 시트에서 실행 가능한 엑셀 수식을 찾았습니다."
                        )
                    value = cell.value
                    value_text = text(value)
                    if EMAIL_PATTERN.search(value_text) or PHONE_PATTERN.search(value_text):
                        raise PublicMappingSecurityError(
                            f"{worksheet.title} 시트에서 이메일 또는 전화번호 패턴을 찾았습니다."
                        )
    finally:
        workbook.close()


def validate_xlsx_archive(payload: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise PublicMappingSecurityError("XLSX 내부 파일 수가 안전 제한을 넘었습니다.")
            names = {member.filename for member in members}
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise PublicMappingSecurityError("유효한 XLSX 구조가 아닙니다.")
            total_uncompressed = 0
            for member in members:
                if member.flag_bits & 0x1:
                    raise PublicMappingSecurityError("암호화된 XLSX는 처리할 수 없습니다.")
                if member.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise PublicMappingSecurityError("XLSX 내부 파일 크기가 안전 제한을 넘었습니다.")
                total_uncompressed += member.file_size
                if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise PublicMappingSecurityError("XLSX 압축 해제 크기가 안전 제한을 넘었습니다.")
                if member.file_size >= MIN_RATIO_CHECK_BYTES:
                    ratio = member.file_size / max(member.compress_size, 1)
                    if ratio > MAX_XLSX_COMPRESSION_RATIO:
                        raise PublicMappingSecurityError("XLSX 압축률이 안전 제한을 넘었습니다.")
    except (zipfile.BadZipFile, OSError) as exc:
        raise PublicMappingSecurityError("유효한 XLSX ZIP 파일이 아닙니다.") from exc


def upload_signature(selected_channel: str, files: Iterable[tuple[str, bytes]]) -> str:
    digest = hashlib.sha256()
    digest.update(selected_channel.encode("utf-8"))
    for filename, payload in files:
        encoded_name = text(filename).encode("utf-8")
        digest.update(len(encoded_name).to_bytes(8, "big"))
        digest.update(encoded_name)
        digest.update(len(payload).to_bytes(8, "big"))
        digest.update(hashlib.sha256(payload).digest())
    return digest.hexdigest()


def build_public_zip(files: Iterable[tuple[str, bytes]]) -> bytes:
    buffer = io.BytesIO()
    seen: set[str] = set()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for filename, payload in files:
            safe_name = Path(filename).name
            if not safe_name.lower().endswith(".xlsx"):
                raise PublicMappingSecurityError(f"공개 ZIP에는 xlsx만 허용됩니다: {safe_name}")
            if safe_name in seen:
                raise PublicMappingSecurityError(f"공개 ZIP 파일명이 중복됩니다: {safe_name}")
            validate_public_workbook(payload)
            archive.writestr(safe_name, payload)
            seen.add(safe_name)
    return buffer.getvalue()
