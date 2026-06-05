from __future__ import annotations

import hashlib
import io
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import load_workbook

from cleaning_rules import text
from settlement_adapters import REGISTRY, _header_score, _sheet_in_scope


KST = timezone(timedelta(hours=9))
SCHEMA_VERSION = "adapter_failure.v1"
MAX_SNAPSHOT_ROWS = 10
MAX_HEADER_SCAN_ROWS = 100
MAX_SNAPSHOT_COLUMNS = 24
MAX_CELL_TEXT = 80


def read_uploaded_file_bytes(source: object) -> bytes:
    if source is None:
        return b""
    if isinstance(source, bytes):
        return source
    if isinstance(source, bytearray):
        return bytes(source)
    if hasattr(source, "getvalue"):
        payload = source.getvalue()
    elif hasattr(source, "read"):
        original_position = None
        try:
            original_position = source.tell()
        except Exception:
            original_position = None
        try:
            if hasattr(source, "seek"):
                source.seek(0)
            payload = source.read()
        finally:
            if original_position is not None and hasattr(source, "seek"):
                try:
                    source.seek(original_position)
                except Exception:
                    pass
    else:
        return b""
    if isinstance(payload, str):
        return payload.encode("utf-8")
    return bytes(payload or b"")


def sha256_hex(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _cell_text(value: object) -> str:
    raw = text(value).replace("\r", " ").replace("\n", " ")
    if len(raw) > MAX_CELL_TEXT:
        return raw[:MAX_CELL_TEXT - 1] + "…"
    return raw


def _json_safe(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and pd.isna(value):
            return ""
        return value
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    if pd.isna(value):
        return ""
    return text(value)


def _frame_records(value: object) -> list[dict[str, object]]:
    if isinstance(value, pd.DataFrame):
        return [
            {str(key): _json_safe(item) for key, item in row.items()}
            for row in value.fillna("").to_dict("records")
        ]
    if isinstance(value, list):
        return [dict(item) for item in value if isinstance(item, Mapping)]
    return []


def _safe_attr_payload(value: object, names: tuple[str, ...]) -> dict[str, object]:
    payload: dict[str, object] = {}
    for name in names:
        if hasattr(value, name):
            payload[name] = _json_safe(getattr(value, name))
    return payload


def collect_workbook_diagnostics(source_bytes: bytes, platform: str) -> dict[str, object]:
    if not source_bytes:
        return {"sheet_names": [], "sheets": [], "error": "source bytes unavailable"}

    try:
        workbook = load_workbook(io.BytesIO(source_bytes), data_only=True, read_only=True)
    except Exception as exc:
        return {"sheet_names": [], "sheets": [], "error": f"{type(exc).__name__}: {exc}"}

    spec = REGISTRY.get(platform)
    sheets: list[dict[str, object]] = []
    try:
        for sheet in workbook.worksheets:
            top_rows: list[list[str]] = []
            header_candidates: list[dict[str, object]] = []
            best_candidate: dict[str, object] | None = None
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = list(row)
                compact_values = [_cell_text(value) for value in values[:MAX_SNAPSHOT_COLUMNS]]
                if row_index <= MAX_SNAPSHOT_ROWS:
                    top_rows.append(compact_values)
                if spec is not None and row_index <= MAX_HEADER_SCAN_ROWS:
                    score = _header_score(values, spec)
                    if score > 0:
                        candidate = {
                            "row": row_index,
                            "score": score,
                            "cells": [cell for cell in compact_values if cell],
                        }
                        header_candidates.append(candidate)
                        if best_candidate is None or score > int(best_candidate.get("score", 0)):
                            best_candidate = candidate
                if row_index >= MAX_HEADER_SCAN_ROWS:
                    break

            sheet_payload: dict[str, object] = {
                "sheet": sheet.title,
                "in_scope": _sheet_in_scope(spec, sheet.title) if spec is not None else "",
                "top_rows": top_rows,
                "header_candidates": header_candidates[:10],
                "best_header_candidate": best_candidate or {},
            }
            sheets.append(sheet_payload)
    finally:
        workbook.close()

    return {"sheet_names": [str(sheet["sheet"]) for sheet in sheets], "sheets": sheets}


def classify_failure(result: Mapping[str, Any], workbook_diagnostics: Mapping[str, object]) -> str:
    platform = text(result.get("platform"))
    s2_channel = text(result.get("s2_sales_channel"))
    error = text(result.get("error"))
    blocking_messages = " | ".join(text(item) for item in result.get("blocking_messages", []) or [])
    combined_message = f"{error} | {blocking_messages}"
    audit_rows = _frame_records(result.get("audit_df"))
    audit_statuses = [text(row.get("status")) for row in audit_rows]

    if not platform or not s2_channel or "플랫폼을 감지하지 못했습니다" in combined_message:
        return "channel_detection_failed"
    if "금액 정책" in combined_message:
        return "amount_policy_unlocked"
    if any(status == "header_not_found" for status in audit_statuses) or "헤더를 찾지 못한 시트" in combined_message:
        return "header_not_found"
    if audit_statuses and all(status == "excluded_sheet" for status in audit_statuses):
        return "sheet_name_mismatch"
    if "어댑터가 데이터 행을 만들지 못했습니다" in combined_message:
        return "zero_parsed_rows"
    if "S2 매핑으로 보낼 입력 행이 없습니다" in combined_message:
        return "zero_default_feed_rows"

    adapter_summary = result.get("adapter_summary") or {}
    if isinstance(adapter_summary, Mapping):
        if int(adapter_summary.get("parsed_rows") or 0) <= 0:
            return "zero_parsed_rows"
        if int(adapter_summary.get("default_feed_rows") or 0) <= 0:
            return "zero_default_feed_rows"

    s2_filter = result.get("s2_channel_filter")
    if s2_filter is not None:
        active = bool(getattr(s2_filter, "active", False))
        after_rows = int(getattr(s2_filter, "after_rows", 0) or 0)
        if active and after_rows <= 0:
            return "s2_filter_empty"

    if workbook_diagnostics.get("error"):
        return "unexpected_exception"
    return "unexpected_exception"


def build_adapter_failure_payload(
    *,
    result: Mapping[str, Any],
    source_file: object | None = None,
    source_bytes: bytes | None = None,
    selected_s2_channel: str = "",
    app_commit_sha: str = "",
    app_version: str = "",
    app_url: str = "",
    created_at: datetime | None = None,
) -> dict[str, object]:
    source_name = text(result.get("source_name")) or text(getattr(source_file, "name", "")) or "uploaded.xlsx"
    payload_bytes = source_bytes if source_bytes is not None else read_uploaded_file_bytes(source_file)
    source_hash = sha256_hex(payload_bytes) if payload_bytes else ""
    created_at = created_at or datetime.now(KST)
    created_label = created_at.astimezone(KST).isoformat(timespec="seconds")
    platform = text(result.get("platform"))
    workbook_diagnostics = collect_workbook_diagnostics(payload_bytes, platform)
    failure_category = classify_failure(result, workbook_diagnostics)
    event_seed = "|".join(
        [
            source_hash,
            source_name,
            platform,
            text(result.get("s2_sales_channel")),
            failure_category,
            app_commit_sha,
        ]
    )
    event_id = hashlib.sha256(event_seed.encode("utf-8")).hexdigest()[:16]
    adapter_result = result.get("adapter_result")
    s2_filter = result.get("s2_channel_filter")

    return {
        "schema_version": SCHEMA_VERSION,
        "event_id": event_id,
        "created_at_kst": created_label,
        "app_commit_sha": app_commit_sha,
        "app_version": app_version,
        "streamlit_app_url": app_url,
        "source_name": source_name,
        "source_size": len(payload_bytes),
        "source_sha256": source_hash,
        "selected_s2_channel": selected_s2_channel,
        "detected_s2_channel": text(result.get("s2_sales_channel")),
        "effective_platform": platform,
        "status": text(result.get("status")),
        "failure_category": failure_category,
        "failure_reason": text(result.get("error")),
        "blocking_messages": [_json_safe(item) for item in result.get("blocking_messages", []) or []],
        "warning_messages": [_json_safe(item) for item in result.get("warning_messages", []) or []],
        "info_messages": [_json_safe(item) for item in result.get("info_messages", []) or []],
        "sheet_names": workbook_diagnostics.get("sheet_names", []),
        "sheet_diagnostics": workbook_diagnostics.get("sheets", []),
        "workbook_error": workbook_diagnostics.get("error", ""),
        "sheet_audits": _frame_records(result.get("audit_df")),
        "adapter_summary": _json_safe(result.get("adapter_summary") or {}),
        "adapter_spec": _safe_attr_payload(
            getattr(adapter_result, "spec", None),
            ("platform", "source_sheet_rule", "parser_contract", "s2_gate", "final_class"),
        ),
        "s2_channel_filter": _safe_attr_payload(
            s2_filter,
            ("active", "sales_channel", "before_rows", "after_rows", "reason"),
        ),
        "stage_seconds": _json_safe(result.get("stage_seconds") or {}),
        "clickup_task_id": "",
        "clickup_task_url": "",
        "github_issue_url": "",
    }


def payload_json_bytes(payload: Mapping[str, object]) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True).encode("utf-8")


def _md(value: object) -> str:
    return text(value).replace("|", "\\|").replace("\n", " ")


def render_failure_report_markdown(payload: Mapping[str, object]) -> str:
    title = f"[긴급][정산서 어댑터 실패] {_md(payload.get('detected_s2_channel')) or '판매채널 확인 필요'} / {_md(payload.get('source_name'))}"
    lines = [
        f"# {title}",
        "",
        "## Summary",
        "",
        f"- failure_category: `{_md(payload.get('failure_category'))}`",
        f"- status: `{_md(payload.get('status'))}`",
        f"- failure_reason: {_md(payload.get('failure_reason')) or '확인 필요'}",
        f"- platform: {_md(payload.get('effective_platform')) or '확인 필요'}",
        f"- s2_sales_channel: {_md(payload.get('detected_s2_channel')) or '확인 필요'}",
        f"- source_sha256: `{_md(payload.get('source_sha256'))}`",
        f"- source_size: {_md(payload.get('source_size'))}",
        f"- app_commit_sha: `{_md(payload.get('app_commit_sha')) or 'unknown'}`",
        "",
        "## Sheet Audits",
        "",
        "| sheet | status | header_row | data_start | parsed_rows | title_present_rows | note |",
        "|---|---|---:|---:|---:|---:|---|",
    ]
    for row in payload.get("sheet_audits", []) or []:
        if not isinstance(row, Mapping):
            continue
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.get("sheet")),
                    _md(row.get("status")),
                    _md(row.get("header_row")),
                    _md(row.get("data_start")),
                    _md(row.get("parsed_rows")),
                    _md(row.get("title_present_rows")),
                    _md(row.get("note")),
                ]
            )
            + " |"
        )

    lines.extend(["", "## Header Candidates", "", "| sheet | best_row | score | cells |", "|---|---:|---:|---|"])
    for sheet in payload.get("sheet_diagnostics", []) or []:
        if not isinstance(sheet, Mapping):
            continue
        best = sheet.get("best_header_candidate") if isinstance(sheet.get("best_header_candidate"), Mapping) else {}
        cells = " / ".join(text(item) for item in (best or {}).get("cells", [])[:12]) if isinstance(best, Mapping) else ""
        lines.append(
            f"| {_md(sheet.get('sheet'))} | {_md((best or {}).get('row'))} | {_md((best or {}).get('score'))} | {_md(cells)} |"
        )

    lines.extend(["", "## Top 10 Row Snapshot", ""])
    for sheet in payload.get("sheet_diagnostics", []) or []:
        if not isinstance(sheet, Mapping):
            continue
        lines.append(f"### {_md(sheet.get('sheet'))}")
        for index, row in enumerate(sheet.get("top_rows", []) or [], start=1):
            if isinstance(row, list):
                compact = " / ".join(text(cell) for cell in row if text(cell))
                lines.append(f"- {index}: {_md(compact)}")
        lines.append("")

    return "\n".join(lines).strip() + "\n"


def _sanitize_cell_for_github(value: object) -> str:
    raw = text(value)
    if not raw:
        return ""
    digit_count = sum(1 for char in raw if char.isdigit())
    if isinstance(value, (int, float)) or (digit_count >= 4 and digit_count >= len(raw) // 2):
        return "<number>"
    return _cell_text(raw)


def sanitize_payload_for_github(payload: Mapping[str, object]) -> dict[str, object]:
    sanitized_sheets: list[dict[str, object]] = []
    for sheet in payload.get("sheet_diagnostics", []) or []:
        if not isinstance(sheet, Mapping):
            continue
        best = sheet.get("best_header_candidate") if isinstance(sheet.get("best_header_candidate"), Mapping) else {}
        cells = []
        if isinstance(best, Mapping):
            cells = [_sanitize_cell_for_github(cell) for cell in best.get("cells", [])[:12]]
        sanitized_sheets.append(
            {
                "sheet": sheet.get("sheet", ""),
                "in_scope": sheet.get("in_scope", ""),
                "best_header_candidate": {
                    "row": (best or {}).get("row", "") if isinstance(best, Mapping) else "",
                    "score": (best or {}).get("score", "") if isinstance(best, Mapping) else "",
                    "cells": [cell for cell in cells if cell],
                },
            }
        )

    return {
        "schema_version": payload.get("schema_version", ""),
        "event_id": payload.get("event_id", ""),
        "created_at_kst": payload.get("created_at_kst", ""),
        "app_commit_sha": payload.get("app_commit_sha", ""),
        "source_name": payload.get("source_name", ""),
        "source_size": payload.get("source_size", ""),
        "source_sha256": payload.get("source_sha256", ""),
        "selected_s2_channel": payload.get("selected_s2_channel", ""),
        "detected_s2_channel": payload.get("detected_s2_channel", ""),
        "effective_platform": payload.get("effective_platform", ""),
        "status": payload.get("status", ""),
        "failure_category": payload.get("failure_category", ""),
        "failure_reason": payload.get("failure_reason", ""),
        "blocking_messages": payload.get("blocking_messages", []),
        "warning_messages": payload.get("warning_messages", []),
        "sheet_names": payload.get("sheet_names", []),
        "sheet_audits": payload.get("sheet_audits", []),
        "header_candidates": sanitized_sheets,
        "adapter_summary": payload.get("adapter_summary", {}),
        "clickup_task_url": payload.get("clickup_task_url", ""),
    }


def render_github_issue_body(payload: Mapping[str, object], *, clickup_url: str = "") -> str:
    safe = sanitize_payload_for_github({**dict(payload), "clickup_task_url": clickup_url or payload.get("clickup_task_url", "")})
    lines = [
        "## Adapter Failure",
        "",
        f"- failure_category: `{_md(safe.get('failure_category'))}`",
        f"- status: `{_md(safe.get('status'))}`",
        f"- source_name: `{_md(safe.get('source_name'))}`",
        f"- source_sha256: `{_md(safe.get('source_sha256'))}`",
        f"- platform: {_md(safe.get('effective_platform'))}",
        f"- s2_sales_channel: {_md(safe.get('detected_s2_channel'))}",
        f"- app_commit_sha: `{_md(safe.get('app_commit_sha')) or 'unknown'}`",
    ]
    if safe.get("clickup_task_url"):
        lines.append(f"- clickup: {safe.get('clickup_task_url')}")
    lines.extend(["", "## Failure Reason", "", _md(safe.get("failure_reason")) or "확인 필요", "", "## Sheet Audits", ""])
    lines.extend(["| sheet | status | header_row | parsed_rows | note |", "|---|---|---:|---:|---|"])
    for row in safe.get("sheet_audits", []) or []:
        if not isinstance(row, Mapping):
            continue
        lines.append(
            f"| {_md(row.get('sheet'))} | {_md(row.get('status'))} | {_md(row.get('header_row'))} | {_md(row.get('parsed_rows'))} | {_md(row.get('note'))} |"
        )
    lines.extend(["", "## Header Candidates", "", "| sheet | best_row | score | cells |", "|---|---:|---:|---|"])
    for sheet in safe.get("header_candidates", []) or []:
        if not isinstance(sheet, Mapping):
            continue
        best = sheet.get("best_header_candidate") if isinstance(sheet.get("best_header_candidate"), Mapping) else {}
        cells = " / ".join(text(item) for item in (best or {}).get("cells", []) if text(item)) if isinstance(best, Mapping) else ""
        lines.append(
            f"| {_md(sheet.get('sheet'))} | {_md((best or {}).get('row'))} | {_md((best or {}).get('score'))} | {_md(cells)} |"
        )
    lines.extend(
        [
            "",
            "## Notes",
            "",
            "- 원본 정산서 xlsx는 GitHub에 첨부하지 않는다.",
            "- 원본과 상세 failure payload는 ClickUp task attachment에서 확인한다.",
        ]
    )
    return "\n".join(lines).strip() + "\n"


def default_failure_artifact_stem(payload: Mapping[str, object]) -> str:
    source_stem = Path(text(payload.get("source_name")) or "adapter_failure").stem
    event_id = text(payload.get("event_id")) or "unknown"
    safe = "".join(ch if ch not in r'\/:*?"<>|' else "_" for ch in source_stem).strip() or "adapter_failure"
    return f"{safe}_{event_id}"
